# -*-coding: utf-8-*-


########################################
## 아래 import datetime 유지
########################################
import datetime
import json
import logging.config


################################################
## slack
################################################
import os
import re
########################################
import math
import time
import urllib
from collections import defaultdict
from decimal import *
from itertools import chain
from operator import itemgetter
########################################
from pprint import pprint

from django import template
from django.contrib import messages
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import connections
from django.db.models import Avg, Count, Max, Q
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
# from django.contrib.staticfiles.templatetags.staticfiles import static
from django.templatetags.static import static
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView, DeleteView, DetailView, ListView, UpdateView)
from django.views.generic.edit import FormMixin

# import cx_Oracle
# 구글 시트 연동
import gspread
import matplotlib.pyplot as plt
import numpy as np
import xlwt
from dal import autocomplete
from database.models import *
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from PIL import Image
from slack import WebClient
from slack.errors import SlackApiError
from wordcloud import STOPWORDS, ImageColorGenerator, WordCloud

# from ..forms import *
from .forms import *

logger = logging.getLogger(__name__)

import xml.etree.ElementTree as elemTree

import requests

import inspect


########################################
## excel download
########################################
# import csv
# from django.http import HttpResponse
# from django.contrib.auth.models import User


########################################
## word cloud
########################################
# import matplotlib.pyplot as plt
# from wordcloud import WordCloud
# # import wordcloud
# import numpy as np
# from PIL import Image

def exec_log(message, *value) :

    curframe = inspect.currentframe()
    calframe = inspect.getouterframes(curframe, 2)
    print('# Caller : ', calframe[1][3])

    if value :
        print("{message} : {value}".format(message=message,
                                           value= ','.join(map(str, value))))        
    else :
        print(message)

    print('')

def f_get_datatype(id_stdattr, db_type):
    data = StdAttr.objects.get(id=id_stdattr)

    in_data_type = ""

    if data:

        if db_type == "ORACLE":
            in_data_type = data.id_stddomain.oracle_datatype


        elif db_type in ("MYSQL", "MARIADB"):
            in_data_type = data.id_stddomain.mysql_datatype

    return in_data_type


def f_div_attr(attr):
    list_attr = []

    if ',' in attr:
        attr = attr.replace(' ', '')
        list_attr = attr.split(',')
    elif ' ' in attr:
        list_attr = attr.split()
    else:
        attr = attr.replace(' ', '')
        list_attr = attr.split(',')

    l_lkor = []
    l_leng = []
    l_lcnt = []
    l_domain = []
    l_eng_attr_name = []
    l_domain_flag = []
    l_oracle_domain_datatype = []
    l_mysql_domain_datatype = []
    l_word_find_flag = []
    l_last_num_flag = []
    l_attr_find = []
    l_aleary_req = []

    total_attr = len(list_attr)

    for attr in list_attr:


        clean_attr = (attr.replace('[', '')).replace(']', '')
        attr_exists = StdAttr.objects.filter(std_attr_kor=clean_attr, accept_yn__in=[0, ], use_yn=1)

        find_attr_flag = False

        if len(attr_exists) > 0:
            if attr_exists[0].accept_yn == 1:
                l_attr_find.append(False)
                l_aleary_req.append(True)
            elif attr_exists[0].accept_yn == 0:
                l_attr_find.append(True)
                l_aleary_req.append(False)

                find_attr_flag = True

        else:
            l_attr_find.append(False)
            l_aleary_req.append(False)

        lcnt = []
        leng = []
        lkor = []
        cnt = 1
        lAtEnNm = ""
        lAtKrNm = ""
        word_find_flag = True
        last_num_flag = True


        last_number = ""

        # 공백제거
        linput = attr.replace(' ', '')

        try:
            last_number = re.search("([a-zA-Z_가-힣])(\d+$)", linput).group(2)
            if (len(last_number) > 2):
                last_num_flag = False

            if (int(last_number) > 20):
                last_num_flag = False
        except:
            pass

        linput = linput.replace(last_number, '')

        menual_div = re.findall(r"\[([A-Za-z0-9_가-힣]+)\]", linput)

        if find_attr_flag :
            lkor.append(attr_exists[0].std_attr_kor)
            leng.append(attr_exists[0].std_attr_eng)
            lcnt.append(cnt)
            cnt = cnt + 1

        elif menual_div:

            for data in menual_div:

                lWdKrNm = data
                lFindWD = False

                flag = StdWord.objects.filter(std_wd_kor=data).exists()

                if flag:
                    std_wd_eng = (StdWord.objects.filter(std_wd_kor=data, accept_yn='0').values('std_wd_eng').first())[
                        'std_wd_eng']

                    lAtEnNm = std_wd_eng + "_" + lAtEnNm
                    lAtKrNm = lWdKrNm + "+" + lAtKrNm
                    lkor.append(lWdKrNm)
                    leng.append(std_wd_eng)
                    lcnt.append(cnt)
                    cnt = cnt + 1

                    lFindWD = True

                if lFindWD == False:

                    word_find_flag = False

                    if len(lkor) > 0:
                        if lkor[len(lkor) - 1:][0] != "<font color='red'>*</font>":
                            lkor.append("<font color='red'>*</font>")
                            lcnt.append(cnt)
                            cnt = cnt + 1

                        if leng[len(leng) - 1:][0] != "<font color='red'>*</font>":
                            leng.append("<font color='red'>*</font>")
                    else:
                        lkor.append("<font color='red'>*</font>")
                        leng.append("<font color='red'>*</font>")
                        lcnt.append(cnt)
                        cnt = cnt + 1

        else:

            lTotalLength = len(linput)

            while lTotalLength > 0:

                i = 1
                lFindWD = False

                while i <= lTotalLength:
                    ch = linput[i - 1:]
                    lWdKrNm = ch

                    flag = StdWord.objects.filter(std_wd_kor=lWdKrNm, accept_yn='0').exists()

                    if flag:
                        std_wd_eng = \
                        (StdWord.objects.filter(std_wd_kor=lWdKrNm, accept_yn='0').values('std_wd_eng').first())[
                            'std_wd_eng']

                        lAtEnNm = std_wd_eng + "_" + lAtEnNm
                        lAtKrNm = lWdKrNm + "+" + lAtKrNm
                        lkor.append(lWdKrNm)
                        leng.append(std_wd_eng)
                        lcnt.append(cnt)
                        cnt = cnt + 1

                        # 종료 조건
                        lTotalLength = i - 1

                        lFindWD = True

                    i = i + 1

                lTotalLength = i - 1 - 1
                linput = linput[0:lTotalLength]

                if lFindWD == False:

                    word_find_flag = False

                    if len(lkor) > 0:
                        if lkor[len(lkor) - 1:][0] != "<font color='red'>*</font>":
                            lkor.append("<font color='red'>*</font>")
                            lcnt.append(cnt)
                            cnt = cnt + 1

                        if leng[len(leng) - 1:][0] != "<font color='red'>*</font>":
                            leng.append("<font color='red'>*</font>")
                    else:
                        lkor.append("<font color='red'>*</font>")
                        leng.append("<font color='red'>*</font>")
                        lcnt.append(cnt)
                        cnt = cnt + 1

            lkor.reverse()
            leng.reverse()

        eng_attr_name = '_'.join(leng)
        eng_attr_name = eng_attr_name + last_number

        d_ind = 0
        d_total_length = len(lkor)
        d_str = ""
        domain = ""
        domain_flag = False

        while d_ind < d_total_length:

            l_d_str = []
            d_str = ''.join(lkor[d_ind:])            

            if StdDomain.objects.filter(domain_name=d_str, accept_yn=0, use_yn=1).exists():

                domain_flag = True

                l_d_str.append(d_str)


            else:
                try :
                    in_last_number = re.search("([a-zA-Z_가-힣])(\d+$)", attr).group(2)
                    clear_number_of_attr = attr.replace(in_last_number,'')
                except :
                    clear_number_of_attr = attr
                    pass

                sql = """SELECT DOMAIN
                 FROM CUST_ASIS_ATTR_DOMAIN_MAPPING
                 WHERE ATTR = '{attr}'
              """.format(attr=clear_number_of_attr)

                result_ds = _query_list("default", sql)

                for r in result_ds :
                    # d_str = result_ds[0][0]
                    l_d_str.append(r[0])
                    domain_flag = True

            if domain_flag:
                domain = StdDomain.objects.filter(domain_name__in=l_d_str, accept_yn=0, use_yn=1)

                break
            d_ind = d_ind + 1

        # oracle_domain_datatype = []

        # for d in domain :

        #   str_oracle_leng = ""

        #   if d.oracle_decimal_leng != None :
        #     str_oracle_leng = "(" + str(d.oracle_leng) + ',' + str(d.oracle_decimal_leng) + ")"
        #   elif d.oracle_leng != None :
        #     str_oracle_leng = "(" + str(d.oracle_leng) + ")"

        #   if d.oracle_leng != None :

        #     # oracle_domain_datatype = d.oracle_data_type + str_oracle_leng + "<br>" + oracle_domain_datatype
        #     oracle_domain_datatype.append(d.oracle_data_type + str_oracle_leng)

        #   elif d.oracle_data_type != None :

        #     oracle_domain_datatype.append(d.oracle_data_type)

        # mysql_domain_datatype = []
        l_domain_local = []

        for d in domain:
            # str_mysql_leng = ""

            # if d.mysql_decimal_leng != None :
            #   str_mysql_leng = "(" + str(d.mysql_leng) + ',' + str(d.mysql_decimal_leng) + ")"
            # elif d.mysql_leng != None :
            #   str_mysql_leng = "(" + str(d.mysql_leng) + ")"

            # if d.mysql_leng != None :
            #   mysql_domain_datatype.append(d.mysql_data_type + str_mysql_leng)
            # elif d.mysql_data_type != None :
            #   mysql_domain_datatype.append(d.mysql_data_type)

            sql = """select count(*) CNT
               from cust_std_attr
               where id_stddomain = {id};
            """.format(id=d.id)

            cnt_ds = _query_dict("default", sql)

            l_domain_local.append([d.id, d.info_type, cnt_ds[0]['CNT'], d.domain_name])

        l_lkor.append(lkor)
        l_leng.append(leng)
        l_lcnt.append(lcnt)

        l_eng_attr_name.append(eng_attr_name)
        l_domain_flag.append(domain_flag)

        l_domain_local = sorted(l_domain_local, key=itemgetter(2), reverse=True)


        #  도메인 중 한개만 노출시키기 위한 코드. 일단 주석처리
        # try :
        #     in_last_number = re.search("([a-zA-Z_가-힣])(\d+$)", attr).group(2)
        #     clear_number_of_attr = attr.replace(in_last_number,'')
        # except :
        #     clear_number_of_attr = attr
        #     pass

        # if clear_number_of_attr[-2:] in ('코드', '번호'):
        #     if len(l_domain_local) > 0:
        #         l_domain_local = [l_domain_local[0]]

        
        l_domain.append(l_domain_local)
        # l_oracle_domain_datatype.append(oracle_domain_datatype)
        # l_mysql_domain_datatype.append(mysql_domain_datatype)
        l_word_find_flag.append(word_find_flag)
        l_last_num_flag.append(last_num_flag)

    max_length = 0

    lcnt_ind = 0
    i = 0
    for lcnt in l_lcnt:
        if max_length <= len(lcnt):
            max_length = len(lcnt)
            lcnt_ind = i
        i = i + 1

    lcnt = l_lcnt[lcnt_ind]

    # [, ] 제거

    buffer_attr = []
    for attr in list_attr:
        attr = attr.replace('[', '')
        attr = attr.replace(']', '')
        buffer_attr.append(attr)
    list_attr = buffer_attr




    # return  l_lkor, \
    #         l_leng, \
    #         l_domain, \
    #         l_eng_attr_name, \
    #         l_domain_flag, \
    #         l_oracle_domain_datatype, \
    #         l_mysql_domain_datatype, \
    #         l_word_find_flag, \
    #         l_last_num_flag, \
    #         lcnt, \
    #         total_attr, \
    #         max_length, \
    #         l_attr_find, \
    #         l_aleary_req, \
    #         list_attr

    return l_lkor, \
           l_leng, \
           l_domain, \
           l_eng_attr_name, \
           l_domain_flag, \
           l_word_find_flag, \
           l_last_num_flag, \
           lcnt, \
           total_attr, \
           max_length, \
           l_attr_find, \
           l_aleary_req, \
           list_attr


sql_execute_saved = """
  select writer, id, data_title
  from
  (
  # select '# 나의데이터' writer, a.id, concat(if(a.exp_yn = '1', '★ ', '☆ '), a.data_title) data_title, mod_dtm, b.id user_id
  select '# 나의데이터' writer, a.id, concat(if(a.exp_yn = '1', '☆ ', '☆ '), a.data_title) data_title, mod_dtm, b.id user_id
  from cust_data_list a,
     auth_user b
  where a.id_reg_user = b.id
    and a.id_reg_user = {id}
  # union all
  # select concat('# ',b.first_name,'(',b.last_name,')') writer, a.id, concat(if(a.exp_yn = '1', '★ ', '☆ '), a.data_title) data_title, mod_dtm, b.id user_id
  # from cust_data_list a,
  #    auth_user b
  # where a.id_reg_user = b.id
  #   and a.exp_yn = '1'
  #   and a.id_reg_user != {id}
  )  a
  order by
    ( case when user_id = {id} then 0
           else 1 end ),
     a.mod_dtm desc;
"""


def _orderedSet(list):
    my_set = set()
    res = []
    for e in list:
        if e not in my_set:
            res.append(e)
            my_set.add(e)

    return res


def _loggingVisit(request, menu_cd):
    new_uservisit = UserVisit(
        name=request.user.first_name,
        sabun=request.user.username,
        team_name=request.user.last_name,
        id_menulist=MenuList.objects.get(id=menu_cd),
        reg_dtm=datetime.datetime.now(),
    )

    new_uservisit.save()


def _send_slack(msg):
    client = WebClient(token='xoxp-127066622802-291541382084-705305252917-63304c080942bf49ae37f184eeddb286')

    try:
        response = client.chat_postMessage(
            channel='#gsshop-db',
            text=msg,
            username="SmartDBA",
            icon_emoji=":smartdba2_0:")
    except SlackApiError as e:
        # You will get a SlackApiError if "ok" is False
        print("Got an error: " + e.response['error'])


def _query_sms(tel, msg, slack_yn='Y', sms_yn='Y', team_name=None, email=None):
    try :
        ###############################################################################################

        start = time.time()

        # headers = {
        #     'accept': 'application/json',
        #     'Authorization': 'Bearer wkftodrlsrpchlrhdi!',
        #     'Content-Type': 'application/json',
        # }

        # data = '{"id":"100019968604040","message":"' + "[other] " + msg + '","botType":"SMARTDBA_BOT"}'

        # data = data.encode("utf-8")

        # print("## 1")
        # print("time :", time.time() - start)
        
        # response = requests.post('https://tong.gsshop.com/api/msg/send', headers=headers, data=data)

        # ###############################################################################################

        # print("## 2")
        # print("time :", time.time() - start)

        work_chat_response = ""

        
        headers = {
            'accept': 'application/json',
            'Authorization': 'Bearer wkftodrlsrpchlrhdi!',
        }

        params = (
            ('email', email),
        )

        response = requests.get('https://tong.gsshop.com/api/users/findByEmail', headers=headers, params=params)

        

        if slack_yn == 'Y':
            _send_slack(msg)

        chk_status_code = response.status_code
        exec_log(1, chk_status_code)

        if chk_status_code == 200 :

            workplace_id = response.json()['id']                        

            headers = {
                'accept': 'application/json',
                'Authorization': 'Bearer wkftodrlsrpchlrhdi!',
                'Content-Type': 'application/json',
            }

            exec_log(workplace_id)
            exec_log(msg)
            data = '{"id":"' + workplace_id + '","message":"' + msg + '","botType":"SMARTDBA_BOT"}'
            
            data = data.encode("utf-8")

            response = requests.post('https://tong.gsshop.com/api/msg/send', headers=headers, data=data)

            work_chat_response = response.status_code

            exec_log(work_chat_response)

        exec_log(2, chk_status_code)
        exec_log(3, work_chat_response)
        if chk_status_code != 200 or work_chat_response != 201 :
        # if True :
            if team_name == "Offshore개발팀":
                # Off
                tel = "010-9943-4003"            

            if sms_yn == 'Y':
                exec_log(4, 'sms 발송')
                with connections["sms"].cursor() as cur:
                    sql = """insert into LGORA_OWN.SC_TRAN  (TR_NUM,
                                    TR_SENDDATE,
                                    TR_SENDSTAT,
                                    TR_MSGTYPE,
                                    TR_PHONE,
                                    TR_CALLBACK,
                                    TR_MSG,
                                    TR_CODE)
                                values (
                                        LGORA_OWN.SC_SEQUENCE.NEXTVAL,
                                    sysdate,
                                    '0',
                                    '0',
                                    '{tel}',
                                    '0220070910',
                                    '{msg}',
                                    '008')""".format(tel=tel,
                                                     msg=msg)
                    exec_log(5, sql)
                    cur.execute(sql)
                    connections["sms"].commit()

    except Exception as e :
        exec_log('에러 : ', str(e))

def _query_dict(db_nm, sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    try:
        with connections[db_nm].cursor() as cur:
            cur.execute(sql)

            columnNames = [d[0] for d in cur.description]
            dataset = [dict(zip(columnNames, row)) for row in cur]

    except:
        dataset = [{'EMP_NM': '?', 'NAME': '?', 'MOBILE': '?', 'SUB_MAIL': '?'}]
        # dataset = []

    return dataset


def _query_list(db_nm, sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    with connections[db_nm].cursor() as cur:
        cur.execute(sql)
        dataset = cur.fetchall()

    return dataset


def _query_commit(db_nm, sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    with connections[db_nm].cursor() as cur:
        cur.execute(sql)
        connections[db_nm].commit()

        return 1


def _query_column_name(db_nm, sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    with connections[db_nm].cursor() as cur:
        cur.execute(sql)

        columnNames = [d[0] for d in cur.description]

    return columnNames


def _query_mysql_dict(sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    with connections['default'].cursor() as cur:
        cur.execute(sql)

        columnNames = [d[0] for d in cur.description]
        dataset = [dict(zip(columnNames, row)) for row in cur]

    return dataset


def _query_mysql_list(sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    with connections['default'].cursor() as cur:
        cur.execute(sql)
        dataset = cur.fetchall()

    return dataset


def _query_mssql_list(sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    with connections['homenet'].cursor() as cur:
        cur.execute(sql)
        dataset = cur.fetchall()

    return dataset


def _query_mssql_dict(sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    try:
        with connections['homenet'].cursor() as cur:
            cur.execute(sql)

            columnNames = [d[0] for d in cur.description]
            dataset = [dict(zip(columnNames, row)) for row in cur]
    except:
        dataset = ""

    return dataset


def _commit_mysql(sql):
    """
    Queries mysql and returns a cursor to the results.
    """
    with connections['default'].cursor() as cur:
        cur.execute(sql)
        connections['default'].commit()

        return cur.lastrowid


def _dashboard_data(request):
    sql = """
    select count(distinct sabun) cnt, count(*) total_cnt
    from cust_user_visit a
    where reg_dtm >= DATE(SYSDATE())
      AND reg_dtm < DATE_ADD(DATE(SYSDATE()), INTERVAL 1 DAY)
  """

    today_visit = _query_mysql_dict(sql)

    sql = """
    select id
    from cust_menu_list
  """

    user_permission_all_list = _query_list("default", sql)
    user_permission_all = []
    for item in user_permission_all_list:
        user_permission_all.append(item[0])

    sql = """
    select id_menulist
    from cust_menu_permission
    where id_grantee_user = {user_id}
  """.format(user_id=request.user.id)

    user_permission_list = _query_list("default", sql)
    user_permission = []
    for item in user_permission_list:
        user_permission.append(item[0])

    user_permission_minus = list(set(user_permission_all) - set(user_permission))

    # curr_path = "http://"+request.META['HTTP_HOST']
    # request.get_full_path()

    dashboard_context = {
        'today_visit': today_visit,
        'user_permission': user_permission,
        'user_permission_minus': user_permission_minus,

    }

    return dashboard_context


# emp_no, emp_nm, sub_mail, mobile, name
iamSql = """select A.emp_no, A.emp_nm, A.sub_mail, A.mobile, c.name
      from iam.EX_GW_EMP_TOTAL A,
           iam.ORG_USER B,
           IAM.ORG_GROUP C
      where A.EMP_NO = B.CODE
        AND B.GROUP_ID = C.GROUP_ID
        AND EMP_STATUS IN ('H','C')
        AND SUB_MAIL IS NOT NULL
        AND A.EMP_NO = '{emp_no}'
    """

dataSearchSql = """
SELECT *
FROM
(
  SELECT  a.id,
      c.db_use,
      d.domain_name,
      a.col_comments,
      ifnull(if(a.column_explain='',a.col_comments,a.column_explain),a.col_comments) column_explain,
      ifnull(a.column_explain,' ') column_explain_keyword,
      (
        select      concat('SELECT', char(10),
              concat('   ',IFNULL(replace(GROUP_CONCAT(concat(rpad(in_ccl2.column_name,40,' '),' /*', in_ccl2.col_comments, ' - PK */', char(10)) SEPARATOR " ,"),',',','),concat('/* PK 없음 */',char(10)))),
              concat(IF(in_ccl2.pk_yn is null, '   ','  ,'),rpad(in_ccl1.column_name,40,' '),' "',substr(in_ccl1.col_comments,1,10),'"'), char(10),
              'FROM ', rpad(concat(in_ctl.owner, ".", in_ctl.table_name),40,' '), ' /*', in_ctl.comments, '*/'
                        )
        from cust_table_list in_ctl
            left outer join cust_column_list in_ccl2
            on (
              in_ctl.id = in_ccl2.id_tablelist
            and in_ccl2.pk_yn = 1
            ),
           cust_column_list in_ccl1
        where in_ctl.id = in_ccl1.id_tablelist
          and in_ccl1.id = a.id
      ) sql_text,
      concat(b.owner,'.',b.table_name) table_name,
      replace(b.comments,'스테이징_','') comments,
      a.column_name,
      a.column_id,
      au1.username it_manager_sabun,
      au1.first_name it_manager_name,
      au1.last_name it_manager_team,
      au2.username dev_manager_sabun,
      au2.first_name dev_manager_name,
      column_type,
      CASE WHEN column_type = 0 THEN '일반 속성'
           WHEN column_type = 1 THEN '공통코드 속성'
           WHEN column_type = 2 THEN '개벌코드 속성'
      END column_type_value,
      CASE WHEN column_type = 0 THEN ''
           WHEN column_type = 1 THEN code_key
           WHEN column_type = 2 THEN bb.table_name
      END code_key,
      {data_type} data_type,
      0 avg_t,
      'Y' prov_yn,
      b.id_dblist id_dblist,
      a.secu_yn privacy_yn,
      '수행이력없음' t,
      concat('데이터 확인 횟수 : ', '0회') execute_cnt
  FROM cust_column_list a
       LEFT OUTER JOIN cust_table_list bb
        ON (
                a.id_tablelist_code_table = bb.id
            and bb.oper_cd = 3
        ),
       cust_table_list b
       LEFT OUTER JOIN auth_user au1
         ON (
        au1.id = b.id_user_it_manager
            )
       LEFT OUTER JOIN auth_user au2
         ON (
        au2.id = b.id_user_dev_manager
            ),
       cust_db_list c,
       cust_domain d
  WHERE a.id_tablelist = b.id
    AND b.id_dblist = c.id
    AND d.id = b.id_domain
    AND a.id = {column_id}
    AND b.oper_cd = 3
    -- exp_order : 0 노출하지 않음
    AND B.EXP_ORDER > 0
    AND {data_type} = 0 -- 0 : COLUMN / 1 : DATA
  UNION ALL
  SELECT  a.id,
      c.db_use,
      d.domain_name,
      a.data_title,
      a.data_explain,
      a.data_explain column_explain_keyword,
      a.sql_text,
      '' table_name,
      d.domain_name comments,
      '',
      '',
      au1.username it_manager_sabun,
      au1.first_name it_manager_name,
      au1.last_name it_manager_team,
      '' dev_manager_sabun,
      '' dev_manager_name,
      '',
      '',
      '',
      {data_type} data_type,
      IF(AVG_T.t is null, '수행 이력 없음',concat(AVG_T.t,' 초')) t,
      case when {super_user} = 1 then 'Y'
              when id_dblist = 39 and a.privacy_yn = 1 and cduiru.user_id is null then 'C' -- 요청자가 아님
              when id_dblist = 39 then 'Y'
          when id_dblist <> 39 and ( prov_yn2 = 0 ) then 'B' -- 승인이 나지 않음
          when id_dblist <> 39 and a.privacy_yn = 1 and cduiru.user_id is null then 'C' -- 요청자가 아님
          when id_dblist <> 39 and a.privacy_yn = 1 and cduiru.user_id is not null and cdr.id_datalist is null then 'D' -- 조회 기간이 지남
          when id_dblist <> 39 and a.privacy_yn = 1 and cduiru.user_id is not null and cdr.id_datalist is not null then 'Y'
          # when id_dblist <> 39 and prov_yn1 = 1 and prov_yn2 = 1 and cduiru.user_id is not null then 'Y'
          else 'Y' END prov_yn,
      a.id_dblist id_dblist,
      a.privacy_yn privacy_yn,
      IF(CUES.T IS NULL, '수행 이력 없음',CONCAT(CUES.T,' 초')) t,
      concat('데이터 확인 횟수 : ',a.execute_cnt,'회') execute_cnt
  FROM cust_data_list a
     LEFT OUTER JOIN
       (
      SELECT ID_DATALIST,ROUND(AVG(EXECUTE_TIME),2) T
      FROM CUST_USER_EXECUTE_SQL
      WHERE ID_DATALIST IS NOT NULL
      GROUP BY ID_DATALIST
       ) CUES
       ON (
          A.ID = CUES.ID_DATALIST
      )
      left outer join cust_data_request cdr
      on (
        a.id = cdr.id_datalist
        and cdr.id_reg_user = {userid}
        and cdr.poss_view_dtm > now()
      )
     LEFT OUTER JOIN CUST_DATALIST_USER_ID_REQ_USERS CDUIRU
     ON (
      A.ID = CDUIRU.DATALIST_ID
      AND CDUIRU.USER_ID = {userid}
     )
     LEFT OUTER JOIN
        (
          SELECT ID_DATALIST,ROUND(AVG(EXECUTE_TIME),2) T
          FROM cust_user_execute_sql
          WHERE ID_DATALIST IS NOT NULL
          GROUP BY ID_DATALIST
        ) AVG_T
        ON (
      A.ID = AVG_T.ID_DATALIST
     )
       LEFT OUTER JOIN auth_user au1
         ON (
        au1.id = a.id_reg_user
            )
       LEFT OUTER JOIN CUST_DOMAIN D
        ON (
               D.ID = A.ID_DOMAIN
        ),
       cust_db_list c
  WHERE a.id_dblist = c.id
    AND a.id = {column_id}
    -- exp_order : 0 노출하지 않음
    AND A.EXP_YN = '1'
    AND {data_type} = 1 -- 0 : COLUMN / 1 : DATA
) A
ORDER BY domain_name, comments, column_id;"""


def makeDictFactory(cursor):
    columnNames = [d[0] for d in cursor.description]

    def createRow(*args):
        return dict(zip(columnNames, args))

    return createRow


# def dashboard(request) :
#   # Login 한 user 만 접속 허용
#   # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
#   if not request.user.is_authenticated :
#     return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

#   template_parent = "dashboard"

#   sql = """
#     SELECT cast(@rownum:=@rownum+1 as unsigned) ROWNUM,
#            A.*
#     FROM
#     (
#       SELECT   ID_DATA,
#            -- IF(DATA_TYPE=0,'컬럼','데이터') DATA_TYPE,
#            DATA_TYPE,
#            DATA DATA_DESC,
#            COUNT(*) CNT
#       FROM CUST_USER_CLICK_DATA
#       GROUP BY ID_DATA, DATA_TYPE, DATA
#       ORDER BY CNT DESC
#     ) A, (SELECT @rownum:=0) TMP
#     LIMIT 10
#   """

#   dataset_user_click_data = _query_mysql_dict(sql)

#   sql = """
#     SELECT CAST(@ROWNUM := @ROWNUM + 1 AS unsigned) ROWNUM , A.*
#     FROM
#     (
#       SELECT KEYWORD, count(*) CNT
#       FROM CUST_USER_SEARCH_KEYWORD
#       GROUP BY KEYWORD
#       ORDER BY CNT DESC
#     ) A, ( SELECT @ROWNUM := 0 ) TMP
#     LIMIT 10
#   """

#   dataset_user_search_keyword = _query_mysql_dict(sql)

#   sql = """
#     select replace(data_title,'/','!_!') data_title_m, data_title data_title_o, reg_dtm,
#            (select concat(first_name,'(',last_name,')') from auth_user where id =  id_reg_user) id_reg_user
#     from cust_data_list
#     order by reg_dtm desc
#     limit 10;
#   """

#   dataset_datalist = _query_mysql_dict(sql)

#   sql = """
#     select new_data.cnt new_data,
#          request_data.cnt + request_word.cnt request,
#            request_tuning.cnt tuning
#     from
#     (
#       select count(*) cnt
#       from cust_data_list
#       where reg_dtm >= date_add(now(), interval - 7 day)
#     ) new_data,
#     (
#       select count(*) cnt
#       from cust_user_request_data
#       where complete_yn = 0
#         and id_reg_user_id = {id}
#     ) request_data,
#     (
#     select count(*) cnt
#     from cust_user_request_word
#     where complete_yn = 0
#       and id_reg_user_id = {id}
#     ) request_word,
#     (
#       select count(*) cnt
#       from cust_tuning_list
#       where id_tuningstatus = 1
#         and id_reg_user = {id}
#     ) request_tuning;
#   """.format(id=request.user.id)

#   dashboard_cnt = _query_mysql_dict(sql)
#   dashboard_cnt = dashboard_cnt[0]
#   return render(request, 'dashboard.html', {
#                           'template_parent':template_parent,
#                           'dataset_user_click_data':dataset_user_click_data,
#                           'dashboard_data':_dashboard_data(request),
#                           'dataset_datalist':dataset_datalist,
#                           'dataset_user_search_keyword':dataset_user_search_keyword,
#                           'dashboard_cnt':dashboard_cnt,
#                         })

# 팀변경
def main(request):
    try:
        team = request.user.last_name
    except:
        team = ""
    if request.user.is_superuser:
        return redirect('DBAWorkLV')
    # elif team in ['인프라클라우드팀', 'IT개발지원팀', 'MicroSVC팀', 'MC 개발지원팀', 'IT매트릭스팀', '보안센터', 'IT개발팀']:
    # elif team in ['클라우드팀', 'IT운영지원팀', '회원Product팀', '전시Product팀', 'Mobile Native-TFT', '보안2팀', '상품정산IT팀', '주문방송IT팀','물류IT팀', 'MSA-TFT']:
    #elif team in ['클라우드팀', 'IT운영지원팀', '회원Product팀', '전시Product팀', 'Mobile Native-TFT', '보안2팀', '상품정산IT팀', '주문방송IT팀','물류IT팀', 'MSA-TFT']:
        return redirect('QuickLinkLV')
    else:
        return redirect('QuickLinkLV')

def howDatabaseConn(request):
    # Login 한 user 만 접속 허용
    # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
    if not request.user.is_authenticated:
        return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

    template_parent = "dev"
    template_child = "howDatabaseConn"

    return render(request, 'how_database_conn.html', {
        'template_parent': template_parent,
        'template_child': template_child,
    })


def databaseMap(request):
    # Login 한 user 만 접속 허용
    # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
    if not request.user.is_authenticated:
        return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

    template_parent = "dev"
    template_child = "databaseMap"

    _loggingVisit(request, 25)

    with connections['default'].cursor() as cursor:
        sql = """
    SELECT a.id,
           a.db_use db_use,
           a.map_coords,
           a.db_desc,
           CASE WHEN a.privacy_yn = 1 THEN '개인정보 데이터 보유'
                WHEN a.privacy_yn = 0 THEN ''
           END privacy_yn,
           MAX(CASE WHEN b1.oper_cd = '3' AND a.id_dbtype = '1' THEN b1.tns_desc
        WHEN b1.oper_cd = '3' AND a.id_dbtype <> '1' THEN CONCAT( 'IP : ', b1.svc_ip, ' / ', 'PORT : ', b1.port)
                    ELSE '운영DB 없음' END) p_tns_desc,
           MAX(CASE WHEN b2.oper_cd='4' AND a.id_dbtype = '1' THEN b2.tns_desc
        WHEN b2.oper_cd='4' AND a.id_dbtype <> '1' THEN CONCAT( 'IP : ', b2.svc_ip, ' / ', 'PORT : ', b2.port)
                    ELSE '개발DB 없음' END) d_tns_desc,
           CASE WHEN MAX(a.id_dbtype) = '1' THEN 'ORACLE'
                WHEN MAX(a.id_dbtype) = '2' THEN 'MySQL'
                WHEN MAX(a.id_dbtype) = '3' THEN 'MariaDB'
                WHEN MAX(a.id_dbtype) = '4' THEN 'PostgreSQL'
                WHEN MAX(a.id_dbtype) = '5' THEN 'MSSQL'
                WHEN MAX(a.id_dbtype) = '6' THEN 'SybaseIQ'
           END db_typ
    FROM cust_db_list a
         LEFT OUTER JOIN cust_db_detail b1
      -- ('1','관리'),
      -- ('2','현행화'),
      -- ('3','운영'),
      -- ('4','개발')
         ON ( a.id = b1.id_dblist
              AND b1.oper_cd IN ('3')  )
         LEFT OUTER JOIN cust_db_detail b2
         ON ( a.id = b2.id_dblist
              AND b2.oper_cd IN ('4') )
    WHERE 1=1
      AND a.MAP_COORDS IS NOT NULL
    GROUP BY A.id, a.db_use, a.map_coords, a.db_desc
    ORDER BY a.ID ;
    """

        cursor.execute(sql)

        columnNames = [d[0] for d in cursor.description]
        dataset = [dict(zip(columnNames, row)) for row in cursor]

    return render(request, 'database_map.html', {
        'template_parent': template_parent,
        'template_child': template_child,
        'dataset': dataset,
    })


def page404(request):
    return render(request, 'page-404.html')


def word_cloud(sql):
    dataset_search_keyword = _query_mysql_dict(sql)

    keyword_dict = {}
    for keyword in dataset_search_keyword:
        keyword_dict[keyword['keyword']] = keyword['cnt']

    font_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/static/custom/font/NanumBarunGothic.ttf'

    wordcloud = WordCloud(font_path=font_path,
                          width=1600,
                          height=800,
                          background_color="rgb(236,237,240)",
                          )
    # wordcloud = wordcloud.generate_from_text(text)
    wordcloud = wordcloud.generate_from_frequencies(keyword_dict)

    plt.figure(figsize=(8, 8))
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis('off')

    plt.savefig(os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/static/custom/img/search_wordcloud.png',
                transparent=True, bbox_inches='tight')


def dataView(request):
    # Login 한 user 만 접속 허용
    # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
    if not request.user.is_authenticated:
        return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

    template_parent = "data"
    template_child = "dataView"

    # 방문 기록
    _loggingVisit(request, 2)

    sql = """
      select replace(if(b.id = {user_id}, '나의데이터', last_name),' ','') last_name
      from cust_data_list a,
         auth_user b,
         cust_datalist_user_id_req_users c
      where
          a.id = c.datalist_id
        and b.id = c.user_id
        and a.exp_yn = 1
        # and a.prov_yn1 = 1
        # and a.prov_yn2 = 1
      union
      select b.last_name
      from cust_data_list a,
         auth_user b,
         cust_datalist_user_id_mod_psbl_users c
      where
          a.id_reg_user = b.id
        and a.id = c.datalist_id
        and c.user_id = {user_id}
        and a.exp_yn = 1
      union
      select '나의데이터'
      from cust_data_list a,
         auth_user b
      where
         a.id_reg_user = b.id
       and b.id = {user_id}
      union
      select '전체'
      from dual
      where {super_user} = 1
      order by ( case when last_name = '나의데이터' then 0
                      when last_name = '전체' then -1
                    else last_name end );
      """.format(user_id=request.user.id,
                 super_user=request.user.is_superuser)

    ds_data_tree_dict = _query_dict('default', sql)

    return render(request, 'data_view.html', {
        'template_parent': template_parent,
        'ds_data_tree_dict': ds_data_tree_dict,
        'template_child': template_child,
    })


# def dataSearchKeyword(request, keyword='', domain_id=-1) :
#   # Login 한 user 만 접속 허용
#   # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
#   if not request.user.is_authenticated :
#     return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))



#   template_parent = "data"
#   template_child = "dataSearchKeyword"


#   if keyword :
#     keyword = keyword.replace("!_!","/")
#     keyword = "!" + keyword
#   else :
#     keyword = ''

#   if domain_id > -1 :
#     keyword = ''


#   select_db = ""
#   select_db_str = ""
#   select_column_type = ""
#   select_column_type_str = ""
#   search_dataset_len = 0
#   search_dataset = None
#   word_meaning = ""
#   dataset_user_search_keyword = ""
#   select_data_type = ""
#   select_domain_id = ""
#   order_by_str = ""
#   chk_tab_comments = ""
#   chk_table_name = ""
#   chk_column_name = ""
#   chk_db_name = ""
#   chk_data_type = ""


#   wordcloud = ""


#   if request.method == "POST" or domain_id > -1:


#     form = searchKeywordForm(request.POST)

#     if form.is_valid() or domain_id > -1:


#       if domain_id == -1 :
#         keyword = form.cleaned_data['keyword']
#         sql = """
#           SELECT TOP 1 *, 1 EXP_ORDER
#           FROM IKEP4_CV_WORD
#           WHERE WORD_NAME = '{keyword}'
#           UNION ALL
#           SELECT TOP 1 *, 1
#           FROM IKEP4_CV_WORD
#           WHERE WORD_NAME_READ = '{keyword}'
#           UNION ALL
#           SELECT TOP 1 *, 2
#           FROM IKEP4_CV_WORD
#           WHERE WORD_NAME_READ like '%{keyword}%'
#           UNION ALL
#           SELECT TOP 1 *, 3
#           FROM IKEP4_CV_WORD
#           WHERE WORD_ENGLISH_NAME like '%{keyword}%'
#           UNION ALL
#           SELECT TOP 1 *, 4
#           FROM IKEP4_CV_WORD
#           WHERE WORD_NAME like '%{keyword}%'
#           ORDER BY exp_order
#           ;
#         """.format(keyword=keyword)

#         word_meaning = _query_mssql_dict(sql)

#         if len(word_meaning) > 0 :
#           word_meaning = word_meaning[0]


#         new_usersearchkeyword = UserSearchKeyword(
#               keyword = keyword,
#               id_reg_user = request.user,
#               reg_dtm = datetime.datetime.now(),
#           )

#         new_usersearchkeyword.save()

#         # select_db = request.POST.getlist('select_db','')

#         # if "multiselect-all" in select_db :
#         #   select_db.remove("multiselect-all")
#         # select_db_str = (', '.join(select_db))
#         # if select_db_str :
#         #   select_db_str = "in ( " + select_db_str + " ) "
#         # else :
#         #   select_db_str = "in ( '' ) "
#         chk_tab_comments = request.POST.get('chk_tab_comments','')
#         chk_table_name = request.POST.get('chk_table_name','')
#         chk_column_name = request.POST.get('chk_column_name','')
#         chk_db_name = request.POST.get('chk_db_name','')
#         chk_data_type = request.POST.get('chk_data_type','')

#         # 속성 한글명, 속성 영문명, 테이블 영문명, 테이블 한글명 등
#         select_column_type = request.POST.get('select_column_type','')

#         if select_column_type == "1" :
#           select_column_type_str = "AND upper(replace(C.COL_COMMENTS,' ','')) LIKE '%{keyword}%'".format(keyword=keyword.replace(" ",""))
#         elif select_column_type == "2" :
#           select_column_type_str = "AND upper(replace(C.COLUMN_NAME,' ','')) LIKE '%{keyword}%'".format(keyword=keyword.replace(" ",""))
#         elif select_column_type == "3" :
#           select_column_type_str = "AND upper(replace(B.COMMENTS,' ','')) LIKE '%{keyword}%'".format(keyword=keyword.replace(" ",""))
#         elif select_column_type == "4" :
#           select_column_type_str = "AND upper(replace(B.TABLE_NAME,' ','')) LIKE '%{keyword}%'".format(keyword=keyword.replace(" ",""))

#         if select_column_type == "1" :
#           order_by_str = "ORDER BY DATA_TYPE desc, exp_order, data_type_exp_order, data_type2,COL_COMMENTS,COMMENTS"
#         elif select_column_type == "2" :
#           order_by_str = "ORDER BY DB_USE, OWNER, TABLE_NAME, COLUMN_ID"
#         elif select_column_type == "3" :
#           order_by_str = "ORDER BY DB_USE, OWNER, TABLE_NAME, COLUMN_ID"
#         elif select_column_type == "4" :
#           order_by_str = "ORDER BY DB_USE, OWNER, TABLE_NAME, COLUMN_ID"


#         # select_data_type = request.POST.getlist('select_data_type','')
#         # select_data_type_str = (', '.join(select_data_type))
#         # select_domain_id = "AND 1 = 1"
#         # if select_data_type_str :
#         #   select_data_type_str = "in ( " + select_data_type_str + " ) "
#         # else :
#         #   select_data_type_str = "in ( '' ) "

#       else :
#         keyword = ""
#         select_column_type = 1
#         # select_db_str = "in ( 39 ) "
#         # select_data_type_str = "in ( 1,2,3,4 ) "
#         # select_domain_id = "AND d.id = {domain_id}".format(domain_id=domain_id)


#       with connections['default'].cursor() as cursor :
#         sql = """
#         SELECT
#                domain_name,
#              upper(substr(COMMENTS,1,15)) comments,
#                replace(upper(col_comments),'{keyword}','<font color="red"><b>{keyword}</b></font>') col_comments,
#                upper(col_comments) tooltip_col_comments,
#                id,
#                db_use,
#                upper(substr(TABLE_NAME,1,15)) table_name,
#                upper(substr(COLUMN_NAME,1,15)) column_name,
#                data_type2,
#                data_type,
#                exp_order,
#                data_type_exp_order,
#                user,
#                data_explain,
#                column_id,
#                owner
#         FROM
#         (
#         SELECT
#              D.DOMAIN_NAME,
#              B.COMMENTS,
#                C.COL_COMMENTS,
#                C.ID,
#                A.DB_USE,
#                B.TABLE_NAME,
#                C.COLUMN_NAME,
#                e.data_type DATA_TYPE2,
#                0 DATA_TYPE, -- 0 : COLUMN / 1 : DATA
#                B.EXP_ORDER,
#                E.EXP_ORDER DATA_TYPE_EXP_ORDER,
#                CONCAT(AU.FIRST_NAME,'(',AU.LAST_NAME,')') USER,
#                c.column_explain data_explain,
#                C.COLUMN_ID,
#                C.OWNER
#         FROM CUST_DB_LIST A,
#              CUST_TABLE_LIST B
#              LEFT OUTER JOIN AUTH_USER AU
#              ON (
#               B.ID_USER_IT_MANAGER = AU.ID
#              ),
#              CUST_COLUMN_LIST C,
#              CUST_DOMAIN D,
#              CUST_DATA_TYPE E
#         WHERE A.ID = B.ID_DBLIST
#           AND B.ID = C.ID_TABLELIST
#           AND D.ID = B.ID_DOMAIN
#           AND E.ID = B.ID_DATATYPE
#           AND B.EXP_ORDER > 0
#           AND B.OPER_CD = 3
#           AND B.COMMENTS IS NOT NULL
#           AND C.COL_COMMENTS IS NOT NULL
#           {select_column_type_str}
#         UNION ALL
#         SELECT
#                D.domain_name,
#                '' COMMENTS,
#                C.DATA_TITLE,
#                C.ID,
#                A.DB_USE,
#                '' TABLE_NAME,
#                '' COLUMN_NAME,
#                E.DATA_TYPE,
#                1 DATA_TYPE, -- 0 : COLUMN / 1 : DATA
#                C.EXP_ORDER,
#                E.EXP_ORDER DATA_TYPE_EXP_ORDER,
#                CONCAT(AU.FIRST_NAME,'(',AU.LAST_NAME,')') USER,
#                c.data_explain data_explain,
#                9999,
#                ''
#         FROM CUST_DATA_LIST C
#              LEFT OUTER JOIN CUST_DB_LIST A
#              ON (
#                     A.ID = C.ID_DBLIST
#              )
#              LEFT OUTER JOIN AUTH_USER AU
#              ON (
#               C.ID_MOD_USER = AU.ID
#              )
#              LEFT OUTER JOIN CUST_DOMAIN D
#              ON (
#                 C.ID_DOMAIN = D.ID
#              ),
#              CUST_DATA_TYPE E
#         WHERE
#               upper(replace(C.DATA_TITLE,' ','')) LIKE '%{keyword}%'
#          AND  E.ID = 4
#          AND  C.EXP_YN = '1'
#          AND  1=1 -- SQL 데이터는 노출 안되도록
#         ) A
#         WHERE 1=1
#         {order_by_str}
#         ;
#         """.format(keyword=(keyword.upper()).replace(' ',''),
#                  select_column_type_str = select_column_type_str,
#                  order_by_str = order_by_str)

#         cursor.execute(sql)


#         columnNames = [d[0] for d in cursor.description]
#         search_dataset = [dict(zip(columnNames, row)) for row in cursor]

#         search_dataset_len = len(search_dataset)
#         search_dataset = search_dataset[0:100]


#   else :
#     form = searchKeywordForm()

#     select_column_type = 1
#     select_data_type = 'all'


#     # text = open("alice.txt").read()

#     # sql = """
#     #     select keyword, count(*) cnt
#     #     from cust_user_search_keyword
#     #     group by keyword
#     #     having count(*) > 1
#     # """
#     # wordcloud = word_cloud(sql)


#   sql = """
#   SELECT channel_name, db_use, id
#   FROM
#   (
#   SELECT A.channel_name, b.id, B.db_use, a.exp_order a_exp_order, b.exp_order b_exp_order
#   FROM CUST_channel A,
#      CUST_DB_LIST B
#   WHERE A.ID = B.ID_channel
#     AND A.EXP_ORDER > 0
#     AND B.EXP_ORDER > 0
#   ) a
#   ORDER BY a_exp_order, b_exp_order;"""

#   db_dataset = _query_list('default',sql)

#   db_dict = defaultdict(list)

#   for upper_domain_name, \
#       db_use, \
#       id \
#       in db_dataset:
#       db_dict[upper_domain_name].append([db_use,str(id)])

#   db_dict = dict(db_dict)


#   sql = """
#   SELECT id,data_type
#   FROM cust_data_type
#   ORDER BY id;"""

#   data_type_list = _query_list('default',sql)

#   sql = """
#     select replace(a.xy,'/',',') xy, b.id db_id, ifnull(c.id,0) domain_id, b.db_use, ifnull(c.domain_name, a.etc1) domain_name
#     from cust_database_map a
#        left outer join cust_db_list b
#          on (
#          a.id_dblist = b.id
#          )
#          left outer join cust_domain c
#          on (
#          a.id_domain = c.id
#          )
#     where a.map_type = 0;
#   """

#   data_database_map_dict = _query_dict('default',sql)

#   sql = """
#     SELECT CAST(@ROWNUM := @ROWNUM + 1 AS unsigned) ROWNUM , A.KEYWORD EXP_NAME, replace(replace(A.KEYWORD,'/','!_!'),' ','') DATA_TITLE, A.CNT
#     FROM
#     (
#       SELECT upper(KEYWORD) KEYWORD, count(*) CNT
#       FROM CUST_USER_SEARCH_KEYWORD
#       GROUP BY upper(KEYWORD)
#       ORDER BY CNT DESC
#     ) A, ( SELECT @ROWNUM := 0 ) TMP
#     LIMIT 10;
#   """

#   dataset_user_search_keyword = _query_mysql_dict(sql)

#   sql = """
#     SELECT CAST(@ROWNUM := @ROWNUM + 1 AS unsigned) ROWNUM, A.EXP_NAME, replace(replace(A.DATA_TITLE,'/','!_!'),' ','') DATA_TITLE, REG_DTM, USERNAME
#     FROM
#     (
#     -- SELECT concat(DATE_FORMAT(A.MOD_DTM,'%Y/%m/%d'),' - ',A.DATA_TITLE) EXP_NAME, DATA_TITLE
#     SELECT A.DATA_TITLE EXP_NAME, DATA_TITLE, A.REG_DTM, B.FIRST_NAME USERNAME
#     FROM CUST_DATA_LIST A,
#          AUTH_USER B
#     WHERE B.ID = A.ID_REG_USER
#       AND EXP_YN = 1
#     ORDER BY REG_DTM DESC
#     ) A, ( SELECT @ROWNUM := 0 ) TMP
#     LIMIT 10;
#   """

#   dataset_latest_datalist = _query_mysql_dict(sql)

#   sql = """
#     SELECT CAST(@ROWNUM := @ROWNUM + 1 AS unsigned) ROWNUM, A.DATA_TITLE EXP_NAME, replace(replace(A.DATA_TITLE,'/','!_!'),' ','') DATA_TITLE, REG_DTM
#     FROM
#     (
#       SELECT C.DATA_TITLE, REG_DTM
#       FROM AUTH_USER A,
#          cust_datalist_user_id_req_users B,
#          CUST_DATA_LIST C
#       WHERE A.USERNAME = '{username}'
#       AND   A.ID = B.USER_ID
#       AND   C.ID = B.DATALIST_ID
#       AND   C.EXP_YN = 1
#       UNION ALL
#       SELECT C.DATA_TITLE, C.REG_DTM
#       FROM  AUTH_USER A,
#           CUST_DATA_LIST C
#       WHERE  A.USERNAME = '{username}'
#       AND  ( C.ID_REG_USER = A.ID OR C.ID_MOD_USER = A.ID)
#       AND   C.EXP_YN = 1
#       ORDER BY REG_DTM DESC
#     )A, ( SELECT @ROWNUM := 0 ) TMP;
#   """.format(username=request.user.username)

#   dataset_my_datalist = _query_mysql_dict(sql)


#   return render(request, 'data_search_keyword.html', {
#                         'template_parent':template_parent,
#                         'template_child':template_child,
#                         'keyword':keyword,
#                         'search_dataset':search_dataset,
#                         'search_dataset_len':search_dataset_len,
#                         'db_dict':db_dict,
#                         # 'select_db':select_db,
#                         'select_column_type':select_column_type,
#                         # 'select_data_type':select_data_type,
#                         # 'data_type_list':data_type_list,
#                         'form':form,
#                         'wordcloud':wordcloud,
#                         'word_meaning':word_meaning,
#                         'dashboard_data':_dashboard_data(request),
#                         'dataset_user_search_keyword':dataset_user_search_keyword,
#                         'dataset_latest_datalist':dataset_latest_datalist,
#                         'data_database_map_dict':data_database_map_dict,
#                         'dataset_my_datalist':dataset_my_datalist,
#                         'chk_tab_comments':chk_tab_comments,
#                         'chk_table_name':chk_table_name,
#                         'chk_column_name':chk_column_name,
#                         'chk_db_name':chk_db_name,
#                         'chk_data_type':chk_data_type,
#                          })


# def dataSearchDomain(request, domain_id) :

#   # Login 한 user 만 접속 허용
#   # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
#   if not request.user.is_authenticated :
#     return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

#   # _loggingVisit(request, 4)

#   template_parent = "dataSearch"
#   template_mid = "dataSearchDomain"
#   template_child = domain_id
#   dataset = ""
#   domain_name = ""

#   sql = """
#     select data_type,
#            comments,
#            table_id,
#            column_or_data,
#            concat("a",domain_order) domain_order
#     from
#     (
#       select c.data_type,
#            b.comments,
#            b.id table_id,
#            0 column_or_data,
#            c.exp_order,
#            c.exp_order domain_order
#       from cust_domain a,
#          cust_table_list b,
#          cust_data_type c
#       where a.id = b.id_domain
#         and c.id = b.id_datatype
#         and a.id = {id}
#         and b.exp_order <= 20
#         and b.comments is not null
#       union all
#       select c.data_type,
#            b.data_title,
#            b.id table_id,
#            1 column_or_data,
#            c.exp_order,
#            c.exp_order domain_order
#       from cust_domain a,
#          cust_data_list b,
#          cust_data_type c
#       where a.id = b.id_domain
#         and c.id = 4
#         and a.id = {id}
#         and b.exp_yn = '1'
#     order by domain_order, exp_order, data_type, comments
#     ) a
#   """.format(id=domain_id)


#   dataset_domain_1 = _query_mysql_list(sql)

#   dataset_domain_2 = defaultdict(list)

#   for data_type, \
#       comments, \
#       table_id, \
#       column_or_data, \
#       domain_order \
#       in dataset_domain_1:
#       dataset_domain_2[data_type].append([comments,str(table_id), column_or_data, domain_order])

#   dataset_domain_2 = dict(dataset_domain_2)


#   return render(request, 'data_search_domain.html', {
#                         'template_parent':template_parent,
#                         'template_mid':template_mid,
#                         'template_child':template_child,
#                         'dataset_domain_2':dataset_domain_2,
#                         'dashboard_data':_dashboard_data(request),
#                         # 'dataset_dict':dataset_dict,
#                         # 'dataList':dataList,
#                         # 'domain_name':domain_name,
#                         # 'tab_nm':tab_nm,
#                         # 'col_nm':col_nm,
#                         # 'dataset':dataset,
#                         # 'data_type':data_type
#                          })


# def dataSearchFlow(request, column_id, domain_id) :


#   # Login 한 user 만 접속 허용
#   # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
#   if not request.user.is_authenticated :
#     return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))


#   template_parent = "dataWhat"
#   template_mid = "dataSearch"
#   template_child = "dataSearchFlow"


#   buffer_domain_id = domain_id
#   dataset = ""
#   domain_name = ""
#   dataset = ""


#   with connections['default'].cursor() as cursor :

#     if column_id > 0 :


#       sql = dataSearchSql.format(column_id=column_id,
#                                data_type=1,
#                                userid = request.user.id,
#                                super_user = request.user.is_superuser)


#       cursor.execute(sql)

#       columnNames = [d[0] for d in cursor.description]
#       dataset = [dict(zip(columnNames, row)) for row in cursor]


#       dataset = dataset[0]

#       tab_nm = dataset['comments']
#       col_nm = dataset['col_comments']
#     else :
#       tab_nm = "None"
#       col_nm = "None"


#     dataset_dict = defaultdict(list)


#     sql = """
#     SELECT process_nm,
#            data_title,
#            data_id,
#            process_id
#     FROM
#     (
#       SELECT
#              B.process_nm,
#              A.DATA_TITLE,
#              A.ID DATA_ID,
#              B.ID PROCESS_ID,
#              B.ORDERING
#       FROM
#            CUST_DATA_LIST A,
#            cust_process_list B
#       WHERE B.ID = A.id_processlist
#         -- AND A.data_search_psbl_yn = 1
#         and A.EXP_YN = '1'
#         AND B.ID = {id}
#     ) A
#     ORDER BY ORDERING,DATA_TITLE;
#     """.format(id=domain_id)


#     cursor.execute(sql)
#     dataset_buffer = cursor.fetchall()

#     dataset_dict = defaultdict(list)

#     for process_nm, \
#         data_title, \
#         data_id, \
#         process_id \
#         in dataset_buffer:
#         dataset_dict[process_nm].append([data_title, data_id, process_id])


#   if buffer_domain_id == 0 :
#     dataset_dict = None
#   else :
#     dataset_dict  = dict(dataset_dict)

#   return render(request, 'data_search_flow.html', {
#                         'template_parent':template_parent,
#                         'template_mid':template_mid,
#                         'template_child':template_child,
#                         'dataset_dict':dataset_dict,
#                         'domain_name':domain_name,
#                         'tab_nm':tab_nm,
#                         'col_nm':col_nm,
#                         'dataset':dataset,
#                         'dashboard_data':_dashboard_data(request),
#                          })


# def tableSearch(request) :

#   # Login 한 user 만 접속 허용
#   # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
#   if not request.user.is_authenticated :
#     return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

#   template_parent = "dev"
#   template_child = "tableSearch"

#   select_db = ""
#   select_data_type = ""
#   select_column_type = ""
#   ds_searchtable_dict = ""
#   keyword = ""

#   if request.method == "POST" :

#     form = searchKeywordForm(request.POST)

#     if form.is_valid() :

#       keyword = form.cleaned_data['keyword']

#       select_db = request.POST.getlist('select_db','')


#       if "multiselect-all" in select_db :
#         select_db.remove("multiselect-all")
#       select_db_str = (', '.join(select_db))
#       if select_db_str :
#         select_db_str = "in ( " + select_db_str + " ) "
#       else :
#         select_db_str = "in ( '' ) "

#       # 속성 한글명, 속성 영문명, 테이블 영문명, 테이블 한글명 등
#       select_column_type = request.POST.get('select_column_type','')

#       if select_column_type == "1" :
#         select_column_type_str = "AND C.COL_COMMENTS LIKE '%{keyword}%'".format(keyword=keyword)
#       elif select_column_type == "2" :
#         select_column_type_str = "AND C.COLUMN_NAME LIKE '%{keyword}%'".format(keyword=keyword)
#       elif select_column_type == "3" :
#         select_column_type_str = "AND B.COMMENTS LIKE '%{keyword}%'".format(keyword=keyword)
#       elif select_column_type == "4" :
#         select_column_type_str = "AND B.TABLE_NAME LIKE '%{keyword}%'".format(keyword=keyword)


#       select_data_type = request.POST.getlist('select_data_type','')
#       select_data_type_str = (', '.join(select_data_type))
#       if select_data_type_str :
#         select_data_type_str = "in ( " + select_data_type_str + " ) "
#       else :
#         select_data_type_str = "in ( '' ) "


#       sql = """
#         select a.id,
#              a.owner,
#              a.table_name,
#                a.comments,
#                a.total_mb,
#                a.created,
#                a.last_ddl_time,
#                a.cdc_yn,
#                a.db_use,
#                b.privacy_yn,
#                c.channel_name,
#                e.data_type,
#                ifnull(concat(au1.first_name,'(',au1.username,')'),'') it_rep,
#                ifnull(cdm.domain_name,'') domain_name
#         from cust_table_list a
#                left outer join cust_domain cdm
#                on ( a.id_domain = cdm.id )
#                left outer join auth_user au1
#                on ( a.id_user_it_manager = au1.id ),
#              cust_db_list b,
#              cust_channel c,
#              cust_data_type e
#         where a.id_dblist = b.id
#           and b.id_channel = c.id
#           and a.id_datatype = e.id
#           and b.id = 10
#           and a.table_name like '%{keyword}%';
#       """.format(
#             keyword=keyword
#             )

#       ds_searchtable_dict = _query_mysql_dict(sql)
#       ds_searchtable_dict = ds_searchtable_dict[0:500]


#   else :
#     form = searchKeywordForm()

#     # 검색 조건 :
#     # 1 - 컬럼 한글
#     # 2 - 컬럼 영문
#     # 3 - 테이블 한글
#     # 4 - 테이블 영문
#     select_column_type = 4

#     # 데이터 유형 : all - 전체 조회
#     select_data_type = 'all'

#   sql = """
#     SELECT channel_name, db_use, id db_id
#     FROM
#     (
#     SELECT A.channel_name, b.id, B.db_use, a.exp_order a_exp_order, b.exp_order b_exp_order
#     FROM CUST_CHANNEL A,
#        CUST_DB_LIST B
#     WHERE A.ID = B.ID_channel
#       AND A.EXP_ORDER > 0
#       AND B.EXP_ORDER > 0
#     ) a
#     ORDER BY a_exp_order, b_exp_order;"""

#   ds_db_list = _query_mysql_list(sql)

#   ds_db_dict = defaultdict(list)

#   for channel_name, \
#       db_use, \
#       db_id \
#       in ds_db_list:
#       ds_db_dict[channel_name].append([db_use,str(db_id)])

#   ds_db_dict = dict(ds_db_dict)

#   sql = """
#     SELECT id,data_type
#     FROM cust_data_type
#     ORDER BY id;""
#   """

#   ds_datatype_list = _query_mysql_list(sql)


#   return render(request, 'table_search.html', {
#                         'template_parent':template_parent,
#                         'template_child':template_child,
#                         'form':form,
#                         'ds_db_dict':ds_db_dict,
#                         'select_db':select_db,
#                         'ds_datatype_list':ds_datatype_list,
#                         'select_column_type': int(select_column_type),
#                         'select_data_type':select_data_type,
#                         'ds_searchtable_dict':ds_searchtable_dict,
#                         'keyword':keyword
#                          })


def executeSQL(request):
    if not request.user.is_authenticated:
        return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

    _loggingVisit(request, 5)

    template_parent = "executeSQL"
    template_child = "executeSQL"

    sql_text = request.POST.get('text_sql', '')
    sql_id = request.POST.get('sql_id', '')
    data_type = request.POST.get('data_type', '')
    avg_t = request.POST.get('avg_t', '')

    # 데이터 타입이 컬럼이라면 id_datalist id 를 저장할 필요가 없으므로 sql_id 는 Null 처리
    if data_type == "0":
        sql_id = None

    col_comments = request.POST.get('col_comments', '')

    sql_title_list = _query_list('default', sql_execute_saved.format(id=request.user.id))

    sql_title_dict = defaultdict(list)

    for writer, \
        id, \
        data_title \
            in sql_title_list:
        sql_title_dict[writer].append([str(id), data_title])

    sql_title_dict = dict(sql_title_dict)

    sql = """
  SELECT channel_name, db_use, id
  FROM
  (
  SELECT A.channel_name, b.id, B.db_use, a.exp_order a_exp_order, b.exp_order b_exp_order
  FROM CUST_channel A,
     CUST_DB_LIST B
  WHERE A.ID = B.ID_channel
    AND A.EXP_ORDER > 0
    AND B.EXP_ORDER > 0
  ) a
  ORDER BY a_exp_order, b_exp_order;"""

    db_dataset = _query_list('default', sql)

    db_dict = defaultdict(list)

    for upper_domain_name, \
        db_use, \
        id \
            in db_dataset:
        db_dict[upper_domain_name].append([db_use, str(id)])

    db_dict = dict(db_dict)

    sql = """
  SELECT ID, DOMAIN_NAME
  FROM CUST_DOMAIN
  ORDER BY EXP_ORDER"""

    data_domain_list = _query_dict('default', sql)

    return render(request, 'execute_sql.html', {
        'template_parent': template_parent,
        'template_child': template_child,
        'sql_text': sql_text,
        'col_comments': col_comments,
        'sql_title_list': sql_title_list,
        'sql_id': sql_id,
        'db_dict': db_dict,
        'data_domain_list': data_domain_list,
        'avg_t': avg_t,
        'sql_title_dict': sql_title_dict,
    })


def json_default(value):
    if isinstance(value, datetime.date):
        return value.strftime('%Y-%m-%d')


def ajaxCreateDDLLVErdDownload(request) :
    try :
        db_case = ""
        if request.FILES.__len__() == 0:
            messages.info(request, "업로드된 파일이 없습니다")
            return HttpResponseRedirect('../CreateDDLLV')

        uploadFile = request.FILES['file_erd'];


        if uploadFile.name.find('.dbs') < 0 :
            messages.info(request, "DBSchema (*.dbs) 파일형식만 가능합니다")
            return HttpResponseRedirect('../CreateDDLLV')

        lFile_name = (uploadFile.name).split('.')

        read = uploadFile.read().decode('utf8')

        root = elemTree.fromstring(read)

        
        
        for element in root.iter(tag="project"):        
            db_case = element.attrib["database"].lower()

        for element in root.iter(tag="table"):

            try:
                for col_element in element.findall("column"):

                    full_data_type = ""
                    div_data_type = col_element.attrib["type"] if 'type' in col_element.attrib else ''
                    div_length = col_element.attrib["length"] if 'length' in col_element.attrib else ''
                    div_decimal = col_element.attrib["decimal"] if 'decimal' in col_element.attrib else ''
                    bFind = False

                    col_name = col_element.attrib["name"]

                    sql = """SELECT A.STD_ATTR_ENG, 
                                    LOWER(B.MYSQL_DATATYPE) MYSQL_DATATYPE, 
                                    LOWER(B.ORACLE_DATATYPE) ORACLE_DATATYPE, 
                                    B.INFO_TYPE
                        from cust_std_attr a,
                             cust_std_domain b
                        where b.id = a.id_stddomain
                        and a.std_attr_kor = '{kor_column_name}'
                        and   a.use_yn = 1
                        and   b.use_yn = 1;
                    """.format(kor_column_name=col_name)

                    std_info = _query_dict("default", sql)
                    

                    
                    if len(std_info) > 0 :
                        eng_column_name = std_info[0]['STD_ATTR_ENG']

                        if db_case == 'oracle' :
                            full_data_type = std_info[0]['ORACLE_DATATYPE']
                        elif db_case == 'mysql' :
                            full_data_type = std_info[0]['MYSQL_DATATYPE']
                        else :
                            full_data_type = std_info[0]['MYSQL_DATATYPE']

                        col_p1 = re.compile("^([a-zA-Z0-9]+)$", re.M)
                        col_p2 = re.compile("^([a-zA-Z0-9]+)\(([0-9]+)\)$", re.M)
                        col_p3 = re.compile("^([a-zA-Z0-9]+)\(([0-9]+),([0-9]+)\)$", re.M)

                        col_m = col_p1.findall(full_data_type)
                      

                        if len(col_m) == 0 :
                            col_m = col_p2.findall(full_data_type)
                        else :                            
                            div_data_type = col_m[0]
                            div_length = None
                            div_decimal = None
                            bFind = True

                        if bFind == False and len(col_m) == 0 :
                            col_m = col_p3.findall(full_data_type)
                        elif bFind == False :
                            div_data_type = col_m[0][0]
                            div_length = col_m[0][1]
                            div_decimal = None
                            bFind = True


                        if bFind == False and len(col_m) > 0 :
                            div_data_type = col_m[0][0]
                            div_length = col_m[0][1]
                            div_decimal = col_m[0][2]


                    else :
                        eng_column_name = col_name

                    if col_element.find("comment") is not None :
                        col_element.find("comment").text = eng_column_name
                    else :
                        elemTree.SubElement(col_element,"comment").text = eng_column_name                    
                    
                    
                    col_element.attrib["type"] = div_data_type
                    
                    if div_length is not None :
                        col_element.attrib["length"] = div_length

                    if div_decimal is not None :
                        col_element.attrib["decimal"] = div_decimal





            except Exception as e:
                print(e)

        
        read = elemTree.tostring(root, encoding='utf8', method='xml')

        
        


        new_title = lFile_name[0]+'_논리'+'_'+datetime.datetime.now().strftime("%Y%m%d%H%M%S")+'.'+lFile_name[1]

        file_name = urllib.parse.quote(new_title.encode('utf-8'))
        
        response = HttpResponse(read, content_type='dbs')
        response['Content-Disposition'] = 'attachment;filename*=UTF-8\'\'%s' % file_name        

       
        return response
    except Exception as e :
        print(e)

def ajaxCreateDDLLVErdChangeDownload(request) :
    

    swich_type = ""
    swich_name = ""
    p = ""
    p_err = ""
    p_limit_count = 1000
    
    if 'btn_erd_ptol_download' in request.POST:
        swich_type = 'ptol'
        swich_name = "논리"            
        p = re.compile('[가-힣]+')
        p_err = "한글로 명명된 컬럼명이 {}개 이상 발견되었습니다.\n물리ERD -> 논리ERD로 변환하는 화면입니다.\n물리ERD가 맞는지 확인해주세요.".format(p_limit_count)
    elif 'btn_erd_ltop_download' in request.POST:
        swich_type = 'ltop'
        swich_name = "물리"
        p = re.compile('[A-Za-z]+')
        p_err = "알파벳로 명명된 컬럼명이 {}개 이상 발견되었습니다.\n논리ERD -> 물리ERD로 변환하는 화면입니다.\n논리ERD가 맞는지 확인해주세요.".format(p_limit_count) 

    err_count = 0
    

    if request.FILES.__len__() == 0:
        messages.info(request, "업로드된 파일이 없습니다")
        return HttpResponseRedirect('../CreateDDLLV')

    uploadFile = request.FILES['file_erd_change'];


    if uploadFile.name.find('.dbs') < 0 :
        messages.info(request, "DBSchema (*.dbs) 파일형식만 가능합니다")
        return HttpResponseRedirect('../CreateDDLLV')

    lFile_name = (uploadFile.name).split('.')

    read = uploadFile.read().decode('utf8')

    root = elemTree.fromstring(read)

    dict_tab = {}
    dict_col = {}

    for element in root.iter(tag="table"):
        try:

            name = element.attrib["name"]

            if "(참조)" in name :
                continue

            comment = element.find("comment").text

            name = name.replace(' ','')
            comment = comment.replace(' ','')

            dict_tab[name] = comment

            element.set('name', comment)
            element.find("comment").text = name

            if err_count > p_limit_count :
                messages.info(request, p_err)
                return HttpResponseRedirect('../CreateDDLLV')                    

            for col_element in element.iter(tag="column"):
                col_name = col_element.attrib["name"]
                col_comment = col_element.find("comment").text

                col_name = col_name.replace(' ','')
                col_comment = col_comment.replace(' ','')
                
                m = p.match(col_name)          

                if m :
                    err_count += 1                    

                col_element.set('name', col_comment)
                col_element.find("comment").text = col_name

                dict_col[col_name] = col_comment
        except Exception:
            pass

    
    for element in root.iter(tag="table"):
        # fk가 여러개 존재할 수 있으므로
        for fk in element.findall("fk"):
            to_table = fk.attrib["to_table"]
            
            if "(참조)" in to_table :
                continue
            
            fk.set("to_table", dict_tab[to_table])            

    for entity_element in root.iter(tag="entity"):
        try:
            name = entity_element.attrib['name']
            
            entity_element.set('name', dict_tab[name])
        except Exception:

            pass

    for element in root.iter(tag="column"):
        try:
            name = element.attrib["name"]
            element.set("name", dict_col[name])
        except Exception:
            pass

    for element in root.iter(tag="fk_column"):
        try:
            name = element.attrib["name"]
            element.set("name", dict_col[name])

            name = element.attrib["pk"]
            element.set("pk", dict_col[name])
        except Exception:
            print("Error : fk_column")
            pass

    # for element in root.findall("callout"):
    #     root.remove(element)

    for element in root.findall("layout"):
        for sub_element in element.findall("query"):
            element.remove(sub_element)
        for sub_element in element.findall("browser"):
            element.remove(sub_element)
        for sub_element in element.findall("script"):
            element.remove(sub_element)
 
    
    read = elemTree.tostring(root, encoding='utf8', method='xml')              



    # print(read)

    new_title = lFile_name[0]+'_'+swich_name+'_'+datetime.datetime.now().strftime("%Y%m%d%H%M%S")+'.'+lFile_name[1]

    file_name = urllib.parse.quote(new_title.encode('utf-8'))
    
    response = HttpResponse(read, content_type='dbs')
    response['Content-Disposition'] = 'attachment;filename*=UTF-8\'\'%s' % file_name        

   
    return response

         
def ajaxCreateDDLLVHtmlChangeDownload(request) :
    try :

        uploadFile = request.FILES['file_erd_change'];

        lFile_name = (uploadFile.name).split('.')

        html = uploadFile.read().decode('utf8')


        dict_tab = {}
        dict_col = {}                

        # print(html)

        # TABLE 치환
        # p = re.compile("(<text.+?>)(.+?)(</text><title>Table.+?\.)(.+?)(\n)(.+?)(</title>)")
        p = re.compile("(<text.+?)(>)(.+?)(</text>)(<title>Table.+?\.)(.+?)(\n)(.+?)(</title>)")

        m = p.findall(html)
        for n in m:
            dict_tab[n[2].replace(' ','')] = n[7].replace(' ','')

        # html = p.sub('\g<1>\g<6>\g<3>\g<4>\g<5>\g<6>\g<7>', html)
        html = p.sub('\g<1> class="p_model" style="display: none" \g<2>\g<3>\g<4>' +
                     '\g<1> class="l_model"                       \g<2>\g<8>\g<4>' +
                     '\g<5>\g<6>\g<7>\g<8>\g<9>'
                     , html)

        # 컬럼 치환
        # p = re.compile("(<text id='.+?\..+?\..+?'.+?>)(.+?)(</text><title>)(.+?)(\n.+?\n)(.+?)(</title>)")
        p = re.compile("(<text id='.+?\..+?\..+?'.+?)(>)(.+?)(</text>)(<title>)(.+?)(\n.+?\n)(.+?)(</title>)")

        m = p.findall(html)
        for n in m:
            dict_col[n[2].replace(' ','')] = n[7].replace(' ','')

        # html = p.sub('\g<1>\g<6>\g<3>\g<4>\g<5>\g<6>\g<7>', html)        
        html = p.sub('\g<1> class="p_model" style="display: none" \g<2>\g<3>\g<4>' +
                     '\g<1> class="l_model"                       \g<2>\g<8>\g<4>' +
                     '\g<5>\g<6>\g<7>\g<8>\g<9>'
                     , html)

        # fk명 제거
        p = re.compile("(<title>)(Fk FK.+?)(\n)")
        html = p.sub('\g<1>\g<3>', html)

        # fk 컬럼 논리명으로 치환
        p = re.compile("(<title>\n)(.+?)( ref )(.+?)( \( )(.+?)( \)</title>)")

        m = p.findall(html)
        for n in m:
            from_eng_table = n[1]
            to_eng_table = n[3]

            from_eng_table = from_eng_table.replace(' ','')
            to_eng_table = to_eng_table.replace(' ','')

            from_kor_table = dict_tab[from_eng_table]
            to_kor_table = dict_tab[to_eng_table]


            eng_columns = n[5]

            l_eng_columns = eng_columns.split(' ')
            l_kor_columns = []
            for l in l_eng_columns:
                try:
                    if l.find(",") > 0:
                        l_kor_columns.append(dict_col[l.replace(",", "")] + ",")
                    else:
                        l_kor_columns.append(dict_col[l])
                except Exception:
                    l_kor_columns.append(l)

            kor_columns = ' '.join(l_kor_columns)
            # print(eng_columns)
            # print(kor_columns)
            # print()

            imsi_p = re.compile("(<title>\n.+? ref .+? \( )(" + eng_columns + ")( \)</title>)")
            html = imsi_p.sub('\g<1>' + kor_columns + '\g<3>', html)

            imsi_p = re.compile("(<title>\n)(.+?)( ref )(.+?)( \( )(.+?)( \)</title>)")
            html = imsi_p.sub('\g<1>' + from_kor_table + '\g<3>' + to_kor_table + '\g<5>\g<6>\g<7>', html)


        new_title = lFile_name[0]+'(게시용)_'+datetime.datetime.now().strftime("%Y%m%d%H%M%S")+'.'+lFile_name[1]

        file_name = urllib.parse.quote(new_title.encode('utf-8'))

        html = html + "n" + \
        """
        <script src="https://code.jquery.com/jquery-latest.min.js"></script>
        <script>
            document.onkeyup = function(e) {
                if (e.which == 80 ) {
                    $('.l_model').css('display','none');
                    $('.p_model').css('display','block');
                }
                if (e.which == 76 ) {
                    $('.l_model').css('display','block');
                    $('.p_model').css('display','none');
                }
            }       
        </script>
        """
        
        response = HttpResponse(html, content_type='html')
        response['Content-Disposition'] = 'attachment;filename*=UTF-8\'\'%s' % file_name        

       
        return response
    except Exception as e :
        print(e)        


def ajaxCreateDDLLVErd(request):

    
    
    message = ""
    err_message = []

    if request.FILES.__len__() == 0:
        message = "업로드된 파일이 없습니다"
        return JsonResponse({"message": message})

    uploadFile = request.FILES['file'];

    if uploadFile.name.find('dbs') < 0 :
        message = "DBSchema (*.dbs) 파일형식만 가능합니다"
        return JsonResponse({"message": message})

    read = uploadFile.read().decode('utf8')

    root = elemTree.fromstring(read)

    err_message = []
    meta_dict = {}
    meta_list = []
    col_list = []
    tab_start_yn = ""
    tab_name = ""
    col_name = ""

    err_count = 0
    p = re.compile('[A-Za-z]+')


    TOTAL_DDL = ""

    befor_table_name = "START"   
    first_table_name = "NEW|START"

    no = 1
    column_id = 1
    table_cnt = 0
    table_chg = 1
    limit_count = 1000

    db_case = ""

    for element in root.iter(tag="project"):        
        db_case = element.attrib["database"].upper()

    for element in root.iter(tag="table"):
        try:
            kor_tab_name = element.attrib["name"]


            if kor_tab_name.find("(참조)") > 0 :                
                continue

            if err_count > limit_count :
                message = "알파벳으로 명명된 컬럼명이 {}개 이상 발견되었습니다.\n논리ERD에 대한 표준체크를 하는 화면입니다.\n논리ERD가 맞는지 확인해주세요.".format(limit_count)
                break

            eng_tab_name = ""

            if element.find("comment") is not None :
                    eng_tab_name = element.find("comment").text

            pk_column = []           


            for ind_element in element.iter(tag="index"):
                if ind_element.attrib["unique"] == "PRIMARY_KEY" :
                    for pk_element in ind_element.iter(tag="column") :
                        pk_column.append(pk_element.attrib["name"])

            for col_element in element.findall("column"):



                if first_table_name != kor_tab_name:
                    tab_start_yn = 'Y'
                    first_table_name = kor_tab_name
                    column_id = 1
                    table_cnt = table_cnt + 1

                    if table_chg == 1 :
                        table_chg = 0
                    else :
                        table_chg = 1
                else:
                    tab_start_yn = 'N'                    
                
                col_list = []
                
                kor_col_name = col_element.attrib["name"]


                m = p.match(kor_col_name)          

                if m :
                    err_count += 1

                not_null = ""
                if 'mandatory' in col_element.attrib:
                    not_null = col_element.attrib["mandatory"].upper()

                
                default = ""
                
                if col_element.find("defo") is not None :
                    if len(col_element.find("defo").text) > 18 :
                        default = col_element.find("defo").text[0:15] + "..."
                    else :
                        default = col_element.find("defo").text

                eng_col_name = ""
                data_type = ""
                domain = ""
                stardard_yn = ""
                pk = ""

                if kor_col_name in pk_column :
                    pk = "Y"

                sql = """SELECT A.STD_ATTR_ENG, 
                                LOWER(B.MYSQL_DATATYPE) MYSQL_DATATYPE, 
                                LOWER(B.ORACLE_DATATYPE) ORACLE_DATATYPE,
                                B.INFO_TYPE
                    from cust_std_attr a,
                         cust_std_domain b
                    where b.id = a.id_stddomain
                    and a.std_attr_kor = '{kor_column_name}'
                    and   a.use_yn = 1
                    and   b.use_yn = 1;
                """.format(kor_column_name=kor_col_name)

                std_info = _query_dict("default", sql)

                if len(std_info) > 0:
                    eng_col_name = std_info[0]['STD_ATTR_ENG'].lower()
                    domain = std_info[0]['INFO_TYPE'].lower()
                    
                    if db_case == 'ORACLE' :
                        data_type = std_info[0]['ORACLE_DATATYPE'].lower()
                    elif db_case == 'MYSQL' :
                        data_type = std_info[0]['MYSQL_DATATYPE'].lower()
                    else :
                        data_type = std_info[0]['MYSQL_DATATYPE'].lower()
                    
                    stardard_yn = "<font color='blue'>표준</font>"
                else :

                    if col_element.find("comment") is not None :
                        eng_col_name = "<del>"+col_element.find("comment").text+"</del>"
                    
                    stardard_yn = "<font color='red'>비표준</font>"
                    err_message.append("<li>비표준 컬럼 존재 : #{no} 번째 라인</li>".format(no=no))




                col_list.append(no)
                col_list.append(column_id)
                col_list.append(kor_tab_name)
                col_list.append(eng_tab_name)
                col_list.append(kor_col_name)
                col_list.append(pk)
                col_list.append(not_null)
                col_list.append(default)
                col_list.append(eng_col_name)
                col_list.append(data_type)
                col_list.append(domain)
                col_list.append(stardard_yn)
                col_list.append(table_chg)
                col_list.append(tab_start_yn)

                # meta_dict['NO'] = no
                # meta_dict['COLUMN_ID'] = column_id
                # meta_dict['KOR_TABLE_NAME'] = tab_name
                # meta_dict['ENG_TABLE_NAME'] = ""
                # meta_dict['KOR_COLUMN_NAME'] = col_name
                # meta_dict['PK'] = ""
                # meta_dict['NOT_NULL'] = ""
                # meta_dict['DEFAULT'] = ""

                meta_list.append(col_list)
                no += 1
                column_id += 1
                

        except Exception as e:
            print(e)    
    
    ddl = makeDDL(meta_list, db_case=db_case)    

    context = {
        'meta_list': meta_list,
        'table_cnt': table_cnt,
        'message': message,
        'err_message' :err_message,
        'ddl' :ddl
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")

def ajaxCreateDDLLVExcel(request):

    sheet_name = request.POST.get('sheet_name', '')    

    message = ""



    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive',
    ]    

    json_file_name = settings.BASE_DIR + '/static/custom/json/chiho-test-22b79f37651a.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
    gc = gspread.authorize(credentials)
    spreadsheet_url = 'https://docs.google.com/spreadsheets/d/1SDjOVFaM_zGFObbsBSvofFkcHtT51siQxp9LdZtXBQQ/edit#gid=0'
    # 스프레스시트 문서 가져오기
    doc = gc.open_by_url(spreadsheet_url)
    # 시트 선택하기
    # worksheet = doc.worksheet('DDL정보입력(시트이름수정불가)')
    worksheet = ""

    try:
        worksheet = doc.worksheet(sheet_name)
    except:
        message = "존재하지 않는 시트입니다. 시트명을 확인해주세요.(대소문자 구분)"
        return JsonResponse({"message": message})        

    all_meta_list = worksheet.get_all_values()    

    x_len = len(all_meta_list[4])

    err_message = []

    meta_dict = {}
    meta_list = []

    TOTAL_DDL = ""
    tab_start_yn = ""

    befor_table_name = "START"

    owner = ""
    if all_meta_list[0][1] != "":
        owner = all_meta_list[0][1].lower() + "."
    data_tablespace = all_meta_list[1][1]
    index_tablespace = all_meta_list[2][1]
    db_case = all_meta_list[3][1]



    first_table_name = "NEW|START"

    no = 1
    column_id = 1
    table_cnt = 0
    table_chg = 1



    for meta in all_meta_list[6:]:

        i = 0
        meta_dict = {}    
        col_list = []    
        stardard_yn = ""

        if first_table_name != meta[1]:
            tab_start_yn = 'Y'
            first_table_name = meta[1]
            column_id = 1
            table_cnt = table_cnt + 1

            if table_chg == 1 :
                table_chg = 0
            else :
                table_chg = 1
        else:
            tab_start_yn = 'N'

        sql = """SELECT A.STD_ATTR_ENG, LOWER(B.MYSQL_DATATYPE) DATA_TYPE, B.INFO_TYPE
          from cust_std_attr a,
               cust_std_domain b
          where b.id = a.id_stddomain
          and a.std_attr_kor = '{kor_column_name}'
          and   a.use_yn = 1
          and   b.use_yn = 1;
      """.format(kor_column_name=meta[2])

        std_info = _query_dict("default", sql)

        eng_col_name = ""
        domain = ""
        data_type = ""

        if len(std_info) > 0:
            eng_col_name = std_info[0]['STD_ATTR_ENG'].lower()
            domain = std_info[0]['INFO_TYPE'].lower()
            data_type = std_info[0]['DATA_TYPE'].lower()    
            stardard_yn = '<font color="blue">표준</font>'

        eng_tab_name = meta[1].lower()
        kor_tab_name = meta[0].replace(' ','')
        kor_col_name = meta[2]
        pk = meta[3].upper()
        not_null = meta[4].upper()
        default = meta[5].lower()                    
        uk = meta[10].upper()

        # if column_id == 1 and pk != "Y":
        #     err_message.append("<li>PK없음 : #{no} 번째 라인</li>".format(no=no))

        if len(std_info) == 0:  

            eng_col_name = meta[6]            
            data_type = meta[7]

            stardard_yn = '<b><font color="red">비표준</font></b>'
            err_message.append("<li>비표준 컬럼 존재 : #{no} 번째 라인</li>".format(no=no))

        col_list.append(no)
        col_list.append(column_id)
        col_list.append(kor_tab_name)
        col_list.append(eng_tab_name)
        col_list.append(kor_col_name)
        col_list.append(pk)
        col_list.append(not_null)
        col_list.append(default)
        col_list.append(eng_col_name)
        col_list.append(data_type)
        col_list.append(domain)
        col_list.append(stardard_yn)
        col_list.append(table_chg)
        col_list.append(tab_start_yn)
        col_list.append(uk)

        meta_list.append(col_list)
        
        no += 1
        column_id += 1    

    
    ddl = makeDDL(meta_list, db_case=db_case, owner=owner)
    

    context = {
        'meta_list': meta_list,
        'table_cnt': table_cnt,
        'err_message': err_message,
        'ddl': ddl,
        'message':message,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")    

def ajaxDatabaseListSyncWithGoogleSheet(request):

    sheet_name = "DB서버리스트(마스터)"

    message = "동기화 성공"



    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive',
    ]    

    json_file_name = settings.BASE_DIR + '/static/custom/json/gs-db-list-c659a42a94c6.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
    gc = gspread.authorize(credentials)
    spreadsheet_url = 'https://docs.google.com/spreadsheets/d/15sLsy8eW42Hk0kPA3baFRCTLJbQyqDiPjScfzzJEMDY/edit#gid=140860747'
    # 스프레스시트 문서 가져오기
    doc = gc.open_by_url(spreadsheet_url)
    # 시트 선택하기
    # worksheet = doc.worksheet('DDL정보입력(시트이름수정불가)')
    worksheet = ""

    try:
        worksheet = doc.worksheet(sheet_name)
    except:
        message = '"DB서버리스트(마스터)" 시트가 존재하지 않습니다.'
        return JsonResponse({"message": message})        

    all_meta_list = worksheet.get_all_values()    

    # print(all_meta_list[2][0])
    l_row = []

    for row in all_meta_list[2:] :
        l_row.append(row[0])
        if not row[0].isdigit() :
            print("##################")
            print(row[0])
            message = 'A열에 숫자가 아닌 값이 존재합니다.'
            return JsonResponse({"message": message})        


    if len(l_row) != len(list(set(l_row))) :
        message = 'A열에 중복된 값이 존재합니다.'
        return JsonResponse({"message": message})                

    exist_col = []
    exist_col_v = []

    # 존재하는 열 인덱스 get
    for row in all_meta_list[0:] :
        for i, v in enumerate(row) :
            if i > 0 :
                if len(v) > 0 :
                    exist_col.append(i)
                    exist_col_v.append(v)

        break

    # print(exist_col)

    d_row = {}
    d_in = {}
    

    for row in all_meta_list[2:] :

        d_in = {}

        for i, v in enumerate(exist_col) :
            d_in[exist_col_v[i]] = row[v]            

        d_row[row[0]] = d_in

        

    for k, v_list in d_row.items() :

        
        
        dblist = DbList.objects.filter(db_use=v_list['sync_db_use']).first()
    
        if dblist :

            
            obj, flag = DbDetail.objects.get_or_create(id = k)            

            if flag:  ## 생성            
            
                print("1")
                id_dbposition = None

                if v_list['sync_db_pos'] == "IDC(인천)" :
                    id_dbposition = 1
                elif v_list['sync_db_pos'] == "클라우드(AWS)" :
                    id_dbposition = 2
                elif v_list['sync_db_pos'] == "클라우드(Azure)" :
                    id_dbposition = 3
                elif v_list['sync_db_pos'] == "클라우드(NCP)" :
                    id_dbposition = 4    
                elif v_list['sync_db_pos'] == "클라우드(Toast)" :
                    id_dbposition = 5    
                elif v_list['sync_db_pos'] == "클라우드(GCP)" :
                    id_dbposition = 6    

                oper_cd = None
                if v_list['sync_oper_cd'] == "현행화" :
                    oper_cd = 2
                elif v_list['sync_oper_cd'] == "운영" :
                    oper_cd = 3
                elif v_list['sync_oper_cd'] == "개발" :
                    oper_cd = 4
                elif v_list['sync_oper_cd'] == "테스트" :
                    oper_cd = 5                

                id_hacase = None
                if v_list['sync_hacase'] == "Primary" :
                    id_hacase = 1
                elif v_list['sync_hacase'] == "Secondary" :
                    id_hacase = 2
                elif v_list['sync_hacase'] == "DRBD" :
                    id_hacase = 3
                elif v_list['sync_hacase'] == "HA" :
                    id_hacase = 4    

                obj.db_order = 1

                obj.id = k

                print(DbList.objects.get(id=dblist.id))
                obj.id_dblist = DbList.objects.get(id=dblist.id)
                obj.use_yn = 1
                obj.alert_yn = 1
                obj.oper_cd = OperCd.objects.get(id=oper_cd)
                obj.db_nm = v_list['sync_db_nm']
                obj.inst_nm = v_list['sync_db_nm']
                obj.host_nm = v_list['sync_hostnm']
                obj.id_hacase = HaCase.objects.get(id=id_hacase)
                # obj.id_monitoritemlist
                obj.svr_ip = v_list['sync_srv_ip']
                obj.svc_ip = v_list['sync_svc_ip']
                obj.port = v_list['sync_db_port']
                obj.reg_dtm = datetime.datetime.now()
                obj.mod_dtm = datetime.datetime.now()                

                obj.id_dbposition = DbPosition.objects.get(id=id_dbposition)

                obj.sync_db_pos = v_list['sync_db_pos']
                obj.sync_oper_cd = v_list['sync_oper_cd']
                obj.sync_hacase = v_list['sync_hacase']
                obj.sync_db_order = v_list['sync_db_order']                
                obj.sync_hostnm = v_list['sync_hostnm']
                obj.sync_srv_ip = v_list['sync_srv_ip']
                obj.sync_svc_ip = v_list['sync_svc_ip']
                obj.sync_db_port = v_list['sync_db_port']
                obj.sync_db_nm = v_list['sync_db_nm']
                obj.sync_db_type = v_list['sync_db_type']
                obj.sync_db_edition = v_list['sync_db_edition']
                obj.sync_db_version = v_list['sync_db_version']
                obj.sync_db_eos = v_list['sync_db_eos']
                obj.sync_os_version = v_list['sync_os_version']
                obj.sync_os_patch = v_list['sync_os_patch']
                obj.sync_os_cpu_type = v_list['sync_os_cpu_type']
                obj.sync_os_cpu_performance = v_list['sync_os_cpu_performance']
                obj.sync_os_server_active = v_list['sync_os_server_active']
                obj.sync_os_cpu_socket = v_list['sync_os_cpu_socket']
                obj.sync_os_cpu_socket_core = v_list['sync_os_cpu_socket_core']
                obj.sync_os_cpu_total_core = v_list['sync_os_cpu_total_core']
                obj.sync_os_real_use_core = v_list['sync_os_real_use_core']
                obj.sync_os_thread_factor = v_list['sync_os_thread_factor']
                obj.sync_os_real_use_thread = v_list['sync_os_real_use_thread']
                obj.sync_license_std = v_list['sync_license_std']
                obj.sync_license_use = v_list['sync_license_use']
                obj.sync_os_memory = v_list['sync_os_memory']
                obj.sync_license_charge_yn = v_list['sync_license_charge_yn']
                obj.sync_server_manage = v_list['sync_server_manage']
                obj.save()

                id_dbtype = dblist.id_dbtype                



                monitoritemlist = ""
                # Secondary
                if id_hacase == 2 :
                    monitoritemlist = MonitorItemList.objects.filter(id_dbtype__db_type=id_dbtype.db_type)
                else :
                    monitoritemlist = MonitorItemList.objects.filter(id_dbtype__db_type=id_dbtype.db_type, id_hacase=HaCase.objects.get(id=id_hacase)) 


                # for item in monitoritemlist :

                #     new_obj = DBDetailMonitorItemList(
                #         id_dbdetail= obj,
                #         id_monitoritemlist= item,
                #         compare_case= item.compare_case,
                #         limit_value= item.limit_value,
                #         risk_level= item.risk_level,
                #         reg_dtm= datetime.datetime.now(),
                #         id_reg_user= request.user,
                #     )
                #     new_obj.save()
            
            else:  # 존재

                obj.mod_dtm = datetime.datetime.now()

                obj.sync_db_pos = v_list['sync_db_pos']
                obj.sync_oper_cd = v_list['sync_oper_cd']
                obj.sync_hacase = v_list['sync_hacase']
                obj.sync_db_order = v_list['sync_db_order']                
                obj.sync_hostnm = v_list['sync_hostnm']
                obj.sync_srv_ip = v_list['sync_srv_ip']
                obj.sync_svc_ip = v_list['sync_svc_ip']
                obj.sync_db_port = v_list['sync_db_port']
                obj.sync_db_nm = v_list['sync_db_nm']
                obj.sync_db_type = v_list['sync_db_type']
                obj.sync_db_edition = v_list['sync_db_edition']
                obj.sync_db_version = v_list['sync_db_version']
                obj.sync_db_eos = v_list['sync_db_eos']
                obj.sync_os_version = v_list['sync_os_version']
                obj.sync_os_patch = v_list['sync_os_patch']
                obj.sync_os_cpu_type = v_list['sync_os_cpu_type']
                obj.sync_os_cpu_performance = v_list['sync_os_cpu_performance']
                obj.sync_os_server_active = v_list['sync_os_server_active']
                obj.sync_os_cpu_socket = v_list['sync_os_cpu_socket']
                obj.sync_os_cpu_socket_core = v_list['sync_os_cpu_socket_core']
                obj.sync_os_cpu_total_core = v_list['sync_os_cpu_total_core']
                obj.sync_os_real_use_core = v_list['sync_os_real_use_core']
                obj.sync_os_thread_factor = v_list['sync_os_thread_factor']
                obj.sync_os_real_use_thread = v_list['sync_os_real_use_thread']
                obj.sync_license_std = v_list['sync_license_std']
                obj.sync_license_use = v_list['sync_license_use']
                obj.sync_os_memory = v_list['sync_os_memory']
                obj.sync_license_charge_yn = v_list['sync_license_charge_yn']
                obj.sync_server_manage = v_list['sync_server_manage']
                obj.save()

        else :                              
            message = "에러발생\n서비스 레벨의 DB가 존재하지 않음 : "  +  str(v_list['sync_db_use'])
            return JsonResponse({"message": message})        



    # x_len = len(all_meta_list[4])

    # err_message = []

    # meta_dict = {}
    # meta_list = []

    # TOTAL_DDL = ""
    # tab_start_yn = ""

    # befor_table_name = "START"

    # owner = ""
    # if all_meta_list[0][1] != "":
    #     owner = all_meta_list[0][1].lower() + "."
    # data_tablespace = all_meta_list[1][1]
    # index_tablespace = all_meta_list[2][1]

    # first_table_name = "NEW|START"

    # no = 1
    # column_id = 1
    # table_cnt = 0
    # table_chg = 1



    # for meta in all_meta_list[6:]:

    #     i = 0
    #     meta_dict = {}    
    #     col_list = []    
    #     stardard_yn = ""

    #     if first_table_name != meta[1]:
    #         tab_start_yn = 'Y'
    #         first_table_name = meta[1]
    #         column_id = 1
    #         table_cnt = table_cnt + 1

    #         if table_chg == 1 :
    #             table_chg = 0
    #         else :
    #             table_chg = 1
    #     else:
    #         tab_start_yn = 'N'

    #     sql = """SELECT A.STD_ATTR_ENG, LOWER(B.MYSQL_DATATYPE) DATA_TYPE, B.INFO_TYPE
    #       from cust_std_attr a,
    #            cust_std_domain b
    #       where b.id = a.id_stddomain
    #       and a.std_attr_kor = '{kor_column_name}'
    #       and   a.use_yn = 1
    #       and   b.use_yn = 1;
    #   """.format(kor_column_name=meta[2])

    #     std_info = _query_dict("default", sql)

    #     eng_col_name = ""
    #     domain = ""
    #     data_type = ""

    #     if len(std_info) > 0:
    #         eng_col_name = std_info[0]['STD_ATTR_ENG'].lower()
    #         domain = std_info[0]['INFO_TYPE'].lower()
    #         data_type = std_info[0]['DATA_TYPE'].lower()    
    #         stardard_yn = '<font color="blue">표준</font>'

    #     eng_tab_name = meta[1].lower()
    #     kor_tab_name = meta[0].replace(' ','')
    #     kor_col_name = meta[2]
    #     pk = meta[3].upper()
    #     not_null = meta[4].upper()
    #     default = meta[5].lower()                    

    #     # if column_id == 1 and pk != "Y":
    #     #     err_message.append("<li>PK없음 : #{no} 번째 라인</li>".format(no=no))

    #     if len(std_info) == 0:        
    #         stardard_yn = '<b><font color="red">비표준</font></b>'
    #         err_message.append("<li>비표준 컬럼 존재 : #{no} 번째 라인</li>".format(no=no))

    #     col_list.append(no)
    #     col_list.append(column_id)
    #     col_list.append(kor_tab_name)
    #     col_list.append(eng_tab_name)
    #     col_list.append(kor_col_name)
    #     col_list.append(pk)
    #     col_list.append(not_null)
    #     col_list.append(default)
    #     col_list.append(eng_col_name)
    #     col_list.append(data_type)
    #     col_list.append(domain)
    #     col_list.append(stardard_yn)
    #     col_list.append(table_chg)
    #     col_list.append(tab_start_yn)
        
    #     meta_list.append(col_list)
        
    #     no += 1
    #     column_id += 1    

    # ddl = makeDDL(meta_list)
    

    context = {        
        'message':message,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")     

def ajaxMetaHandWork(request):
    script = request.POST.get('script', '')

    error_yn = 'N'
    error_msg = ''

    obj = script.split('\n')

    if (len(obj) >= 200):
        error_yn = 'Y'
        error_msg = '컬럼 개수가 200개를 넘을 수 없습니다. 필요한 경우 DBA에게 문의주세요.'

    data = []

    column_id = 1

    for o in obj:

        line = [column_id] + o.split()

        if len(line) == 2:
            line = line + [''] + ['']
        elif len(line) == 3:
            line = line + ['']

        if (len(line) == 4):
            data.append(line)
            column_id += 1

    context = {
        'data': data,
        'error_yn': error_yn,
        'error_msg': error_msg,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxTuningListDVModifyTuningStatus(request):
    obj_id = request.POST.get('obj_id', '')
    id_tuningstatus = request.POST.get('id_tuningstatus', '')

    tuninglist = TuningList.objects.get(id=obj_id)

    tuninglist.id_tuningstatus = TuningStatus.objects.get(id=id_tuningstatus)
    tuninglist.mod_dtm = datetime.datetime.now()

    tuninglist.save()

    context = {
        'success': 'ok',
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxAllExecuteSql(request):
    sql = request.POST.get('sqltext', '')
    sql_id = request.POST.get('sql_id', '')
    id_dbdetail = request.POST.get('id_dbdetail', '')

    sql2 = sql.replace("\n", "")

    if sql_id:
        obj = DataList.objects.get(id=sql_id)
        sql1 = obj.sql_text.replace("\r\n", "")
        if sql1 != sql2:
            sql_id = ""


    row_count_limit = 2000

    if id_dbdetail != "":

        error_yn = "N"
        message = ""

        org_sql = sql

        sql = sql.replace(";", "")

        sql = """SELECT *
  FROM (
    """ + \
              sql + \
              """
  )
  where rownum <= {row_count_limit}""".format(row_count_limit=row_count_limit)

        obj = DbDetail.objects.get(id=id_dbdetail)

        with connections[obj.inst_nm].cursor() as cursor:

            new_userexecutesql = UserExecuteSQL(
                sql_text=org_sql,
                id_reg_user=request.user,
                reg_dtm=datetime.datetime.now(),
                id_datalist=sql_id,
                id_dbdetail=obj,
            )

            new_userexecutesql.save()

            save_id = new_userexecutesql.id

            #####################################################
            ## 수행 시간 측정
            #####################################################
            start = time.time()

            cursor.execute(sql)

            end = time.time() - start
            #####################################################

            new_userexecutesql = UserExecuteSQL.objects.get(id=save_id)
            new_userexecutesql.execute_time = end

            new_userexecutesql.save()

            columnNames = [[d[0]] for d in cursor.description]

            data = cursor.fetchall()

            row_count = len(data)

            data2 = []

            ###############################
            ## Decimal('53.94') -> str 로 변경
            ###############################
            buff1 = []
            buff2 = []

            ind = 0

            for ds in data:
                for d in ds:

                    if isinstance(d, Decimal):
                        buff1.append(str(d))
                    elif isinstance(d, int) and columnNames[ind][0][-2:].upper() != 'CD' \
                            and columnNames[ind][0][-2:].upper() != 'ID' \
                            and columnNames[ind][0][-2:].upper() != '번호' \
                            and columnNames[ind][0][-2:].upper() != 'NO' \
                            and columnNames[ind][0][-2:] != '코드' \
                            and columnNames[ind][0][-4:] != '고객번호' \
                            and columnNames[ind][0][-6:] != 'CUSTNO' \
                            and columnNames[ind][0][-7:] != 'CUST_NO':
                        buff1.append(format(d, ','))
                    else:
                        buff1.append(d)
                    ind = ind + 1

                buff2.append(buff1)
                buff1 = []
                ind = 0

            data = buff2

        if len(columnNames) > 35:
            error_yn = "Y"
            message = """[안내] 조회되는 컬럼 수가 너무 많습니다. [SELECT *] 보다는 컬럼명을 명시해주세요 """

        if error_yn == "N":
            if row_count == row_count_limit:
                message = """성능 확보를 위해 조회 건수는 {row_count_limit}건으로 제한 됩니다.""".format(row_count_limit=row_count_limit)

        prohibition_tab = ['CSR_CST_CUST_ADDR_D', ]

        for tab in prohibition_tab:

            if tab in sql.upper():
                error_yn = "Y"
                message = """[안내] 해당 테이블은 조회할 수 없습니다. """
                break




    else:
        data = ""
        columnNames = ""
        end = 0
        error_yn = 'Y'
        message = "[에러] Database를 먼저 선택하세요"

    context = {
        'data': data,
        'columns': columnNames,
        'time': end,
        'error_yn': error_yn,
        'message': message
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxDatabaseDetail(request):
    id_dblist = request.POST.get('id_dblist', '')

    obj = DbDetail.objects.filter(id_dblist=id_dblist).order_by("-oper_cd", "db_order")

    dblist = []

    # db_dict = defaultdict(list)

    # for upper_domain_name, \
    #     db_use, \
    #     id \
    #     in db_dataset:
    #     db_dict[upper_domain_name].append([db_use,str(id)])

    # db_dict = dict(db_dict)

    all_dblist = []

    # id_dblist = ['DB용도']
    oper_cd = ['운영/개발']
    db_order = ['노드번호']
    db_nm = ['DB명']
    inst_nm = ['인스턴스명']
    host_nm = ['HOSTNAME']
    svr_ip = ['서버IP']
    svc_ip = ['서비스IP']
    port = ['서비스PORT']
    tns_desc = ['접속정보']

    for o in obj:

        # id_dblist.append(o.id_dblist.db_use)
        db_order.append(o.db_order)
        oper_cd.append(o.oper_cd.oper_cd)
        db_nm.append(o.db_nm)
        inst_nm.append(o.inst_nm)
        host_nm.append(o.host_nm)
        svr_ip.append(o.svr_ip)
        svc_ip.append(o.svc_ip)
        port.append(o.port)

        tns = ""
        if o.tns_desc:
            tns = o.tns_desc.upper()
            tns = tns.replace(' ', '').replace('(DESCRIPTION', '<br>(DESCRIPTION').replace('(HOST',
                                                                                           '<br>(HOST').replace(
                '(SERVICE_NAME', '<br>(SERVICE_NAME')
        else:
            tns = ""
        tns_desc.append(tns)

    # all_dblist.append(id_dblist)
    all_dblist.append(oper_cd)
    all_dblist.append(db_order)
    all_dblist.append(db_nm)
    all_dblist.append(inst_nm)
    all_dblist.append(host_nm)
    all_dblist.append(svr_ip)
    all_dblist.append(svc_ip)
    all_dblist.append(port)
    all_dblist.append(tns_desc)

    context = {
        'result': "ok",
        'all_dblist': all_dblist
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxSaveDbaScript(request):
    id_metareq = request.POST.get('id_metareq', '')
    sql = request.POST.get('sql', '')

    meta = MetaReq.objects.get(id=id_metareq)
    meta.script = sql
    meta.script_dtm = datetime.datetime.now()
    meta.id_script_reg_user = request.user
    meta.save()

    script_dtm = str(datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S"))

    context = {
        'script_dtm': script_dtm,
        'script_reg_user': request.user.first_name
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")

def ajaxDbAlertOff(request):
    id_dbdetail = request.POST.get('id_dbdetail', '')

    dbdetail = DbDetail.objects.get(id=id_dbdetail)

    curr_alert_yn = dbdetail.alert_yn

    if curr_alert_yn == '1' :
        dbdetail.alert_yn = '0' 
    else :
        dbdetail.alert_yn = '1' 

    dbdetail.save()

    context = {
        'result': 'ok',
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")    


def ajaxMetaCreate(request):
    data_tbs = request.POST.get('data_tbs', '')
    ind_tbs = request.POST.get('ind_tbs', '')
    id_metareq = request.POST.get('meta_req', '')
    grant = request.POST.get('grant', '')
    owner = request.POST.get('owner', '')
    id_tablelist = request.POST.get('id_tablelist', '')

    id_tablelist_table_name = ""
    id_tablelist_owner = ""

    if id_tablelist:
        id_tablelist_obj = TableList.objects.get(id=id_tablelist)
        id_tablelist_table_name = id_tablelist_obj.table_name
        id_tablelist_owner = id_tablelist_obj.owner

    if data_tbs:
        data_tbs = DBTablespace.objects.get(id=data_tbs)
        data_tbs = data_tbs.tablespace_name
    else:
        data_tbs = "[data_tablespace]"

    if ind_tbs:
        ind_tbs = DBTablespace.objects.get(id=ind_tbs)
        ind_tbs = ind_tbs.tablespace_name
    else:
        ind_tbs = "[index_tablespace]"

    if owner:
        owner = DBOwner.objects.get(id=owner)
        owner = owner.owner + "."
    else:
        if id_tablelist:
            owner = id_tablelist_owner + "."
        else:
            owner = "[owner]."

    obj_meta = MetaReq.objects.get(id=id_metareq)

    table_name = ""
    table_comments = ""
    TOTAL_DDL = ""
    # owner = ""

    if obj_meta.id_tablelist:
        table_name = obj_meta.id_tablelist.table_name
        table_comments = obj_meta.id_tablelist.comments
        # owner = obj_meta.id_tablelist.owner + "."

    else:
        table_name = obj_meta.table_name
        table_comments = obj_meta.table_comments
        # owner = "[owner]."

    if obj_meta.id_dblist.id_dbtype.db_type == 'ORACLE' :  # ORACLE

        DDL_STR_CREATE_TABLE = "CREATE TABLE {owner}{table_name} ( \n"
        DDL_END_CREATE_TABLE = ")"
        DDL_IN_TABLE_COLUMN = "\t{column_name} {data_type}{data_default}{not_null}"
        DDL_DATA_TABLESPACE = " TABLESPACE {data_tbs};"
        DDL_IND_TABLESPACE = " TABLESPACE {ind_tbs};"
        DDL_STR_CREATE_PK = "CREATE UNIQUE INDEX {owner}PK_{table_name} ON {owner}{table_name} ( \n"
        DDL_END_CREATE_PK = ")"

        DDL_STR_CREATE_CONST_PK = "ALTER TABLE {owner}{table_name} ADD CONSTRAINT PK_{table_name} PRIMARY KEY ( \n"
        DDL_END_CREATE_CONST_PK = ");"

        DDL_IN_INDEX_COLUMN = "{pk_index_column}"

        DDL_SYNONYM = "CREATE OR REPLACE PUBLIC SYNONYM {table_name} FOR {owner}{table_name};"

        if owner == "SDHUB_OWN.":
            DDL_SYNONYM += "\n" + "CREATE OR REPLACE SYNONYM SDHUB_ETL.{table_name} FOR {owner}{table_name};"

        DDL_GRANT = "GRANT {crud} ON {owner}{table_name} TO {role};\n"

        DDL_TABLE_COMMENT = "COMMENT ON TABLE {owner}{table_name} IS '{comment}';"
        DDL_COLUMN_COMMENT = "COMMENT ON COLUMN {owner}{table_name}.{column_name} IS '{comment}';"

        ERD_GRANT = "-- ERD 권한\nGRANT SELECT ON {owner}{table_name} TO {erd_user};\n"
        erd_user = ""

        
        if obj_meta.id_domainanddblist.erd_user :
            ERD_GRANT = ERD_GRANT.format(owner=owner, table_name=table_name, erd_user=obj_meta.id_domainanddblist.erd_user)
        else :
            ERD_GRANT = "-- ERD 계정이 생성되어 있지 않음"            

        ddl_grants = ""
        if grant:
            grant = MetaGrantDetail.objects.filter(id_metagrantlist=grant)
            for g in grant:
                ddl_grants = ddl_grants + DDL_GRANT.format(crud=g.crud, owner=owner, table_name=table_name, role=g.role)

        else:
            ddl_grants = "[grant]"


        obj = MetaReqList.objects.filter(id_metareq=id_metareq).order_by('column_id','id')

        pk_index_column = ""

        for o in obj:
            if o.change_list in ['+ ┌ ASIS', '+ 삭제']:
                continue

            if o.pk_yn == "1":
                pk_index_column = pk_index_column + "\t" + o.column_name + ",\n"

        pk_index_column = pk_index_column + "|"
        pk_index_column = pk_index_column.replace(",\n|", "\n")

        table_column = ""
        table_column_comment = ""

        for o in obj:

            if o.change_list in ['+ ┌ ASIS', '+ 삭제']:
                continue

            default_keyword = ['SYSDATE', ]
            org_default = o.data_default.rstrip(' ')
            org_default = org_default.lstrip(' ')
            in_default = ""

            if org_default.upper() in default_keyword:
                in_default = " DEFAULT '" + o.data_default.replace("'", "") + "'"
            elif org_default != "":
                in_default = " DEFAULT '" + o.data_default.replace("'", "") + "'"

            table_column = table_column + DDL_IN_TABLE_COLUMN.format(
                column_name=o.column_name,
                data_type=o.data_type,
                not_null=' NOT NULL' if o.not_null == "1" else '',
                data_default=in_default,
            ) + ",\n"
            table_column_comment = table_column_comment + DDL_COLUMN_COMMENT.format(
                owner=owner,
                table_name=table_name,
                column_name=o.column_name,
                comment=o.col_comments,
            ) + "\n"

        table_column = table_column + "|"

        table_column = table_column.replace(",\n|", "\n")

        table_column_comment = table_column_comment + "|"
        table_column_comment = table_column_comment.replace("\n|", "\n")

        TOTAL_DDL = DDL_STR_CREATE_TABLE.format(owner=owner, table_name=table_name) + \
                    table_column + \
                    DDL_END_CREATE_TABLE + \
                    DDL_DATA_TABLESPACE.format(data_tbs=data_tbs) + \
                    "\n\n" + \
                    DDL_STR_CREATE_PK.format(owner=owner, table_name=table_name) + \
                    pk_index_column + \
                    DDL_END_CREATE_PK + \
                    DDL_IND_TABLESPACE.format(ind_tbs=ind_tbs) + \
                    "\n\n" + \
                    DDL_STR_CREATE_CONST_PK.format(owner=owner, table_name=table_name) + \
                    pk_index_column + \
                    DDL_END_CREATE_CONST_PK + \
                    "\n\n" + \
                    DDL_SYNONYM.format(owner=owner, table_name=table_name) + \
                    "\n\n" + \
                    ddl_grants + \
                    "\n" + \
                    ERD_GRANT + \
                    "\n\n" + \
                    DDL_TABLE_COMMENT.format(owner=owner, table_name=table_name, comment=table_comments) + \
                    "\n" + \
                    table_column_comment

    
    elif obj_meta.id_dblist.id_dbtype.db_type in ['MYSQL', 'MARIADB'] : # MYSQL, MARIADB

        DDL_STR_CREATE_TABLE = "create table {owner}{table_name} ( \n"
        DDL_END_CREATE_TABLE = ")\ncomment '{comment}' \nengine=innodb default charset=utf8mb4 collate=utf8mb4_general_ci;"
        DDL_IN_TABLE_COLUMN = "\t{column_name} {data_type}{not_null}{auto_increment}{data_default} comment '{comment}'"
        DDL_CREATE_PK = "\tprimary key ({pk_index_column})\n"

        obj = MetaReqList.objects.filter(id_metareq=id_metareq).order_by('column_id', 'id')

        table_column = ""
        table_column_comment = ""
        pk_index_column = ""
        pk_exists = "N"

        for o in obj:

            if o.change_list in ['+ ┌ ASIS', '+ 삭제']:
                continue

            if o.pk_yn == "1":
                pk_index_column = pk_index_column + o.column_name + ", "
                pk_exists = "Y"

            default_keyword = ['NOW()', 'CURRENT_TIMESTAMP']
            org_default = o.data_default.rstrip(' ')
            org_default = org_default.lstrip(' ')
            in_default = ""

            if org_default.upper() in default_keyword:
                in_default = " default " + o.data_default
            elif org_default != "":
                in_default = " default '" + o.data_default + "'"

            table_column = table_column + DDL_IN_TABLE_COLUMN.format(
                column_name=o.column_name.lower(),
                data_type=o.data_type.lower(),
                not_null=' not null' if o.not_null == "1" else '',
                data_default=in_default,
                comment=o.col_comments,
                auto_increment=' auto_increment' if o.column_name.lower() == "id" else '', ) + ",\n"

        if pk_exists == "N":
            table_column = table_column + "|"
            table_column = table_column.replace(",\n|", "\n")

        pk_index_column = pk_index_column + "|"
        pk_index_column = pk_index_column.replace(", |", "")

        TOTAL_DDL = DDL_STR_CREATE_TABLE.format(owner=owner, table_name=table_name) + \
                    table_column + \
                    DDL_CREATE_PK.format(pk_index_column=pk_index_column) + \
                    DDL_END_CREATE_TABLE.format(comment=table_comments)

    context = {
        'result': 'ok',
        'total_ddl': TOTAL_DDL
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxMetaReqOK(request):
    comment = request.POST.get('comment', '')
    div_ok = request.POST.get('div_ok', '')
    dist_dtm = request.POST.get('dist_dtm', '')
    id_metareq = request.POST.get('id_metareq', '')
    ok_yn = request.POST.get('ok_yn', '')

    ok_yn_msg = ''
    error_msg = ''

    if ok_yn == '1':
        ok_yn_msg = "반영"
    elif ok_yn == '2':
        ok_yn_msg = "반려"

    obj = MetaReq.objects.get(id=id_metareq)

    sabun = obj.id_pl_prod.username

    if comment == "":
        comment = "승인합니다"

    # 39 : 데이터허브
    only_prod_db_list = [39, ]

    if div_ok == "pl_dev":
        obj.id_pl_dev = request.user
        obj.pl_dev_comment = comment
        obj.pl_dev_yn = ok_yn
        obj.pl_dev_dtm = datetime.datetime.now()
    elif div_ok == "da_dev":
        obj.id_da_dev = request.user
        obj.da_dev_comment = comment
        obj.da_dev_yn = ok_yn
        obj.da_dev_dtm = datetime.datetime.now()
    elif div_ok == "dba_dev":

        obj.id_dba_dev = request.user

        if comment == "승인합니다":
            comment = "개발DB 반영 완료 되었습니다"

        obj.dba_dev_comment = comment
        obj.dba_dev_yn = ok_yn
        obj.dba_dev_dtm = datetime.datetime.now()

        if obj.id_dblist.id in only_prod_db_list:
            if ok_yn == '1':
                # obj.id_pl_prod = request.user
                obj.pl_prod_comment = "운영DB만 있으므로 자동 승인"
                obj.pl_prod_yn = ok_yn
                obj.pl_prod_dtm = datetime.datetime.now()
                obj.id_dba_prod = request.user
                obj.dba_prod_comment = "운영DB만 있으므로 자동 승인"
                obj.dba_prod_yn = ok_yn
                obj.dba_prod_dtm = datetime.datetime.now()
                # obj.id_req_prod = request.user
                obj.req_prod_comment = "운영DB만 있으므로 자동 승인"
                obj.req_prod_yn = ok_yn
                obj.req_prod_dtm = datetime.datetime.now()
            else:
                obj.script = None
                obj.script_dtm = None

            ds_emp = _query_dict('iam', iamSql.format(emp_no=obj.id_reg_user.username))[0]
            msg = """[SmartDBA] 변경 요청건이 DB에 {ok_yn_msg}되었습니다. 메타번호: {no}""".format(no=id_metareq,
                                                                                   ok_yn_msg=ok_yn_msg)
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])
        else:
            ds_emp = _query_dict('iam', iamSql.format(emp_no=obj.id_reg_user.username))[0]

            if ok_yn == '1':
                msg = """[SmartDBA] 변경 요청건이 개발DB에 반영되었습니다. 메타번호: {no}""".format(no=id_metareq, )
            else:
                msg = """[SmartDBA] 변경 요청건이 반려되었습니다. 반려 사유 확인부탁드립니다. 메타번호: {no}""".format(no=id_metareq, )

            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])


    elif div_ok == "pl_prod":
        obj.id_pl_prod = request.user
        obj.pl_prod_comment = comment
        obj.pl_prod_yn = ok_yn
        obj.pl_prod_dtm = datetime.datetime.now()
    elif div_ok == "dba_prod":
        obj.id_dba_prod = request.user

        if comment == "승인합니다":
            comment = "운영DB 반영 완료 되었습니다"
        obj.dba_prod_comment = comment
        obj.dba_prod_yn = ok_yn
        obj.dba_prod_dtm = datetime.datetime.now()
    elif div_ok == "req_prod":
        obj.id_req_prod = request.user

        if comment == "승인합니다":
            comment = "운영DB 반영 요청 합니다"

        # 테이블일 경우에만 미신청 건 유무 확인
        if obj.obj_class == '2':

            not_prov_meta_yn = MetaReq.objects.filter(Q(id_tablelist__isnull=False), 
                                                      Q(id_tablelist=obj.id_tablelist),
                                                      Q(req_prod_yn='0'), 
                                                      Q(id__lt=obj.id),
                                                      ~Q(dba_dev_yn='2'),
                                                      ).exists()


            if not_prov_meta_yn:
                not_prov_meta = MetaReq.objects.filter(Q(id_tablelist__isnull=False), 
                                                       Q(id_tablelist=obj.id_tablelist),
                                                       Q(req_prod_yn='0'), 
                                                       Q(id__lt=obj.id),
                                                       ~Q(dba_dev_yn='2'),
                                                       )                

                not_prov_id = ""
                for not_prov in not_prov_meta:
                    not_prov_id += "[" + str(not_prov.id) + "번] "

                error_msg = "순서대로 요청해주세요.\n선 요청이 필요한 메타번호 : {not_prov_id}".format(not_prov_id=not_prov_id)

        obj.req_prod_comment = comment
        obj.req_prod_yn = ok_yn
        obj.dist_dtm = dist_dtm
        obj.req_prod_dtm = datetime.datetime.now()

        obj.pl_prod_yn = '0'
        obj.dba_prod_yn = '0'

    if error_msg == '':
        obj.save()

        if div_ok == "req_prod":
            ds_emp = _query_dict('iam', iamSql.format(emp_no=sabun))[0]
            msg = "[SmartDBA] 테이블 운영DB 반영 요청건이 있습니다. 승인해주세요. 메타번호: {no}".format(no=id_metareq)

            # SMS 만 발송
            _query_sms(ds_emp['MOBILE'], msg, slack_yn='N', sms_yn='Y', team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

    id_ok_person = str(request.user.first_name)
    ok_dtm = str(datetime.datetime.now().strftime("%m-%d %H:%M"))

    context = {
        'id_ok_person': id_ok_person,
        'ok_dtm': ok_dtm,
        'error_msg': error_msg,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxMetaCancel(request):
    div_cancel = request.POST.get('div_cancel', '')
    id_metareq = request.POST.get('id_metareq', '')

    error_msg = ""

    obj = MetaReq.objects.get(id=id_metareq)

    if div_cancel == "dba_dev":
        obj.id_dba_dev = None
        obj.dba_dev_comment = None
        obj.dba_dev_yn = "0"
        obj.dba_dev_dtm = None

        obj.pl_dev_comment = None
        obj.pl_dev_yn = "0"
        obj.pl_dev_dtm = None

        obj.da_dev_comment = None
        obj.da_dev_yn = "0"
        obj.da_dev_dtm = None


    elif div_cancel == "req_prod":
        obj.id_req_prod = None
        obj.req_prod_comment = None
        obj.req_prod_yn = "0"
        obj.dist_dtm = None
        obj.req_prod_dtm = None

    elif div_cancel == "pl_prod":
        obj.pl_prod_comment = None
        obj.pl_prod_yn = "0"
        obj.pl_prod_dtm = None

    elif div_cancel == "dba_prod":
        obj.id_dba_prod = None
        obj.dba_prod_comment = None
        obj.dba_prod_yn = "0"
        obj.dba_prod_dtm = None

    obj.save()

    context = {
        'error_msg': error_msg,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


# def ajaxMetaCreateColumnDdl(request) :


#   id_metareq = request.POST.get('meta_req','')
#   c_id = request.POST.get('c_id','')
#   table_name = request.POST.get('table_name','')
#   owner = request.POST.get('owner','')
#   id_tablelist = request.POST.get('id_tablelist','')

#   id_tablelist_table_name = ""
#   id_tablelist_owner = ""
#   if id_tablelist :
#     id_tablelist_obj = TableList.objects.get(id=id_tablelist)
#     id_tablelist_table_name = id_tablelist_obj.table_name
#     id_tablelist_owner = id_tablelist_obj.owner

#   obj_meta = MetaReq.objects.get(id=id_metareq)

#   if obj_meta.id_tablelist :
#     table_name = obj_meta.id_tablelist.table_name
#   else :
#     if id_tablelist :
#       table_name = id_tablelist_table_name
#     else :
#       table_name = table_name

#   if owner :
#     owner = DBOwner.objects.get(id=owner)
#     owner = owner.owner + "."
#   else :
#     if id_tablelist :
#       owner = id_tablelist_owner + "."
#     else :
#       owner = "[owner]."


#   obj = MetaReqList.objects.get(id=c_id)


#   not_null = ' NOT NULL' if obj.not_null == "1" else ' NULL'
#   data_default = " DEFAULT '"+obj.data_default.replace("'","")+"'" if obj.data_default.replace("'","") != "" else ''

#   DDL_ALTER_DROP = ""
#   DDL_ALTER_RENAME = ""
#   DDL_ALTER_MODIFY = ""
#   DDL_ALTER_ADD = ""
#   DDL_COLUMN_COMMENT = ""
#   TMP_DDL_ALTER_DROP = ""
#   TMP_DDL_ALTER_RENAME = ""
#   TMP_DDL_ALTER_MODIFY = ""
#   TMP_DDL_ALTER_ADD = ""
#   TMP_DDL_COLUMN_COMMENT = ""
#   column_ddl = ""

#   if obj_meta.id_dblist.id_dbtype.id == 1 : # ORACLE

#     DDL_ALTER_DROP = "ALTER TABLE {owner}{table_name} DROP COLUMN {column_name};\n"
#     DDL_ALTER_RENAME = "ALTER TABLE {owner}{table_name} RENAME COLUMN [org] TO {column_name};\n"
#     DDL_ALTER_MODIFY = "ALTER TABLE {owner}{table_name} MODIFY ( {column_name} {data_type}{data_default}{not_null} );\n"
#     DDL_ALTER_ADD = "ALTER TABLE {owner}{table_name} ADD ( {column_name} {data_type}{data_default}{not_null} );\n"
#     DDL_COLUMN_COMMENT = "COMMENT ON COLUMN {owner}{table_name}.{column_name} IS '{comment}';"

#     TMP_DDL_ALTER_DROP = DDL_ALTER_DROP.format(owner=owner, table_name=table_name, column_name=obj.column_name)
#     TMP_DDL_ALTER_RENAME = DDL_ALTER_RENAME.format(owner=owner, table_name=table_name, column_name=obj.column_name)
#     TMP_DDL_ALTER_MODIFY = DDL_ALTER_MODIFY.format(owner=owner, table_name=table_name, column_name=obj.column_name, data_type=obj.data_type, not_null=not_null, data_default=data_default)
#     TMP_DDL_ALTER_ADD = DDL_ALTER_ADD.format(owner=owner, table_name=table_name, column_name=obj.column_name, data_type=obj.data_type, not_null=not_null, data_default=data_default)
#     TMP_DDL_COLUMN_COMMENT = DDL_COLUMN_COMMENT.format(owner=owner, table_name=table_name, column_name=obj.column_name, comment=obj.col_comments)

#     if obj.change_list == "+ 신규" :
#       column_ddl = TMP_DDL_ALTER_ADD + TMP_DDL_COLUMN_COMMENT
#     elif obj.change_list == "+ 삭제" :
#       column_ddl = TMP_DDL_ALTER_DROP
#     elif obj.change_list == "+ └ TOBE" :

#       before_obj = MetaReqList.objects.filter(id_metareq=obj.id_metareq.id,
#                                               column_id=int(obj.column_id)-1)
#       before_obj = before_obj[0]


#       if obj.column_name.lower() == before_obj.column_name.lower() :
#         column_ddl = TMP_DDL_ALTER_MODIFY + TMP_DDL_COLUMN_COMMENT
#       else :
#         column_ddl = TMP_DDL_ALTER_RENAME.replace("[org]",before_obj.column_name) + TMP_DDL_ALTER_MODIFY + TMP_DDL_COLUMN_COMMENT

#   elif obj_meta.id_dblist.id_dbtype.id in [2,3] : # MYSQL, MARIADB
#     DDL_ALTER_ADD = "ALTER TABLE {owner}{table_name} ADD COLUMN {column_name} {data_type}{data_default}{not_null} COMMENT '{comment}';\n"
#     DDL_ALTER_MODIFY = "ALTER TABLE {owner}{table_name} CHANGE COLUMN [org] {column_name} {data_type}{data_default}{not_null} COMMENT '{comment}';\n"
#     DDL_ALTER_DROP = "ALTER TABLE {owner}{table_name} DROP COLUMN {column_name};\n"


#     TMP_DDL_ALTER_ADD = DDL_ALTER_ADD.format(owner=owner, table_name=table_name, column_name=obj.column_name, data_type=obj.data_type, not_null=not_null, data_default=data_default, comment=obj.col_comments)
#     TMP_DDL_ALTER_MODIFY = DDL_ALTER_MODIFY.format(owner=owner, table_name=table_name, column_name=obj.column_name, data_type=obj.data_type, not_null=not_null, data_default=data_default, comment=obj.col_comments)
#     TMP_DDL_ALTER_DROP = DDL_ALTER_DROP.format(owner=owner, table_name=table_name, column_name=obj.column_name)

#     if obj.change_list == "+ 신규" :
#       column_ddl = TMP_DDL_ALTER_ADD + TMP_DDL_COLUMN_COMMENT
#     elif obj.change_list == "+ 삭제" :
#       column_ddl = TMP_DDL_ALTER_DROP
#     elif obj.change_list == "+ └ TOBE" :
#       before_obj = MetaReqList.objects.filter(id_metareq=obj.id_metareq.id,
#                                               column_id=int(obj.column_id)-1)
#       before_obj = before_obj[0]
#       if obj.column_name.lower() == before_obj.column_name.lower() :
#         column_ddl = TMP_DDL_ALTER_MODIFY.replace("[org]",obj.column_name) + TMP_DDL_COLUMN_COMMENT
#       else :
#         column_ddl = TMP_DDL_ALTER_MODIFY.replace("[org]",before_obj.column_name) + TMP_DDL_COLUMN_COMMENT


#   context = {
#               'result' : 'ok',
#               'column_ddl' : column_ddl
#             }

#   param = json.dumps(context, default=json_default)

#   return HttpResponse(param, content_type="application/json")


def ajaxMetaAlter(request):
    id_metareq = request.POST.get('meta_req', '')
    owner = request.POST.get('owner', '')
    c_id = request.POST.get('c_id', '')

    obj_meta = MetaReq.objects.get(id=id_metareq)

    if obj_meta.id_tablelist:
        table_name = obj_meta.id_tablelist.table_name
    else:
        table_name = obj_meta.table_name
    # if id_tablelist :
    #   table_name = id_tablelist_table_name
    # else :
    #   table_name = table_name

    if owner:
        owner = DBOwner.objects.get(id=owner)
        owner = owner.owner + "."
    else:
        if obj_meta.id_tablelist:
            owner = obj_meta.id_tablelist.owner + "."
        else:
            owner = "[owner]."

    if c_id == "":
        data_obj = MetaReqList.objects.filter(id_metareq=id_metareq).order_by('column_id', 'id')
    else:
        data_obj = MetaReqList.objects.filter(id=c_id)

    total_column_ddl = ""
    column_ddl = ""

    for obj in data_obj:
        if obj.change_list in ['+ 신규', '+ 삭제', '+ └ TOBE']:

            not_null = ' NOT NULL' if obj.not_null == "1" else ' NULL'
            data_default = " DEFAULT '" + obj.data_default.replace("'", "") + "'" if obj.data_default.replace("'",
                                                                                                              "") != "" else ''

            DDL_ALTER_DROP = ""
            DDL_ALTER_RENAME = ""
            DDL_ALTER_MODIFY = ""
            DDL_ALTER_ADD = ""
            DDL_COLUMN_COMMENT = ""
            TMP_DDL_ALTER_DROP = ""
            TMP_DDL_ALTER_RENAME = ""
            TMP_DDL_ALTER_MODIFY = ""
            TMP_DDL_ALTER_ADD = ""
            TMP_DDL_COLUMN_COMMENT = ""

            DDL_REMARK = ""
            if len(obj.change_reason) > 0:
                DDL_REMARK = "/* " + obj.change_reason + " */" + "\n"

            
            exec_log(obj_meta.id_dblist.id_dbtype.db_type)
            if obj_meta.id_dblist.id_dbtype.db_type == 'ORACLE':  # ORACLE

                DDL_ALTER_DROP = "ALTER TABLE {owner}{table_name} DROP COLUMN {column_name};\n"
                DDL_ALTER_RENAME = "ALTER TABLE {owner}{table_name} RENAME COLUMN [org] TO {column_name};\n"
                DDL_ALTER_MODIFY = "ALTER TABLE {owner}{table_name} MODIFY ( {column_name} {data_type}{data_default}{not_null} );\n"
                DDL_ALTER_ADD = "ALTER TABLE {owner}{table_name} ADD ( {column_name} {data_type}{data_default}{not_null} );\n"
                DDL_COLUMN_COMMENT = "COMMENT ON COLUMN {owner}{table_name}.{column_name} IS '{comment}';"

                TMP_DDL_ALTER_DROP = DDL_ALTER_DROP.format(owner=owner, table_name=table_name,
                                                           column_name=obj.column_name)
                TMP_DDL_ALTER_RENAME = DDL_ALTER_RENAME.format(owner=owner, table_name=table_name,
                                                               column_name=obj.column_name)
                TMP_DDL_ALTER_MODIFY = DDL_ALTER_MODIFY.format(owner=owner, table_name=table_name,
                                                               column_name=obj.column_name, data_type=obj.data_type,
                                                               not_null=not_null, data_default=data_default)
                TMP_DDL_ALTER_ADD = DDL_ALTER_ADD.format(owner=owner, table_name=table_name,
                                                         column_name=obj.column_name, data_type=obj.data_type,
                                                         not_null=not_null, data_default=data_default)
                TMP_DDL_COLUMN_COMMENT = DDL_COLUMN_COMMENT.format(owner=owner, table_name=table_name,
                                                                   column_name=obj.column_name,
                                                                   comment=obj.col_comments)

                if obj.change_list == "+ 신규":
                    column_ddl = TMP_DDL_ALTER_ADD + TMP_DDL_COLUMN_COMMENT
                elif obj.change_list == "+ 삭제":
                    column_ddl = TMP_DDL_ALTER_DROP
                elif obj.change_list == "+ └ TOBE":

                    before_obj = MetaReqList.objects.filter(id_metareq=obj.id_metareq.id,
                                                            column_id=int(obj.column_id) - 1)
                    before_obj = before_obj[0]

                    if obj.column_name.lower() == before_obj.column_name.lower():
                        column_ddl = TMP_DDL_ALTER_MODIFY + TMP_DDL_COLUMN_COMMENT
                    else:
                        column_ddl = TMP_DDL_ALTER_RENAME.replace("[org]",
                                                                  before_obj.column_name) + TMP_DDL_ALTER_MODIFY + TMP_DDL_COLUMN_COMMENT



            elif obj_meta.id_dblist.id_dbtype.db_type in ['MYSQL', 'MARIADB']:  # MYSQL, MARIADB

                DDL_ALTER_ADD = "ALTER TABLE {owner}{table_name} ADD COLUMN {column_name} {data_type}{data_default}{not_null} COMMENT '{comment}';\n"
                DDL_ALTER_MODIFY = "ALTER TABLE {owner}{table_name} CHANGE COLUMN [org] {column_name} {data_type}{data_default}{not_null} COMMENT '{comment}';\n"
                DDL_ALTER_DROP = "ALTER TABLE {owner}{table_name} DROP COLUMN {column_name};\n"

                TMP_DDL_ALTER_ADD = DDL_ALTER_ADD.format(owner=owner, table_name=table_name,
                                                         column_name=obj.column_name, data_type=obj.data_type,
                                                         not_null=not_null, data_default=data_default,
                                                         comment=obj.col_comments)
                TMP_DDL_ALTER_MODIFY = DDL_ALTER_MODIFY.format(owner=owner, table_name=table_name,
                                                               column_name=obj.column_name, data_type=obj.data_type,
                                                               not_null=not_null, data_default=data_default,
                                                               comment=obj.col_comments)
                TMP_DDL_ALTER_DROP = DDL_ALTER_DROP.format(owner=owner, table_name=table_name,
                                                           column_name=obj.column_name)

                if obj.change_list == "+ 신규":
                    column_ddl = TMP_DDL_ALTER_ADD + TMP_DDL_COLUMN_COMMENT
                elif obj.change_list == "+ 삭제":
                    column_ddl = TMP_DDL_ALTER_DROP
                elif obj.change_list == "+ └ TOBE":
                    before_obj = MetaReqList.objects.filter(id_metareq=obj.id_metareq.id,
                                                            column_id=int(obj.column_id) - 1)
                    before_obj = before_obj[0]
                    if obj.column_name.lower() == before_obj.column_name.lower():
                        column_ddl = TMP_DDL_ALTER_MODIFY.replace("[org]", obj.column_name) + TMP_DDL_COLUMN_COMMENT
                    else:
                        column_ddl = TMP_DDL_ALTER_MODIFY.replace("[org]",
                                                                  before_obj.column_name) + TMP_DDL_COLUMN_COMMENT

            total_column_ddl += DDL_REMARK + column_ddl + "\n\n"

    context = {
        'result': 'ok',
        'column_ddl': total_column_ddl
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxMetaLoadReq(request):
    id_metareq = request.POST.get('id_metareq', '')

    meta = MetaReq.objects.get(id=id_metareq)
    script = meta.req_script
    db_type = meta.id_dblist.id_dbtype.id_db.id

    obj = MetaReqList.objects.filter(id_metareq=id_metareq).order_by("column_id", 'id')

    columns = []
    for o in obj:
        column = [o.div,
                  o.column_id,
                  o.col_comments,
                  o.pk_yn,
                  o.not_null,
                  o.data_default if o.data_default is not None else '',
                  o.column_name,
                  o.data_type,
                  o.change_list,
                  o.change_reason,
                  o.get_privacy_list_display(),
                  ]
        columns.append(column)

    context = {
        'columns': columns,
        'script': script,
        'db_type': db_type
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxMetaReqSave(request):
    # data = request.POST.get('data','')
    saveOrupdate = request.POST.get('saveOrupdate')
    id_metareq = request.POST.get('id_metareq')

    id_tablelist = request.POST.get('id_tablelist')
    id_dblist = request.POST.get('id_dblist')
    id_pl_dev = request.POST.get('id_pl_dev')
    title = request.POST.get('title')
    csr = request.POST.get('csr')
    data = request.POST.getlist('data[]')
    l_data_type = request.POST.getlist('l_data_type[]')
    script = request.POST.get('script')
    obj_class = request.POST.get('obj_class')
    req_contents = request.POST.get('req_contents')
    obj_new = request.POST.get('obj_new')
    table_name = request.POST.get('table_name')
    table_comments = request.POST.get('table_comments')
    storage_cycle = request.POST.get('storage_cycle')
    storage_cycle_column = request.POST.get('storage_cycle_column')
    id_domainanddblist = request.POST.get('id_domainanddblist')

    dbms_type = (DbList.objects.get(id=id_dblist)).id_dbtype.id

    if obj_new == "2":
        id_tablelist = ""

    ds = []

    for d in data:
        buf = d.split('|*|')
        ds.append(buf)

    ind = 0

    script = script.replace('/* 반영 할 Script를 입력해주세요 */', '')

    metareq = ""

    if saveOrupdate == "save":

        metareq = MetaReq(
            id_dblist=DbList.objects.get(id=id_dblist),
            id_tablelist=TableList.objects.get(id=id_tablelist) if id_tablelist != "" else None,
            id_pl_dev=User.objects.get(id=id_pl_dev) if id_pl_dev != "" else None,
            id_pl_prod=User.objects.get(id=id_pl_dev) if id_pl_dev != "" else None,
            title=title,
            csr=csr,
            reg_dtm=datetime.datetime.now(),
            id_reg_user=request.user,
            req_script=script,
            obj_class=obj_class,
            req_contents=req_contents,
            pl_dev_yn='1',
            da_dev_yn='1',
            obj_new=obj_new,
            table_name=table_name if dbms_type == 1 else table_name.lower(),
            table_comments=table_comments,
            mod_dtm=datetime.datetime.now(),
            storage_cycle=storage_cycle if storage_cycle != "" else None,
            storage_cycle_column=storage_cycle_column,
            id_domainanddblist=DomainAndDbList.objects.get(id=id_domainanddblist) if id_domainanddblist != "" else None,
        )
        metareq.save()



    elif saveOrupdate == "update":
        metareq = MetaReq.objects.get(id=id_metareq)

        metareq.id_dblist = DbList.objects.get(id=id_dblist)
        metareq.id_tablelist = TableList.objects.get(id=id_tablelist) if id_tablelist != "" else None
        metareq.id_pl_dev = User.objects.get(id=id_pl_dev) if id_pl_dev != "" else None
        metareq.id_pl_prod = User.objects.get(id=id_pl_dev) if id_pl_dev != "" else None
        metareq.title = title
        metareq.csr = csr
        metareq.req_script = script
        metareq.obj_class = obj_class
        metareq.req_contents = req_contents
        metareq.obj_new = obj_new

        metareq.table_name = table_name
        metareq.table_comments = table_comments
        metareq.mod_dtm = datetime.datetime.now()
        metareq.storage_cycle = storage_cycle if storage_cycle != "" else None
        metareq.storage_cycle_column = storage_cycle_column
        metareq.id_domainanddblist = DomainAndDbList.objects.get(
            id=id_domainanddblist) if id_domainanddblist != "" else None

        # metareq.id_pl_dev = None
        # metareq.id_pl_prod = None
        metareq.id_dba_dev = None
        metareq.id_dba_prod = None
        metareq.id_reqw_prod = None
        metareq.id_da_dev = None

        metareq.pl_dev_yn = '1'
        metareq.pl_prod_yn = '0'
        metareq.dba_dev_yn = '0'
        metareq.dba_prod_yn = '0'
        metareq.req_prod_yn = '0'
        metareq.da_dev_yn = '1'

        metareq.pl_dev_comment = None
        metareq.pl_prod_comment = None
        metareq.dba_dev_comment = None
        metareq.dba_prod_comment = None
        metareq.req_prod_comment = None
        metareq.da_dev_comment = None

        metareq.pl_dev_dtm = None
        metareq.pl_prod_dtm = None
        metareq.dba_dev_dtm = None
        metareq.dba_prod_dtm = None
        metareq.req_prod_dtm = None
        metareq.da_dev_dtm = None

        metareq.dist_dtm = None
        metareq.script = None
        metareq.script_dtm = None
        metareq.id_script_reg_user = None

        metareq.save()

        MetaReqList.objects.filter(id_metareq=id_metareq).delete()

    save_id = metareq.id

    for d in ds:

        data_type = d[7]

        if d[7][0] == "<":
            data_type = l_data_type[ind]
            ind = ind + 1

        metareqlist = MetaReqList(
            div=d[0],
            id_metareq=metareq,
            id_dblist=DbList.objects.get(id=id_dblist),
            id_tablelist=TableList.objects.get(id=id_tablelist) if id_tablelist != "" else None,
            column_id=d[1] if d[1] != '' else 999999,
            col_comments=d[2],
            pk_yn=1 if d[3] == "Y" else 0,
            not_null=1 if d[4] == "Y" else 0,
            data_default=d[5],
            column_name=d[6] if dbms_type == 1 else d[6].lower(),
            data_type=data_type,
            change_list=d[8],
            change_reason=d[9],
            privacy_list=0 if d[10] == '' else 1,
        )
        metareqlist.save()

    context = {
        "result": "ok",
        "save_id": save_id
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxMetaReqUpdate(request):
    # data = request.POST.get('data','')
    id_tablelist = request.POST.get('id_tablelist')
    id_metareq = request.POST.get('id_metareq')
    title = request.POST.get('title')
    data = request.POST.getlist('data[]')
    l_data_type = request.POST.getlist('l_data_type[]')

    ds = []

    for d in data:
        buf = d.split('|*|')
        ds.append(buf)

    metareq = MetaReq.objects.get(id=id_metareq)
    metareq.id_tablelist = TableList.objects.get(id=id_tablelist)
    metareq.title = title
    metareq.mod_dtm = datetime.datetime.now()
    metareq.id_reg_user = request.user
    metareq.save()

    MetaReqList.objects.filter(id_metareq=id_metareq).delete()

    ind = 0

    for d in ds:

        data_type = d[6]

        if d[6][0] == "<":
            data_type = l_data_type[ind]

        # reg_dtm = datetime.datetime.now(),
        # id_reg_user = request.user,

        # id_metareq
        # save_id = metareq.id

        metareqlist = MetaReqList(
            div=d[0],
            id_metareq=metareq,
            id_tablelist=TableList.objects.get(id=id_tablelist),
            column_id=d[1],
            col_comments=d[2],
            pk_yn=1 if d[3] == "Y" else 0,
            not_null=1 if d[4] == "Y" else 0,
            column_name=d[5],
            data_type=data_type,
            change_list=d[7],
            change_reason=d[8],
        )
        metareqlist.save()

    context = {
        "result": "ok"
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxMetaRegCheckColumn(request):


    # data = request.POST.get('data','')
    data = request.POST.getlist('data[]')

    org_kor = []

    l_lkor = []
    l_leng = []
    l_domain = []
    l_eng_attr_name = []
    l_domain_flag = []
    l_oracle_domain_datatype = []
    l_mysql_domain_datatype = []
    l_word_find_flag = []
    l_last_num_flag = []
    lcnt = []
    total_attr = []
    max_length = []
    list_attr = []

    ds = []

    for d in data:
        buf = d.split(',')
        ds.append(buf)

    attrs = []

    # 주문번호, 상품코드
    # ['주문번호', '상품코드']
    # [True, True]
    # [True, True]
    # ['ORD_NO', 'PRD_CD']
    # [['NUMBER(15)'], ['NUMBER(15)']]
    # [['BIGINT'], ['BIGINT']]

    for d in ds:
        if d[6] != "삭제":
            attrs.append(d[1])
            # org_kor.append(d[1])

    if len(attrs) > 0:

        for d in attrs:

            dataset = StdAttr.objects.filter(std_attr_kor=d, accept_yn=0, use_yn=1)

            oracle_datatype = []
            mysql_datatype = []

            if dataset:
                org_kor.append(dataset[0].std_attr_kor)
                l_eng_attr_name.append(dataset[0].std_attr_eng)
                l_domain_flag.append(True)
                l_word_find_flag.append(True)
            else:
                org_kor.append(d)
                l_eng_attr_name.append("<b><font color=red>매핑 용어 없음</fong></b>")
                l_domain_flag.append(False)
                l_word_find_flag.append(False)

            final_data_type = []
            print("#############")
            print(len(dataset))
            print("#############")
            if len(dataset) > 1:

                for in_d in dataset:
                    sql = """SELECT {id} ID, data_type DATA_TYPE, count(*) CNT
                    FROM CUST_COLUMN_LIST
                    WHERE COLUMN_NAME = '{column_name}'
                    AND DATA_TYPE = '{data_type}'
                    AND DB_USE = 'SMTC'
                    AND DROP_YN = 0
                    group by data_type
                    order by 2 desc
                """.format(column_name=in_d.std_attr_eng,
                           data_type=in_d.id_stddomain.oracle_datatype,
                           id=in_d.id)

                    cnt_ds = _query_dict("default", sql)

                    



                    if len(cnt_ds) > 0:
                        final_data_type.append([cnt_ds[0]['ID'], cnt_ds[0]['CNT'], len(cnt_ds[0]['DATA_TYPE'])])

                if len(final_data_type) > 0:
                    final_data_type = sorted(final_data_type, key=itemgetter(1, 2), reverse=True)

                    final_data_type_id = final_data_type[0][0]

                    in_dataset = StdAttr.objects.get(id=final_data_type_id)

                    oracle_datatype.append(in_dataset.id_stddomain.oracle_datatype)
                    mysql_datatype.append(in_dataset.id_stddomain.mysql_datatype)

            if len(dataset) == 1 or len(final_data_type) == 0:
                for data in dataset:
                    # str_oracle_leng = ""

                    # if data.id_stddomain.oracle_decimal_leng != None :
                    #   if data.id_stddomain.oracle_decimal_leng == 0 :
                    #     str_oracle_leng = "(" + str(data.id_stddomain.oracle_leng) + ")"
                    #   else :
                    #     str_oracle_leng = "(" + str(data.id_stddomain.oracle_leng) + ',' + str(data.id_stddomain.oracle_decimal_leng) + ")"

                    # elif data.id_stddomain.oracle_leng != None :
                    #   str_oracle_leng = "(" + str(data.id_stddomain.oracle_leng) + ")"

                    # if data.id_stddomain.oracle_leng != None :

                    #   oracle_datatype.append(data.id_stddomain.oracle_data_type + str_oracle_leng)

                    # elif data.id_stddomain.oracle_data_type != None :

                    #   oracle_datatype.append(data.id_stddomain.oracle_data_type)

                    oracle_datatype.append(data.id_stddomain.oracle_datatype)

                    # str_mysql_leng = ""

                    # if data.id_stddomain.mysql_decimal_leng != None :
                    #   if data.id_stddomain.mysql_decimal_leng == 0 :
                    #     str_mysql_leng = "(" + str(data.id_stddomain.mysql_leng) + ")"
                    #   else :
                    #     str_mysql_leng = "(" + str(data.id_stddomain.mysql_leng) + ',' + str(data.id_stddomain.mysql_decimal_leng) + ")"

                    # elif data.id_stddomain.mysql_leng != None :
                    #   str_mysql_leng = "(" + str(data.id_stddomain.mysql_leng) + ")"

                    # if data.id_stddomain.mysql_leng != None :
                    #   mysql_datatype.append(data.id_stddomain.mysql_data_type + str_mysql_leng)
                    # elif data.id_stddomain.mysql_data_type != None :
                    #   mysql_datatype.append(data.id_stddomain.mysql_data_type)
                    
                    mysql_datatype.append(data.id_stddomain.mysql_datatype)


            # 중복제거
            oracle_datatype = list(set(oracle_datatype))
            oracle_datatype.sort()
            l_oracle_domain_datatype.append(oracle_datatype)

            # 중복제거
            mysql_datatype = list(set(mysql_datatype))
            mysql_datatype.sort()
            l_mysql_domain_datatype.append(list(set(mysql_datatype)))

            print("#############")
            print(mysql_datatype)
            print("#############")

    context = {
        'org_kor': org_kor,
        'l_domain_flag': l_domain_flag,
        'l_word_find_flag': l_word_find_flag,
        'l_eng_attr_name': l_eng_attr_name,
        'l_oracle_domain_datatype': l_oracle_domain_datatype,
        'l_mysql_domain_datatype': l_mysql_domain_datatype
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxMetaRegShowColumn(request):
    id_tablelist = request.POST.get('id_tablelist', '')

    obj = ColumnList.objects.filter(id_tablelist=id_tablelist, drop_yn=0).order_by("column_id")

    id_dbtype = 0
    id_domain = 0

    if obj:
        id_dbtype = obj[0].id_dblist.id_dbtype.id_db.id

        if obj[0].id_tablelist.id_domain:
            id_domain = obj[0].id_tablelist.id_domain.id
        else:
            id_domain = None

        id_dblist = obj[0].id_tablelist.id_dblist.id

        id_domainanddblist = DomainAndDbList.objects.filter(id_dblist=id_dblist,
                                                            id_domain=id_domain).first()

        if id_domainanddblist:
            text_domainanddblist = str(id_domainanddblist)
            id_domainanddblist = id_domainanddblist.id
        else:
            text_domainanddblist = ""
            id_domainanddblist = -1

    columns = []
    for o in obj:

        short_col_comments = ""
        if o.col_comments is not None:
            if len(o.col_comments) > 30:
                short_col_comments = o.col_comments[0:30] + "..."
            else:
                short_col_comments = o.col_comments

        column = [o.column_id,
                  re.sub('[^0-9a-zA-Zㄱ-힗_. ]', '', short_col_comments if short_col_comments is not None else ''),
                  o.pk_yn,
                  o.not_null,
                  o.data_default if o.data_default is not None else '',
                  o.column_name,
                  o.data_type,
                  '개인정보' if o.secu_yn == "1" else ''
                  ]
        columns.append(column)

    context = {
        'columns': columns,
        'id_domain': id_domain,
        'id_dbtype': id_dbtype,
        'text_domainanddblist': text_domainanddblist,
        'id_domainanddblist': id_domainanddblist,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxGetDBMSCase(request):
    id_dblist = request.POST.get('id_dblist', '')

    obj = DbList.objects.get(id=id_dblist)

    id_dbtype = obj.id_dbtype.id_db.id

    context = {
        'id_dbtype': id_dbtype
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxGetDupTable(request):
    id_tablelist = request.POST.get('id_tablelist', '')

    err_message = ""

    sql = """
    SELECT ID
    FROM cust_meta_req
    WHERE ID_TABLELIST = {id_tablelist}
    AND not (
    --    dba_prod_yn = 1
    -- OR pl_dev_yn = 2
    -- OR pl_prod_yn = 2
    -- OR dba_dev_yn = 2
    -- OR dba_prod_yn = 2
    -- OR da_dev_yn = 2
    -- OR req_prod_yn = 2
    dba_dev_yn = 2
    or dba_dev_yn = 1

    )
  """.format(id_tablelist=id_tablelist)

    dataset = _query_dict("default", sql)

    if dataset:
        dataset = dataset[0]
        err_message = "* 선택하신 테이블은 아직 메타 변경 진행중(메타번호 : {id})입니다. 완료 된 후 진행해주세요".format(id=dataset["ID"])

    context = {
        'err_message': err_message,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxViewColumnDetail(request):
    id_tablelist = request.POST.get('id_tablelist', '')

    data_set = ColumnList.objects.filter(id_tablelist=id_tablelist, drop_yn="0").order_by("column_id")

    # attr_data = list(attr_data.values())

    attr_data = []
    table_name = ""

    for data in data_set:
        table_name = data.table_name + " ( " + (data.comments if data.comments is not None else '한글명 없음') + " )"

        attr_data.append(
            [
                data.column_id,
                data.col_comments,
                'Y' if data.pk_yn == '1' else '',
                'Y' if data.not_null == '1' else '',
                data.data_default,
                data.column_name,
                data.data_type,
                'Y' if data.secu_yn == '1' else '',
            ]
        )

    context = {
        'attr_data': attr_data,
        'table_name': table_name,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxCreateUser(request):
    user_id = request.POST.get('user_id', '')
    username = request.POST.get('username', '')

    obj, flag = SabunDBUserMapping.objects.get_or_create(
        id_userlist=User.objects.get(id=user_id),
        db_username=username,
    )

    sabun = User.objects.get(id=user_id)
    sabun = sabun.username

    # 패스워드 같은 로직으로 생성
    cal = 0
    for c in sabun:
        if c.isalpha():
            c = ord(c)

        cal = cal + int(c) * 111

    v1 = "gsshop"
    v2 = str(cal % 1000)

    passwd = v1 + "_" + v2

    ds_emp = _query_dict('iam', iamSql.format(emp_no=sabun))[0]
    msg = "[SmartDBA] 계정 생성 완료 {name} / {username} / {passwd}".format(name=ds_emp['EMP_NM'],
                                                                      username=username,
                                                                      passwd=passwd)
    _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

    context = {
        'result': 'OK',
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxCreateUserFromSabun(request):
    sabun = request.POST.get('sabun', '')
    sabun = sabun.replace(" ", "")
    sabun = sabun.upper()

    username = ""

    # ds_user = User.objects.filter(username=sabun.upper())

    # for ds in ds_user:

    exist_user_obj = UserList.objects.filter(id_user_name__username=sabun.upper(), drop_yn=0).order_by("db_use")

    exist_user = []

    for obj in exist_user_obj:
        buf_dict = {}
        buf_dict['db_use'] = obj.db_use
        buf_dict['oper_cd'] = obj.oper_cd.oper_cd
        buf_dict['username'] = obj.username
        buf_dict['name'] = str(obj.id_user_name)
        exist_user.append(buf_dict)

    user_obj = User.objects.filter(username=sabun.upper())

    if user_obj:
        user_obj = user_obj[0]

        user_id = user_obj.id
        name = user_obj.first_name
        team = user_obj.last_name

        check_flag = User.objects.filter(Q(username=sabun.upper()),
                                         ~Q(sabundbusermapping__db_username=None)).exists()

        if check_flag:
            get_username = User.objects.get(Q(username=sabun.upper()),
                                            ~Q(sabundbusermapping__db_username=None))

            username = get_username.sabundbusermapping.db_username

        else:
            get_email = User.objects.get(username=sabun.upper())
            get_email = get_email.email

            p = re.compile(r"(\w.+)[@](\w+)+")
            m = p.search(get_email)
            username = m.group(1)

            username = 'I_' + username.replace(".", "").upper()

            check_dup = SabunDBUserMapping.objects.filter(db_username=username).exists()

            # 동일 username 이 존재한다면...
            if check_dup:
                check_cnt = SabunDBUserMapping.objects.filter(db_username=username)
                username = username + str(check_cnt.count() + 1)

        # 패스워드 생성
        cal = 0
        for c in sabun:
            if c.isalpha():
                c = ord(c)

            cal = cal + int(c) * 111

        v1 = "gsshop"
        v2 = str(cal % 1000)

        passwd = v1 + "_" + v2

        sql = """/* {name} ({team}) */
/* ORACLE */
CREATE USER {username_lower} IDENTIFIED BY "{passwd}" PROFILE GS_DEV;
GRANT RL_SM_SEL TO {username_lower};
GRANT CONNECT TO {username_lower};

/* MYSQL */
CREATE USER {username_lower}@'10.53.%' IDENTIFIED BY '{passwd}';
GRANT SELECT,UPDATE,DELETE,INSERT ON db_nm.* TO {username_lower}@'10.53.%';
""" \
            .format(
            username_lower=username.lower(),
            passwd=passwd,
            name=name,
            team=team
        )

        context = {
            'sql': sql,
            'user_id': user_id,
            'username': username.lower(),
            'name': name + ' (' + team + ') ',
            'exist_user': exist_user
        }
    else:
        context = {
            'sql': "없는 사번입니다",
            'user_id': '',
            'username': '',
            'name': '',
            'exist_user': ''
        }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxModifyMeta(request):
    req_no = request.POST.get('req_no', '')
    choice_owner = request.POST.get('choice_owner', '')
    choice_role = request.POST.get('choice_role', '')
    choice_tab_tbs = request.POST.get('choice_tab_tbs', '')
    choice_ind_tbs = request.POST.get('choice_ind_tbs', '')

    sql = """
            SELECT A.REQ_NO req_no, --요청번호
                 DECODE(SYS_NM,   '[N]DHUB','DHUB',
                                  '[N]TCPROD','TCPROD',
                                  '[N]NEWINSA','ORAINSA',
                                  '[N]MySQL','MYSQL',
                                  '[N]SAP','PRD',
                                  '[N]WEBDB2','WEBDB',
                                  '[N]TISTM','TISTM',
                                  '[N]OASPRD','OASPRD',
                                  '[N]SMTC','SMTCPRD',
                                  '[N]LGHS002N','ORANBS',
                                  '[N]ETC','ETC',
                                  '[N]BaseITSM','ORAITSM',
                                  '[N]D&Shop','D&SHOP', SYS_NM) db_nm, --시스템
                 decode(c.apc_type, '0001','신규','0002','삭제','0003',decode(c.apc_obj_type,'0002','컬럼추가','변경'),'0004','컬럼변경') req_div, -- 신청구분
                 decode(c.status, '0001','검토중','1001','PL승인','1002','PL반려','0002','승인','0003','반려','0004','DB반영요청중','0005','테스트DB반영완료','0006','테스트DB반영오류','0007','운영DB반영요청중', '0008','운영DB반영완료','0009','운영DB반영오류','운영DB미반영','0010') procss_status, -- 처리상태
                 (select user_nm from meta.usr_user where user_id = c.usr_id) req_user, -- 신청자
                 to_char(to_date(c.create_date,'yyyymmddhh24miss'),'yyyy-mm-dd') req_dt, -- 신청일자
                 trim(replace(c.remark1,'\n','\r\n')) req_explain, -- 설명
                 REPLACE(A.SCRIPT,'CREDATE TABLE','CREATE TABLE') ||CHR(10)||CHR(10)||
                 '/* END */' ddl, -- DDL
                 SYSDATE curr_dt, -- 현재일자
                 trim(replace(c.reason0,'\n','\r\n')) dba_explain, -- DBA설명
                 C.STATUS system_status -- 10
            FROM META.GS_MD_APC_SCRIPT A, META.GS_MD_SYSTEM B, META.GS_MD_APC C
            WHERE
                A.REQ_NO = C.REQ_NO
            AND C.SYS_ID = B.SYS_ID
            AND A.REQ_NO = {req_no}
        """.format(req_no=req_no)

    ds = _query_dict("metadb", sql)

    ds = ds[0]

    before_sql = str(ds['DDL'])
    # req_no
    # choice_owner
    # choice_tab_tbs
    # choice_ind_tbs

    if choice_owner:
        owner = choice_owner + "."
    else:
        owner = ""
    tablespace_name_data = choice_tab_tbs
    tablespace_name_index = choice_ind_tbs

    before_sql = before_sql.replace(owner, "")

    after_sql = """/****************************************************************
 * SQL 명 : {db_nm}_{req_no}.sql
 * 요청번호 : {req_no}
 * DB용도 : {db_nm}
 * 신청자 : {req_user}
 * 신청일자 : {req_dt}
 ****************************************************************
 * 요청자 메모
 {req_explain}
 ****************************************************************
 * DA 메모
 {dba_explain}
 ****************************************************************/

""".format(
        db_nm=ds['DB_NM'],
        req_no=ds['REQ_NO'],
        req_user=ds['REQ_USER'],
        req_dt=ds['REQ_DT'],
        req_explain=ds['REQ_EXPLAIN'],
        dba_explain=ds['DBA_EXPLAIN'],
    )

    ddl = before_sql
    change_ddl = after_sql

    if len(owner) > 0:
        ddl = ddl.replace(owner, "")

    # last_ddl = ddl

    # ddl = ddl.replace("\n"," ")
    # ddl = ddl.replace("\t"," ")

    DDL_CREATE_TABLE_REX = re.compile(r'''(
    (CREATE\s+TABLE)+  # 1 : CREATE TABLE
    (\s)+              # 2 : 구분자
    (\w+\.)?           # 3 : OWNER 무시
    (\w+)              # 4 : TABLE_NAME
    (.*?)              # 5 : COLUMNS
    (;)                # 6:  ;
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

    for groups in DDL_CREATE_TABLE_REX.findall(ddl):

        # deb(groups[1])
        # deb(groups[4])

        DDL_CREATE_TABLE_COLUMN_REX = re.compile(r'''(
    \(
    (.*)              # 1 : COLUMNS
    \)
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

        # deb((DDL_CREATE_TABLE_COLUMN_REX.findall(groups[5]))[0][1])

        # deb(groups[6])

        change_ddl = change_ddl + groups[1]
        change_ddl = change_ddl + ' '
        change_ddl = change_ddl + owner
        change_ddl = change_ddl + groups[4]
        change_ddl = change_ddl + "(\n"
        change_ddl = change_ddl + (DDL_CREATE_TABLE_COLUMN_REX.findall(groups[5]))[0][1]
        change_ddl = change_ddl + "\n)" + "TABLESPACE " + tablespace_name_data
        change_ddl = change_ddl + groups[6]
        change_ddl = change_ddl + '\n\n'

        change_ddl = change_ddl + "CREATE PUBLIC SYNONYM " + groups[4] + " FOR " + owner + groups[4] + ";"
        change_ddl = change_ddl + '\n\n'

        ds_grant = MetaGrantDetail.objects.filter(id_metagrantlist__app_service=choice_role)

        cr_grant = []
        for grant in ds_grant:
            cr_grant.append("GRANT {crud} ON TABNM TO {role};".format(
                crud=grant.crud,
                role=grant.role
            ))
        for grant in cr_grant:
            grant_ddl = grant
            change_ddl = change_ddl + grant_ddl.replace("TABNM", owner + groups[4]) + '\n'

        change_ddl = change_ddl + '\n\n'

    DDL_CREATE_UNIQUE_INDEX_REX = re.compile(r'''(
    (CREATE\s*UNIQUE\s*INDEX)+    # 1 : index
    (\s)+                         # 2 : 구분자
    (\w+\.)?                      # 3 : OWNER 무시
    (\w+)                         # 4 : INDEX_NAME
    (\s)+                         # 5 : 구분자
    ON                            #   : ON
    (\s)+                         # 6 : 구분자
    (\w+\.)?                      # 7 : OWNER 무시
    (\w+)                         # 8 : TABLE_NAME
    (.*?)                         # 9 : COLUMNS
    (;)                           # 10  ;
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

    for groups in DDL_CREATE_UNIQUE_INDEX_REX.findall(ddl):
        # deb(groups[1])
        # deb(groups[4])
        # deb(groups[8])

        DDL_CREATE_UNIQUE_INDEX_REX = re.compile(r'''(
    \(
    (.*)              # 1 : COLUMNS
    \)
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

        # deb((DDL_CREATE_UNIQUE_INDEX_REX.findall(groups[9]))[0][1])

        # deb(groups[10])

        change_ddl = change_ddl + groups[1]
        change_ddl = change_ddl + ' '
        change_ddl = change_ddl + owner
        change_ddl = change_ddl + groups[4]
        change_ddl = change_ddl + '\n'
        change_ddl = change_ddl + 'ON '
        change_ddl = change_ddl + owner
        change_ddl = change_ddl + groups[8]
        change_ddl = change_ddl + '\n'
        change_ddl = change_ddl + "(\n"
        change_ddl = change_ddl + (DDL_CREATE_UNIQUE_INDEX_REX.findall(groups[9]))[0][1]
        change_ddl = change_ddl + "\n)\n" + "TABLESPACE " + tablespace_name_index
        change_ddl = change_ddl + groups[10]
        change_ddl = change_ddl + '\n\n'

    DDL_CREATE_INDEX_REX = re.compile(r'''(
    (CREATE\s*INDEX)+             # 1 : index
    (\s)?                         # 구분자
    (\w+\.)?                      # OWNER 무시
    (\w+)                         # 4 : INDEX_NAME
    (\s)+                         # 구분자
    ON
    (\s)+                         # 구분자
    (\w+\.)?                      # OWNER 무시
    (\w+)                         # 8 : TABLE_NAME
    (.*?)                         # 9 : COLUMNS
    (;)                           # 10  ;
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

    for groups in DDL_CREATE_INDEX_REX.findall(ddl):
        # deb(groups[1])
        # deb(groups[4])
        # deb(groups[8])

        DDL_CREATE_INDEX_COLUMN_REX = re.compile(r'''(
    \(
    (.*)              # 1 : COLUMNS
    \)
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

        # deb((DDL_CREATE_INDEX_COLUMN_REX.findall(groups[9]))[0][1])

        # deb(groups[10])

        change_ddl = change_ddl + groups[1]
        change_ddl = change_ddl + ' '
        change_ddl = change_ddl + owner
        change_ddl = change_ddl + groups[4]
        change_ddl = change_ddl + '\n'
        change_ddl = change_ddl + 'ON '
        change_ddl = change_ddl + owner
        change_ddl = change_ddl + groups[8]
        change_ddl = change_ddl + '\n'
        change_ddl = change_ddl + "(\n"
        change_ddl = change_ddl + (DDL_CREATE_INDEX_COLUMN_REX.findall(groups[9]))[0][1]
        change_ddl = change_ddl + "\n)\n" + "TABLESPACE " + tablespace_name_index
        change_ddl = change_ddl + groups[10]
        change_ddl = change_ddl + '\n\n'

    DDL_ALTER_TABLE_REX = re.compile(r'''(
    (ALTER\s*TABLE)+              # 1 : alter table
    (\s)+                         # 2 : 구분자
    (\w+\.)?                      # 3 : OWNER 무시
    (\w+)                         # 4 : TABLE_NAME
    (.*?)                         # 5 : TABLE ALTER
    (;)                           # 6  ;
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

    for groups in DDL_ALTER_TABLE_REX.findall(ddl):
        # deb(groups[1])
        # deb(groups[4])
        # deb(groups[5])
        # deb(groups[6])

        change_ddl = change_ddl + groups[1]
        change_ddl = change_ddl + ' '
        change_ddl = change_ddl + owner
        change_ddl = change_ddl + groups[4]
        change_ddl = change_ddl + groups[5]
        change_ddl = change_ddl + groups[6]
        change_ddl = change_ddl + '\n\n'

    DDL_COMMENT_TABLE_REX = re.compile(r'''(
    (COMMENT\s*ON)+               # 1 : COMMENT ON
    (\s)+                         # 2 : 구분자
    (TABLE|COLUMN)+               # 3 : TABLE 또는 COLUMNS
    (\s)+                         # 4 : 구분자
    (\w+)                         # 5 : TABLE_NAME
    (.*?)                         # 6 : COMMENT OPTION
    (;)                           # 7  ;
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

    for groups in DDL_COMMENT_TABLE_REX.findall(ddl):
        # deb(groups[1])
        # deb(groups[3])
        # deb(groups[5])
        # deb(groups[6])
        # deb(groups[7])

        change_ddl = change_ddl + groups[1]
        change_ddl = change_ddl + ' '
        change_ddl = change_ddl + groups[3]
        change_ddl = change_ddl + ' '
        change_ddl = change_ddl + owner
        change_ddl = change_ddl + groups[5]
        change_ddl = change_ddl + groups[6]
        change_ddl = change_ddl + groups[7]
        change_ddl = change_ddl + '\n\n'

    DDL_CREATE_SEQ_REX = re.compile(r'''(
    (CREATE\s*SEQUENCE)+          # 1 : index
    (\s)+                         # 구분자
    (\w+\.)?                      # OWNER 무시
    (\w+)                         # 4 : SEQUENCE_NAME
    (.*?)                         # 5 : SEQUENCE_OPTION
    (;)                           # 6  ;
    )''', re.IGNORECASE | re.DOTALL | re.VERBOSE)

    for groups in DDL_CREATE_SEQ_REX.findall(ddl):
        # deb(groups[1])
        # deb(groups[4])
        # deb(groups[5])
        # deb(groups[6])

        change_ddl = change_ddl + groups[1]
        change_ddl = change_ddl + ' '
        change_ddl = change_ddl + owner
        change_ddl = change_ddl + groups[4]
        change_ddl = change_ddl + groups[5]
        change_ddl = change_ddl + groups[6]
        change_ddl = change_ddl + '\n'

    context = {'after_sql': change_ddl}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxExecuteMeta(request):
    req_no = request.POST.get('req_no', '')

    sql = """
    update META.GS_MD_APC
    set reason3 = '* 작업완료'||chr(10)||'* 작업자 : '||'{worker}',
        remark3 = to_char(sysdate,'YYYY-MM-DD'),
        STATUS = '0008',
        MODIFY_DATE = to_char(sysdate,'YYYYMMDDHH24MISS')
    where req_no = {req_no}
  """.format(
        worker=request.user,
        req_no=req_no
    )

    _query_commit("metadb", sql)

    # ds_emp = _query_dict('iam',iamSql.format(emp_no=request.user.username))[0]
    msg = "[SmartDBA] 변경관리 개발DB 반영 완료, 작업번호 : {req_no}".format(req_no=req_no)
    # _query_sms(ds_emp['MOBILE'],msg,team_name=ds_emp['NAME'])
    _send_slack(msg)

    context = {'reulst': "OK"}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxExecuteUserPriv(request):
    chk_sms_dup = []

    sql = """
      INSERT INTO stg_dba_tab_privs
      (
        id_dblist,
        oper_cd,
        db_use,
        table_name,
        privilege,
        owner,
        grantee
      )
      VALUES
      (
        {id_dblist},
        {oper_cd},
        '{db_use}',
        '{table_name}',
        '{privilege}',
        '{owner}',
        '{grantee}'
      )
  """

    obj_priv = request.POST.get('obj_priv', '')

    # 1,2,3,4,5,6,7, -->
    # ['1', '2', '3', '4', '5', '6', '7', ''] -->
    # ['1', '2', '3', '4', '5', '6', '7']
    obj_priv = obj_priv.strip().split(',')[:-1]

    for obj in obj_priv:
        priv = UserRequestTabPrivHist.objects.get(id=int(obj),
                                                  use_yn='1')
        priv.id_approver = request.user
        priv.approv_dtm = datetime.datetime.now()
        priv.approv_yn = "1"

        priv.save()

        list_priv = priv.priv.strip().split(',')
        for p in list_priv:
            insert_sql = sql.format(
                id_dblist=priv.id_objectlist.id_dblist.id,
                oper_cd=priv.id_objectlist.oper_cd.id,
                db_use=priv.id_objectlist.db_use,
                table_name=priv.id_objectlist.object_name,
                privilege=p.strip(),
                owner=priv.id_objectlist.owner,
                grantee=priv.id_userlist.username,
            )

            _query_commit('default', insert_sql)

        # 중복 메시지 안나가도록
        sabun = priv.id_userlist.id_user_name.username

        if sabun not in chk_sms_dup:
            chk_sms_dup.append(sabun)
            ds_emp = _query_dict('iam', iamSql.format(emp_no=sabun))[0]
            msg = "[SmartDBA] DB 권한 반영 완료되었습니다. 요청자 : {name}".format(name=ds_emp['EMP_NM'], )
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

    context = {'reulst': "OK"}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxExecuteUserActivation(request):
    chk_sms_dup = []

    user_acc = request.POST.get('user_acc', '')

    # 1,2,3,4,5,6,7, -->
    # ['1', '2', '3', '4', '5', '6', '7', ''] -->
    # ['1', '2', '3', '4', '5', '6', '7']
    user_acc = user_acc.strip().split(',')[:-1]

    for obj in user_acc:
        user = UserList.objects.get(id=int(obj))
        user.status = "0"
        user.request_status = "0"
        user.mod_dtm = datetime.datetime.now()

        user.save()

        hist = UserRequestAccountOpenHist.objects.filter(id_userlist=int(obj),
                                                         approv_dtm=None)
        for h in hist:
            h.id_approver = request.user
            h.approv_dtm = datetime.datetime.now()
            h.save()

        sabun = user.id_user_name.username

        if sabun not in chk_sms_dup:
            chk_sms_dup.append(sabun)
            username = user.username

            # 패스워드 같은 로직으로 생성
            cal = 0
            for c in sabun:
                if c.isalpha():
                    c = ord(c)

                cal = cal + int(c) * 111

            v1 = "gsshop"
            v2 = str(cal % 1000)

            passwd = v1 + "_" + v2

            ds_emp = _query_dict('iam', iamSql.format(emp_no=sabun))[0]
            msg = "[SmartDBA] 계정 활성화 완료 {username} / {passwd}".format(username=username.lower(), passwd=passwd)
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

    context = {'reulst': "OK"}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxExecuteUserDrop(request):
    user_acc_delete = request.POST.get('user_acc_delete', '')

    # 1,2,3,4,5,6,7, -->
    # ['1', '2', '3', '4', '5', '6', '7', ''] -->
    # ['1', '2', '3', '4', '5', '6', '7']
    user_acc_delete = user_acc_delete.strip().split(',')[:-1]

    for obj in user_acc_delete:
        user = UserList.objects.get(id=int(obj))

        msg = "[SmartDBA] 퇴사자 계정 삭제 {oper_cd} {db_use} {username}".format(oper_cd=user.oper_cd.oper_cd,
                                                                          db_use=user.db_use,
                                                                          username=user.username)

        user.drop_yn = "1"
        user.mod_dtm = datetime.datetime.now()

        user.save()

        _send_slack(msg)

    context = {'reulst': "OK"}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxReqestPriv(request):
    req_reason = request.POST.get('req_reason', '')
    table_id = request.POST.get('table_id', '')
    user_id = request.POST.get('user_id', '')
    object_type = request.POST.get('object_type', '')
    req_priv_list = request.POST.getlist('req_priv_list[]')

    str_priv_list = (', ').join(req_priv_list)

    if object_type == "TABLE":
        obj = TableList.objects.get(id=table_id)

        obj = ObjectList.objects.filter(id_dblist=obj.id_dblist.id,
                                        oper_cd=obj.oper_cd.id,
                                        owner=obj.owner,
                                        object_name=obj.table_name,
                                        object_type='TABLE')
        obj = obj[0]
    else:
        obj = ObjectList.objects.get(id=table_id)

    userrequesttabprivhist = UserRequestTabPrivHist(
        id_userlist=UserList.objects.get(id=int(user_id)),
        id_reg_user=request.user,
        id_objectlist=obj,
        reg_dtm=datetime.datetime.now(),
        priv=str_priv_list,
        object_type=object_type,
        req_reason=req_reason,
        use_yn='1',
    )

    userrequesttabprivhist.save()

    # delete_datalist = DataList.objects.get(id=sql_id)

    # delete_datalist.delete()

    # sql_title_list = _query_list('default',sql_execute_saved.format(id=request.user.id))

    # sql_title_dict = defaultdict(list)

    # for writer, \
    #     id, \
    #     data_title \
    #     in sql_title_list:
    #     sql_title_dict[writer].append([str(id),data_title])

    # sql_title_dict = dict(sql_title_dict)

    context = {'data_list': '1'}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxCheckTabPriv(request):
    table_id = request.POST.get('table_id', '')
    id_dblist = request.POST.get('id_dblist', '')
    object_type = request.POST.get('object_type', '')
    username = ""
    err_msg = ""
    err_yn = "N"
    ds = ""
    user_id = ""

    sql = """
    select id, username
    from cust_user_list
    where id_user_name = {user_id}
      and id_dblist = {id_dblist}
      and oper_cd = 3;
  """.format(user_id=request.user.id,
             id_dblist=id_dblist)

    ds = _query_dict("default", sql)

    obj = []
    table_name = ""

    if len(ds) == 0:
        err_msg = "[안내] 개인계정이 없습니다. 개인계정을 먼저 신청해주세요"
        err_yn = "Y"

    else:
        ds = ds[0]
        username = ds['username']
        user_id = ds['id']

        if object_type == "TABLE":
            obj = TableList.objects.get(id=table_id)

            obj = ObjectList.objects.filter(id_dblist=obj.id_dblist.id,
                                            oper_cd=obj.oper_cd.id,
                                            owner=obj.owner,
                                            object_name=obj.table_name,
                                            object_type='TABLE')
            obj = obj[0]
        else:
            obj = ObjectList.objects.get(id=table_id)

        sql = """
      SELECT distinct privilege
      FROM STG_DBA_TAB_PRIVS A
      WHERE A.ID_DBLIST = {id_dblist}
        AND A.OPER_CD = {oper_cd}
        AND A.OWNER = '{owner}'
        AND A.TABLE_NAME = '{table_name}'
        AND A.GRANTEE = '{username}'
    """.format(
            id_dblist=obj.id_dblist.id,
            oper_cd=obj.oper_cd.id,
            owner=obj.owner,
            table_name=obj.object_name,
            username=username,
        )

        ds = _query_dict("default", sql)

    context = {'priv': ds,
               'err_msg': err_msg,
               'err_yn': err_yn,
               'user_id': user_id}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxShowClickedSQL(request):
    sql_id = request.POST.get('sql_id', '')

    obj = DataList.objects.get(id=sql_id)

    data = {}
    data['id'] = obj.id
    data['data_title'] = obj.data_title
    data['sql_text'] = obj.sql_text

    obj = UserExecuteSQL.objects.filter(id_datalist=sql_id).aggregate(Avg('execute_time'))

    execute_time = ""
    if obj['execute_time__avg']:
        execute_time = "( 평균 응답시간 : {}초 )".format(round(obj['execute_time__avg'], 2))
    else:
        execute_time = "( 실행 이력 없음 )"

    context = {'data': data,
               'execute_time': execute_time}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxShareSQL(request):
    sql_id = request.POST.get('sql_id', '')

    sql = """
    SELECT EXP_YN
    FROM CUST_DATA_LIST
    WHERE ID = {sql_id}
  """.format(sql_id=sql_id)

    ds = _query_dict("default", sql)
    ds = ds[0]

    #

    if ds['EXP_YN'] == '1':
        sql = """
      UPDATE CUST_DATA_LIST
      SET EXP_YN = '0'
      WHERE ID = {sql_id}
    """.format(
            sql_id=sql_id
        )
    elif ds['EXP_YN'] == '0':
        sql = """
      UPDATE CUST_DATA_LIST
      SET EXP_YN = '1'
      WHERE ID = {sql_id}
    """.format(
            sql_id=sql_id
        )

    apply_cnt = _query_commit('default', sql)

    sql = """
    SELECT EXP_YN
    FROM CUST_DATA_LIST
    WHERE ID = {sql_id}
  """.format(sql_id=sql_id)

    ds = _query_dict("default", sql)
    ds = ds[0]

    text = []

    if ds['EXP_YN'] == '0':
        text = ["동료와 공유", "동료와 SQL을 공유하시겠습니까?"]
    elif ds['EXP_YN'] == '1':
        text = ["공유 해제", "정말 공유를 해제하시겠습니까?"]

    sql_title_list = _query_list('default', sql_execute_saved.format(id=request.user.id))

    sql_title_dict = defaultdict(list)

    for writer, \
        id, \
        data_title \
            in sql_title_list:
        sql_title_dict[writer].append([str(id), data_title])

    sql_title_dict = dict(sql_title_dict)

    context = {'data': text, 'data_list': sql_title_dict}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxOtherSaveNameSQL(request):
    savedname = request.POST.get('savedname', '')
    savedsql = request.POST.get('savedsql', '')
    id_dblist = request.POST.get('id_dblist', '')

    # t = str(datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S") )
    # t = "["+t+"] "

    if id_dblist == 39:
        id_dblist = 318

    id_dblist = DbDetail.objects.get(id=id_dblist)

    new_datalist = DataList(
        reg_dtm=datetime.datetime.now(),
        mod_dtm=datetime.datetime.now(),
        id_dblist=DbList.objects.get(id=id_dblist.id_dblist.id),
        id_domain=None,
        id_reg_user=request.user,
        id_mod_user=request.user,
        data_title=savedname,
        data_explain=""""SQL실행" 메뉴에서 저장된 SQL""",
        sql_text=savedsql,
        exp_yn='1'
    )

    new_datalist.save()

    obj = DataList.objects.filter(id_reg_user=request.user).order_by("-reg_dtm")

    my_data = []

    for o in obj:
        d = [o.id, o.data_title, o.reg_dtm.strftime("%Y/%m/%d %H:%M:%S")]
        my_data.append(d)

    context = {'my_data': my_data}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxDeleteSQL(request):
    sql_id = request.POST.get('sql_id', '')

    delete_datalist = DataList.objects.get(id=sql_id)

    delete_datalist.delete()

    obj = DataList.objects.filter(id_reg_user=request.user).order_by("-reg_dtm")

    my_data = []

    for o in obj:
        d = [o.id, o.data_title, o.reg_dtm.strftime("%Y/%m/%d %H:%M:%S")]
        my_data.append(d)

    context = {'my_data': my_data}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")

    # sql_title_list = _query_list('default',sql_execute_saved.format(id=request.user.id))

    # sql_title_dict = defaultdict(list)

    # for writer, \
    #     id, \
    #     data_title \
    #     in sql_title_list:
    #     sql_title_dict[writer].append([str(id),data_title])

    # sql_title_dict = dict(sql_title_dict)

    # context = {'data_list':sql_title_dict}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


# 표준 도메인과 관련 없음
def ajaxDomain(request):
    table_id = request.POST.get('table_id', '')

    sql = """
    select b.id column_id, if(b.pk_yn = 1, concat(b.col_comments,' *'), b.col_comments) col_comments
    from cust_table_list a,
         cust_column_list b
    where a.id = b.id_tablelist
      and a.id = {id}
      and col_comments is not null
    order by b.column_id;
  """.format(id=table_id)

    dataset_column = _query_mysql_dict(sql)

    context = {'columnList': dataset_column}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxSearchKeywordAutoComplete(request):
    # domain_id = request.POST.get('domain_id','')
    keyword = request.POST.get('keyword', '')

    #
    # data_type = request.POST.get('data_type','')

    sql = """
    select data
    from
    (
      select distinct 1 no, substr(col_comments,1,150) data
      from cust_column_list
      -- where replace(col_comments,' ','') like replace('%{keyword}%',' ','')
      where col_comments like '{keyword}%'
        union all
        select distinct 0, substr(data_title,1,150)
      from cust_data_list
      -- where replace(data_title,' ','') like replace('%{keyword}%',' ','')
      where data_title like '{keyword}%'
        order by no
    ) a
    limit 5;
  """.format(keyword=keyword)

    dataset_col_comments = _query_mysql_dict(sql)

    param = json.dumps(dataset_col_comments, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxExecuteSQL(request):
    sql = request.POST.get('sql', '')
    sql_id = request.POST.get('sql_id', '')


    error_yn = "N"
    message = ""

    org_sql = sql

    sql = sql.replace(";", "")

    sql = """SELECT *
FROM (
  """ + \
          sql + \
          """
)
where rownum <= 10000"""

    # try :

    with connections['DHUB11'].cursor() as cursor:

        try:
            id_datalist = DataList.objects.get(id=sql_id).id
        except:
            id_datalist = None

        new_userexecutesql = UserExecuteSQL(
            sql_text=org_sql,
            id_reg_user=request.user,
            reg_dtm=datetime.datetime.now(),
            id_datalist=id_datalist,
        )

        new_userexecutesql.save()

        save_id = new_userexecutesql.id

        #####################################################
        ## 수행 시간 측정
        #####################################################
        start = time.time()

        cursor.execute(sql)

        end = time.time() - start
        #####################################################

        new_userexecutesql = UserExecuteSQL.objects.get(id=save_id)
        new_userexecutesql.execute_time = end

        new_userexecutesql.save()

        columnNames = [[d[0]] for d in cursor.description]

        # data = [dict(zip(columnNames, row)) for row in cursor]
        data = cursor.fetchall()
        data2 = []

        ###############################
        ## Decimal('53.94') -> str 로 변경
        ###############################
        buff1 = []
        buff2 = []

        ind = 0

        for ds in data:
            for d in ds:

                if isinstance(d, Decimal):
                    buff1.append(str(d))
                elif isinstance(d, int) and columnNames[ind][0][-2:].upper() != 'CD' \
                        and columnNames[ind][0][-2:].upper() != 'ID' \
                        and columnNames[ind][0][-2:] != '코드':
                    buff1.append(format(d, ','))
                else:
                    buff1.append(d)
                ind = ind + 1

            buff2.append(buff1)
            buff1 = []
            ind = 0

        data = buff2

        ###############################

        # for d in data :
        #   data2.append(list(d))

        # # data = [item[0] for item in cursor.fetchall()]

        # ind = 0
        # list_ind = []
        # for columnName in columnNames :
        #   if columnName[0] in ("CUST_NO",
        #                      "고객번호",
        #                      "고객_번호",
        #                      "CUSTNO",
        #                      "고객 번호") :
        #     list_ind.append(ind)

        #   ind = ind + 1

        # # if len(columnNames) == ind :
        # #   ind = -1

        # for i in list_ind :
        #   for d in data2 :
        #     d[i] = "*******"

        # data = data2

    # except :
    #   data = [["a","b"]]
    #   columnNames = ["a","b"]

    # data = cursor.fetchall()

    if len(columnNames) > 35:
        error_yn = "Y"
        message = """[안내] 조회되는 컬럼 수가 너무 많습니다. [SELECT *] 보다는 컬럼명을 명시해주세요 """

    context = {'data': data,
               'columns': columnNames,
               'time': end,
               'error_yn': error_yn,
               'message': message
               }

    param1 = json.dumps(context, default=json_default)

    return HttpResponse(param1, content_type="application/json")


def ajaxMetaDataDetail(request):
    dataset2 = ""
    columnNames2 = ""

    column_id = request.POST.get('column_id', '')
    data_type = 1

    sql = dataSearchSql.format(column_id=column_id,
                               data_type=data_type,
                               userid=request.user.id,
                               super_user=request.user.is_superuser)

    dataset = _query_dict('default', sql)

    dataset = dataset[0]

    sql_text = dataset['sql_text']

    sql_text = sql_text.replace(":MM", "")
    sql_text = sql_text.replace(":mm", "")
    sql_text = sql_text.replace(":DD", "")
    sql_text = sql_text.replace(":dd", "")
    sql_text = sql_text.replace(":HH", "")
    sql_text = sql_text.replace(":hh", "")
    sql_text = sql_text.replace(":MI", "")
    sql_text = sql_text.replace(":mi", "")
    sql_text = sql_text.replace(":SS", "")
    sql_text = sql_text.replace(":ss", "")
    sql_text = sql_text.replace(":digit", "")
    sql_text = sql_text.replace(":DIGIT", "")

    regex = re.compile(":\w+")
    bind_value = regex.findall(sql_text)

    bind_value = _orderedSet(bind_value)

    new_userclickdata = UserClickData(
        data=dataset['col_comments'],
        id_data=dataset['id'],
        data_type=dataset['data_type'],
        id_reg_user=request.user,
        reg_dtm=datetime.datetime.now(),
    )

    new_userclickdata.save()

    # emp_no, emp_nm, sub_mail, mobile, name
    ds_emp = _query_dict('iam', iamSql.format(emp_no=dataset['it_manager_sabun']))

    if len(ds_emp) > 0:
        ds_emp = ds_emp[0]

    # 데이터타입이 컬럼일 경우만
    if data_type == "0":
        #
        sql = """
        SELECT
             B.COLUMN_ID "#"
           , A.DB_USE "DB용도"
             , A.TABLE_NAME "테이블 영문명"
             , SUBSTR(A.COMMENTS,1,40) "테이블 한글명"
           , B.COLUMN_NAME "속성 영문명"
           , SUBSTR(B.COL_COMMENTS,1,40) "속성 한글명"
           , B.DATA_TYPE "데이터 타입"
           , IF(B.PK_YN = 1,'Y','') "PK여부"
           , IF(B.SECU_YN = 1, 'Y', '') "개인정보 여부"
        FROM CUST_TABLE_LIST A
            ,CUST_COLUMN_LIST B
            ,CUST_COLUMN_LIST C
        WHERE A.ID = C.ID_TABLELIST
          AND A.ID = B.ID_TABLELIST
          AND C.ID = {column_id}
          AND A.OPER_CD = 3
        ORDER BY B.COLUMN_ID;""".format(column_id=column_id)

        dataset2 = _query_list('default', sql)
        columnNames2 = _query_column_name('default', sql)

    context = {
        'dataset': dataset,
        'bind_value': bind_value,
        'columnNames2': columnNames2,
        'dataset2': dataset2,
        'ds_emp': ds_emp,
    }

    param1 = json.dumps(context, default=json_default)

    return HttpResponse(param1, content_type="application/json")


# def dataSearchCategory(request, column_id, domain_id, data_type) :
# def dataSearchCategory(request) :


# metaStdWord 와 관련 없음
def ajaxUserRequestWord(request):
    word_add_name = request.POST.get('word_add_name', '')
    word_add_explain = request.POST.get('word_add_explain', '')

    new_userrequestword = UserRequestWord(
        word_name=word_add_name,
        word_text=word_add_explain,
        id_reg_user=request.user,
        reg_dtm=datetime.datetime.now(),
    )

    new_userrequestword.save()

    context = {
        'success': 'ok',
    }

    param1 = json.dumps(context, default=json_default)

    return HttpResponse(param1, content_type="application/json")


def ajaxUserRequestData(request):
    data_add_name = request.POST.get('data_add_name', '')
    data_add_explain = request.POST.get('data_add_explain', '')

    new_userrequestdata = UserRequestData(
        data_name=data_add_name,
        data_text=data_add_explain,
        id_reg_user=request.user,
        reg_dtm=datetime.datetime.now(),
    )

    new_userrequestdata.save()

    context = {
        'success': 'ok',
    }

    param1 = json.dumps(context, default=json_default)

    return HttpResponse(param1, content_type="application/json")


def ajaxDataview(request):
    team_name = request.POST.get('team_name', '')

    sql = """
    select
         a.id,
         b.id user_id,
         b.username,
         b.first_name,
         replace(b.last_name,' ','') last_name,
         a.data_title,
         a.data_explain,
         a.sql_text,
         a.reg_dtm,
         a.mod_dtm,
         ( case when '{team_name}' = '나의데이터' then  a.data_title
                    else concat(a.data_title) end ) tree_title,
         a.id_reg_user,
         a.id_mod_user,
         '{team_name}' team_name
    from cust_data_list a,
       auth_user b,
       cust_datalist_user_id_req_users c
    where
        a.id = c.datalist_id
      and a.exp_yn = 1
      and b.id = c.user_id
      # and a.prov_yn1 = 1
      # and a.prov_yn2 = 1
      and ( case when '{team_name}' = '나의데이터' then b.id = {user_id}
            else replace(b.last_name,' ','') = '{team_name}' end )
    UNION
    select
         a.id,
         b.id user_id,
         b.username,
         b.first_name,
         replace(b.last_name,' ','') last_name,
         a.data_title,
         a.data_explain,
         a.sql_text,
         a.reg_dtm,
         a.mod_dtm,
         concat(a.data_title) tree_title,
         a.id_reg_user,
         a.id_mod_user,
         '{team_name}' team_name
    from cust_data_list a,
       auth_user b,
       cust_datalist_user_id_mod_psbl_users c
    where
         a.id_reg_user = b.id
      and a.id = c.datalist_id
      and c.user_id = {user_id}
      and a.exp_yn = 1
      and replace(b.last_name,' ','') = '{team_name}'
    UNION
    select
         a.id,
         b.id user_id,
         b.username,
         b.first_name,
         replace(b.last_name,' ','') last_name,
         a.data_title,
         a.data_explain,
         a.sql_text,
         a.reg_dtm,
         a.mod_dtm,
         case when a.exp_yn = 1
            then concat(a.data_title)
          when a.exp_yn = 0
            then concat(a.data_title, ' [only you]')
         end tree_title,
         a.id_reg_user,
         a.id_mod_user,
         '{team_name}' team_name
    from cust_data_list a,
       auth_user b
    where
        a.id_reg_user = b.id
       and ( case when '{team_name}' = '나의데이터' then b.id = {user_id}
            else 1=0 end )
    union
    select
         a.id,
         b.id user_id,
         b.username,
         b.first_name,
         replace(b.last_name,' ','') last_name,
         a.data_title,
         a.data_explain,
         a.sql_text,
         a.reg_dtm,
         a.mod_dtm,
         case when a.exp_yn = 1
            then concat(a.data_title)
          when a.exp_yn = 0
            then concat(a.data_title, ' [only you]')
         end tree_title,
         a.id_reg_user,
         a.id_mod_user,
         '{team_name}' team_name
    from cust_data_list a,
       auth_user b
    where
        a.id_reg_user = b.id
       and '{team_name}' = '전체'
       and b.id <> 2
    # order by ( case when user_id = {user_id} then 0
    #                 else 1 end ), username, mod_dtm desc;
    order by reg_dtm desc,username desc
  """.format(team_name=team_name, user_id=request.user.id)

    ds_data_list = _query_dict("default", sql)

    context = {'ds_data_list': ds_data_list}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxDataviewClick(request):
    data_id = request.POST.get('data_id', '')
    id_dblist = request.POST.get('id_dblist', '')

    sql = """
    select distinct
           a.id,
         a.id_dblist,
         if(a.realoretl = 0, '약 1시간 이전 버전 데이터', '실시간 데이터') realoretl,
         concat('데이터 확인 횟수 : ',a.execute_cnt,'회') execute_cnt,
         d.db_use,
         b.username,
         b.first_name reg_name,
         b.last_name reg_team,
         a.data_title,
         a.data_explain,
         a.sql_text,
         date_format(a.reg_dtm,'%Y/%m/%d') reg_dtm,
         date_format(a.mod_dtm,'%Y/%m/%d') mod_dtm,
         concat(a.data_title) tree_title,
         if(a.prov_yn1=0,'N','Y') prov_yn1,
         if(a.prov_yn2=0,'N','Y') prov_yn2,
         case when {super_user} = 1 then 'Y'
              when id_dblist = 39 and a.privacy_yn = 1 and cduiru.user_id is null then 'C' -- 요청자가 아님
              when id_dblist = 39 then 'Y'
          when id_dblist <> 39 and ( prov_yn2 = 0 ) then 'B' -- 승인이 나지 않음
          when id_dblist <> 39 and a.privacy_yn = 1 and cduiru.user_id is null then 'C' -- 요청자가 아님
          when id_dblist <> 39 and a.privacy_yn = 1 and cduiru.user_id is not null and cdr.id_datalist is null then 'D' -- 조회 기간이 지남
          when id_dblist <> 39 and a.privacy_yn = 1 and cduiru.user_id is not null and cdr.id_datalist is not null then 'Y'
          # when id_dblist <> 39 and prov_yn1 = 1 and prov_yn2 = 1 and cduiru.user_id is not null then 'Y'
          else 'Y' END prov_yn,
         a.privacy_yn,
         IF(cues.t is null, '수행 이력 없음',concat(cues.t,' 초')) t,
         a.id_reg_user,
         a.id_mod_user,
         (select concat(first_name,'(',last_name,')') from auth_user where id =  a.id_reg_user) reg_n,
         (select concat(first_name,'(',last_name,')') from auth_user where id =  a.id_mod_user) mod_n,
         a.id_mod_user,
         (
            select group_concat(concat(bb.first_name,'(',bb.last_name,')') SEPARATOR ',')
            from cust_datalist_user_id_req_users aa,
                 auth_user bb
            where aa.user_id = bb.id
              and aa.datalist_id = a.id
         ) req_user,
         (
            select group_concat(concat(bb.first_name,'(',bb.last_name,')') SEPARATOR ',')
            from cust_datalist_user_id_mod_psbl_users aa,
                 auth_user bb
            where aa.user_id = bb.id
              and aa.datalist_id = a.id
         ) edit_user
    from cust_data_list a
       left outer join cust_datalist_user_id_req_users cduiru
       on (
        a.id = cduiru.datalist_id
        and cduiru.user_id = {userid}
       )
       left outer join cust_data_request cdr
       on (
        a.id = cdr.id_datalist
        and cdr.id_reg_user = {userid}
        and cdr.poss_view_dtm > now()
       )
       left outer join
        (
        select id_datalist,round(avg(execute_time),2) t
        from cust_user_execute_sql
        where id_datalist is not null
        group by id_datalist
        ) cues
        on (
            a.id = cues.id_datalist
        ),
         auth_user b,
         cust_db_list d
      where 1=1
          and a.id_mod_user = b.id
        and a.id_dblist = d.id
        and a.id = {data_id}""".format(data_id=data_id,
                                       userid=request.user.id,
                                       super_user=request.user.is_superuser)

    ds_data_list = _query_dict("default", sql)

    sql_text = ds_data_list[0]['sql_text']

    # : 로 시작하는 단어를 찾는다.

    sql_text = sql_text.replace(":MM", "")
    sql_text = sql_text.replace(":mm", "")
    sql_text = sql_text.replace(":DD", "")
    sql_text = sql_text.replace(":dd", "")
    sql_text = sql_text.replace(":HH", "")
    sql_text = sql_text.replace(":hh", "")
    sql_text = sql_text.replace(":MI", "")
    sql_text = sql_text.replace(":mi", "")
    sql_text = sql_text.replace(":SS", "")
    sql_text = sql_text.replace(":ss", "")
    sql_text = sql_text.replace(":digit", "")
    sql_text = sql_text.replace(":DIGIT", "")

    regex = re.compile(":\w+")
    bind_value = regex.findall(sql_text)

    bind_value = _orderedSet(bind_value)

    sql = """
    select user_id
    from cust_datalist_user_id_mod_psbl_users
    where datalist_id = {data_id}
    union
    select id_reg_user
    from cust_data_list
    where id = {data_id}
  """.format(data_id=data_id)

    ds_psbl_users = _query_list("default", sql)

    context = {'ds_data_list': ds_data_list,
               'bind_value': bind_value,
               'ds_psbl_users': ds_psbl_users,
               }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxDataviewExecuteSQL(request):
    data = ""
    columnNames = ""
    data_title = ""
    end = ""
    error_yn = ""
    error_message = ""
    col_limit_cnt = 200
    row_limit_cnt = 50000

    sql_id = request.POST.get('sql_id', '')
    id_dblist = request.POST.get('id_dblist', '')
    data_title = request.POST.get('data_title', '')

    regex = re.compile("\w*")
    data_title = "".join(regex.findall(data_title))
    now = datetime.datetime.now()
    now = now.strftime("%Y%m%d")
    data_title = data_title + "_" + now

    txt_inputs = request.POST.getlist('txt_inputs[]')
    txt_binds = request.POST.getlist('txt_binds[]')

    error_yn = "N"
    error_message = ""
    excel_down_expyn = "N"

    sql = """
    select sql_text
    from cust_data_list a
    where a.id = {sql_id}
  """.format(sql_id=sql_id)

    ds_sql = _query_dict("default", sql)

    sql_text = ds_sql[0]['sql_text']

    org_sql = sql_text

    sql_text = sql_text.replace(";", "")

    i = 0
    in_param = ""
    txt_in_param = ""
    for txt in txt_inputs:
        if txt_binds[i] == "in_param":
            txt = txt.replace(' ', '')
            txt = txt.replace('\t', '')
            in_param = txt.split(",")

            cnt = len(in_param)

            if cnt > col_limit_cnt:
                error_yn = "Y"
                error_message = "멀티 입력값 개수는 총 {col_limit_cnt}개 까지 가능합니다. 입력된 개수는 {cnt}개 입니다.".format(cnt=cnt,
                                                                                                     col_limit_cnt=col_limit_cnt)
                break

            for v in in_param:
                txt_in_param = txt_in_param + "'" + v + "',"

            txt_in_param = txt_in_param + "|"
            txt_in_param = txt_in_param.replace(",|", "")

            sql_text = sql_text.replace(":" + txt_binds[i], txt_in_param)

        else:
            sql_text = sql_text.replace(":" + txt_binds[i], "'" + txt + "'")
        i = i + 1

    org_sql_cnt = sql_text

    if error_yn == "N":

        sql = """
      select svc_ip, port, inst_nm
      from cust_db_list a,
           cust_db_detail b
      where a.id = {id_dblist}
        and b.oper_cd = 3
        and b.id_dblist = a.id
        and b.db_order = 1;
    """.format(id_dblist=id_dblist)

        inst_nm = _query_dict("default", sql)
        inst_nm = inst_nm[0]['inst_nm']

        try:
            with connections[inst_nm].cursor() as cursor:

                sql_text = sql_text.replace(";", "")

                sql_text = """SELECT *
               FROM (
              """ + \
                           sql_text + \
                           """
            )
            where rownum <= {row_limit_cnt}""".format(row_limit_cnt=row_limit_cnt)

                new_userexecutesql = UserExecuteSQL(
                    sql_text=org_sql,
                    id_reg_user=request.user,
                    reg_dtm=datetime.datetime.now(),
                    id_datalist=sql_id,
                )

                new_userexecutesql.save()

                save_id = new_userexecutesql.id

                #####################################################
                ## 수행 시간 측정
                #####################################################
                start = time.time()

                cursor.execute(sql_text)

                end = time.time() - start
                #####################################################

                # try :
                #   id_datalist = DataList.objects.get(id=sql_id).id
                # except :
                #   id_datalist = None

                new_userexecutesql = UserExecuteSQL.objects.get(id=save_id)
                new_userexecutesql.execute_time = end

                new_userexecutesql.save()

                # new_userexecutesql = UserExecuteSQL(
                #       sql_text = org_sql,
                #       id_reg_user = request.user,
                #       reg_dtm = datetime.datetime.now(),
                #       execute_time = end,
                #       id_datalist = sql_id,
                # )

                # new_userexecutesql.save()

                execute_plus_datalist = DataList.objects.get(id=sql_id)
                execute_plus_datalist.execute_cnt += 1
                execute_plus_datalist.save()

                columnNames = [[d[0]] for d in cursor.description]

                # data = [dict(zip(columnNames, row)) for row in cursor]
                data = cursor.fetchall()
                data2 = []

                ###############################
                ## Decimal('53.94') -> str 로 변경
                ###############################
                buff1 = []
                buff2 = []

                ind = 0

                for ds in data:
                    for d in ds:

                        if isinstance(d, Decimal):
                            buff1.append(str(d))
                        elif isinstance(d, int) and columnNames[ind][0][-2:].upper() != 'CD' \
                                and columnNames[ind][0][-2:].upper() != 'ID' \
                                and columnNames[ind][0][-2:].upper() != '번호' \
                                and columnNames[ind][0][-2:].upper() != 'NO' \
                                and columnNames[ind][0][-2:] != '코드' \
                                and columnNames[ind][0][-4:] != '고객번호' \
                                and columnNames[ind][0][-6:] != 'CUSTNO' \
                                and columnNames[ind][0][-7:] != 'CUST_NO':
                            buff1.append(format(d, ','))
                        else:
                            buff1.append(d)
                        ind = ind + 1

                    buff2.append(buff1)
                    buff1 = []
                    ind = 0

                data = buff2

                ###############################

                if len(data) == row_limit_cnt:
                    sql_text = """SELECT count(*) cnt
               FROM (
              """ + \
                               org_sql_cnt + \
                               """
            ) """

                    cursor.execute(sql_text)
                    data_cnt = cursor.fetchall()[0][0]
                    data_cnt = format(data_cnt, ',')

                    error_yn = "Y"
                    # error_message = "[안내] 전체 {data_cnt}건 중 {row_limit_cnt}건만 보여줍니다. 전체 데이터는 아래 엑셀 다운로드 버튼을 통해 다운 받으세요".format(data_cnt=data_cnt,row_limit_cnt=row_limit_cnt)
                    # 비활성화 함
                    error_message = "[안내] 전체 {data_cnt}건 중 {row_limit_cnt}건만 표시합니다.".format(data_cnt=data_cnt,
                                                                                            row_limit_cnt=row_limit_cnt)
                    # excel_down_expyn = "Y"

                #################################################
                ## 고객번호 마스킹 부분 시작
                #################################################
                # for d in data :
                #   data2.append(list(d))

                # ind = 0
                # list_ind = []
                # for columnName in columnNames :
                #   if columnName[0] in ("CUST_NO",
                #                      "고객번호",
                #                      "고객_번호",
                #                      "CUSTNO",
                #                      "고객 번호") :
                #     list_ind.append(ind)

                #   ind = ind + 1

                # for i in list_ind :
                #   for d in data2 :
                #     d[i] = "*******"

                # data = data2

                #################################################
                ## 고객번호 마스킹 부분 끝
                #################################################

        except Exception as ex:
            error_yn = "Y"
            error_message = "[안내] 오류가 발생되었습니다. 관리자에 문의해주세요."

    context = {
        'ds_result': data,
        'ds_result_leng': len(data),
        'ds_columns': columnNames,
        'data_title': data_title,
        'time': end,
        'error_yn': error_yn,
        'error_message': error_message,
        'excel_down_expyn': excel_down_expyn,
        'row_limit_cnt': row_limit_cnt
    }

    # ds_result = [[1,2,3],[4,5,6]]
    # ds_columns = ['a','b','c']

    # context = {'ds_result': ds_result,
    #            'ds_columns': ds_columns}

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


# 데이터 수행 시 SQL 튜닝 요청
def ajaxRequestTuning(request):
    sql_id = request.POST.get('sql_id', '')
    tuning_title = request.POST.get('tuning_title', '')
    tuning_tobe_time = request.POST.get('tuning_tobe_time', '')
    tuning_asis_time = request.POST.get('tuning_asis_time', '')
    tuning_explain = request.POST.get('tuning_explain', '')

    new_tuninglist = TuningList(

        title=tuning_title,
        sql_type='2',
        id_tuningstatus=TuningStatus.objects.get(id=1),
        id_datalist=DataList.objects.get(id=sql_id),
        asis_sql_text=DataList.objects.get(id=sql_id).sql_text,
        asis_elapsed_time=tuning_asis_time,
        tobe_elapsed_time=tuning_tobe_time,
        id_reg_user=request.user,
        id_mod_user=request.user,
        reg_dtm=datetime.datetime.now(),
        mod_dtm=datetime.datetime.now(),
    )

    new_tuninglist.save()

    context = {
        'success': 'ok',
    }

    param1 = json.dumps(context, default=json_default)

    return HttpResponse(param1, content_type="application/json")


def databaseList(request):
    if not request.user.is_authenticated:
        return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

    template_parent = "dev"
    template_child = "databaseList"

    _loggingVisit(request, 15)

    header = [
        'DB용도',
        '고객정보',
        '임직원정보',
        'DBMS종류',
        'DBMS버전',
        'DBMS위치',
        '담당부서',
        '담당자',
        'EOS',
        'EOS날짜',
    ]
    ds_db_detail = DbDetail.objects.filter(oper_cd__in=['3'], use_yn="1", db_order=1,
                                           id_dblist__infra_op_yn=1).all().order_by('-oper_cd', 'id_dblist__exp_order',
                                                                                    'id_dblist__db_use', 'db_order')
    # ds_db_detail = DbDetail.objects.filter(oper_cd__in=['3'], use_yn="1", db_order=1).all().order_by('oper_cd','id_dblist__db_use','db_order')

    dbms_type = DbList.objects.filter(use_yn="1").values('id_dbtype__db_type').order_by('id_dbtype').annotate(
        count=Count('id_dbtype'))

    all_db = DbDetail.objects.filter( use_yn=1,
                                      sync_svc_ip__isnull=False).order_by('oper_cd', 'id_dblist__exp_order', 'id_dblist__db_use', 'db_order')

    approval_db = DbDetail.objects.filter(use_yn=1,
                                          db_order=1,
                                          oper_cd__id__in = [3,4],
                                          id_hacase=1,
                                          sync_svc_ip__isnull=False).order_by('oper_cd', 'id_dblist__exp_order', 'id_dblist__db_use', 'db_order')

    return render(request, 'database_list.html', {
        'template_parent': template_parent,
        'template_child': template_child,
        'all_db': all_db,
        'ds_db_detail': ds_db_detail,
        'approval_db': approval_db,
        'dbms_type': dbms_type,
        'header': header
    })


def msaTableTest(request):
    template_parent = "dba"
    template_child = "msaTableTest"

    return render(request, 'msaTableTest.html', {
        'template_parent': template_parent,
        'template_child': template_child,
    })

def makeDDL(table_info, db_case="ORACLE", owner="[owner].") :



    # no
    # column_id
    # eng_tab_name
    # kor_tab_name
    # kor_col_name
    # pk
    # not_null
    # default
    # eng_col_name
    # data_type
    # domain
    # stardard_yn
    # table_chg    
    # table_start
    # uk

    

    #########################################
    # MYSQL DDL MAKE
    #########################################
    if db_case == 'MYSQL' :

        table_column = ""
        table_column_comment = ""
        pk_index_column = ""
        uk_index_column = ""
        pk_exists = "N"
        uk_exists = "N"

        DDL_STR_CREATE_TABLE = ""
        DDL_END_CREATE_TABLE = ""
        DDL_IN_TABLE_COLUMN = ""
        DDL_CREATE_PK = ""
        DDL_CREATE_UK = ""
        START_YN = 'Y'
        TOTAL_DDL = ""

        table_name = ""
        table_comments = ""

        for table in table_info:


            no = table[0]
            column_id = table[1]
            kor_tab_name = table[2]
            eng_tab_name = table[3]
            kor_col_name = table[4]
            pk = table[5]
            not_null = table[6]
            default = table[7]
            eng_col_name = table[8]
            data_type = table[9]
            domain = table[10]
            stardard_yn = table[11]
            table_chg     = table[12]
            table_start     = table[13]
            uk     = table[14]

            if table_start == 'Y':

                DDL_STR_CREATE_TABLE = "create table {owner}{table_name} ( \n"
                DDL_END_CREATE_TABLE = ")\ncomment '{comment}' \nengine=innodb default charset=utf8mb4 collate=utf8mb4_general_ci;"
                DDL_IN_TABLE_COLUMN = "\t{column_name} {data_type}{not_null}{auto_increment}{data_default} comment '{comment}'"
                DDL_CREATE_PK = "\tprimary key ({pk_index_column})\n"
                DDL_CREATE_UK = "\t,unique index ux_{table_name}_01 ({uk_index_column})\n"




                if START_YN == 'N':
                    if pk_exists == "N":
                        table_column = table_column + "|"
                        table_column = table_column.replace(",\n|", "\n")

                    pk_index_column = pk_index_column + "|"
                    pk_index_column = pk_index_column.replace(", |", "")


                    uk_index_column = uk_index_column + "|"
                    uk_index_column = uk_index_column.replace(", |", "")

                    TOTAL_DDL += '\n\n' + \
                                 DDL_STR_CREATE_TABLE.format(table_name=table_name,owner=owner) + \
                                 table_column + \
                                 DDL_CREATE_PK.format(pk_index_column=pk_index_column) + \
                                 (DDL_CREATE_UK.format(table_name=table_name, uk_index_column=uk_index_column) if uk_exists == "Y" else "" )+ \
                                 DDL_END_CREATE_TABLE.format(comment=table_comments)

                    table_column = ""
                    table_column_comment = ""
                    pk_index_column = ""
                    pk_exists = "N"

                    uk_index_column = ""
                    uk_exists = "N"


            table_name = eng_tab_name
            table_comments = kor_tab_name

            if pk == "Y":
                pk_index_column = pk_index_column + eng_col_name.lower() + ", "
                pk_exists = "Y"

            if uk == "Y":
                uk_index_column = uk_index_column + eng_col_name.lower() + ", "
                uk_exists = "Y"

            default_keyword = ['NOW()', 'CURRENT_TIMESTAMP']
            org_default = default.rstrip(' ')
            org_default = org_default.lstrip(' ')
            in_default = ""

            if org_default.upper() in default_keyword:
                in_default = " default " + default
            elif org_default != "":
                in_default = " default '" + default + "'"

            table_column = table_column + DDL_IN_TABLE_COLUMN.format(
                column_name=eng_col_name.lower(),
                data_type=data_type.lower(),
                not_null=' not null' if not_null == "Y" else '',
                data_default=in_default,
                comment=kor_col_name,
                auto_increment=' auto_increment' if eng_col_name.lower() == "id" else '', ) + ",\n"
            START_YN = "N"

        if pk_exists == "N":
            table_column = table_column + "|"
            table_column = table_column.replace(",\n|", "\n")

        pk_index_column = pk_index_column + "|"
        pk_index_column = pk_index_column.replace(", |", "")

        uk_index_column = uk_index_column + "|"
        uk_index_column = uk_index_column.replace(", |", "")


        

        TOTAL_DDL += '\n\n' + \
                     DDL_STR_CREATE_TABLE.format(table_name=table_name,owner=owner) + \
                     table_column + \
                     DDL_CREATE_PK.format(pk_index_column=pk_index_column) + \
                     (DDL_CREATE_UK.format(table_name=table_name, uk_index_column=uk_index_column)if uk_exists == "Y" else "") + \
                     DDL_END_CREATE_TABLE.format(comment=table_comments)




    #########################################
    # ORACLE DDL MAKE
    #########################################
    if db_case == 'ORACLE' :

        table_column = ""
        table_column_comment = ""
        pk_index_column = ""
        pk_exists = "N"

        DDL_STR_CREATE_TABLE = ""
        DDL_END_CREATE_TABLE = ""
        DDL_IN_TABLE_COLUMN = ""
        DDL_CREATE_PK = ""
        START_YN = 'Y'
        TOTAL_DDL = ""

        table_name = ""
        table_comments = ""

        # owner = "[owner]."
        data_tbs = "[data_tbs]"
        ind_tbs = "[ind_tbs]"
        ddl_grants = "GRANT SELECT ON {owner}{table_name} TO RL_SM_SEL;\n" + \
                     "GRANT SELECT,UPDATE,DELETE,INSERT ON {owner}{table_name} TO [grantee];\n"

        for table in table_info:


            no = table[0]
            column_id = table[1]
            kor_tab_name = table[2]
            eng_tab_name = table[3]
            kor_col_name = table[4]
            pk = table[5]
            not_null = table[6]
            default = table[7]
            eng_col_name = table[8]
            data_type = table[9]
            domain = table[10]
            stardard_yn = table[11]
            table_chg     = table[12]
            table_start     = table[13]
            
            if table_start == 'Y':

                DDL_STR_CREATE_TABLE = "\nCREATE TABLE {owner}{table_name} ( \n"
                DDL_END_CREATE_TABLE = ")"
                DDL_IN_TABLE_COLUMN = "\t{column_name} {data_type}{data_default}{not_null}"
                DDL_DATA_TABLESPACE = " TABLESPACE {data_tbs};"
                DDL_IND_TABLESPACE = " TABLESPACE {ind_tbs};"
                DDL_STR_CREATE_PK = "CREATE UNIQUE INDEX {owner}PK_{table_name} ON {owner}{table_name} ( \n"
                DDL_END_CREATE_PK = ")"

                DDL_STR_CREATE_CONST_PK = "ALTER TABLE {owner}{table_name} ADD CONSTRAINT PK_{table_name} PRIMARY KEY ( \n"
                DDL_END_CREATE_CONST_PK = ");"

                DDL_IN_INDEX_COLUMN = "{pk_index_column}"

                DDL_SYNONYM = "CREATE OR REPLACE PUBLIC SYNONYM {table_name} FOR {owner}{table_name};"

                DDL_GRANT = "GRANT {crud} ON {owner}{table_name} TO {role};\n"

                DDL_TABLE_COMMENT = "COMMENT ON TABLE {owner}{table_name} IS '{comment}';"
                DDL_COLUMN_COMMENT = "COMMENT ON COLUMN {owner}{table_name}.{column_name} IS '{comment}';"




                if START_YN == 'N':
                    table_column = table_column + "|"
                    table_column = table_column.replace(",\n|", "\n")

                    pk_index_column = pk_index_column + "|"
                    pk_index_column = pk_index_column.replace(", |", "")

                    TOTAL_DDL += DDL_STR_CREATE_TABLE.format(owner=owner, table_name=table_name) + \
                                table_column + \
                                DDL_END_CREATE_TABLE + \
                                DDL_DATA_TABLESPACE.format(data_tbs=data_tbs) + \
                                "\n\n" + \
                                DDL_STR_CREATE_PK.format(owner=owner, table_name=table_name) + \
                                pk_index_column + \
                                DDL_END_CREATE_PK + \
                                DDL_IND_TABLESPACE.format(ind_tbs=ind_tbs) + \
                                "\n\n" + \
                                DDL_STR_CREATE_CONST_PK.format(owner=owner, table_name=table_name) + \
                                pk_index_column + \
                                DDL_END_CREATE_CONST_PK + \
                                "\n\n" + \
                                DDL_SYNONYM.format(owner=owner, table_name=table_name) + \
                                "\n\n" + \
                                ddl_grants.format(owner=owner, table_name=table_name) + \
                                "\n" + \
                                DDL_TABLE_COMMENT.format(owner=owner, table_name=table_name, comment=table_comments) + \
                                "\n" + \
                                table_column_comment

                    # DDL_CREATE_PK.format(pk_index_column=pk_index_column) + \
                    
                    table_column = ""
                    table_column_comment = ""
                    pk_index_column = ""


            table_name = eng_tab_name
            table_comments = kor_tab_name

            if pk == "Y":
                pk_index_column = pk_index_column + eng_col_name.lower() + ", "

            default_keyword = ['SYSDATE', ]
            org_default = default.rstrip(' ')
            org_default = org_default.lstrip(' ')
            in_default = ""

            if org_default.upper() in default_keyword:
                in_default = " default " + default
            elif org_default != "":
                in_default = " default '" + default + "'"

            table_column = table_column + DDL_IN_TABLE_COLUMN.format(
                column_name=eng_col_name.lower(),
                data_type=data_type.lower(),
                not_null=' not null' if not_null == "Y" else '',
                data_default=in_default,
            ) + ",\n"
                # comment=kor_col_name,
                # auto_increment=' auto_increment' if eng_col_name.lower() == "id" else '', 

            
            table_column_comment = table_column_comment + DDL_COLUMN_COMMENT.format(
                owner=owner,
                table_name=table_name,
                column_name=eng_col_name.lower(),
                comment=kor_col_name,
            ) + "\n"
            START_YN = "N"

        table_column = table_column + "|"
        table_column = table_column.replace(",\n|", "\n")

        pk_index_column = pk_index_column + "|"
        pk_index_column = pk_index_column.replace(", |", "")



        
        # TOTAL_DDL += '\n\n' + \
        #              DDL_STR_CREATE_TABLE.format(table_name=table_name) + \
        #              table_column + \
        #              DDL_END_CREATE_TABLE + \
        #              DDL_CREATE_PK.format(pk_index_column=pk_index_column) + \
        #              DDL_END_CREATE_TABLE.format(comment=table_comments)

        
        TOTAL_DDL += '\n' + \
                    DDL_STR_CREATE_TABLE.format(owner=owner, table_name=table_name) + \
                    table_column + \
                    DDL_END_CREATE_TABLE + \
                    DDL_DATA_TABLESPACE.format(data_tbs=data_tbs) + \
                    "\n\n" + \
                    DDL_STR_CREATE_PK.format(owner=owner, table_name=table_name) + \
                    pk_index_column + \
                    DDL_END_CREATE_PK + \
                    DDL_IND_TABLESPACE.format(ind_tbs=ind_tbs) + \
                    "\n\n" + \
                    DDL_STR_CREATE_CONST_PK.format(owner=owner, table_name=table_name) + \
                    pk_index_column + \
                    DDL_END_CREATE_CONST_PK + \
                    "\n\n" + \
                    DDL_SYNONYM.format(owner=owner, table_name=table_name) + \
                    "\n\n" + \
                    ddl_grants.format(owner=owner, table_name=table_name) + \
                    "\n" + \
                    DDL_TABLE_COMMENT.format(owner=owner, table_name=table_name, comment=table_comments) + \
                    "\n" + \
                    table_column_comment

        

    
    return TOTAL_DDL


class CreateDDLLV(LoginRequiredMixin, FormMixin, ListView):
    model = MetaReq
    template_name = 'CreateDDLLV.html'
    context_object_name = 'objects'
    form_class = CreateDDLLVForm
    paginate_by = 99999999

    def get_initial(self):


        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################

        sheet_name = self.request.session['sheet_name'] if 'sheet_name' in self.request.session else ""
        return {
            'sheet_name': sheet_name,            
        }

   

    def get_queryset(self):

        # if self.request.method == 'POST' :
        #     print(self.request.FILES['file_erd'].read())

        self.form = self.get_form(self.form_class)        

        context = {}
        meta_list = ""
        TOTAL_DDL = ""
        err_message = ""
        table_cnt = ""
        std_eng_column_name = ""   
        sheet_name = ""     


        if self.form.is_valid():
            
            sheet_name = self.form.cleaned_data['sheet_name']

            if len(sheet_name) > 0 and len(self.request.FILES) == 0:

                ###################################################
                ## 화면에 바인딩 된 값을 유지하기 위한 처리
                ###################################################
                self.request.session['sheet_name'] = sheet_name if sheet_name is not None else ""

                scope = [
                    'https://spreadsheets.google.com/feeds',
                    'https://www.googleapis.com/auth/drive',
                ]

                

                json_file_name = settings.BASE_DIR + '/static/custom/json/chiho-test-22b79f37651a.json'
                credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
                gc = gspread.authorize(credentials)
                spreadsheet_url = 'https://docs.google.com/spreadsheets/d/1SDjOVFaM_zGFObbsBSvofFkcHtT51siQxp9LdZtXBQQ/edit#gid=0'
                # 스프레스시트 문서 가져오기
                doc = gc.open_by_url(spreadsheet_url)
                # 시트 선택하기
                # worksheet = doc.worksheet('DDL정보입력(시트이름수정불가)')
                worksheet = ""

                try:
                    worksheet = doc.worksheet(sheet_name)
                except:
                    worksheet = doc.worksheet("기본템플릿(수정금지)")
                    pass

                all_meta_list = worksheet.get_all_values()               

                x_len = len(all_meta_list[4])

                err_message = []

                meta_dict = {}
                meta_list = []

                TOTAL_DDL = ""

                befor_table_name = "START"

                owner = ""
                if all_meta_list[0][1] != "":
                    owner = all_meta_list[0][1].lower() + "."
                data_tablespace = all_meta_list[1][1]
                index_tablespace = all_meta_list[2][1]

                first_table_name = "NEW|START"

                no = 1
                column_id = 1
                table_cnt = 0

                for meta in all_meta_list[6:]:

                    i = 0
                    meta_dict = {}

                    # ## 반영된 변경분은 스킵
                    # if meta[10] != "" :
                    #   continue

                    # while i < x_len :
                    # meta_dict[all_meta_list[0][i]] = meta[i]

                    if first_table_name != meta[1]:
                        meta_dict['START'] = 'Y'
                        first_table_name = meta[1]
                        column_id = 1
                        table_cnt = table_cnt + 1
                    else:
                        meta_dict['START'] = 'N'

                    sql = """SELECT A.STD_ATTR_ENG, LOWER(B.MYSQL_DATATYPE) DATA_TYPE, B.INFO_TYPE
                      from cust_std_attr a,
                           cust_std_domain b
                      where b.id = a.id_stddomain
                      and a.std_attr_kor = '{kor_column_name}'
                      and   a.use_yn = 1
                      and   b.use_yn = 1;
                  """.format(kor_column_name=meta[2])

                    std_info = _query_dict("default", sql)

                    std_column_name = ""
                    std_domain = ""
                    std_data_type = ""

                    if len(std_info) > 0:
                        std_eng_column_name = std_info[0]['STD_ATTR_ENG'].lower()
                        std_domain = std_info[0]['INFO_TYPE'].lower()
                        std_data_type = std_info[0]['DATA_TYPE'].lower()

                    meta_dict['NO'] = no
                    meta_dict['COLUMN_ID'] = column_id
                    meta_dict['KOR_TABLE_NAME'] = meta[0]
                    meta_dict['ENG_TABLE_NAME'] = meta[1].lower()
                    meta_dict['KOR_COLUMN_NAME'] = meta[2]
                    meta_dict['PK'] = meta[3].upper()
                    meta_dict['NOT_NULL'] = meta[4].upper()
                    meta_dict['DEFAULT'] = meta[5].lower()

                    if meta[6].lower() == "":
                        meta_dict['ENG_COLUMN_NAME'] = std_eng_column_name
                    else:
                        meta_dict['ENG_COLUMN_NAME'] = meta[6].lower()

                    if meta[7].lower() == "":
                        meta_dict['DATA_TYPE'] = std_data_type
                    else:
                        meta_dict['DATA_TYPE'] = meta[7].lower()

                    if re.sub(r'[A-Za-z0-9()_가-힣 ]', '', (meta_dict['KOR_TABLE_NAME'])) != "" or \
                            re.sub(r'[A-Za-z0-9_가-힣]', '', (meta_dict['ENG_TABLE_NAME'])) != "" or \
                            re.sub(r'[A-Za-z0-9_가-힣]', '', (meta_dict['KOR_COLUMN_NAME'])) != "" or \
                            re.sub(r'[A-Za-z0-9_가-힣]', '', (meta_dict['ENG_COLUMN_NAME'])) != "" or \
                            re.sub(r'[A-Za-z0-9_()가-힣]', '', (meta_dict['DATA_TYPE'])) != "":
                        err_message.append("<li>특수문자사용 됨 : #{no} 번째 라인</li>".format(no=no))

                    if meta_dict['KOR_TABLE_NAME'] == "" or \
                            meta_dict['ENG_TABLE_NAME'] == "" or \
                            meta_dict['KOR_COLUMN_NAME'] == "":
                        err_message.append("<li>정보 부족 : #{no} 번째 라인</li>".format(no=no))

                    if column_id == 1 and meta_dict['PK'] != "Y":
                        err_message.append("<li>PK없음 : #{no} 번째 라인</li>".format(no=no))

                    # 표준
                    if std_eng_column_name != "" and \
                            std_data_type != "" and \
                            std_eng_column_name == meta_dict['ENG_COLUMN_NAME'].lower() and \
                            std_data_type == meta_dict['DATA_TYPE'].lower():
                        meta_dict['STD_YN'] = '<font color="blue">표준</font>'
                        meta_dict['INFO_TYPE'] = std_domain
                    # 비표준
                    else:
                        meta_dict['STD_YN'] = '<b><font color="red">비표준</font></b>'
                        err_message.append("<li>비표준 컬럼 존재 : #{no} 번째 라인</li>".format(no=no))

                    # meta_dict['INDEX'] =
                    meta_list.append(meta_dict)
                    no += 1
                    column_id += 1

               

                

        context['meta_list'] = meta_list
        context['ddl'] = TOTAL_DDL
        context['err_message'] = err_message
        
        context['table_cnt'] = table_cnt

        obj = []
        obj.append(context)

        return obj

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context['template_parent'] = "dba"
        context['template_child'] = "msaTable"

        return context

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)


def msaTable(request):
    template_parent = "dba"
    template_child = "msaTable"

    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive',
    ]

    # json_file_name = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))+'/static/custom/json/chiho-test-22b79f37651a.json'

    json_file_name = settings.BASE_DIR + '/static/custom/json/chiho-test-22b79f37651a.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
    gc = gspread.authorize(credentials)
    spreadsheet_url = 'https://docs.google.com/spreadsheets/d/1SDjOVFaM_zGFObbsBSvofFkcHtT51siQxp9LdZtXBQQ/edit#gid=0'
    # 스프레스시트 문서 가져오기
    doc = gc.open_by_url(spreadsheet_url)
    # 시트 선택하기
    worksheet = doc.worksheet('DDL정보입력(시트이름수정불가)')

    # cell_data = worksheet.acell('E2').value

    # row_data = worksheet.row_values(1)

    # column_data = worksheet.col_values(1)

    # # 범위(셀 위치 리스트) 가져오기
    # range_list = worksheet.range('A1:D2')

    # # 범위에서 각 셀 값 가져오기
    # for cell in range_list:
    #

    # row_number = cell.row

    # column_number = cell.col

    # last_row = worksheet.row_count
    # last_row = 5
    # range_list = worksheet.range('A1:H'+str(last_row))

    # for cell in range_list :
    #   if cell.value is not None :

    all_meta_list = worksheet.get_all_values()

    x_len = len(all_meta_list[4])

    err_message = []

    meta_dict = {}
    meta_list = []

    TOTAL_DDL = ""

    befor_table_name = "START"

    owner = ""
    if all_meta_list[0][1] != "":
        owner = all_meta_list[0][1].lower() + "."
    data_tablespace = all_meta_list[1][1]
    index_tablespace = all_meta_list[2][1]

    first_table_name = "NEW|START"

    no = 1
    column_id = 1
    table_cnt = 0
    for meta in all_meta_list[6:]:
        i = 0
        meta_dict = {}

        # ## 반영된 변경분은 스킵
        # if meta[10] != "" :
        #   continue

        # while i < x_len :
        # meta_dict[all_meta_list[0][i]] = meta[i]

        if first_table_name != meta[1]:
            meta_dict['START'] = 'Y'
            first_table_name = meta[1]
            column_id = 1
            table_cnt = table_cnt + 1
        else:
            meta_dict['START'] = 'N'

        sql = """SELECT A.STD_ATTR_ENG, LOWER(B.MYSQL_DATATYPE) DATA_TYPE, B.INFO_TYPE
              from cust_std_attr a,
                   cust_std_domain b
              where b.id = a.id_stddomain
              and a.std_attr_kor = '{kor_column_name}'
              and   a.use_yn = 1
              and   b.use_yn = 1;
          """.format(kor_column_name=meta[2])

        std_info = _query_dict("default", sql)

        meta_dict['NO'] = no
        meta_dict['COLUMN_ID'] = column_id
        meta_dict['KOR_TABLE_NAME'] = meta[0]
        meta_dict['ENG_TABLE_NAME'] = meta[1].lower()
        meta_dict['KOR_COLUMN_NAME'] = meta[2]
        meta_dict['PK'] = meta[3].upper()
        meta_dict['NOT_NULL'] = meta[4].upper()
        meta_dict['DEFAULT'] = meta[5].lower()

        if meta[6].lower() == "":
            meta_dict['ENG_COLUMN_NAME'] = std_info[0]['STD_ATTR_ENG']
        else:
            meta_dict['ENG_COLUMN_NAME'] = meta[6].lower()

        if meta[7].lower() == "":
            meta_dict['DATA_TYPE'] = std_info[0]['DATA_TYPE']
        else:
            meta_dict['DATA_TYPE'] = meta[7].lower()

        if re.sub(r'[A-Za-z0-9()_가-힣]', '', (meta_dict['KOR_TABLE_NAME'])) != "" or \
                re.sub(r'[A-Za-z0-9_가-힣]', '', (meta_dict['ENG_TABLE_NAME'])) != "" or \
                re.sub(r'[A-Za-z0-9_가-힣]', '', (meta_dict['KOR_COLUMN_NAME'])) != "" or \
                re.sub(r'[A-Za-z0-9_가-힣]', '', (meta_dict['ENG_COLUMN_NAME'])) != "" or \
                re.sub(r'[A-Za-z0-9_()가-힣]', '', (meta_dict['DATA_TYPE'])) != "":
            err_message.append("<li>특수문자사용 됨 : {no} 번째 라인</li>".format(no=no))

        if meta_dict['KOR_TABLE_NAME'] == "" or \
                meta_dict['ENG_TABLE_NAME'] == "" or \
                meta_dict['KOR_COLUMN_NAME'] == "":
            err_message.append("<li>정보 부족 : {no} 번째 라인</li>".format(no=no))

        if column_id == 1 and meta_dict['PK'] != "Y":
            err_message.append("<li>PK없음 : {no} 번째 라인</li>".format(no=no))

        # 표준
        if std_info[0]['STD_ATTR_ENG'].lower() == meta_dict['ENG_COLUMN_NAME'].lower() and std_info[0][
            'DATA_TYPE'].lower() == meta_dict['DATA_TYPE'].lower():
            meta_dict['STD_YN'] = '<font color="blue">표준</font>'
            meta_dict['INFO_TYPE'] = std_info[0]['INFO_TYPE']
        # 비표준
        else:
            meta_dict['STD_YN'] = '<b><font color="red">비표준</font></b>'
            err_message.append("<li>비표준 컬럼 존재 : {no} 번째 라인</li>".format(no=no))

        # meta_dict['INDEX'] =
        meta_list.append(meta_dict)
        no += 1
        column_id += 1

    # 첫줄 제거
    # all_meta = all_meta[1:]

    # for table in all_meta :
    #   if

    # if request.method == "POST" :

    # for table in meta_list :

    # drop_table = "DROP TABLE {table_name};\n"
    # start_create_table = "CREATE TABLE {table_name} (\n"
    # close_create_table = ") ENGINE=InnoDB DEFAULT CHARSET=utf8;\n\n"
    # create_column = "\t{column_name} {data_type} {not_null},\n"
    # primary_key_ddl = "\tPRIMARY KEY ({primary_key})\n"
    # not_null_ddl = ""

    table_column = ""
    table_column_comment = ""
    pk_index_column = ""
    pk_exists = "N"

    DDL_STR_CREATE_TABLE = ""
    DDL_END_CREATE_TABLE = ""
    DDL_IN_TABLE_COLUMN = ""
    DDL_CREATE_PK = ""
    START_YN = 'Y'

    table_name = ""
    table_comments = ""

    for table in meta_list:

        if table['START'] == 'Y':

            DDL_STR_CREATE_TABLE = "create table {owner}{table_name} ( \n"
            DDL_END_CREATE_TABLE = ")\ncomment '{comment}' \nengine=innodb default charset=utf8mb4 collate=utf8mb4_general_ci;"
            DDL_IN_TABLE_COLUMN = "\t{column_name} {data_type}{not_null}{auto_increment}{data_default} comment '{comment}'"
            DDL_CREATE_PK = "\tprimary key ({pk_index_column})\n"

            if START_YN == 'N':
                if pk_exists == "N":
                    table_column = table_column + "|"
                    table_column = table_column.replace(",\n|", "\n")

                pk_index_column = pk_index_column + "|"
                pk_index_column = pk_index_column.replace(", |", "")

                TOTAL_DDL += '\n\n' + \
                             DDL_STR_CREATE_TABLE.format(owner=owner, table_name=table_name) + \
                             table_column + \
                             DDL_CREATE_PK.format(pk_index_column=pk_index_column) + \
                             DDL_END_CREATE_TABLE.format(comment=table_comments)

                table_column = ""
                table_column_comment = ""
                pk_index_column = ""
                pk_exists = "N"

        table_name = table['ENG_TABLE_NAME']
        table_comments = table['KOR_TABLE_NAME']

        if table['PK'] == "Y":
            pk_index_column = pk_index_column + table['ENG_COLUMN_NAME'].lower() + ", "
            pk_exists = "Y"

        default_keyword = ['NOW()', 'CURRENT_TIMESTAMP']
        org_default = table['DEFAULT'].rstrip(' ')
        org_default = org_default.lstrip(' ')
        in_default = ""

        if org_default.upper() in default_keyword:
            in_default = " default " + table['DEFAULT']
        elif org_default != "":
            in_default = " default '" + table['DEFAULT'] + "'"

        table_column = table_column + DDL_IN_TABLE_COLUMN.format(
            column_name=table['ENG_COLUMN_NAME'].lower(),
            data_type=table['DATA_TYPE'].lower(),
            not_null=' not null' if table['NOT_NULL'] == "Y" else '',
            data_default=in_default,
            comment=table['KOR_COLUMN_NAME'],
            auto_increment=' auto_increment' if table['ENG_COLUMN_NAME'].lower() == "id" else '', ) + ",\n"
        START_YN = "N"

    if pk_exists == "N":
        table_column = table_column + "|"
        table_column = table_column.replace(",\n|", "\n")

    pk_index_column = pk_index_column + "|"
    pk_index_column = pk_index_column.replace(", |", "")

    TOTAL_DDL += '\n\n' + \
                 DDL_STR_CREATE_TABLE.format(owner=owner, table_name=table_name) + \
                 table_column + \
                 DDL_CREATE_PK.format(pk_index_column=pk_index_column) + \
                 DDL_END_CREATE_TABLE.format(comment=table_comments)

    #   table_name = table['ENG_TABLE_NAME']

    #   if table_name != befor_table_name :

    #     # primary_key = ""

    #     if befor_table_name != "START" :
    #       primary_key = primary_key + "|"
    #       primary_key = primary_key.replace(",|","")
    #       ddl = ddl + primary_key_ddl.format(primary_key = primary_key)
    #       ddl = ddl + close_create_table

    #       primary_key = ""

    #     if table["PK"].upper() == "Y" :
    #       primary_key += table["KOR_COLUMN_NAME"] + ","

    #     ddl = ddl + drop_table.format(table_name = table_name)
    #     ddl = ddl + start_create_table.format(table_name = table_name)

    #   else :

    #     if table["PK"].upper() == "Y" :
    #       primary_key += table["ENG_TABLE_NAME"] + ","

    #   if table['NOT_NULL'] == "Y" :
    #     not_null_ddl = "NOT NULL"
    #   else :
    #     not_null_ddl = ""

    #   ddl = ddl + create_column.format(column_name = table['ENG_COLUMN_NAME'],
    #                                  data_type = table['DATA_TYPE'],
    #                                  not_null = not_null_ddl)

    #   befor_table_name = table_name

    # primary_key = primary_key + "|"
    # primary_key = primary_key.replace(",|","")
    # ddl = ddl + primary_key_ddl.format(primary_key = primary_key)
    # ddl = ddl + close_create_table

    return render(request, 'msa_table.html', {
        'template_parent': template_parent,
        'template_child': template_child,
        'meta_list': meta_list,
        'ddl': TOTAL_DDL,
        'err_message': err_message,
        'table_cnt': table_cnt
    })


# def export_users_csv(request):
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="users.csv"'

#     writer = csv.writer(response)
#     writer.writerow(['Username', 'First name', 'Last name', 'Email address'])

#     users = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
#     for user in users:
#         writer.writerow(user)

#     return response


class OneDBMonitorLV(LoginRequiredMixin, FormMixin, ListView):

    model = MonitorItemLog
    template_name = 'OneDBMonitorLV.html'
    context_object_name = 'objects'
    form_class = MonitorForm

    def get_initial(self):



        end_date = datetime.datetime.now()
        str_date = end_date - datetime.timedelta(hours=12)

        return {
            'str_date': str_date,
            'end_date': end_date,
        }

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):


        start_time = time.time()                    

        id_dbdetail = self.kwargs['id']

        end_date = self.request.POST.get('end_date', '')
        str_date = self.request.POST.get('str_date', '')
        interval = self.request.POST.get('interval', '')

        if interval == "":
            interval = 1

        if end_date == "":
            end_date = datetime.datetime.now()
            str_date = end_date - datetime.timedelta(hours=12)
            end_date = datetime.datetime.now() + datetime.timedelta(minutes=5)

        dbdetail = DbDetail.objects.get(id=id_dbdetail)
        db_nm = dbdetail.id_dblist.db_use + '(' + str(dbdetail.db_order) + ')'
        id_db_type = dbdetail.id_dblist.id_dbtype.id
        host_nm = dbdetail.host_nm
        db_order = dbdetail.db_order
        alert_yn = dbdetail.alert_yn

        if isinstance(str_date, datetime.datetime) == False :
            str_date = datetime.datetime.strptime(str_date,'%Y-%m-%d %H:%M')

        if isinstance(end_date, datetime.datetime) == False :
            end_date = datetime.datetime.strptime(end_date,'%Y-%m-%d %H:%M')

        delta = end_date - str_date      
        delta_h = delta.total_seconds()/60/60

        if delta_h <= 3*24 :
            interval = 1
        elif delta_h > 3*24 and  delta_h <= 7*24 :
            interval = 10
        elif delta_h > 7*24 :
            interval = 60        
        

        sql = """
        SELECT
          ITEM_NM,
          RELATIVE_CASE,
          MONITOR_VALUE,
          DATE_FORMAT(concat(DT,' ',H,':',m*{interval}),'%Y-%m-%d %H:%i') REG_DTM,
          OPER_CD,
          DB_USE,
          DB_ORDER,
          HOST_NM,
          SVC_IP,
          ID_DBDETAIL,
          ID_MONITORITEMLIST
        FROM
        (
          SELECT
            E.ITEM_NM,
            C.RELATIVE_CASE,
            E.EXP_ORDER,
            MAX(A.MONITOR_VALUE) MONITOR_VALUE,
            DATE(A.REG_DTM) DT, 
            HOUR(A.REG_DTM) H, 
            FLOOR(MINUTE(A.REG_DTM)/{interval}) M,
            -- DATE_FORMAT(A.REG_DTM,'%Y.%m.%d %H:%i') REG_DTM,
            MAX(D.OPER_CD) OPER_CD,
            MAX(B.DB_USE) DB_USE,
            MAX(D.DB_ORDER) DB_ORDER,
            MAX(D.HOST_NM) HOST_NM,
            MAX(D.SVC_IP) SVC_IP,
            MAX(D.ID) ID_DBDETAIL,
            MAX(C.ID) ID_MONITORITEMLIST
          FROM CUST_MONITOR_ITEM_LOG A USE INDEX (ix_cust_monitor_item_log_03),
           CUST_DB_LIST B,
           CUST_MONITOR_ITEM_LIST C,
           CUST_DB_DETAIL D,
           CUST_MONITOR_ITEM E,
           cust_dbdetail_monitoritemlist f
          -- WHERE A.REG_DTM >= DATE_ADD(STR_TO_DATE(DATE_FORMAT(NOW(), '%Y%m%d %H%i'),'%Y%m%d %H%i'), INTERVAL -1 hour)
          WHERE A.REG_DTM between DATE_FORMAT('{str_date}','%Y-%m-%d %H:%i') and DATE_FORMAT('{end_date}','%Y-%m-%d %H:%i')
          AND D.ID = {id_dbdetail}
          AND A.id_dbdetailmonitoritemlist = f.id
          AND f.id_monitoritemlist = c.id
          AND c.id_monitoritem = e.id
          AND f.id_dbdetail = d.id
          AND d.id_dblist = b.id
          AND C.GRAPH_YN = 1            
          AND E.ID <> 1 -- Health Check
          GROUP BY E.ITEM_NM, E.EXP_ORDER, C.RELATIVE_CASE, DATE(A.REG_DTM), HOUR(A.REG_DTM), FLOOR(MINUTE(A.REG_DTM)/{interval})          
          ) A
        ORDER BY A.EXP_ORDER, A.DT, A.H, A.M, A.ID_MONITORITEMLIST
      """.format(id_dbdetail=id_dbdetail,
                 str_date=str_date,
                 end_date=end_date,
                 interval=interval)

        print(sql)

        queryset = _query_dict("default", sql)
        
        elapsed_time = time.time() - start_time                            


        graph_list = defaultdict(list)

        for graph in queryset:
            graph_list[graph['ITEM_NM']].append(graph)

        graph_draw = []

        for key, value in graph_list.items():

            ind = 0

            label = []
            labels = []
            data = defaultdict(list)
            backgroundColor = []
            item_nm = ""

            id_dbdetail = 0
            id_monitoritemlist = 0
            data_lengh = len(value)

            
            before_monitor_value = value[0]['MONITOR_VALUE']
            before_reg_dtm = value[0]['REG_DTM'].split('.')[0]
            

            for team in value:                

                id_dbdetail = team['ID_DBDETAIL']
                id_monitoritemlist = team['ID_MONITORITEMLIST']

                item_nm = team['ITEM_NM']
                label.append(item_nm)

                reg_dtm = team['REG_DTM'].split('.')[0]

                if str(str_date).split()[0] == str(end_date).split()[0] :
                    
                    labels.append(reg_dtm.split()[1])
                elif str(str_date).split('-')[0] == str(end_date).split('-')[0] :
                    labels.append(reg_dtm[8:])  
                else :
                    labels.append(reg_dtm[5:])  
                
                if value[0]['RELATIVE_CASE'] == '0' :
                    data[item_nm].append(team['MONITOR_VALUE'])
                else :
                    if before_monitor_value == team['MONITOR_VALUE'] :
                        data[item_nm].append(0)
                    else :                
                        try :        
                            time_diff = datetime.datetime.strptime(reg_dtm,'%Y-%m-%d %H:%M') - datetime.datetime.strptime(before_reg_dtm,'%Y-%m-%d %H:%M')
                            time_diff = time_diff.total_seconds()                                                        

                            
                            diff = int(team['MONITOR_VALUE']) - int(before_monitor_value)
                            # div = int(time_diff/60) # 분당
                            div = int(time_diff) # 초당
                            data[item_nm].append(math.ceil(diff/div))
                            # if diff > 60 :
                            #     data[item_nm].append(round(diff/time_diff))
                            # else :
                            #     data[item_nm].append(1)
                        except Exception as e:
                            print(str(e))
                            data[item_nm].append(0)

                
                backgroundColor.append('rgba(54, 162, 235, 1)',)

                if len(totalColor) == ind + 1:
                    ind = 0
                else:
                    ind += 1

                before_monitor_value = team['MONITOR_VALUE']
                before_reg_dtm = reg_dtm

            labels = list(set(labels))
            labels.sort()
            label = list(set(label))

            backgroundColor = list(set(backgroundColor))

            item = []

            ind = 0

            for t in label:
                inner_dict = {}
                inner_dict["data"] = data[t]
                inner_dict["label"] = t
                inner_dict["borderColor"] = backgroundColor[ind]                
                inner_dict["radius"] = 0
                inner_dict["fill"] = False
                inner_dict["pointStyle"] = "line"
                item.append(inner_dict)
                ind += 1

            limitValue = DBDetailMonitorItemList.objects.filter(id_dbdetail=id_dbdetail,
                                                                id_monitoritemlist=id_monitoritemlist)
            limit = limitValue[0].limit_value
            
            if len(limit) > 0 :
                inner_dict = {}
                inner_dict["data"] = [limit] * data_lengh
                inner_dict["label"] = "임계치"
                inner_dict["borderColor"] = 'rgba(255, 99, 132, 0.5)'
                inner_dict["radius"] = 0
                inner_dict["fill"] = False
                inner_dict["pointStyle"] = "line"
                item.append(inner_dict)

            graph_draw.append([item_nm, item, labels])

        context = super().get_context_data(**kwargs)

        db_detail = DbDetail.objects.get(id=id_dbdetail)

        print("## 일반 모니터링 임계치 응답시간 ",time.time()-start_time)
        

        context['id_dbdetail'] = self.kwargs['id']

        context['id_db_type'] = id_db_type
        context['db_order'] = db_order
        context['db_nm'] = db_nm
        context['db_detail'] = db_detail

        context['graph_draw'] = graph_draw
        context['alert_yn'] = alert_yn

        context['template_parent'] = "dba"
        context['template_child'] = "DBAWorkLV"

        return context


def ajaxOneDBMonitorLV(request):
    activeTab = request.POST.get('activeTab', '')
    id_dbdetail = request.POST.get('id_dbdetail', '')

    ymax = 100
    ymin = 0

    start_time = time.time()                    

    if activeTab == 'filesystem' :
        sql = """
            select DATE_FORMAT(a.REG_DTM,'%Y/%m/%d') REG_DTM, B.NOTE1 FILESYSTEM, A.MONITOR_VALUE USAGE_PERCENT
            from cust_monitor_item_log  a use index(ix_cust_monitor_item_log_02),
                 cust_dbdetail_monitoritemlist b,
                 cust_monitor_item_list c
            where b.id_dbdetail = {id_dbdetail}
            AND a.curr_yn = '1'
            AND a.id_dbdetailmonitoritemlist = b.id
            AND c.id = b.id_monitoritemlist
            and c.id = 20
            union all
            select DATE_FORMAT(a.REG_DTM,'%Y/%m/%d') dt, b.note1, max(a.monitor_value) usage_percent
            from cust_monitor_item_log a  use index(ix_cust_monitor_item_log_03),
                 cust_dbdetail_monitoritemlist b,
                 cust_monitor_item_list c
            where b.id_dbdetail = {id_dbdetail}
            AND a.id_dbdetailmonitoritemlist = b.id
            AND (  a.reg_dtm between  DATE_FORMAT('{day1} 23:58:00', '%Y-%m-%d %H:%i:%s') and DATE_FORMAT('{day1} 23:59:59', '%Y-%m-%d %H:%i:%s')
                or a.reg_dtm between  DATE_FORMAT('{day2} 23:58:00', '%Y-%m-%d %H:%i:%s') and DATE_FORMAT('{day2} 23:59:59', '%Y-%m-%d %H:%i:%s')                
                or a.reg_dtm between  DATE_FORMAT('{day3} 23:58:00', '%Y-%m-%d %H:%i:%s') and DATE_FORMAT('{day3} 23:59:59', '%Y-%m-%d %H:%i:%s')                
                or a.reg_dtm between  DATE_FORMAT('{day4} 23:58:00', '%Y-%m-%d %H:%i:%s') and DATE_FORMAT('{day4} 23:59:59', '%Y-%m-%d %H:%i:%s')                
                or a.reg_dtm between  DATE_FORMAT('{day5} 23:58:00', '%Y-%m-%d %H:%i:%s') and DATE_FORMAT('{day5} 23:59:59', '%Y-%m-%d %H:%i:%s')                
                )
            AND c.id = b.id_monitoritemlist
            and c.id = 20            
            group by DATE_FORMAT(a.REG_DTM,'%Y/%m/%d'), b.note1
            order by 1
      """.format(id_dbdetail=id_dbdetail,
                 day1=(datetime.datetime.today() - datetime.timedelta(1)).strftime('%Y-%m-%d'),
                 day2=(datetime.datetime.today() - datetime.timedelta(2)).strftime('%Y-%m-%d'),                 
                 day3=(datetime.datetime.today() - datetime.timedelta(3)).strftime('%Y-%m-%d'),                 
                 day4=(datetime.datetime.today() - datetime.timedelta(4)).strftime('%Y-%m-%d'),                 
                 day5=(datetime.datetime.today() - datetime.timedelta(5)).strftime('%Y-%m-%d'),                 
                )


        print(sql)
        queryset = _query_dict("default", sql)

        ind = 0

        label = []
        labels = []
        data = defaultdict(list)
        backgroundColor = []

        for team in queryset:

            label.append(team['FILESYSTEM'])
            labels.append(team['REG_DTM'])
            data[team['FILESYSTEM']].append(team['USAGE_PERCENT'])
            backgroundColor.append(totalBorderColor[ind])

            if len(totalColor) == ind + 1:
                ind = 0
            else:
                ind += 1

        labels = list(set(labels))
        labels.sort()
        label = list(set(label))
        backgroundColor = list(set(backgroundColor))

        item = []

        ind = 0

        for t in label:
            inner_dict = {}
            inner_dict["data"] = data[t]
            inner_dict["label"] = t
            inner_dict["borderColor"] = backgroundColor[ind]
            inner_dict["radius"] = 2
            inner_dict["fill"] = False
            inner_dict["pointStyle"] = "line"
            item.append(inner_dict)

            if len(totalColor) == ind + 1:
                ind = 0
            else:
                ind += 1  

        print("## 파일 시스템 모니터링 임계치 응답시간 ",time.time()-start_time)             

    elif activeTab == 'tablespace' :
        sql = """
            SELECT REG_DTM,
                   TABLESPACE_NAME,
                   CASE WHEN USAGE_PERCENT < 80 THEN 80
                   ELSE USAGE_PERCENT
                   END USAGE_PERCENT
            FROM
            (
              SELECT DATE_FORMAT(REG_DTM,'%Y/%m/%d') REG_DTM,
                 TABLESPACE_NAME,
                 max(USAGE_PERCENT) USAGE_PERCENT
              FROM cust_monitor_tablespace_log a use index(ix_cust_monitor_tablespace_log_02)
              where a.id_dbdetail = {id_dbdetail}
              AND REG_DTM > DATE_ADD(NOW(), INTERVAL -7 DAY )
              group by DATE_FORMAT(REG_DTM,'%Y/%m/%d'), TABLESPACE_NAME
            ) A
            WHERE TABLESPACE_NAME NOT LIKE '%UNDO%'
            AND TABLESPACE_NAME NOT LIKE '%TEMP%'
            AND TABLESPACE_NAME NOT LIKE '%TMP%'            
            ORDER BY REG_DTM
      """.format(id_dbdetail=id_dbdetail)

        queryset = _query_dict("default", sql)

        ind = 0

        label = []
        labels = []
        data = defaultdict(list)
        backgroundColor = []

        for team in queryset:

            label.append(team['TABLESPACE_NAME'])
            labels.append(team['REG_DTM'])
            data[team['TABLESPACE_NAME']].append(team['USAGE_PERCENT'])
            backgroundColor.append(totalBorderColor[ind])

            if len(totalColor) == ind + 1:
                ind = 0
            else:
                ind += 1

        labels = list(set(labels))
        labels.sort()
        label = list(set(label))
        backgroundColor = list(set(backgroundColor))

        item = []

        ind = 0

        for t in label:
            inner_dict = {}
            inner_dict["data"] = data[t]
            inner_dict["label"] = t
            inner_dict["borderColor"] = backgroundColor[ind]
            inner_dict["radius"] = 2
            inner_dict["fill"] = False
            inner_dict["pointStyle"] = "line"
            item.append(inner_dict)

            if len(totalColor) == ind + 1:
                ind = 0
            else:
                ind += 1

        ymax = 100
        ymin = 80

        print("## 테이블스페이스 모니터링 임계치 응답시간 ",time.time()-start_time)             

    
    context = {
        'result': 'OK',
        'labels': labels,
        'item': item,
        'ymax' : ymax,
        'ymin' : ymin,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def DataViewExportData(request):
    sql_id = request.POST.get('div_sql_id', '')
    id_dblist = request.POST.get('div_id_dblist', '')
    data_title = request.POST.get('div_data_title', '')

    regex = re.compile("\w*")
    data_title = "".join(regex.findall(data_title))
    now = datetime.datetime.now()
    now = now.strftime("%Y%m%d")
    data_title = data_title + "_" + now + '.xlsx'

    txt_inputs = request.POST.getlist('txt_inputs[]')
    txt_binds = request.POST.getlist('txt_binds[]')

    org_sql = sql_id

    sql = """
    select sql_text
    from cust_data_list a
    where a.id = {sql_id}
  """.format(sql_id=sql_id)

    ds_sql = _query_dict("default", sql)

    sql_text = ds_sql[0]['sql_text']

    sql_text = sql_text.replace(";", "")

    i = 0
    for txt in txt_inputs:
        sql_text = sql_text.replace(":" + txt_binds[i], "'" + txt + "'")
        i = i + 1

    sql = """
    select svc_ip, port, inst_nm
    from cust_db_list a,
         cust_db_detail b
    where a.id = {id_dblist}
      and b.oper_cd = 3
      and b.id_dblist = a.id
      and b.db_order = 1;
  """.format(id_dblist=id_dblist)

    inst_nm = _query_dict("default", sql)
    inst_nm = inst_nm[0]['inst_nm']

    try:
        with connections[inst_nm].cursor() as cursor:

            #####################################################
            ## 수행 시간 측정
            #####################################################
            start = time.time()

            cursor.execute(sql_text)

            end = time.time() - start
            #####################################################

            try:
                id_datalist = DataList.objects.get(id=sql_id).id
            except:
                id_datalist = None

            new_userexecutesql = UserExecuteSQL(
                sql_text=org_sql,
                id_reg_user=request.user,
                reg_dtm=datetime.datetime.now(),
                execute_time=end,
                id_datalist=id_datalist,
            )

            new_userexecutesql.save()

            columnNames = [[d[0]] for d in cursor.description]

            data = cursor.fetchall()
            data2 = []

            ###############################
            ## Decimal('53.94') -> str 로 변경
            ###############################
            # buff1 = []
            # buff2 = []

            # ind = 0

            # for ds in data :
            #   for d in ds :

            #     if isinstance(d,Decimal) :
            #       buff1.append(str(d))
            #     elif isinstance(d,int) and columnNames[ind][0][-2:].upper() != 'CD' \
            #                            and columnNames[ind][0][-2:].upper() != 'ID' \
            #                            and columnNames[ind][0][-2:] != '코드' :
            #       buff1.append(format(d, ','))
            #     else :
            #       buff1.append(d)
            #     ind = ind + 1

            #   buff2.append(buff1)
            #   buff1 = []
            #   ind = 0

            # data = buff2

            ###############################

            for d in data:
                data2.append(list(d))

            ind = 0
            list_ind = []
            for columnName in columnNames:
                if columnName[0] in ("CUST_NO",
                                     "고객번호",
                                     "고객_번호",
                                     "CUSTNO",
                                     "고객 번호"):
                    list_ind.append(ind)

                ind = ind + 1

            for i in list_ind:
                for d in data2:
                    d[i] = "*******"

            data = data2
    except:
        data = [['a', 'b', 'c'], ['e', 'f', 'g']]
        columnNames = ['a', 'b', 'c']
        end = 3

    file_name = urllib.parse.quote(data_title.encode('utf-8'))
    # response = HttpResponse(content_type='application/ms-excel')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment;filename*=UTF-8\'\'%s' % file_name

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # wb = xlwt.Workbook(encoding='utf-8')
    # ws = wb.add_sheet('Sheet1')

    # Sheet header, first row
    row_num = 1

    # font_style = xlwt.XFStyle()
    # font_style.font.bold = True

    for col_num in range(len(columnNames)):
        cell = ws.cell(row_num, col_num + 1)
        cell.value = columnNames[col_num][0]

    for row in data:
        row_num += 1

        for col_num in range(len(row)):
            cell = ws.cell(row=row_num, column=col_num + 1)
            cell.value = row[col_num]

    wb.save(response)
    return response


class MetaDV(LoginRequiredMixin, FormMixin, DetailView):
    model = TuningList
    template_name = 'MetaDV.html'
    context_object_name = 'objects'
    form_class = MetaForm

    def get_object(self):
        sql = """
          SELECT A.REQ_NO req_no, --요청번호
               DECODE(SYS_NM,   '[N]DHUB','DHUB',
                                '[N]TCPROD','TCPROD',
                                '[N]NEWINSA','ORAINSA',
                                '[N]MySQL','MYSQL',
                                '[N]SAP','PRD',
                                '[N]WEBDB2','WEBDB',
                                '[N]TISTM','TISTM',
                                '[N]OASPRD','OASPRD',
                                '[N]SMTC','SMTCPRD',
                                '[N]LGHS002N','ORANBS',
                                '[N]ETC','ETC',
                                '[N]BaseITSM','ORAITSM',
                                '[N]D&Shop','D&SHOP', SYS_NM) db_nm, --시스템
               decode(c.apc_type, '0001','신규','0002','삭제','0003',decode(c.apc_obj_type,'0002','컬럼추가','변경'),'0004','컬럼변경') req_div, -- 신청구분
               decode(c.status, '0001','검토중','1001','PL승인','1002','PL반려','0002','승인','0003','반려','0004','DB반영요청중','0005','테스트DB반영완료','0006','테스트DB반영오류','0007','운영DB반영요청중', '0008','운영DB반영완료','0009','운영DB반영오류','운영DB미반영','0010') procss_status, -- 처리상태
               (select user_nm from meta.usr_user where user_id = c.usr_id) req_user, -- 신청자
               to_char(to_date(c.create_date,'yyyymmddhh24miss'),'yyyy-mm-dd') req_dt, -- 신청일자
               trim(replace(c.remark1,'\n','\r\n')) req_explain, -- 설명
               REPLACE(A.SCRIPT,'CREDATE TABLE','CREATE TABLE') ||CHR(10)||CHR(10)||
               '/* END */' ddl, -- DDL
               SYSDATE curr_dt, -- 현재일자
               trim(replace(c.reason0,'\n','\r\n')) dba_explain, -- DBA설명
               C.STATUS system_status -- 10
          FROM META.GS_MD_APC_SCRIPT A, META.GS_MD_SYSTEM B, META.GS_MD_APC C
          WHERE
              A.REQ_NO = C.REQ_NO
          AND C.SYS_ID = B.SYS_ID
          AND A.REQ_NO = {req_no}
      """.format(req_no=self.kwargs['id'])

        ds = _query_dict("metadb", sql)

        return ds[0]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['template_parent'] = "dba"
        context['template_child'] = "DBAWorkLV"

        return context


def TotalScriptDownload(request):
    sql = """
      SELECT ID,
             TABLE_NAME,
             COLUMN_NAME,
             IF(PK_YN=1, 'PK','') PK,
             IF(NOT_NULL=1, 'NOT NULL','') NOTNULL,
             COL_COMMENTS,
             DATA_TYPE
      FROM CUST_COLUMN_LIST
      WHERE DB_USE = 'SMTC'
      AND OPER_CD = 3
      AND OWNER = 'SMTC_OWN'
      -- AND TABLE_NAME LIKE 'PRD_PRD%'
      -- AND TABLE_NAME IN ('PRD_ADD_GRP_D','PRD_ATACH_FILE_D')
      AND TABLE_NAME IN
      (
      'PRD_PRD_M',
      'PRD_BRAND_M',
      'PRD_CLS_M',
      'PRD_ITEM_M',
      'PRD_PRD_M',
      'PRD_SPEC_D',
      'PRD_ATTR_PRD_M',
      'PRD_STOCK_D',
      'PRD_PRC_H',
      'PRD_CHANL_D',
      'PRD_MULTI_CD_D',
      'PRD_ORD_PSBL_QTY_D',
      'PRD_NM_CHG_H'
      )
      ORDER BY TABLE_NAME,COLUMN_ID
  """

    dataset = _query_dict("default", sql)

    table_name = dataset[0]['TABLE_NAME']
    col_ddl = ""
    pk_column = ""
    total_ddl = ""

    for data in dataset:

        if table_name != data['TABLE_NAME']:
            total_ddl += "\n"
            total_ddl += "CREATE TABLE {table_name}\n (".format(table_name=table_name)
            total_ddl += col_ddl
            total_ddl += ");\n"
            total_ddl += "ALTER TABLE {table_name} ADD CONSTRAINT `PK_{table_name}` PRIMARY KEY ( {pk_column} );\n".format(
                table_name=table_name,
                pk_column=pk_column)

            col_ddl = ""
            pk_column = ""
            table_name = data['TABLE_NAME']

        col_ddl += "{column_name} {data_type} {notnull} comment '{comment}',\n".format(
            column_name=data['COLUMN_NAME'],
            data_type=data['DATA_TYPE'],
            notnull=data['NOTNULL'],
            comment=data['COL_COMMENTS']
        )
        if data['PK'] == "PK":
            pk_column += data['COLUMN_NAME'] + ", "

    total_ddl += "\n"
    total_ddl += "CREATE TABLE {table_name} (".format(table_name=table_name)
    total_ddl += col_ddl
    total_ddl += ");\n"
    total_ddl += "ALTER TABLE {table_name} ADD CONSTRAINT `PK_{table_name}` PRIMARY KEY ( {pk_column} );\n".format(
        table_name=table_name,
        pk_column=pk_column)

    res = HttpResponse(total_ddl)

    file_name = "DDL_" + datetime.datetime.now().strftime("%Y%m%d")
    file_name = urllib.parse.quote(file_name.encode('utf-8'))

    res['Content-Disposition'] = 'attachment; filename={file_name}.sql'.format(file_name=file_name)
    return res


def MetaScriptDownload(request):
    # notice = get_object_or_404(Notice, pk=pk)
    # url = notice.upload_files.url[1:]
    # file_url = urllib.parse.unquote(url)

    # if os.path.exists(file_url):
    # with open(file_url, 'rb') as fh:
    #     quote_file_url = urllib.parse.quote(notice.filename.encode('utf-8'))
    #     response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_url)[0])
    #     response['Content-Disposition'] = 'attachment;filename*=UTF-8\'\'%s' % quote_file_url
    #     return response
    # raise Http404
    obj = MetaReq.objects.filter(check_prod_exec_yn="1").order_by('id_dblist', 'id')

    text_header = """[HEADER]
VERSION=7.0.0
TOOL=
CHAR=UNI16_LE
DATE={date}
[WORKSPACE]
COUNT={cnt}
""".format(
        date=datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        cnt=len(obj)
    )

    text_contents = ""
    req_contents = """/***********************************************************************
* 적용 데이터베이스 : {id_dblist}
* 요청자 : {id_reg_user}
* 승인자 : {id_pl_prod}
* 요청 내용
* {req_contents}
************************************************************************
*/
/* 스크립트 시작 */
"""

    end_msg = """
/* 스크립트 끝 */
"""

    f_text_contents = """NAME={name}
STMT={stmt}
{req_contents}
{sql}
{end_msg}
VAR=0"""

    for o in obj:
        req_contents_tmp = req_contents.format(req_contents=o.req_contents,
                                               id_dblist=o.id_dblist,
                                               id_reg_user=o.id_reg_user,
                                               id_pl_prod=o.id_pl_prod,
                                               )
        text_contents += f_text_contents.format(
            name=o.id_dblist.db_use + "_" + "No." + str(o.id),
            stmt=len(req_contents_tmp.split("\n")) + len(o.script.split("\n")) + len(end_msg.split("\n")),
            req_contents=req_contents_tmp,
            sql=o.script,
            end_msg=end_msg,
        )
        text_contents += "\n"

    fileContent = text_header + text_contents
    # BOM, 한글 깨짐 방지
    fileContent = fileContent.encode('utf-8-sig')

    res = HttpResponse(fileContent)

    file_name = "정기반영_" + datetime.datetime.now().strftime("%Y%m%d")
    file_name = urllib.parse.quote(file_name.encode('utf-8'))

    res['Content-Disposition'] = 'attachment; filename={file_name}.sqd'.format(file_name=file_name)
    return res


class ApiTest(ListView):
    template_name = 'ApiTest.html'

    def get_queryset(self):
        prd_cd = self.request.GET.get("fname")

        with connections['DEVEC'].cursor() as cursor:
            # sql = """select prd_cd from prd_prd_m where rownum<1000000"""
            sql = """
      SELECT /*+ ORDERED  USE_NL(PNCH PPH PSCD PCD PPD PCDDD LB PND SDD SDD2 SSM PMD PBD T19 T24 ILNDSDD JEJUSDD ILNDRTPSDD JEJURTPSDD DAWNSDD) PUSH_PRED(T24) PUSH_PRED(W) */
             PPM.PRD_CD                         AS PRD_CD                             
            ,PPM.USE_YN                         AS USE_YN                             
            ,PPM.MNFC_CO_NM                     AS MNFC_CO_NM                         
            ,PPM.PRD_TYP_CD                     AS PRD_TYP_CD                         
            ,PPM.OPER_MD_ID                     AS OPER_MD_ID                         
            ,PPM.REG_DTM                        AS REG_DTM                            
            ,PPM.EXPOS_ST_CD                    AS EXPOS_ST_CD                        
            ,PPM.ORGP_NM                        AS NATIVE_COUNTRY                     
            ,PPM.SUP_CD                         AS SUP_CD                             
            ,SSM.SUP_NM                         AS SUP_NM                             
            ,NVL(SSM.BZ_REG_NO, '')             AS BZ_REG_NO                          
                ,NVL((SELECT NVL((ROAD_NM_BASE_ADDR || ROAD_NM_DTL_ADDR), (BASE_ADDR || DTL_ADDR))
                      FROM SUP_ADDR_D
                       WHERE SUP_CD = PPM.SUP_CD
                       AND CNTC_TYP_CD = 'BASE'
                       AND USE_YN = 'Y'
                       AND ROWNUM = 1), '')       AS SUP_ADDR                           
            ,PPM.SUB_SUP_CD             AS SUB_SUP_CD                         
            ,SSM.B2B_SUP_CD                     AS B2B_SUP_CD                   
            ,PPM.ORD_PRD_TYP_CD                 AS ORD_PRD_TYP_CD                     
            ,PPM.PRD_APRV_ST_CD                 AS PRD_APRV_ST_CD                     
            ,PPM.EC_APRV_ST_CD                  AS EC_APRV_ST_CD                      
            ,PPM.BASE_ACCM_LIMIT_YN             AS BASE_ACCM_LIMIT_YN                 
            ,PPM.CARD_USE_LIMIT_YN              AS CARD_USE_LIMIT_YN                  
            ,PPM.SALE_END_DTM                   AS SALE_END_DTM                       
            ,PPM.PRCH_TYP_CD                    AS PRCH_TYP_CD                        
            ,NVL(SUBSTR(PPM.PRD_CLS_CD,1,3),'') AS PRD_LRG_CLS_CD                     
            ,NVL(SUBSTR(PPM.PRD_CLS_CD,4,2),'') AS PRD_MID_CLS_CD                     
            ,NVL(SUBSTR(PPM.PRD_CLS_CD,6,2),'') AS PRD_SML_CLS_CD                     
            ,NVL(SUBSTR(PPM.PRD_CLS_CD,8,2),'') AS PRD_DTL_CLS_CD                     
            ,PPM.PRD_CLS_CD                     AS PRD_CLS_CD                         
            ,PPM.DLV_PICK_MTHOD_CD              AS DLV_PICK_MTHOD_CD                  
            ,NVL(PPM.DTCT_CD, 'DC01')       AS DTCT_CD                  
            ,PPM.GFT_TYP_CD                     AS GFT_TYP_CD                         
            ,PPM.SUP_PRD_CD                     AS SUP_PRD_CD                         
            ,PPM.FRMLES_PRD_TYP_CD              AS FRMLES_PRD_TYP_CD                  
            ,PPM.PRC_COMPR_TNS_CD               AS PRC_COMPR_TNS_CD                   
            ,PPD.CLS_CHK_AFT_APRV_CD            AS CLS_CHK_AFT_APRV_CD                
            ,PPM.IST_TYP_CD                     AS IST_TYP_CD                         
            ,PPM.PRD_GBN_CD                     AS PRD_GBN_CD                         
            ,PPM.BUNDL_PRD_GBN_CD               AS BUNDL_PRD_GBN_CD                   
            ,NVL(PPM.FLGD_PRC_EXPOS_YN,'Y')     AS FLGD_PRC_EXPOS_YN                  
            ,PPM.GSNPNT_NO_GIV_YN               AS GSNPNT_NO_GIV_FLG                  
            ,PPM.REP_PRD_CD                     AS REP_PRD_CD                         
            ,PSCD.SAFE_CERT_NO                  AS SAFE_CERT_NO                       
            ,PSCD.SAFE_CERT_GBN_CD              AS SAFE_CERT_GBN_CD                   
            ,PPM.RFN_TYP_CD                     AS RFN_TYP_CD                         
            ,PPM.ARFN_TIME_CD                   AS ARFN_TYP_CD                        
            ,'N'                          AS DITTO_FLG                          
              ,NVL(PPM.DITTO_YN,'N')              AS DITTO_PRD_FLG                      
            ,PPM.IMG_CNF_YN                     AS IMG_CNF_YN                         
            ,PPM.DLV_DT_GUIDE_CD                AS DLV_DT_GUIDE_CD                    
            ,PPM.CHR_DLV_YN                     AS CHR_DLV_YN                         
            ,PPM.INSTL_DLV_PRD_YN               AS INSTL_DLV_PRD_YN                   
            ,PPH.INSTL_COST                     AS INSTL_COST                         
            ,PPM.REP_PRD_YN                     AS REP_PRD_YN                         
            ,PPM.SEL_ACCM_APPLY_YN              AS SEL_ACCM_APPLY_YN                  
            ,PNCH.EXPOS_PMO_NM                  AS EXPOS_PMO_NM                       
            ,PNCH.EXPOS_PRD_NM                  AS EXPOS_PRD_NM                       
            ,PNCH.EXPOS_PR_SNTNC_NM             AS EXPOS_PR_SNTNC_NM                  
            ,NVL (PPM.ATTR_TYP_EXPOS_CD, 'L')   AS ATTR_TYP_EXPOS_CD                  
            ,PPM.OPEN_AFT_RTP_NOADMT_YN         AS OPEN_AFT_RTP_NOADMT_YN             
            ,( CASE WHEN RSRV_SALE_PRD_YN = 'Y'
                  AND (SELECT 'Y'
                         FROM PRD_SCHD_D SC
                        WHERE SC.PRD_CD = PPM.PRD_CD
                          AND SC.PRD_SCHD_GBN_CD = '10'
                          AND SYSDATE BETWEEN SC.VALID_STR_DTM AND SC.VALID_END_DTM
                          AND ROWNUM = 1) = 'Y' THEN 'Y'
                    ELSE 'N' END )          AS RSRV_SALE_PRD_YN
            ,PPM.ORD_MNFC_YN                    AS ORD_MNFC_YN                        
            ,PPM.ONAIR_SALE_PSBL_YN             AS ONAIR_SALE_PSBL_YN                 
            ,PPM.COMPR_PRC_MRK_YN               AS COMPR_PRC_MRK_YN                   
            ,PPM.BRAND_CD                       AS BRAND_CD                           
            ,(SELECT BRAND_NM
                     FROM PRD_BRAND_M
                     WHERE BRAND_CD = PPM.BRAND_CD
                       AND TRIM(BRAND_ST_CD) IN ('A','H')  
              )                 AS BRAND_NM 
            ,PPM.ALIA_SPCLSAL_LIMIT_YN          AS ALIA_SPCLSAL_LIMIT_YN              
            ,NVL(PPM.TEMPOUT_YN,'N')            AS TEMPOUT_FLG                        
            ,PPM.TEMPOUT_CHG_DTM                AS DTM_OUTSTK_CHG_DTM                 
            ,NVL(PCD.STD_RELS_DDCNT, 0)         AS PLAN_DLVPRD_TERM_DDCNT             
            ,PPM.PRE_ORD_TYP_CD                 AS PRE_ORD_TYP_CD                     
            ,PPM.PRE_ORD_BROAD_DT               AS PRE_ORD_BROAD_DT                   
            ,NVL(PPM.ATTR_TYP_NM1, 'COLOR')     AS ATTR_TYP_NM_1                      
            ,NVL(PPM.ATTR_TYP_NM2, 'SIZE')      AS ATTR_TYP_NM_2                      
            ,NVL(PPM.ATTR_TYP_NM3, '스타일')      AS ATTR_TYP_NM_3                     
            ,NVL(PPM.ATTR_TYP_NM4, '사은품')      AS ATTR_TYP_NM_4                     
            ,TO_CHAR(NVL(PCDDD.AVG_DLV_DDCNT_VAL, 0)) AS AVG_DLV_DDCNT_VAL            
            ,TO_CHAR(NVL(PCDDD.LONG_DLV_DDCNT, 0))    AS LONG_DLV_DDCNT               
            ,NVL(PCDDD.EXPOS_YN, 'Y')                 AS DLV_DT_EXPOS_YN              
            ,TO_CHAR(NVL(PCDDD.DD3_IN_AVG_DLV_RT, 0)) AS DD3_IN_AVG_DLV_DDCNT         
            ,NVL(PCD.STD_RELS_DDCNT, 999)       AS STD_RELS_DDCNT                     
            ,(CASE WHEN NVL(PCD.STD_RELS_DDCNT, 999) = 0 THEN
                       NVL((SELECT UDA_VAL
                  FROM PRD_UDA_D
                 WHERE UDA_GBN_CD = '10'
                   AND UDA_NO = 27
                   AND SYSDATE BETWEEN VALID_STR_DTM
                   AND VALID_END_DTM
                   AND USE_YN = 'Y'
                   AND PRD_CD = PPM.PRD_CD
                   AND ROWNUM = 1), '0')
              ELSE '0'  END)                AS THEDAY_RELS_ORD_DEDLN_TIME         
            ,NVL(LB.MEDIA,'PA')           AS MED_TYP                            
            ,NVL(PPM.IMM_ACCM_DC_LIMIT_YN, 'N') AS IMM_ACCM_DC_LIMIT_YN               
            ,PPM.ORD_MNFC_TYP_CD                AS ORD_MNFC_TYP                       
            ,NVL(PPM.SEPAR_ORD_NOADMT_YN,'N')   AS SEPAR_ORD_NOADMT_YN                
            ,PPM.SALE_PSBL_APRV_YN              AS SALE_PSBL_APRV_YN                  
            ,PCD.SALE_PSBL_YN                   AS SALE_PSBL_YN                       
            ,CASE WHEN NVL(PPM.SALE_PSBL_APRV_YN, 'N') = 'Y'
                       AND NVL(PPM.TEMPOUT_YN, 'N') = 'N'
                       AND NVL(PCD.SALE_PSBL_YN, 'N') = 'Y'
                  THEN 'Y'
                  ELSE 'N'
             END AS CHANL_SALE_PSBL_YN                            
            ,NVL(PPM.OBOX_CD,'N')               AS OBOX_CD                            
            ,PPM.MODEL_NO                       AS MODEL_NO                         
            ,NVL(PPM.ADULT_CERT_YN,'N')         AS ADULT_CERT_FLG                     
            ,NVL(PPM.CHR_DLV_ADD_YN,'N')        AS CHR_DLV_ADD_YN                     
            ,PPM.ZRWON_SALE_YN                  AS ZRWON_SALE_FLG                     
            ,NVL((SELECT ATTR_NUM_VAL_1
                      FROM PRD_META_D
                       WHERE PRD_CD = PPM.PRD_CD
                         AND PRD_META_TYP_CD = '10'),0) AS CELPHN_TRMNL_PRC       
            ,PPM.ORD_LIMIT_QTY                  AS ORD_LIMIT_QTY                      
            ,NVL(PND.UNIT_VAL_1,'')             AS PRD_NUMVAL_VAL                     
            ,NVL(PND.NUMVAL_UNIT_CD,'')         AS PRD_NUMVAL_UNIT_CD                 
            ,PPM.CVS_DLVS_RTP_YN            AS CVS_DLVS_RTP_YN              
            ,DECODE(PNCH.EXPOS_PMO_NM,NULL,'',PNCH.EXPOS_PMO_NM||' ')||
             DECODE(PNCH.EXPOS_PRD_NM,NULL,'',PNCH.EXPOS_PRD_NM||' ')||
             NVL(PNCH.EXPOS_PR_SNTNC_NM,'')     AS PRD_NM 
            ,(CASE WHEN PPM.CPN_APPLY_TYP_CD = '09'
                   THEN 'Y'
                   ELSE 'N'
               END)                             AS CPN_NO_APPLY_YN                
            ,PPM.CPN_APPLY_TYP_CD           AS CPN_APPLY_TYP_CD               
            ,NVL(BUNDL_DLV_CD,'A02')            AS BUNDL_DLV_CD                   
            ,NVL(SDD.STD_AMT_YN,'N')            AS STD_AMT_YN                     
            ,DECODE((SELECT PRD_CD
                       FROM PRD_UDA_D
                      WHERE UDA_GBN_CD = '10'
                        AND UDA_NO = 16
                        AND SYSDATE BETWEEN VALID_STR_DTM
                                        AND VALID_END_DTM
                        AND USE_YN = 'Y'
                        AND PRD_CD = PPM.PRD_CD
                        AND ROWNUM = 1),PPM.PRD_CD,'N','Y') AS ALIA_CARD_ACC_LIMIT_YN  
             ,NVL((SELECT UDA_VAL
                  FROM PRD_UDA_D
                    WHERE UDA_GBN_CD = '10'
                      AND UDA_NO = 19
                      AND PRD_CD = PPM.PRD_CD
                      AND SYSDATE BETWEEN VALID_STR_DTM AND VALID_END_DTM
                      AND USE_YN = 'Y'
                      AND ROWNUM = 1),'X') AS EMP_EXCLU_PRD_GBN_VAL                
            ,NVL(TO_NUMBER((SELECT VAL
                              FROM EC_REFSETVAL
                             WHERE REFSETID = 5417
                               AND CODE = NVL(PRDM.ATTR_CHAR_VAL_2,'')
                             )), 0)    AS PRD_CHAR_CPN_LIMIT_QTY     
                , NVL((SELECT 'Y' FROM (
                                      SELECT /*+ NO_REWRITE LEADING(B) USE_NL(A B)  */
                                            B.GOODS_CODE
                                      FROM LGEC_PGM A
                                         , LGEC_PGM_DTL B
                                      WHERE A.BROAD_DATE     <= SYSDATE  + 1/48
                                        AND A.CLOSING_DATE   >= SYSDATE  - 1/48
                                        AND A.BROAD_DATE   = B.BROAD_DATE
                                      UNION ALL
                                      SELECT /*+ NO_REWRITE LEADING(A) USE_NL(A B)  */
                                             B.GOODS_CODE
                                      FROM GS_BROAD_FORM_DTC_M A         
                                         , GS_BROAD_FORM_PRD_DTC_D B     
                                      WHERE A.BROAD_DATE    <= SYSDATE + 1/48
                                        AND A.CLOSING_DATE   >= SYSDATE - 1/48
                                        AND A.BROAD_DATE   = B.BROAD_DATE
                                      ) MAIN
                                 WHERE MAIN.GOODS_CODE = ppm.prd_cd
                                   AND ROWNUM = 1
                   ),'N') AS MIN30_BEF_AFT_LIVBROD_YN   
                , NVL((SELECT 'Y' FROM (
                                      SELECT /*+ NO_REWRITE LEADING(B) USE_NL(A B)  */
                                            B.GOODS_CODE
                                      FROM LGEC_PGM A
                                         , LGEC_PGM_DTL B
                                      WHERE A.BROAD_DATE     <= SYSDATE
                                        AND A.CLOSING_DATE   >= SYSDATE
                                        AND A.BROAD_DATE   = B.BROAD_DATE
                                      UNION ALL
                                      SELECT /*+ NO_REWRITE LEADING(A) USE_NL(A B)  */
                                             B.GOODS_CODE
                                      FROM GS_BROAD_FORM_DTC_M A         
                                         , GS_BROAD_FORM_PRD_DTC_D B     
                                      WHERE A.BROAD_DATE     <= SYSDATE
                                        AND A.CLOSING_DATE   >= SYSDATE
                                        AND A.BROAD_DATE   = B.BROAD_DATE
                                      ) MAIN
                                 WHERE MAIN.GOODS_CODE = ppm.prd_cd
                                   AND ROWNUM = 1
                   ),'N')  AS LIVBROD_YN        
            ,(SELECT /*+ NO_REWRITE */ P.PGM_ID AS PGM_ID
              FROM LGEC_PGM    P
                      ,LGEC_PGM_DTL M
              WHERE M.GOODS_CODE  = PPM.PRD_CD
                  AND P.BROAD_DATE   = M.BROAD_DATE
                  AND P.BROAD_DATE - 1/48   <= SYSDATE
                AND P.CLOSING_DATE + 1/48   > SYSDATE
                  AND ROWNUM           = 1 )        AS PGM_ID        
            ,(SELECT /*+ INDEX (CAL GS_COMM_CALENDAR_IX3) */   
                     CAL.HDATE
                FROM GS_COMM_CALENDAR CAL
                    ,PRD_SCHD_D       SC
               WHERE CAL.HDATE BETWEEN TRUNC(SC.RSRV_DLV_END_DT)+1
                                   AND TRUNC(SC.RSRV_DLV_END_DT)+20
                 AND CAL.DAY NOT IN ('Sat', 'Sun')
                 AND CAL.HOLIDAY_FLAG = 'N'
                 AND SC.PRD_CD = PPM.PRD_CD
                 AND SC.PRD_SCHD_GBN_CD ='10'
                 AND SYSDATE BETWEEN SC.VALID_STR_DTM
                                 AND SC.VALID_END_DTM
                 AND ROWNUM = 1)         AS RSRV_DLV_DT                           
            ,DECODE(PPM.PRD_GBN_CD,'30',(SELECT MULTI_CD
                                          FROM DMC_DM_SEQNO_PRD_D
                                         WHERE PRD_CD      = PPM.PRD_CD
                                           AND USE_YN      = 'Y'
                                           AND EC_EXPOS_YN = 'Y'
                                           AND ROWNUM = 1), '8'||PPM.PRD_CD) AS MULTI_CD   
            ,(CASE WHEN PPM.PRD_RELSP_ADDR_CD IS NULL
                   THEN (SELECT NVL(TX.SUP_ADDR_CD,'00')
                           FROM SUP_ADDR_D  TX
                          WHERE TX.SUP_CD        = PPM.SUP_CD
                            AND TX.USE_YN        = 'A'
                            AND TX.BASE_RELSP_YN = 'Y'
                            AND ROWNUM           = 1)
                   ELSE PPM.PRD_RELSP_ADDR_CD
               END)                                  AS PRD_RELSP_ADDR_CD             
            ,(SELECT STD_RELS_DDCNT
                FROM PRD_CHANL_D
               WHERE PRD_CD = PPM.PRD_CD
                 AND CHANL_CD = 'P'
                 AND SALE_PSBL_YN = 'Y')             AS DM_STD_RELS_DDCNT             
            ,(SELECT COUNT(PRD_CD)
                FROM PRD_DLV_REGON_D PDRD
               WHERE PDRD.PRD_CD = PPM.PRD_CD
                 AND  NVL(USE_YN,'N') = 'Y'
                 AND ROWNUM = 1)                     AS DLV_NOADMT_REGON_CNT          
            ,NVL((SELECT TRIM(ER.VAL)
                    FROM EC_REFSETVAL ER
                   WHERE ER.REFSETID = '10083'
                     AND ER.CODE = PPM.SUP_CD
                     AND ER.DBSTS = 'A'
                     AND ROWNUM = 1),'')             AS ALIA_DPAT_NM                  
            ,NVL((SELECT ER.DSCR
                    FROM EC_REFSETVAL ER
                   WHERE ER.REFSETID = '10083'
                     AND ER.CODE = PPM.SUP_CD
                     AND ER.DBSTS = 'A'
                     AND ROWNUM = 1),'')             AS ALIA_DPAT_SUP_NM              
            ,NVL((SELECT TRIM(ER.CODE)
                    FROM EC_REFSETVAL ER
                   WHERE ER.REFSETID = '10143'
                     AND ER.DBSTS = 'A'
                     AND ROWNUM = 1),'Y')            AS RTM_STOCK_CHK_YN              
            ,PPM.EXCH_RTP_CHR_YN                     AS EXCH_RTP_CHR_YN               
            ,PPM.RTP_ONEWY_RNDTRP_CD                 AS RTP_ONEWY_RNDTRP_CD           
            ,PPM.EXCH_ONEWY_RNDTRP_CD                AS EXCH_ONEWY_RNDTRP_CD          
            ,SDD.DLVC_AMT                            AS EXCH_RTP_DLVC_AMT             
            ,SDD.DLVC_LIMIT_AMT                      AS EXCH_RTP_DLVC_LIMIT_AMT       
            ,(CASE WHEN NVL(PPM.CHR_DLVC_CD,0)   >   0
                   THEN PPM.CHR_DLVC_CD
                   ELSE SDD2.DLVC_CD
               END)                                  AS CHR_DLVC_CD                   
            ,(CASE WHEN NVL(PPM.CHR_DLVC_CD,0)   >   0
                   THEN SDD2.DLVC_AMT
                   ELSE 0
               END)                                  AS DLVC                          
            ,SDD2.DLVC_LIMIT_AMT                     AS DLVC_LIMIT_AMT                
            ,SDD2.STD_AMT_YN                         AS DLVC_STD_AMT_EXPOS_LIMIT_YN     
            ,(CASE WHEN SUBSTR(PPM.DLV_PICK_MTHOD_CD,0,1) = '3'
                   THEN TRIM(SSM.PRD_OBOX_CD)
               END)                                  AS DIRDLV_OBOX_SUP_YN            
            ,NVL(PMD.ATTR_CHAR_VAL_1,'')             AS ATTR_CHAR_VAL_1               
            ,NVL(PMD.ATTR_CHAR_VAL_2,'')             AS ATTR_CHAR_VAL_2               
            ,NVL(PMD.ATTR_CHAR_VAL_3,'')             AS ATTR_CHAR_VAL_3               
            ,NVL(PMD.ATTR_CHAR_VAL_4,'')             AS ATTR_CHAR_VAL_4               
            ,NVL(PMD.ATTR_CHAR_VAL_5,'')             AS ATTR_CHAR_VAL_5               
            ,NVL(PMD.ATTR_CHAR_VAL_6,'')             AS ATTR_CHAR_VAL_6               
            ,NVL(PMD.ATTR_CHAR_VAL_7,'')             AS ATTR_CHAR_VAL_7               
            ,NVL(PMD.ATTR_CHAR_VAL_8,'')             AS ATTR_CHAR_VAL_8               
            ,NVL(PMD.ATTR_CHAR_VAL_9,'')             AS ATTR_CHAR_VAL_9               
            ,NVL(PMD.ATTR_CHAR_VAL_10,'')            AS ATTR_CHAR_VAL_10              
            ,NVL(PMD.ATTR_CHAR_VAL_11,'')            AS ATTR_CHAR_VAL_11              
            ,NVL(PMD.ATTR_CHAR_VAL_12,'')            AS ATTR_CHAR_VAL_12              
            ,NVL(PMD.ATTR_CHAR_VAL_15,'')            AS ATTR_CHAR_VAL_15              
            ,NVL(T24.ORD_LIMIT, 0)                   AS PERIOD_LIMIT_QTY            
            ,NVL(T24.ORD_LIMIT_DAYS, 30)             AS PERIOD_LIMIT_DAY            
            ,(CASE WHEN PBD.BOOK_PRD_GBN_CD = 'B'
                   THEN PBD.PUBLC_CO_NM||'('||PBD.PAGE_CNT||'P)'
                   ELSE PBD.PUBLC_CO_NM
               END)                                  AS PUBLC_INFO                    
            ,TO_CHAR(PBD.ISSUE_DT, 'YYYY-MM-DD HH24:MI:SS') AS ISSUE_DTM          
            ,PBD.BOOK_PRD_GBN_CD                     AS BOOK_PRD_GBN_CD               
            ,PBD.DMSTC_FIXPRC                        AS DMSTC_FIXPRC                  
            ,NVL(PBD.CENT_STOCK_YN,'N')              AS CENT_STOCK_YN                 
            ,T19.CLASS_1                             AS PRDRPT_LIMIT_CLS_1            
            ,T19.CLASS_2                             AS PRDRPT_LIMIT_CLS_2            
            ,T19.CLASS_3                             AS PRDRPT_LIMIT_CLS_3            
            ,T19.CLASS_4                             AS PRDRPT_LIMIT_CLS_4            
            ,DECODE(T19.PRD_CD,NULL,'N','Y')         AS PRDRPT_WRIT_LIMIT_PRD_FLG     
            ,NVL(W.CUSTOM_AVG,0) AS PRDRPT_AVG_SCORE       
            ,NVL(W.KNOW_CNT,0)   AS PRDRPT_CNT                 
            ,NVL( PRDM.ATTR_CHAR_VAL_3, '') AS AUTO_RFN_YN     
            ,NVL( TO_CHAR(TO_DATE(SUBSTR(PRDM.ATTR_CHAR_VAL_4,0,8)),'YYYY.MM.DD') , '') AS VALID_DTM       
          ,NVL( PRDM.ATTR_CHAR_VAL_5 , '') AS PRD_AUTO_RFN_YN  
            ,NVL( PRDM.ATTR_NUM_VAL_1  , 0)  AS AUTO_RFN_RT      
            ,NVL( PRDM.ATTR_NUM_VAL_4  , 0)  AS VALID_TERM_DDCNT 
            ,0                 AS SALE_PSBL_QTY    
            ,NVL((SELECT 'Y'
              FROM CMM_CMM_C
            WHERE CMM_GRP_CD = 'PMO127' AND USE_YN = 'Y'
              AND (SUBSTR(PPM.PRD_CLS_CD,1,3) = TRIM(CMM_CD)
                     OR SUBSTR(PPM.PRD_CLS_CD,1,5) = TRIM(CMM_CD)
                     OR SUBSTR(PPM.PRD_CLS_CD,1,7) = TRIM(CMM_CD)
                     OR (PPM.PRD_CLS_CD = TRIM(CMM_CD))  
                     )  
                     AND ROWNUM = 1), 'N') AS PMO_BANNER_LIMIT_YN
           ,NVL(PPM.FORGN_DLV_PSBL_YN,'N') AS FORGN_DLV_PSBL_YN 
           ,PPM.GNUIN_YN AS HONEST_GOODS_FLAG  
         ,NVL(PPD.CMPOS_INFO_EXPOS_YN, 'N') AS CMPOS_INFO_EXPOS_YN  
         ,NVL((SELECT 'Y'
                FROM PRD_UDA_D
               WHERE PRD_CD = PPM.PRD_CD
                 AND UDA_GBN_CD = '10'
                 AND UDA_NO = '23'
                 AND UDA_VAL = 'Y'
                 AND USE_YN = 'Y'
                 AND SYSDATE BETWEEN VALID_STR_DTM AND VALID_END_DTM
                 AND ROWNUM = 1), 'N') AS PARALLEL_IMPORT_YN      
           ,NVL((SELECT 'Y'
                 FROM PRD_META_D PMD
                WHERE PMD.PRD_META_TYP_CD = '60'
                  AND NVL(PMD.ATTR_CHAR_VAL_1, 'N') = 'Y'
                  AND PMD.PRD_CD = PPM.PRD_CD
                  AND ROWNUM = 1), 'N') AS DIRECT_PURCHASE_YN     
        ,NVL((SELECT CD_VAL FROM CMM_CMM_C WHERE USE_YN = 'Y' AND CMM_CD = PPM.DLVS_CO_CD AND CMM_GRP_CD = 'PRD008'), '') AS DLV_CO_NM 
        ,NVL(PPD.QUICK_DLV_NOADMT_YN, 'N') AS QUICK_DLV_NOADMT_YN   
        ,NVL(PPD.APNT_DT_DLV_TYP, ' ') AS APNT_DT_DLV_TYP     
        ,( SELECT SUBSTR(NVL(SU.UDA_VAL,''),1,2)
               FROM SUP_UDA_D SU
              WHERE 1 = 1
                AND SU.UDA_NO = 16
                AND SU.SUP_CD = PPM.SUP_CD
                AND SU.SUB_SUP_CD = PPM.SUB_SUP_CD
                AND SU.VALID_STR_DTM   <=   SYSDATE
                AND SU.VALID_END_DTM   >=   SYSDATE
                AND SU.USE_YN = 'Y') AS ADIDAS_CD                   
        ,NVL(PPD.JEJU_CHR_DLV_YN, 'N') AS JEJU_CHR_DLV_YN               
        ,NVL(PPD.ILND_CHR_DLV_YN, 'N') AS ILND_CHR_DLV_YN               
        ,NVL(ILNDSDD.DLVC_AMT, '0') AS ILND_DLVC                        
        ,NVL(JEJUSDD.DLVC_AMT, '0') AS JEJU_DLVC                        
        ,NVL(ILNDRTPSDD.DLVC_AMT,'0') AS ILND_RTP_DLVC                  
        ,NVL(DAWNSDD.DLVC_AMT,'0') AS DAWN_DLVC                   
        ,NVL(DAWNSDD.DLVC_LIMIT_AMT,'0') AS DAWN_DLVC_LIMIT             
        ,NVL(PPD.ILND_EXCH_RTP_CHR_YN,'N') AS ILND_EXCH_RTP_CHR_YN          
        ,NVL(PPD.ILND_RTP_ONEWY_RNDTRP_CD,'0') AS ILND_RTP_ONEWY_RNDTRP_CD      
        ,NVL(PPD.ILND_EXCH_ONEWY_RNDTRP_CD,'0') AS ILND_EXCH_ONEWY_RNDTRP_CD    
        ,NVL(JEJURTPSDD.DLVC_AMT,'0') AS JEJU_RTP_DLVC                
        ,NVL(PPD.JEJU_EXCH_RTP_CHR_YN,'N') AS JEJU_EXCH_RTP_CHR_YN          
        ,NVL(PPD.JEJU_RTP_ONEWY_RNDTRP_CD,'0') AS JEJU_RTP_ONEWY_RNDTRP_CD      
        ,NVL(PPD.JEJU_EXCH_ONEWY_RNDTRP_CD,'0') AS JEJU_EXCH_ONEWY_RNDTRP_CD    
        ,PPD.ILND_DLV_PSBL_YN                           
        ,PPD.JEJU_DLV_PSBL_YN                           
        ,(CASE WHEN SUBSTR(PPM.DLV_PICK_MTHOD_CD,0,1) = '3'
                   THEN (SELECT MAX(UDA_NO)
                 FROM SUP_UDA_D SDD
                WHERE SDD.UDA_NO IN (52,53,54)     
                  AND SDD.SUB_SUP_CD = PPM.SUB_SUP_CD
                  AND SYSDATE BETWEEN SDD.VALID_STR_DTM AND SDD.VALID_END_DTM
                  AND SDD.USE_YN = 'Y'
                  AND SDD.UDA_VAL = 'Y'
                  AND SDD.SUB_SUP_YN = 'Y')
             END) AS GIFT_PACKING_CD               
          ,NVL((SELECT 'Y'
                  FROM SUP_UDA_D SUD
                 WHERE SUD.UDA_NO = 63
                   AND SUD.SUP_CD = PPM.SUP_CD
                   AND SUD.UDA_VAL = 'Y'
                   AND SYSDATE BETWEEN SUD.VALID_STR_DTM AND SUD.VALID_END_DTM
                   AND SUD.USE_YN = 'Y'), 'N') AS DIRDLV_EXCLS_SUP_YN   
                ,NVL((SELECT 'Y'
                  FROM SUP_UDA_D SUD
                   WHERE SUD.UDA_NO = 51
                     AND SUD.SUP_CD = PPM.SUP_CD
                   AND SUD.UDA_VAL = 'Y'
                   AND SYSDATE BETWEEN SUD.VALID_STR_DTM AND SUD.VALID_END_DTM
                   AND SUD.USE_YN = 'Y'
                   AND (SUD.SUB_SUP_YN = 'N' OR (SUD.SUB_SUP_CD = PPM.SUB_SUP_CD AND SUD.SUB_SUP_YN = 'Y'))), 'N') AS DESCD_CAPTR_EXP_SUP_YN  
          ,NVL(PPD.GSCHOICE_YN, 'N')                  AS GSCHOICE_YN  
          ,NVL(PPD.APNT_DT_DLV_YN,'N') AS APNT_DT_DLV_YN  
        FROM (
                 SELECT /*+ NO_MERGE INDEX(PPM PK_PRD_PRD_M) NO_EXPAND */ *
                 FROM PRD_PRD_M    PPM      
                 WHERE 1=1
                   AND PPM.PRD_CD IN
                    (  :prd_cd  )
                   AND PPM.USE_YN = 'Y'
                ) PPM
            ,PRD_NM_CHG_H      PNCH     
            ,PRD_PRC_H         PPH      
            ,PRD_SAFE_CERT_D   PSCD     
            ,PRD_CHANL_D       PCD      
            ,PRD_PRD_D         PPD      
            ,PRD_CLS_DLV_DAY_D PCDDD    
            ,LGEC_BY230        LB       
            ,PRD_NUMVAL_D      PND      
            ,SUP_DLVC_D        SDD      
            ,SUP_DLVC_D        SDD2     
            ,SUP_DLVC_D        ILNDSDD  
            ,SUP_DLVC_D        JEJUSDD  
            ,SUP_DLVC_D        ILNDRTPSDD  
            ,SUP_DLVC_D        JEJURTPSDD  
            ,SUP_DLVC_D      DAWNSDD  
            ,SUP_SUP_M         SSM      
            ,PRD_META_D        PMD      
            ,PRD_BOOK_D        PBD      
            ,(SELECT *
                    FROM (SELECT CLASS_1
                                ,CLASS_2
                                ,CLASS_3
                                ,CLASS_4
                                ,PRD_CD
                                ,ROW_NUMBER() OVER(PARTITION BY PRD_CD ORDER BY SORT_SEQ ) R_IDX
                      FROM (SELECT NVL(SUBSTR(PRD.PRD_CLS_CD, 1,3), '')    AS CLASS_1
                                  ,NVL(SUBSTR(PRD.PRD_CLS_CD, 4,2), '')    AS CLASS_2
                                  ,NVL(SUBSTR(PRD.PRD_CLS_CD, 6,2), '')    AS CLASS_3
                                  ,NVL(SUBSTR(PRD.PRD_CLS_CD, 8,2), '')    AS CLASS_4
                                  ,PRD_CD
                                  ,1 SORT_SEQ
                              FROM PRD_PRD_M PRD
                             WHERE EXISTS (SELECT 'Y'
                                             FROM GS_KNOW_RESTRICT A
                                            WHERE A.LCLASS = SUBSTR(PRD.PRD_CLS_CD, 1,3)
                                              AND DECODE(A.MCLASS, '~', SUBSTR(PRD.PRD_CLS_CD, 4,2), A.MCLASS) = SUBSTR(PRD.PRD_CLS_CD, 4,2)
                                              AND DECODE(A.SCLASS, '~', SUBSTR(PRD.PRD_CLS_CD, 6,2), A.SCLASS) = SUBSTR(PRD.PRD_CLS_CD, 6,2)
                                              AND DECODE(A.SCLASS, '~', SUBSTR(PRD.PRD_CLS_CD, 8,2), A.DCLASS) = SUBSTR(PRD.PRD_CLS_CD, 8,2)
                                              AND ROWNUM = 1
                                          )
                              AND NOT EXISTS (SELECT 1
                                                FROM GS_KNOW_RESTRICT_EXCEPT_PRD  B
                                               WHERE B.PRDID = PRD.PRD_CD
                                                 AND DBSTS='A'
                                                 AND BBS_ID='estimate'
                                             )
                              AND PRD_CD IN
                               (  :prd_cd )
                            UNION ALL
                           SELECT '0'    AS CLASS_1
                                 ,'0'    AS CLASS_2
                                 ,'0'    AS CLASS_3
                                 ,'0'    AS CLASS_4
                                 ,PRD_CD
                                 ,2 SORT_SEQ
                             FROM PRD_PRD_M C
                            WHERE EXISTS (SELECT 1
                                            FROM GS_KNOW_RESTRICT_EXCEPT_PRD B
                                           WHERE B.PRDID = C.PRD_CD
                                             AND DBSTS   ='A'
                                             AND BBS_ID  ='restrict'
                                         )
                              AND PRD_CD IN
                               (   :prd_cd   )
                            )
                      )
               WHERE R_IDX = 1
             ) T19
            , (SELECT /*+ NO_MERGE  */
                 * FROM (
                      SELECT
                          PRD_CD, UDA_VAL AS ORD_LIMIT, UDA_VAL_1 AS ORD_LIMIT_DAYS
                            , ROW_NUMBER() OVER(PARTITION BY PRD_CD ORDER BY PRD_CD) RN
                        FROM PRD_UDA_D
                       WHERE UDA_GBN_CD = '10'
                         AND UDA_NO = 15
                         AND USE_YN = 'Y'
                         AND SYSDATE BETWEEN VALID_STR_DTM AND VALID_END_DTM
                     ) T24
               WHERE RN   <=   1
           ) T24  
               , ( SELECT  /*+ PUSH_PRED */
                           W.PRDID
                          ,NVL(SUM(CUSTOM_CNT), 0) AS KNOW_CNT
                          ,DECODE(NVL(SUM(CUSTOM_CNT), 0), 0, 0, ROUND(SUM(CUSTOM_TOTAL)/SUM(CUSTOM_CNT))) CUSTOM_AVG
                     FROM  ( SELECT /*+ INDEX(X GS_KNOW_SHARE_PRD_IX1)*/
                                    X.ACTIVE_PRDID AS PRDID
                                    , Y.ESTIMATE_COUNT AS CUSTOM_CNT
                                    , DECODE(NVL (Y.ESTIMATE_COUNT, 0), 0, 0, Y.EVAL_PRD_TOTAL) CUSTOM_TOTAL
                               FROM GS_KNOW_SHARE_PRD X,
                                    GS_KNOW_PRD_TOTAL Y
                              WHERE X.PASSIVE_PRDID = Y.PRDID
                                AND X.DBSTS = 'A'
                              UNION ALL
                              SELECT Z.PRDID
                                    , NVL (Z.ESTIMATE_COUNT, 0) CUSTOM_CNT
                                    , DECODE (NVL(Z.ESTIMATE_COUNT, 0), 0, 0, Z.EVAL_PRD_TOTAL) CUSTOM_TOTAL
                                FROM GS_KNOW_PRD_TOTAL Z
                           ) W
                   GROUP BY W.PRDID
                 ) W
                , PRD_META_D PRDM
       WHERE 1=1
           AND PPM.PRD_CD     = PRDM.PRD_CD(+)
           AND PRDM.PRD_META_TYP_CD(+) = '30'
           AND PPM.PRD_CD     = W.PRDID(+)
         AND PPM.USE_YN     = 'Y'
         AND PPM.PRD_CD     = PNCH.PRD_CD(+)
--         AND PNCH.CHANL_CD  = :5
         AND SYSDATE BETWEEN PNCH.VALID_STR_DTM AND PNCH.VALID_END_DTM
         AND PPM.PRD_CD     = PPH.PRD_CD(+)
--         AND PPH.PRD_ATTR_GBN_CD = :6
         AND SYSDATE BETWEEN PPH.VALID_STR_DTM AND PPH.VALID_END_DTM
         AND PPM.PRD_CD     = PSCD.PRD_CD(+)
         AND PPM.PRD_CD     = PCD.PRD_CD(+)
--         AND PCD.CHANL_CD   = :7    
         AND PPM.PRD_CD     = PPD.PRD_CD(+)
         AND PPM.SUP_CD     = PCDDD.SUP_CD(+)
         AND PPM.PRD_CLS_CD = PCDDD.PRD_CLS_CD(+)
         AND PPM.OPER_MD_ID = LB.MD_ID(+)
         AND PPM.PRD_CD     = PND.PRD_CD(+)
         AND '10'           = PND.UNIT_GBN_CD(+)
         AND PPM.RTP_DLVC_CD= SDD.DLVC_CD(+)
         AND PPM.CHR_DLVC_CD= SDD2.DLVC_CD(+)
         AND PPD.ILND_CHR_DLVC_CD = ILNDSDD.DLVC_CD(+)    
         AND PPD.JEJU_CHR_DLVC_CD = JEJUSDD.DLVC_CD(+)    
         AND PPD.ILND_RTP_DLVC_CD = ILNDRTPSDD.DLVC_CD(+) 
         AND PPD.JEJU_RTP_DLVC_CD = JEJURTPSDD.DLVC_CD(+) 
         AND PPD.DAWN_CHR_DLVC_CD = DAWNSDD.DLVC_CD(+)    
         AND PPM.SUP_CD     = SSM.SUP_CD(+)
         AND PPM.PRD_CD     = PMD.PRD_CD(+)
         AND '20'           = PMD.PRD_META_TYP_CD(+)
         AND PPM.PRD_CD     = PBD.PRD_CD(+)
         AND PPM.PRD_CD     = T19.PRD_CD(+)
         AND PPM.PRD_CD     = T24.PRD_CD(+)
         """

            # cursor.prepare(sql)
            cursor.execute(sql, {'prd_cd': prd_cd})
            data = cursor.fetchall()
            # print(data)
        return ''


class MetaReqLV(LoginRequiredMixin, FormMixin, ListView):
    model = MetaReq
    template_name = 'MetaReqLV.html'
    context_object_name = 'objects'
    form_class = MetaReqLVForm
    paginate_by = 60

    def get_initial(self):

        _loggingVisit(self.request, 28)

        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################

        title = self.request.session['title'] if 'title' in self.request.session else ""
        choice_status = self.request.session['choice_status'] if 'choice_status' in self.request.session else ""

        search_my_accept = self.request.session[
            'search_my_accept'] if 'search_my_accept' in self.request.session else ""
        search_my_req = self.request.session['search_my_req'] if 'search_my_req' in self.request.session else ""

        return {
            'title': title,
            'choice_status': choice_status,
            'search_my_accept': search_my_accept,
            'search_my_req': search_my_req,
        }

    def get_queryset(self):

        self.form = self.get_form(self.form_class)

        search_my_accept = self.request.session[
            'search_my_accept'] if 'search_my_accept' in self.request.session else ""
        search_my_req = self.request.session['search_my_req'] if 'search_my_req' in self.request.session else ""

        if self.request.POST.get("search_my_accept") == "search_my_accept":
            if search_my_accept:
                search_my_accept = False
            else:
                search_my_accept = True

        if self.request.POST.get("search_my_req") == "search_my_req":
            if search_my_req:
                search_my_req = False
            else:
                search_my_req = True

        search_week_req = ""
        if self.request.POST.get("search_week_req") == "search_week_req":
            if search_week_req:
                search_week_req = False
            else:
                search_week_req = True

        if self.form.is_valid():

            title = self.form.cleaned_data['title']
            choice_status = self.form.cleaned_data['choice_status']

            ###################################################
            ## 화면에 바인딩 된 값을 유지하기 위한 처리
            ###################################################
            self.request.session['title'] = title if title is not None else ""
            self.request.session['choice_status'] = choice_status if choice_status is not None else ""
            self.request.session['search_my_accept'] = search_my_accept if search_my_accept is not None else ""
            self.request.session['search_my_req'] = search_my_req if search_my_req is not None else ""

            ###################################################
            ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            self.request.session['click_search'] = 'Y'
            ###################################################

        else:
            ###################################################
            ## 1. 처음 진입
            ## 2. 페이징 버튼 클릭 시
            ###################################################

            page_kwarg = self.page_kwarg
            page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
            # if page == 1 :
            #   self.request.session['title'] = ""
            #   self.request.session['choice_status'] = ""

            title = self.request.session['title'] if 'title' in self.request.session else ""
            choice_status = self.request.session['choice_status'] if 'choice_status' in self.request.session else ""
            search_my_accept = self.request.session[
                'search_my_accept'] if 'search_my_accept' in self.request.session else ""
            search_my_req = self.request.session['search_my_req'] if 'search_my_req' in self.request.session else ""

        # obj = MetaReq.objects.all().order_by('-reg_dtm')

        if title is None:
            title = ""

        obj = ""
        obj1 = ""
        obj2 = ""

        if choice_status == '1':
            obj = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                Q(pl_dev_yn=0),
                Q(pl_prod_yn=0),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
            ).order_by('-reg_dtm')

        elif choice_status == '2':
            obj = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(id_tablelist__table_name__icontains=title) | Q(
                    table_name__icontains=title) | Q(id_reg_user__first_name=title),
                Q(pl_dev_yn=1),
                Q(da_dev_yn=0),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
            ).order_by('-reg_dtm')
        elif choice_status == '3':
            obj = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                Q(da_dev_yn=1),
                Q(dba_dev_yn=0),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
            ).order_by('-reg_dtm')
        elif choice_status == '4':
            obj = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                Q(dba_dev_yn=1),
                Q(req_prod_yn=0),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
            ).order_by('-reg_dtm')
        elif choice_status == '5':
            obj = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                Q(req_prod_yn=1),
                Q(pl_prod_yn=0),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
            ).order_by('-reg_dtm')
        elif choice_status == '6':
            obj = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                Q(pl_prod_yn=1),
                Q(dba_prod_yn=0),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
            ).order_by('-reg_dtm')
        elif choice_status == '7':

            obj = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                Q(dba_prod_yn=1),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
            ).order_by('-reg_dtm')
        elif choice_status == '8':
            obj = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                Q(pl_dev_yn=2) | Q(pl_prod_yn=2) | Q(dba_dev_yn=2) | Q(dba_prod_yn=2) | Q(da_dev_yn=2),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
            ).order_by('-reg_dtm')
        else:
            obj1 = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
                Q(dba_prod_yn=0, dist_dtm__gte=datetime.datetime.now())
            ).order_by('-check_prod_exec_yn', '-pl_prod_yn', 'dist_dtm', 'id')
            obj2 = MetaReq.objects.filter(
                Q(id=title if title.isdigit() else -1) | Q(title__icontains=title) | Q(
                    id_tablelist__table_name__icontains=title) | Q(table_name__icontains=title) | Q(
                    id_reg_user__first_name=title),
                *((Q(id_pl_prod=self.request.user),) if search_my_accept else ()),
                *((Q(id_reg_user=self.request.user),) if search_my_req else ()),
                *((Q(check_prod_exec_yn=1),) if search_week_req else ()),
                (Q(dba_prod_yn__in=[1, 2], dist_dtm__gte=datetime.datetime.now()) |
                 Q(dist_dtm__lt=datetime.datetime.now()) |
                 Q(dist_dtm=None))
            ).order_by('-check_prod_exec_yn', 'dba_prod_yn', '-reg_dtm')

        if obj1 != "":
            obj = list(chain(obj1, obj2))

        return obj

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def paginate_queryset(self, queryset, page_size):

        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty())
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            if page == 'last':
                page_number = paginator.num_pages
            else:
                raise Http404(_("Page is not 'last', nor can it be converted to an int."))
        try:

            ###################################################
            ## SEARCH 버튼 클릭 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            if 'click_search' in self.request.session:
                if self.request.session['click_search'] == "N":
                    page = paginator.page(page_number)
                else:
                    page = paginator.page(1)
            else:
                page = paginator.page(page_number)
            ###################################################

            ###################################################
            ## COLUMNLIST 에서 목록으로 클릭 시 원래 페이지 번호를 찾아가기 위한 코드
            ###################################################
            self.request.session['page_number'] = page_number

            return (paginator, page, page.object_list, page.has_other_pages())



        except:
            page = paginator.page(1)
            return (paginator, page, page.object_list, page.has_other_pages())

    def get_context_data(self, **kwargs):

        add_secu_msg = ""

        context = super().get_context_data(**kwargs)

        page_range = ""
        paginator = context['paginator']
        page_numbers_range = 5  # Display only 5 page numbers
        max_index = len(paginator.page_range)

        ###################################################
        ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
        ###################################################
        if 'click_search' in self.request.session:
            if self.request.session['click_search'] == "N":
                page = self.request.GET.get('page')
            else:
                page = 1
                self.request.session['click_search'] = "N"
        else:
            page = self.request.GET.get('page')
        ###################################################

        current_page = int(page) if page else 1

        start_index = int((current_page - 1) / page_numbers_range) * page_numbers_range
        end_index = start_index + page_numbers_range
        if end_index >= max_index:
            end_index = max_index

        page_range = paginator.page_range[start_index:end_index]

        email = ""
        email_contents = ""
        domain_of_new_table = ""
        if self.request.POST.get("search_week_req") == "search_week_req":
            email_contents = """<b>제목 : [공유] DB 정기 반영 작업 공유</b><br>
안녕하세요. 클라우드팀 입니다.

<b><font color="red">내일 새벽 반영될 DB정기반영건 공유</font></b>드립니다.

"{title}" 외 {cnt}건으로 작업 중 Lock 발생이 예상되어 <b>순환방송 시간(02:00)</b>에 반영할 예정입니다.<br>
관련 내용은 SmartDBA 내 DB형상관리 메뉴에서 확인할 수 있습니다.
<b><a href="http://smartdba.gshs.co.kr/database/MetaReqLV/" target="blank">http://smartdba.gshs.co.kr/database/MetaReqLV/</a></b><br>감사합니다.

      """

            add_secu_yn = ""

            email = []
            data_email = SendEmailTargetList.objects.filter(send_case='1')
            for e in data_email:
                email.append(e.id_receive_user.email)

            new_table = MetaReq.objects.filter(check_prod_exec_yn=1,
                                               obj_class=2, # 테이블
                                               # obj_new=2, # 신규
                                              )

            domain_of_new_table = []

            for tab in new_table :
                try :                                    
                    if tab.id_domainanddblist.erd_user :
                        domain_of_new_table.append(tab.id_domainanddblist.id_dblist.db_use+" : "+tab.id_domainanddblist.id_domain.domain_name)
                except Exception as e :
                    pass
            domain_of_new_table = list(set(domain_of_new_table))
            


            objs = MetaReq.objects.filter(check_prod_exec_yn=1)
            for o in objs:

                metareqlist = MetaReqList.objects.filter(id_metareq=o.id)

                for m in metareqlist:
                    if m.privacy_list != "0" and m.change_list in ['+ 신규', '+ └ TOBE']:
                        add_secu_yn = "Y"
                        break

                email.append(o.id_reg_user.email)
                email.append(o.id_pl_prod.email)

                email_contents = email_contents.format(title=o.title,
                                                       cnt=str(len(objs) - 1))

            if add_secu_yn == 'Y':
                secu_email = SendEmailTargetList.objects.filter(send_case='2')
                for e in secu_email:
                    email.append(e.id_receive_user.email)
                add_secu_msg = "<font color=red><b>* 개인정보가 추가되어 보안센터 담장자에게도 메일 드립니다.</b></font>"

            email = list(set(email))

        prod_exec_count = MetaReq.objects.filter(check_prod_exec_yn="1").order_by('id_dblist', 'id')
        prod_exec_count = len(prod_exec_count)

        context['page_range'] = page_range

        context['email'] = '; '.join(email) if email != "" else ""
        context['email_contents'] = email_contents
        context['prod_exec_count'] = prod_exec_count
        context['add_secu_msg'] = add_secu_msg
        context['domain_of_new_table'] = domain_of_new_table
        context['template_parent'] = "MetaReqLV"
        context['template_child'] = "MetaReqLV"

        return context


class MetaReqListUV(LoginRequiredMixin, CreateView):
    model = MetaReq
    form_class = MetaReqForm
    template_name = 'MetaReqListUV.html'
    context_object_name = 'objects'


    def get_initial(self):

        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################

        obj = MetaReq.objects.get(id=self.kwargs['id_metareq'])

        id_dblist = obj.id_dblist.id

        if obj.id_tablelist:
            id_tablelist = obj.id_tablelist.id
            owner = obj.id_tablelist.owner

            id_dbowner = DBOwner.objects.filter(id_dblist=id_dblist,
                                                owner=owner, ).first()
            if id_dbowner:
                owner = id_dbowner.id
            else:
                owner = -1

        else:
            id_tablelist = None
            owner = None

        title = obj.title
        csr = obj.csr
        req_contents = obj.req_contents
        table_name = obj.table_name
        table_comments = obj.table_comments
        obj_class = obj.obj_class
        obj_new = obj.obj_new
        id_domainanddblist = obj.id_domainanddblist
        storage_cycle = obj.storage_cycle
        storage_cycle_column = obj.storage_cycle_column
        id_pl_dev = obj.id_pl_dev

        return {
            'title': title,
            'csr': csr,
            'id_tablelist': id_tablelist,
            'id_dblist': id_dblist,
            'id_pl_dev': id_pl_dev,
            'req_contents': req_contents,
            'table_name': table_name,
            'table_comments': table_comments,
            'obj_class': obj_class,
            'obj_new': obj_new,
            'id_domainanddblist': id_domainanddblist,
            'owner': owner,
            'storage_cycle': storage_cycle,
            'storage_cycle_column': storage_cycle_column,
        }

    def get_queryset(self):

        obj = MetaReqList.objects.filter(id_metareq=self.kwargs['id_metareq']).order_by("column_id", 'id')

        return obj

    def get_context_data(self, **kwargs):

        context = super(CreateView, self).get_context_data(**kwargs)

        obj = MetaReq.objects.get(id=self.kwargs['id_metareq'])

        if obj.id_tablelist:
            id_tablelist = obj.id_tablelist.id
        else:
            id_tablelist = None

        id_dblist = obj.id_dblist.id

        context['template_parent'] = "MetaReqLV"
        context['template_child'] = "MetaReqLV"
        context['id_metareq'] = self.kwargs['id_metareq']
        context['id_tablelist'] = id_tablelist
        context['id_dblist'] = id_dblist

        return context


class MetaReqListLV(LoginRequiredMixin, FormMixin, ListView):
    model = MetaReq
    template_name = 'MetaReqListLV.html'
    context_object_name = 'objects'
    form_class = MetaReqForm

    def get_initial(self):

        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################

        obj = MetaReq.objects.get(id=self.kwargs['id_metareq'])

        if obj:
            id_dblist = obj.id_dblist.id
        else:
            id_dblist = ""

        return {
            'id_dblist': id_dblist,
        }

    def get_queryset(self):

        obj = MetaReqList.objects.filter(id_metareq=self.kwargs['id_metareq']).order_by("column_id","id")
        return obj

    def get_context_data(self, **kwargs):

        meta = MetaReq.objects.get(id=self.kwargs['id_metareq'])

        try :
            ds_emp = _query_dict('iam', iamSql.format(emp_no=meta.id_reg_user.username))[0]
            req_user_tel = ds_emp['MOBILE']
        except :
            req_user_tel = ""

        try :
            ds_emp = _query_dict('iam', iamSql.format(emp_no=meta.id_pl_prod.username))[0]
            pl_prod_tel = ds_emp['MOBILE']
        except :
            pl_prod_tel = ""

        add_secu_yn = ""
        metareqlist = MetaReqList.objects.filter(id_metareq=self.kwargs['id_metareq'])

        for m in metareqlist:
            if m.privacy_list != "0" and m.change_list in ['+ 신규', '+ └ TOBE']:
                add_secu_yn = "+ 개인정보 컬럼 추가"
                break

        context = super().get_context_data(**kwargs)

        context['template_parent'] = "MetaReqLV"
        context['template_child'] = "MetaReqLV"
        context['id_metareq'] = self.kwargs['id_metareq']
        context['meta'] = meta
        context['req_user_tel'] = req_user_tel
        context['pl_prod_tel'] = pl_prod_tel
        context['add_secu_yn'] = add_secu_yn

        return context


class MetaReqListCV(LoginRequiredMixin, CreateView):
    model = MetaReq
    form_class = MetaReqForm
    template_name = 'MetaReqListCV.html'

    # success_url = reverse_lazy('sqlTuningMain')

    def get_context_data(self, **kwargs):
        context = super(CreateView, self).get_context_data(**kwargs)

        context['template_parent'] = "MetaReqLV"
        context['template_child'] = "MetaReqLV"

        return context


# class MetaRegLV(LoginRequiredMixin, FormMixin, ListView) :


#   model = ColumnList
#   template_name = 'MetaRegLV.html'
#   context_object_name = 'objects'
#   form_class = ColumnListForm

#   def get_queryset(self):

#     obj = []

#     return obj


#   def get_context_data(self, **kwargs):


#       context = super().get_context_data(**kwargs)


#       context['template_parent'] = "dba"
#       context['template_child'] = "MetaRegLV"

#       return context


class DBAWorkLV(ListView):
    template_name = 'DBAWorkLV.html'
    context_object_name = 'objects'
    paginate_by = 15

    def get_queryset(self):

        obj = []

        return obj

    def get_context_data(self, **kwargs):        

        context = super().get_context_data(**kwargs)

        obj_priv = UserRequestTabPrivHist.objects.filter(approv_yn=0, use_yn='1').order_by("id_objectlist__db_use",
                                                                                           "id_userlist__username")
        
        user_acc_all = UserList.objects.filter(
                                            Q(request_status=1),
                                            ~Q(id_dblist=39)
                                          ).order_by("oper_cd", "id_dblist", "id_user_name")

        user_acc_dhub = UserList.objects.filter(
                                            Q(request_status=1),
                                            Q(id_dblist=39)
                                          ).order_by("oper_cd", "id_dblist", "id_user_name")

        tuning = TuningList.objects.filter(id_tuningstatus=1)

        # sql = """
        #   SELECT A.REQ_NO req_no, --요청번호
        #        DECODE(SYS_NM,   '[N]DHUB','DHUB',
        #                         '[N]TCPROD','TCPROD',
        #                         '[N]NEWINSA','ORAINSA',
        #                         '[N]MySQL','MYSQL',
        #                         '[N]SAP','PRD',
        #                         '[N]WEBDB2','WEBDB',
        #                         '[N]TISTM','TISTM',
        #                         '[N]OASPRD','OASPRD',
        #                         '[N]SMTC','SMTCPRD',
        #                         '[N]LGHS002N','ORANBS',
        #                         '[N]ETC','ETC',
        #                         '[N]BaseITSM','ORAITSM',
        #                         '[N]D&Shop','D&SHOP', SYS_NM) db_nm, --시스템
        #        decode(c.apc_type, '0001','신규','0002','삭제','0003',decode(c.apc_obj_type,'0002','컬럼추가','변경'),'0004','컬럼변경') req_div, -- 신청구분
        #        decode(c.status, '0001','검토중','1001','PL승인','1002','PL반려','0002','승인','0003','반려','0004','DB반영요청중','0005','테스트DB반영완료','0006','테스트DB반영오류','0007','운영DB반영요청중', '0008','운영DB반영완료','0009','운영DB반영오류','운영DB미반영','0010') procss_status, -- 처리상태
        #        (select user_nm from meta.usr_user where user_id = c.usr_id) req_user, -- 신청자
        #        to_char(to_date(c.create_date,'yyyymmddhh24miss'),'yyyy-mm-dd') req_dt, -- 신청일자
        #        trim(replace(c.remark1,'\n','\r\n')) req_explain, -- 설명
        #        SYSDATE curr_dt, -- 현재일자
        #        C.STATUS system_status -- 10
        #   FROM META.GS_MD_APC_SCRIPT A, META.GS_MD_SYSTEM B, META.GS_MD_APC C
        #   WHERE
        #       A.REQ_NO = C.REQ_NO
        #   AND C.SYS_ID = B.SYS_ID
        #   AND C.STATUS in ('0007') -- 반영요청
        #   -- AND to_date(c.create_date,'yyyymmddhh24miss') > SYSDATE - 3
        #   AND A.REQ_NO > 9113 -- 해당정책 시행이후 건만
        #   ORDER BY A.REQ_NO
        # """

        # req_meta = _query_dict("metadb",sql)

        req_meta = ""

        sql = """
      SELECT a.db_use,
             case when a.oper_cd = 3 then '운영'
                  when a.oper_cd = 4 then '개발' end oper_cd,
             lower(a.username) username,
             a.id,
             a.conn_ip,
             e.db_type
      FROM CUST_USER_LIST A,
           STG_DBA_USERS B,
           CUST_DB_LIST C,
           CUST_DB_TYPE D,
           CUST_DB E
      WHERE A.OPER_CD = B.OPER_CD
        AND A.ID_DBLIST = B.ID_DBLIST
        AND A.USERNAME = B.USERNAME
        AND A.ID_DBLIST = C.ID
        AND C.ID_DBTYPE = D.ID
        AND D.ID_DB = E.ID
        AND A.DROP_YN = 0
        AND C.ID <> 39
        AND A.USERNAME <> 'I_GSSHOP'
      AND not exists
      (
        SELECT 1
          FROM AUTH_USER C
          WHERE C.ID = A.ID_USER_NAME
          AND C.first_name <> '퇴사자'
    )
    order by oper_cd, db_use
      """

        user_acc_delete_all = _query_dict("default", sql)


        sql = """
      SELECT a.db_use,
             case when a.oper_cd = 3 then '운영'
                  when a.oper_cd = 4 then '개발' end oper_cd,
             lower(a.username) username,
             a.id,
             a.conn_ip,
             e.db_type
      FROM CUST_USER_LIST A,
           STG_DBA_USERS B,
           CUST_DB_LIST C,
           CUST_DB_TYPE D,
           CUST_DB E
      WHERE A.OPER_CD = B.OPER_CD
        AND A.ID_DBLIST = B.ID_DBLIST
        AND A.USERNAME = B.USERNAME
        AND A.ID_DBLIST = C.ID
        AND C.ID_DBTYPE = D.ID
        AND D.ID_DB = E.ID
        AND A.DROP_YN = 0
        AND C.ID = 39
      AND not exists
      (
        SELECT 1
          FROM AUTH_USER C
          WHERE C.ID = A.ID_USER_NAME
          AND C.first_name <> '퇴사자'
    )
    order by oper_cd, db_use
      """

        user_acc_delete_dhub = _query_dict("default", sql)

        sql = """
          SELECT MNT_CD
          FROM MNT_MNT_M A
          WHERE A.APRV_ST_CD = '10'
          AND A.use_yn = 'Y'
      """

        bms = _query_dict("BMS", sql)

        sql = """
          SELECT DB_USE,
             EXP_ORDER,
             ID_DBDETAIL,
             HA_CASE,
             EXP_ORDER,
             MAX(RISK_LEVEL) RISK_LEVEL
            FROM
            (
            SELECT  CONCAT(substr(B.DB_USE,1,8),'(', D.DB_ORDER,')') DB_USE,
                    CASE WHEN a.error_yn = '1' then E.RISK_LEVEL
                    WHEN e.compare_case = '0' then
                  (
                    CASE WHEN E.LIMIT_VALUE = A.MONITOR_VALUE THEN -1
                    ELSE E.RISK_LEVEL END
                          )
                   WHEN e.compare_case = '1' then
                  (
                    CASE WHEN CAST(E.LIMIT_VALUE as double) > CAST(A.MONITOR_VALUE as double) THEN -1
                    ELSE E.RISK_LEVEL END
                          )
                   WHEN e.compare_case = '2' then
                  (
                    CASE WHEN CAST(E.LIMIT_VALUE as double) < CAST(A.MONITOR_VALUE as double) THEN -1
                    ELSE E.RISK_LEVEL END
                          )
                END RISK_LEVEL,
              B.EXP_ORDER,
              D.ID ID_DBDETAIL,
              F.ID HA_CASE,
                A.MONITOR_VALUE,
                E.LIMIT_VALUE
            FROM CUST_MONITOR_ITEM_LOG A  use index(ix_cust_monitor_item_log_02),
             CUST_DB_LIST B,
             CUST_MONITOR_ITEM_LIST C,
             CUST_DB_DETAIL D,
             cust_dbdetail_monitoritemlist e,
             cust_ha_case f,
             cust_monitorm_dbdetail g
            -- WHERE A.REG_DTM >= DATE_ADD(STR_TO_DATE(DATE_FORMAT(NOW(), '%Y%m%d %H%i'),'%Y%m%d %H%i'), INTERVAL -1 MINUTE)
            WHERE A.CURR_YN = '1'
                AND A.id_dbdetailmonitoritemlist = E.ID
                AND E.id_monitoritemlist = C.ID
                AND E.ID_DBDETAIL = D.ID
                AND D.ID_DBLIST = B.ID
                AND D.OPER_CD = 4
                AND D.id_hacase = F.id            
                AND D.ID = G.DBDETAIL_ID
                AND G.MONITORMANAGEMENT_ID = 2 -- 2 : DB기본모니터링
                AND B.USE_YN = 1
                AND D.USE_YN = 1
                ORDER BY B.EXP_ORDER, B.DB_USE
            ) A
            GROUP BY EXP_ORDER, DB_USE, ID_DBDETAIL, HA_CASE, EXP_ORDER
            ORDER BY EXP_ORDER, DB_USE
      """

        dev_monitor = _query_dict("default", sql)

        sql = """
          SELECT DB_USE,
             EXP_ORDER,
             ID_DBDETAIL,
             HA_CASE,
             EXP_ORDER,
             MAX(RISK_LEVEL) RISK_LEVEL
            FROM
            (
            SELECT  CONCAT(substr(B.DB_USE,1,8),'(', D.DB_ORDER,')') DB_USE,
                    CASE WHEN a.error_yn = '1' then E.RISK_LEVEL
                    WHEN e.compare_case = '0' then
                  (
                    CASE WHEN E.LIMIT_VALUE = A.MONITOR_VALUE THEN -1
                    ELSE E.RISK_LEVEL END
                          )
                   WHEN e.compare_case = '1' then
                  (
                    CASE WHEN CAST(E.LIMIT_VALUE as double) > CAST(A.MONITOR_VALUE as double) THEN -1
                    ELSE E.RISK_LEVEL END
                          )
                   WHEN e.compare_case = '2' then
                  (
                    CASE WHEN CAST(E.LIMIT_VALUE as double) < CAST(A.MONITOR_VALUE as double) THEN -1
                    ELSE E.RISK_LEVEL END
                          )
                END RISK_LEVEL,
              B.EXP_ORDER,
              D.ID ID_DBDETAIL,
              F.ID HA_CASE,
                A.MONITOR_VALUE,
                E.LIMIT_VALUE
            FROM CUST_MONITOR_ITEM_LOG A  use index(ix_cust_monitor_item_log_02),
             CUST_DB_LIST B,
             CUST_MONITOR_ITEM_LIST C,
             CUST_DB_DETAIL D,
             cust_dbdetail_monitoritemlist e,
             cust_ha_case f,
             cust_monitorm_dbdetail g
            -- WHERE A.REG_DTM >= DATE_ADD(STR_TO_DATE(DATE_FORMAT(NOW(), '%Y%m%d %H%i'),'%Y%m%d %H%i'), INTERVAL -1 MINUTE)
            WHERE A.CURR_YN = '1'
                AND A.id_dbdetailmonitoritemlist = E.ID
                AND E.id_monitoritemlist = C.ID
                AND E.ID_DBDETAIL = D.ID
                AND D.ID_DBLIST = B.ID
                AND D.OPER_CD = 3
                AND D.id_hacase = F.id            
                AND D.ID = G.DBDETAIL_ID
                AND G.MONITORMANAGEMENT_ID = 2 -- 2 : DB기본모니터링
                AND B.USE_YN = 1
                AND D.USE_YN = 1                
                ORDER BY B.EXP_ORDER, B.DB_USE
            ) A
            GROUP BY EXP_ORDER, DB_USE, ID_DBDETAIL, HA_CASE, EXP_ORDER
            ORDER BY EXP_ORDER, DB_USE
      """

        prod_monitor = _query_dict("default", sql)

        sql = """
            select concat(TIMESTAMPDIFF(second, MIN(A.REG_DTM), now()),'초 전') min_dtm
            from cust_monitor_item_log A USE INDEX(IX_CUST_MONITOR_ITEM_LOG_02),
                 cust_dbdetail_monitoritemlist B,
                 cust_db_detail c
            where A.curr_yn = '1'
              and a.id_dbdetailmonitoritemlist = b.id
              and b.id_dbdetail = c.id
              and c.oper_cd = 3  
              and A.error_yn = 0;
            """
        prod_min_dtm = _query_dict("default", sql)

        sql = """
            SELECT A.CNT META_CNT,
                   B.CNT ACCOUNT_CNT,
                   C.CNT PRIV_CNT,
                   D.CNT DOMAIN_CNT,
                   E.CNT ATTR_CNT,
                   F.CNT WORD_CNT,
                   G.CNT TUNING_CNT,
                   H.CNT USER_CREATE_CNT
            FROM
            (
            SELECT COUNT(*) CNT
            FROM CUST_META_REQ
            WHERE reg_dtm > DATE_SUB(now(), interval 7 day)
            AND dba_dev_yn = 1
            ) A,
            (
            SELECT COUNT(*) CNT
            FROM cust_user_request_account_open_hist
            WHERE reg_dtm > DATE_SUB(now(), interval 7 day)
            ) B,
            (
            SELECT COUNT(*) CNT
            FROM cust_user_request_tab_priv_hist
            WHERE reg_dtm > DATE_SUB(now(), interval 7 day)
            ) C,
            (
            SELECT COUNT(*) CNT
            FROM cust_std_domain
            WHERE reg_dtm > DATE_SUB(now(), interval 7 day)
            ) D,
            (
            SELECT COUNT(*) CNT
            FROM cust_std_attr
            WHERE reg_dtm > DATE_SUB(now(), interval 7 day)
            ) E,
            (
            SELECT COUNT(*) CNT
            FROM cust_std_word
            WHERE reg_dtm > DATE_SUB(now(), interval 7 day)
            ) F,
            (
            SELECT COUNT(*) CNT
            FROM cust_tuning_list
            WHERE reg_dtm > DATE_SUB(now(), interval 7 day)
            ) G,
            (
            select COUNT(*) CNT
            from cust_user_list
            WHERE reg_dtm > DATE_SUB(now(), interval 7 day)
            )H;
            """
        dba_work = _query_dict("default", sql)
        dba_work = dba_work[0]

        sql = """
            select concat(TIMESTAMPDIFF(second, MIN(A.REG_DTM), now()),'초 전') min_dtm
            from cust_monitor_item_log A use index(ix_cust_monitor_item_log_02) ,
                 cust_dbdetail_monitoritemlist B,
                 cust_db_detail c
            where A.curr_yn = '1'
              and a.id_dbdetailmonitoritemlist = b.id
              and b.id_dbdetail = c.id
              and c.oper_cd = 4  
              and A.error_yn = 0;
            """
        dev_min_dtm = _query_dict("default", sql)

        sql = """
          SELECT MONITOR_NAME, 
                 CASE WHEN ALERT_YN=1 THEN 'ON'
                      ELSE '<font color="red"><b>OFF</b></font>'
                END ALERT_YN,
                  CASE WHEN MONITOR_YN=1 THEN 'ON'
                      ELSE '<font color="red"><b>OFF</b></font>'
                END MONITOR_YN
          from cust_monitor_management
          order by id
        """

        monitor_yn = _query_dict("default", sql)


        sql = """
            SELECT distinct A.DAG_NAME, A.SERVICE_NAME, A.TEAM, A.REQ_USER, B.STATE, C.SCHEDULE_INTERVAL
            from cust_airflow_manage a,
                 airflow_smartdba.dag_run b,
                 airflow_smartdba.dag c
            where a.dag_id = b.dag_id
            and   a.dag_id = c.dag_id
            and b.execution_date > DATE(NOW())
            and state not in ('success','running')            
        """

        airflow_error = _query_dict("default", sql)


        sql = """
            select  distinct a.id ID,concat(d.oper_cd,' ',c.db_use,' 노드', db_order) NOT_ALERT
            from cust_db_detail a,
                 cust_dbdetail_monitoritemlist b,
                 cust_db_list c,
                 cust_oper_cd d
            where a.id = b.id_dbdetail
              and c.id = a.id_dblist
              and a.oper_cd = d.id
              and a.alert_yn = "0"     
              and a.use_yn = "1"
        """

        alert_yn = _query_dict("default", sql)

        req_word = StdWord.objects.filter(accept_yn=1)
        req_domain = StdDomain.objects.filter(accept_yn=1)
        req_attr = StdAttr.objects.filter(accept_yn=1, use_yn=1)
        

        # metareq = MetaReq.objects.filter(dba_dev_yn=0)
        metareq_all = MetaReq.objects.filter(
                                                Q(dba_dev_yn=0),
                                                ~Q(id_dblist=39)
                                            )
        metareq_dhub = MetaReq.objects.filter(
                                                Q(dba_dev_yn=0),
                                                Q(id_dblist=39)
                                            )

        sql = """
                SELECT *
                FROM
                (
                  SELECT    d.OPER_CD ID_OPERCD,
                            g.OPER_CD,
                            I.SMS_MESSAGE,
                            CONCAT(substr(B.DB_USE,1,8),'(', D.DB_ORDER,')') DB_USE,
                      CASE WHEN a.error_yn = '1' then E.RISK_LEVEL
                      WHEN e.compare_case = '0' then
                      (
                      CASE WHEN E.LIMIT_VALUE = A.MONITOR_VALUE THEN -1
                      ELSE E.RISK_LEVEL END
                          )
                       WHEN e.compare_case = '1' then
                      (
                      CASE WHEN CAST(E.LIMIT_VALUE as double) > CAST(A.MONITOR_VALUE as double) THEN -1
                      ELSE E.RISK_LEVEL END
                          )
                       WHEN e.compare_case = '2' then
                      (
                      CASE WHEN CAST(E.LIMIT_VALUE as double) < CAST(A.MONITOR_VALUE as double) THEN -1
                      ELSE E.RISK_LEVEL END
                          )
                    END RISK_LEVEL,
                    B.EXP_ORDER,
                    D.ID ID_DBDETAIL,
                    F.ID HA_CASE,
                    CASE WHEN A.MONITOR_VALUE='Error' then A.ERROR_MSG
                                        ELSE A.MONITOR_vALUE
                                        END MONITOR_VALUE,
                    E.LIMIT_VALUE,
                    A.ERROR_YN,
                    A.ERROR_MSG,
                    E.NOTE1
                  FROM CUST_MONITOR_ITEM_LOG A use index(ix_cust_monitor_item_log_02),
                   CUST_DB_LIST B,
                   CUST_MONITOR_ITEM_LIST C,
                   CUST_DB_DETAIL D,
                   cust_dbdetail_monitoritemlist e,
                   cust_ha_case f,
                   cust_oper_cd g,
                   cust_monitorm_dbdetail h,
                   CUST_MONITOR_ITEM I
                  -- WHERE A.REG_DTM >= DATE_ADD(STR_TO_DATE(DATE_FORMAT(NOW(), '%Y%m%d %H%i'),'%Y%m%d %H%i'), INTERVAL -1 MINUTE)
                  WHERE A.CURR_YN = '1'
                    AND A.id_dbdetailmonitoritemlist = E.ID
                    AND E.ID_MONITORITEMLIST = C.ID
                    AND E.ID_DBDETAIL = D.ID
                    AND B.ID = D.ID_DBLIST                  
                    AND d.oper_cd = g.id                  
                    AND D.OPER_CD in (3,4)                  
                    AND D.id_hacase = F.id                  
                    AND D.ID = H.DBDETAIL_ID
                    AND H.MONITORMANAGEMENT_ID = 2 -- 2 : DB기본모니터링
                    AND B.USE_YN = 1
                    AND D.USE_YN = 1
                    AND I.ID = C.ID_MONITORITEM
                    ORDER BY B.EXP_ORDER, B.DB_USE
                ) A
                WHERE A.RISK_LEVEL > -1
            """

        monitor_msg = _query_dict("default", sql)

        error_msg = "<strong>[{oper_cd}] {db_use}</strong><br>" + \
                    "<b>{msg}</b> : {monitor_value}"

        prod_error_msg = []
        dev_error_msg = []

        risk_level = ""

        for msg in monitor_msg:

            if msg['RISK_LEVEL'] == "0":
                risk_level = "주의"
            elif msg['RISK_LEVEL'] == "1":
                risk_level = "심각"
            elif msg['RISK_LEVEL'] == "2":
                risk_level = "장애"

            # error_msg2 = ""
            # if len(msg['ERROR_YN']) == 1:
            #     error_msg2 = "<br>" + msg['ERROR_MSG']

            
            if msg['ID_OPERCD'] == 3 :
                prod_error_msg.append(error_msg.format(
                    oper_cd=msg['OPER_CD'],
                    db_use=msg['DB_USE'],
                    # risk_level=risk_level,
                    msg=(msg['SMS_MESSAGE'] if msg['SMS_MESSAGE'] is not None else '') + ("<br>"+msg['NOTE1'] if msg['NOTE1'] is not None else "") ,
                    # id_dbdetail=msg['ID_DBDETAIL'],
                    monitor_value=msg['MONITOR_VALUE'],
                    # limit_value=msg['LIMIT_VALUE'],
                ))

            elif msg['ID_OPERCD'] == 4 :
                dev_error_msg.append(error_msg.format(
                    oper_cd=msg['OPER_CD'],
                    db_use=msg['DB_USE'],
                    # risk_level=risk_level,
                    msg=msg['SMS_MESSAGE'],
                    # id_dbdetail=msg['ID_DBDETAIL'],
                    monitor_value=msg['MONITOR_VALUE'],
                    # limit_value=msg['LIMIT_VALUE'],
                ))

        prod_error_msg = list(set(prod_error_msg))
        prod_error_msg = prod_error_msg[:5]
        
        dev_error_msg = list(set(dev_error_msg))
        dev_error_msg = dev_error_msg[:5]

        context['template_parent'] = "dba"
        context['template_child'] = "DBAWorkLV"
        context['prod_error_msg'] = prod_error_msg
        context['dev_error_msg'] = dev_error_msg
        context['dba_work'] = dba_work
        context['obj_priv'] = obj_priv
        context['obj_priv_len'] = len(obj_priv)
        # context['user_acc'] = user_acc
        # context['user_acc_len'] = len(user_acc)
        context['user_acc_all'] = user_acc_all
        context['user_acc_all_len'] = len(user_acc_all)
        context['user_acc_dhub'] = user_acc_dhub
        context['user_acc_dhub_len'] = len(user_acc_dhub)
        # context['user_acc_delete'] = user_acc_delete
        # context['user_acc_delete_len'] = len(user_acc_delete)
        context['user_acc_delete_all'] = user_acc_delete_all
        context['user_acc_delete_all_len'] = len(user_acc_delete_all)
        context['user_acc_delete_dhub'] = user_acc_delete_dhub
        context['user_acc_delete_dhub_len'] = len(user_acc_delete_dhub)
        context['req_meta'] = req_meta
        context['req_meta_len'] = len(req_meta)
        context['req_word'] = req_word
        context['req_word_len'] = len(req_word)
        context['req_domain_len'] = len(req_domain)
        context['req_attr_len'] = len(req_attr)
        # context['metareq_len'] = len(metareq)
        context['metareq_dhub_len'] = len(metareq_dhub)
        context['metareq_all_len'] = len(metareq_all)
        context['monitor_yn'] = monitor_yn
        context['alert_yn'] = alert_yn

        context['bms_len'] = len(bms)
        context['tuning_len'] = len(tuning)
        context['dev_monitor'] = dev_monitor
        context['prod_monitor'] = prod_monitor
        context['prod_min_dtm'] = prod_min_dtm[0]
        context['dev_min_dtm'] = dev_min_dtm[0]

        context['airflow_error'] = airflow_error

        return context


class QuickLinkLV(LoginRequiredMixin, ListView):
    template_name = 'QuickLinkLV.html'
    context_object_name = 'objects'

    def get_queryset(self):

        obj = []

        return obj

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        sql = """
                    SELECT '데이터 모델 (ERD)' menu_name, 'ModelLV' url, 10000 cnt
                    FROM DUAL
                    UNION ALL
                    SELECT b.menu_name, b.url, count(*) cnt
                    FROM cust_menu_list b
                         left outer join cust_user_visit a     
                        on ( b.id = a.id_menulist         
                             and a.sabun = 22980
                             and a.reg_dtm > date_add(now(), interval -14 day ))
                    where b.use_yn = 1         
                      and b.id <> 1
                    GROUP BY b.menu_name, b.url
                    HAVING CNT >= 10
                    order by 3 desc;
              """.format(sabun=self.request.user.username)

        menu_list1 = _query_dict("default", sql)

        sql = """
                            SELECT b.menu_name, b.url, count(*) cnt
                            FROM cust_menu_list b
                                 left outer join cust_user_visit a     
                                on ( b.id = a.id_menulist         
                                     and a.sabun = '{sabun}'
                                     and a.reg_dtm > date_add(now(), interval -14 day ))
                            where b.use_yn = 1  
                            and b.id <> 1       
                            GROUP BY b.menu_name, b.url
                            HAVING CNT < 10
                            order by 1 ;
                      """.format(sabun=self.request.user.username)

        menu_list2 = _query_dict("default", sql)

        context['menu_list1'] = menu_list1
        context['menu_list2'] = menu_list2
        context['template_parent'] = "quick_link"
        context['template_child'] = "quick_link"

        return context


class executeSQLLV(LoginRequiredMixin, FormMixin, ListView):
    model = TuningList
    template_name = 'executeSQLLV.html'
    context_object_name = 'objects'
    form_class = ExecuteSQLLV

    def get_initial(self):
        return {
            'id_dblist': 39,
        }

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        _loggingVisit(self.request, 5)

        context = super().get_context_data(**kwargs)

        my_data = DataList.objects.filter(id_reg_user=self.request.user).order_by("-reg_dtm")

        context['my_data'] = my_data
        context['template_parent'] = "executeSQLLV"
        context['template_child'] = "executeSQLLV"

        return context


class UserListLV(LoginRequiredMixin, FormMixin, ListView):
    model = UserList
    template_name = 'UserListLV.html'
    context_object_name = 'objects'
    paginate_by = 100
    form_class = UserListForm

    def get_initial(self):

        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################
        username = ""

        username = self.request.session[
            'username'] if 'username' in self.request.session else self.request.user.first_name

        return {'username': username,
                }
        ###################################################

    def get_queryset(self):

        _loggingVisit(self.request, 16)

        self.form = self.get_form(self.form_class)

        username = ""
        id_dblist = ""
        oper_cd = ""

        if self.form.is_valid():
            username = self.form.cleaned_data['username']
            id_dblist = self.form.cleaned_data['id_dblist']
            oper_cd = self.form.cleaned_data['oper_cd']

            ###################################################
            ## 화면에 바인딩 된 값을 유지하기 위한 처리
            ###################################################
            self.request.session['username'] = username if username is not None else ""

            ###################################################
            ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            self.request.session['click_search'] = 'Y'
            ###################################################

            if self.form.cleaned_data["btn_account_active"] != "":
                UserListId = int(self.form.cleaned_data["btn_account_active"])


                obj, flag = UserRequestAccountOpenHist.objects.get_or_create(
                    id_userlist=UserList.objects.get(id=UserListId),
                    id_approver=None,
                    approv_dtm=None,
                    id_reg_user=self.request.user,
                )

                if flag:

                    userlist = UserList.objects.get(id=UserListId)
                    userlist.request_status = "1"
                    userlist.save()

                else:
                    pass

        else:
            page_kwarg = self.page_kwarg
            page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
            if page == 1:
                self.request.session['username'] = ""
            username = self.request.session['username'] if 'username' in self.request.session else ""

        # Q(title__icontains=title)|Q(id_tablelist__table_name__icontains=title)|Q(table_name__icontains=title)|Q(id_reg_user__first_name=title),

        obj = UserList.objects.filter(username__startswith='I_',
                                      id_user_name__isnull=False,
                                      drop_yn=0,
                                      *((Q(id_user_name__first_name__iexact=username) | Q(
                                          username__istartswith=username),) if username else ()),
                                      *((Q(id_dblist=id_dblist),) if id_dblist else ()),
                                      *((Q(oper_cd=oper_cd),) if oper_cd else ()),
                                      ).order_by("oper_cd", "id_dblist__id_dbtype", "db_use")

        return obj

    def post(self, request, *args, **kwargs):

        return self.get(request, *args, **kwargs)

    def paginate_queryset(self, queryset, page_size):

        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty())
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            if page == 'last':
                page_number = paginator.num_pages
            else:
                raise Http404(_("Page is not 'last', nor can it be converted to an int."))
        try:

            ###################################################
            ## SEARCH 버튼 클릭 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            if 'click_search' in self.request.session:
                if self.request.session['click_search'] == "N":
                    page = paginator.page(page_number)
                else:
                    page = paginator.page(1)
            else:
                page = paginator.page(page_number)
            ###################################################

            ###################################################
            ## COLUMNLIST 에서 목록으로 클릭 시 원래 페이지 번호를 찾아가기 위한 코드
            ###################################################
            self.request.session['page_number'] = page_number

            return (paginator, page, page.object_list, page.has_other_pages())



        except:
            page = paginator.page(1)
            return (paginator, page, page.object_list, page.has_other_pages())

    def get_context_data(self, **kwargs):
        header = []

        context = super().get_context_data(**kwargs)

        page_range = ""
        paginator = context['paginator']
        page_numbers_range = 5  # Display only 5 page numbers
        max_index = len(paginator.page_range)

        ###################################################
        ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
        ###################################################
        if 'click_search' in self.request.session:
            if self.request.session['click_search'] == "N":
                page = self.request.GET.get('page')
            else:
                page = 1
                self.request.session['click_search'] = "N"
        else:
            page = self.request.GET.get('page')
        ###################################################

        current_page = int(page) if page else 1

        start_index = int((current_page - 1) / page_numbers_range) * page_numbers_range
        end_index = start_index + page_numbers_range
        if end_index >= max_index:
            end_index = max_index

        page_range = paginator.page_range[start_index:end_index]
        context['page_range'] = page_range

        context['template_parent'] = "dev"
        context['template_child'] = "UserListLV"

        return context


class DataSearchDV(LoginRequiredMixin, DetailView):
    model = DataList
    template_name = 'DataSearchDV.html'
    context_object_name = 'objects'

    def get_object(self):
        obj = get_object_or_404(DataList, id=self.kwargs['id'])

        return obj

    def get_context_data(self, **kwargs):
        _loggingVisit(self.request, 1)

        exec_time = UserExecuteSQL.objects.filter(id_datalist=self.kwargs['id']).aggregate(Avg('execute_time'))

        context = super().get_context_data(**kwargs)

        context['template_parent'] = "data"
        context['template_child'] = "DataSearchLV"
        context['exec_time'] = exec_time

        return context


class DataSearchLV(LoginRequiredMixin, FormMixin, ListView):
    model = ColumnList
    template_name = 'DataSearchLV.html'
    context_object_name = 'objects'
    paginate_by = 10
    form_class = DataSearchForm
    keyword = ""

    def get_initial(self):

        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################
        data_title = ""
        db_type = ""
        my_data = ""

        if 'keyword' not in self.kwargs:
            data_title = self.request.session['data_title'] if 'data_title' in self.request.session else ""
        else:
            data_title = self.kwargs['keyword']
            data_title = data_title.replace("!_!", "/")
        db_type = self.request.session['db_type'] if 'db_type' in self.request.session else ""
        my_data = self.request.session['my_data'] if 'my_data' in self.request.session else ""

        return {'data_title': data_title,
                'db_type': db_type,
                'my_data': my_data,
                }
        ###################################################

    def get_queryset(self):

        _loggingVisit(self.request, 1)

        self.form = self.get_form(self.form_class)

        if self.form.is_valid():

            _loggingVisit(self.request, 26)

            data_title = self.form.cleaned_data['data_title']
            db_type = self.form.cleaned_data['db_type']
            my_data = self.form.cleaned_data['my_data']

            ###################################################
            ## 화면에 바인딩 된 값을 유지하기 위한 처리
            ###################################################
            self.request.session['data_title'] = data_title if data_title is not None else ""
            self.request.session['db_type'] = db_type if db_type is not None else ""
            self.request.session['my_data'] = my_data if my_data is not None else ""

            ###################################################
            ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            self.request.session['click_search'] = 'Y'
            ###################################################

        else:
            ###################################################
            ## 1. 처음 진입
            ## 2. 페이징 버튼 클릭 시
            ###################################################
            page_kwarg = self.page_kwarg
            page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
            if page == 1:
                self.request.session['data_title'] = ""
                self.request.session['db_type'] = ""
                self.request.session['my_data'] = ""

            if 'keyword' not in self.kwargs:
                data_title = self.request.session['data_title'] if 'data_title' in self.request.session else ""
            else:
                # 링크 클릭 시
                data_title = self.kwargs['keyword']
                data_title = data_title.replace("!_!", "/")

            db_type = self.request.session['db_type'] if 'db_type' in self.request.session else ""
            my_data = self.request.session['my_data'] if 'my_data' in self.request.session else ""

        if my_data == "1":
            my_data = self.request.user.id

        obj = DataList.objects.filter(
            *((Q(data_title__icontains=data_title.upper()),) if data_title else ()),
            *((Q(id_dblist=db_type),) if db_type else ()),
            *(
                (
                    (
                            Q(id_req_users__id=my_data) |
                            Q(id_mod_psbl_users__id=my_data) |
                            Q(id_reg_user=my_data) |
                            Q(id_mod_user=my_data)
                    ),
                ) if my_data else ()
            ),
            # exp_yn = 1
        ).order_by('data_title', 'id_domain').distinct()

        return obj

    def paginate_queryset(self, queryset, page_size):

        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty())
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            if page == 'last':
                page_number = paginator.num_pages
            else:
                raise Http404(_("Page is not 'last', nor can it be converted to an int."))
        try:

            ###################################################
            ## SEARCH 버튼 클릭 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            if 'click_search' in self.request.session:
                if self.request.session['click_search'] == "N":
                    page = paginator.page(page_number)
                else:
                    page = paginator.page(1)
            else:
                page = paginator.page(page_number)
            ###################################################

            ###################################################
            ## COLUMNLIST 에서 목록으로 클릭 시 원래 페이지 번호를 찾아가기 위한 코드
            ###################################################
            self.request.session['page_number'] = page_number

            return (paginator, page, page.object_list, page.has_other_pages())



        except:
            page = paginator.page(1)
            return (paginator, page, page.object_list, page.has_other_pages())

    def post(self, request, *args, **kwargs):

        # self.object = None
        # keyword = None
        # self.form = self.get_form(self.form_class)
        # if self.form.is_valid():
        #   # self.object = self.form.save()
        #   keyword = self.form.cleaned_data['keyword']

        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        page_range = ""
        paginator = context['paginator']
        page_numbers_range = 5  # Display only 5 page numbers
        max_index = len(paginator.page_range)

        ###################################################
        ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
        ###################################################
        if 'click_search' in self.request.session:
            if self.request.session['click_search'] == "N":
                page = self.request.GET.get('page')
            else:
                page = 1
                self.request.session['click_search'] = "N"
        else:
            page = self.request.GET.get('page')
        ###################################################

        current_page = int(page) if page else 1

        start_index = int((current_page - 1) / page_numbers_range) * page_numbers_range
        end_index = start_index + page_numbers_range
        if end_index >= max_index:
            end_index = max_index

        page_range = paginator.page_range[start_index:end_index]
        context['page_range'] = page_range

        header_data = [
            'DB용도',
            '도메인',
            '데이터 한글명',
            '설명',
            '개인정보',
            '등록자',
            '등록시간'
        ]

        my_privs = UserRequestTabPrivHist.objects.filter(id_reg_user_id=self.request.user.id,
                                                         approv_yn="0",
                                                         use_yn='1')

        my_privs_list = []
        for priv in my_privs:
            my_privs_list.append(priv.id_objectlist.id)

        sql = """
        SELECT CAST(@ROWNUM := @ROWNUM + 1 AS unsigned) ROWNUM , A.KEYWORD EXP_NAME, replace(A.KEYWORD,'/','!_!') DATA_TITLE, A.CNT
        FROM
        (
          SELECT upper(KEYWORD) KEYWORD, count(*) CNT
          FROM CUST_USER_SEARCH_KEYWORD
          GROUP BY upper(KEYWORD)
          ORDER BY CNT DESC
        ) A, ( SELECT @ROWNUM := 0 ) TMP
        LIMIT 10;
      """

        dataset_user_search_keyword = _query_mysql_dict(sql)

        sql = """
        SELECT CAST(@ROWNUM := @ROWNUM + 1 AS unsigned) ROWNUM, A.EXP_NAME, replace(A.DATA_TITLE,'/','!_!') DATA_TITLE, REG_DTM, USERNAME
        FROM
        (
        -- SELECT concat(DATE_FORMAT(A.MOD_DTM,'%Y/%m/%d'),' - ',A.DATA_TITLE) EXP_NAME, DATA_TITLE
        SELECT A.DATA_TITLE EXP_NAME, DATA_TITLE, A.REG_DTM, B.FIRST_NAME USERNAME
        FROM CUST_DATA_LIST A,
             AUTH_USER B
        WHERE B.ID = A.ID_REG_USER
          AND EXP_YN = 1
        ORDER BY REG_DTM DESC
        ) A, ( SELECT @ROWNUM := 0 ) TMP
        LIMIT 10;
      """

        dataset_latest_datalist = _query_mysql_dict(sql)

        sql = """
        SELECT CAST(@ROWNUM := @ROWNUM + 1 AS unsigned) ROWNUM, A.DATA_TITLE EXP_NAME, replace(A.DATA_TITLE,'/','!_!') DATA_TITLE, REG_DTM
        FROM
        (
          SELECT C.DATA_TITLE, REG_DTM
          FROM AUTH_USER A,
             cust_datalist_user_id_req_users B,
             CUST_DATA_LIST C
          WHERE A.USERNAME = '{username}'
          AND   A.ID = B.USER_ID
          AND   C.ID = B.DATALIST_ID
          AND   C.EXP_YN = 1
          UNION ALL
          SELECT C.DATA_TITLE, C.REG_DTM
          FROM  AUTH_USER A,
              CUST_DATA_LIST C
          WHERE  A.USERNAME = '{username}'
          AND  ( C.ID_REG_USER = A.ID OR C.ID_MOD_USER = A.ID)
          AND   C.EXP_YN = 1
          ORDER BY REG_DTM DESC
        )A, ( SELECT @ROWNUM := 0 ) TMP;
      """.format(username=self.request.user.username)

        dataset_my_datalist = _query_mysql_dict(sql)

        word_meaning = None

        if self.request.POST.get('data_title') != "":

            sql = """
          SELECT TOP 1 *, 1 EXP_ORDER
          FROM IKEP4_CV_WORD
          WHERE WORD_NAME = '{keyword}'
          UNION ALL
          SELECT TOP 1 *, 1
          FROM IKEP4_CV_WORD
          WHERE WORD_NAME_READ = '{keyword}'
          UNION ALL
          SELECT TOP 1 *, 2
          FROM IKEP4_CV_WORD
          WHERE WORD_NAME_READ like '%{keyword}%'
          UNION ALL
          SELECT TOP 1 *, 3
          FROM IKEP4_CV_WORD
          WHERE WORD_ENGLISH_NAME like '%{keyword}%'
          UNION ALL
          SELECT TOP 1 *, 4
          FROM IKEP4_CV_WORD
          WHERE WORD_NAME like '%{keyword}%'
          ORDER BY exp_order
          ;
        """.format(keyword=self.request.POST.get('data_title'))

            word_meaning = _query_mssql_dict(sql)

            if len(word_meaning) > 0:
                word_meaning = word_meaning[0]

        context['template_parent'] = "data"
        context['template_child'] = "DataSearchLV"
        context['header_data'] = header_data
        context['my_privs_list'] = my_privs_list
        context['dataset_user_search_keyword'] = dataset_user_search_keyword
        context['dataset_latest_datalist'] = dataset_latest_datalist
        context['dataset_my_datalist'] = dataset_my_datalist
        context['word_meaning'] = word_meaning

        return context


class ObjectListLV(LoginRequiredMixin, FormMixin, ListView):
    model = ObjectList
    template_name = 'ObjectListLV.html'
    context_object_name = 'objects'
    paginate_by = 10
    form_class = ObjectListForm
    keyword = ""

    def get_initial(self):

        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################
        object_name = ""
        owner = ""
        id_dblist = ""
        oper_cd = ""
        choice_object_type = ""

        object_name = self.request.session['object_name'] if 'object_name' in self.request.session else ""
        owner = self.request.session['owner'] if 'owner' in self.request.session else ""
        id_dblist = self.request.session['id_dblist'] if 'id_dblist' in self.request.session else ""
        oper_cd = self.request.session['oper_cd'] if 'oper_cd' in self.request.session else ""
        choice_object_type = self.request.session[
            'choice_object_type'] if 'choice_object_type' in self.request.session else ""

        return {'object_name': object_name,
                'owner': owner,
                'id_dblist': id_dblist,
                'oper_cd': oper_cd,
                'choice_object_type': choice_object_type,
                }
        ###################################################

    def get_queryset(self):

        _loggingVisit(self.request, 17)

        obj = []

        self.form = self.get_form(self.form_class)

        if self.form.is_valid():

            object_name = self.form.cleaned_data['object_name']
            owner = self.form.cleaned_data['owner']
            if owner:
                owner = owner.id
            id_dblist = self.form.cleaned_data['id_dblist']
            oper_cd = self.form.cleaned_data['oper_cd']
            choice_object_type = self.form.cleaned_data['choice_object_type']

            ###################################################
            ## 화면에 바인딩 된 값을 유지하기 위한 처리
            ###################################################
            self.request.session['object_name'] = object_name if object_name is not None else ""
            self.request.session['owner'] = owner if owner is not None else ""
            self.request.session['id_dblist'] = id_dblist.id if id_dblist is not None else ""
            self.request.session['oper_cd'] = oper_cd.id if oper_cd is not None else ""
            self.request.session['choice_object_type'] = choice_object_type if choice_object_type is not None else ""

            ###################################################
            ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            self.request.session['click_search'] = 'Y'
            ###################################################

        else:
            ###################################################
            ## 1. 처음 진입
            ## 2. 페이징 버튼 클릭 시
            ###################################################

            page_kwarg = self.page_kwarg
            page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
            if page == 1:
                self.request.session['object_name'] = ""
                self.request.session['owner'] = ""
                self.request.session['id_dblist'] = ""
                self.request.session['oper_cd'] = ""
                self.request.session['choice_object_type'] = ""

            object_name = self.request.session['object_name'] if 'object_name' in self.request.session else ""
            owner = self.request.session['owner'] if 'owner' in self.request.session else ""
            id_dblist = self.request.session['id_dblist'] if 'id_dblist' in self.request.session else ""
            oper_cd = self.request.session['oper_cd'] if 'oper_cd' in self.request.session else ""
            choice_object_type = self.request.session[
                'choice_object_type'] if 'choice_object_type' in self.request.session else ""

        dbowner = DBOwner.objects.all()
        list_owner = []

        for o in dbowner:
            list_owner.append(o.owner)

        obj = ObjectList.objects.filter(
            *((Q(object_name__startswith=object_name.upper()),) if object_name else ()),
            *((Q(id_dblist=id_dblist),) if id_dblist else ()),
            *((Q(owner=DBOwner.objects.filter(id=owner).values('owner').first()['owner']),) if owner else ()),
            *((Q(oper_cd=oper_cd),) if oper_cd else ()),
            *((Q(object_type=choice_object_type),) if choice_object_type else ()),
            Q(owner__in=list_owner),
            drop_yn=0,
            object_type__in=['DATABASE LINK', 'FUNCTION', 'INDEX', 'JOB', 'MATERIALIZED VIEW', 'PACKAGE',
                             'PACKAGE BODY', 'PROCEDURE', 'SEQUENCE', 'TRIGGER', 'TYPE', 'TYPE BODY', 'VIEW']
        ).order_by('db_use', 'oper_cd', 'owner', 'object_type', 'object_name')

        return obj

    def paginate_queryset(self, queryset, page_size):

        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty())
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            if page == 'last':
                page_number = paginator.num_pages
            else:
                raise Http404(_("Page is not 'last', nor can it be converted to an int."))
        try:

            ###################################################
            ## SEARCH 버튼 클릭 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            if 'click_search' in self.request.session:
                if self.request.session['click_search'] == "N":
                    page = paginator.page(page_number)
                else:
                    page = paginator.page(1)
            else:
                page = paginator.page(page_number)
            ###################################################

            ###################################################
            ## COLUMNLIST 에서 목록으로 클릭 시 원래 페이지 번호를 찾아가기 위한 코드
            ###################################################
            self.request.session['page_number'] = page_number

            return (paginator, page, page.object_list, page.has_other_pages())



        except:
            page = paginator.page(1)
            return (paginator, page, page.object_list, page.has_other_pages())

    def post(self, request, *args, **kwargs):

        # self.object = None
        # keyword = None
        # self.form = self.get_form(self.form_class)
        # if self.form.is_valid():
        #   # self.object = self.form.save()
        #   keyword = self.form.cleaned_data['keyword']

        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        header = []

        context = super().get_context_data(**kwargs)

        page_range = ""
        paginator = context['paginator']
        page_numbers_range = 5  # Display only 5 page numbers
        max_index = len(paginator.page_range)

        ###################################################
        ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
        ###################################################
        if 'click_search' in self.request.session:
            if self.request.session['click_search'] == "N":
                page = self.request.GET.get('page')
            else:
                page = 1
                self.request.session['click_search'] = "N"
        else:
            page = self.request.GET.get('page')
        ###################################################

        current_page = int(page) if page else 1

        start_index = int((current_page - 1) / page_numbers_range) * page_numbers_range
        end_index = start_index + page_numbers_range
        if end_index >= max_index:
            end_index = max_index

        page_range = paginator.page_range[start_index:end_index]
        context['page_range'] = page_range

        choice_tab_or_col = self.request.session[
            'choice_tab_or_col'] if 'choice_tab_or_col' in self.request.session else "1"

        header = [
            'DB용도',
            '개발/운영',
            'OWNER',
            '오브젝트명',
            '오브젝트 타입',
            '상태',
            '생성일자',
            '변경일자',
            '권한',
            '권한 신청',
        ]

        my_privs = UserRequestTabPrivHist.objects.filter(id_reg_user_id=self.request.user.id,
                                                         approv_yn="0",
                                                         use_yn='1')

        my_privs_list = []
        for priv in my_privs:
            my_privs_list.append(priv.id_objectlist.id)

        context['template_parent'] = "dev"
        context['template_child'] = "ObjectListLV"
        context['header'] = header
        context['my_privs_list'] = my_privs_list

        return context


class ObjectListDV(LoginRequiredMixin, DetailView):
    model = ObjectList
    template_name = 'ObjectListDV.html'
    context_object_name = 'objects'

    def get_object(self):
        obj = get_object_or_404(ObjectList, id=self.kwargs['id'])

        return obj

    def get_context_data(self, **kwargs):


        obj_info = ObjectList.objects.get(id=self.kwargs['id'])

        sql = """
            select CASE WHEN LINE = 1 THEN REPLACE(REPLACE(UPPER(TEXT),NAME,CONCAT(OWNER,'.',NAME)),'"','')
                   ELSE TEXT END test,
                   reg_dtm
            from stg_dba_source
            where id_dblist = {id_dblist}
            and oper_cd = {oper_cd}
            and owner = '{owner}'
            and name = '{name}'
            and type = '{type}'
            order by line
      """.format(
            id_dblist=obj_info.id_dblist.id,
            oper_cd=obj_info.oper_cd.id,
            owner=obj_info.owner,
            name=obj_info.object_name,
            type=obj_info.object_type
        )

        obj_source = _query_list("default", sql)

        context = super().get_context_data(**kwargs)

        context['template_parent'] = "dev"
        context['template_child'] = "ObjectListLV"
        context['obj_source'] = obj_source

        return context


class GrantListLV(LoginRequiredMixin, FormMixin, ListView):
    model = GrantList
    template_name = 'GrantListLV.html'
    context_object_name = 'objects'
    paginate_by = 10
    form_class = GrantListForm
    keyword = ""

    def get_initial(self):

        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################
        table_name = ""
        grantee = ""
        id_dblist = ""
        oper_cd = ""

        table_name = self.request.session['table_name'] if 'table_name' in self.request.session else "ORD_ORD_M"
        grantee = self.request.session['grantee'] if 'grantee' in self.request.session else ""
        id_dblist = self.request.session['id_dblist'] if 'id_dblist' in self.request.session else "10"
        oper_cd = self.request.session['oper_cd'] if 'oper_cd' in self.request.session else "3"

        return {'table_name': table_name,
                'grantee': grantee,
                'id_dblist': id_dblist,
                'oper_cd': oper_cd,
                }
        ###################################################

    def get_queryset(self):

        _loggingVisit(self.request, 18)

        self.form = self.get_form(self.form_class)

        # bb = self.request.POST.get('btn_batch','')

        # if bb == "click" :
        #   user = "gsdba"
        #   pw = "inomuski"
        #   ip = "  165.243.205.158"
        #   port = "1521"
        #   tns = "METADB"

        #   conn = cx_Oracle.connect(user, pw, ip+":"+port+"/"+tns)
        #   cur = conn.cursor()
        #   cur.execute("""
        #   SELECT
        #          ID,
        #          TITLE,
        #          SQL_ID,
        #          CASE WHEN NEW_SQL IS NULL THEN 1
        #               WHEN NEW_SQL = 1 THEN 1
        #               WHEN NEW_SQL = 2 THEN 2
        #               WHEN NEW_SQL = 0 THEN 1
        #          END SQL_TYPE,
        #          CASE WHEN CATEGORY_DETAIL = 'MCPC_JBP' THEN 18
        #               WHEN CATEGORY_DETAIL = 'MCPC_고객/제휴' THEN 1
        #               WHEN CATEGORY_DETAIL = 'MCPC_채널주문' THEN 3
        #               WHEN CATEGORY_DETAIL = 'MCPC_제휴' THEN 7
        #               WHEN CATEGORY_DETAIL = 'MCPC_여행몰' THEN 21
        #               WHEN CATEGORY_DETAIL = 'MCPC_캠페인' THEN 29
        #               WHEN CATEGORY_DETAIL = '기간계_업무지원' THEN 20
        #               WHEN CATEGORY_DETAIL = '기간계_상품' THEN 2
        #               WHEN CATEGORY_DETAIL = '기간계_방송' THEN 15
        #               WHEN CATEGORY_DETAIL = '기간계_주문/결제' THEN 3
        #               WHEN CATEGORY_DETAIL = '기간계_물류/배송' THEN 14
        #               WHEN CATEGORY_DETAIL = '기간계_회계' THEN 31
        #               WHEN CATEGORY_DETAIL = '기간계_분석' THEN 17
        #               WHEN CATEGORY_DETAIL = '기간계_그룹웨어' THEN 20
        #               WHEN CATEGORY_DETAIL = 'MCPC_보험몰' THEN 16
        #               WHEN CATEGORY_DETAIL = 'MCPC_단품' THEN 2
        #               WHEN CATEGORY_DETAIL = 'MCPC_매장' THEN 13
        #               WHEN CATEGORY_DETAIL = '기간계_고객/SR' THEN 1
        #               WHEN CATEGORY_DETAIL = 'MCPC_공통' THEN 10
        #               WHEN CATEGORY_DETAIL = 'MCPC_이벤트' THEN 25
        #               WHEN CATEGORY_DETAIL = '통합보험' THEN 16
        #               WHEN CATEGORY_DETAIL = '기간계_공통' THEN 10
        #               WHEN CATEGORY_DETAIL = '연동형TC' THEN 15
        #               WHEN CATEGORY_DETAIL = '기간계_데이터홈쇼핑' THEN 15
        #               WHEN CATEGORY_DETAIL = '종료된서비스' THEN 38
        #               WHEN CATEGORY_DETAIL = 'MCPC_업무지원' THEN 20
        #               WHEN CATEGORY_DETAIL = '기간계_위드넷/상품' THEN 47
        #               WHEN CATEGORY_DETAIL = 'MCPC_위드넷' THEN 47
        #               WHEN CATEGORY_DETAIL = '기간계_위드넷/물류' THEN 47
        #          ELSE 38 END ID_DOMAIN,
        #          CASE  WHEN V_GS_DB_LIST_D_ID = 1 THEN 10
        #                WHEN V_GS_DB_LIST_D_ID = 2 THEN 3
        #                WHEN V_GS_DB_LIST_D_ID = 20 THEN 40
        #                WHEN V_GS_DB_LIST_D_ID = 18 THEN 39
        #                WHEN V_GS_DB_LIST_D_ID = 3 THEN 19
        #                WHEN V_GS_DB_LIST_D_ID = 4 THEN 25
        #                WHEN V_GS_DB_LIST_D_ID = 5 THEN 24
        #                WHEN V_GS_DB_LIST_D_ID = 6 THEN 13
        #                WHEN V_GS_DB_LIST_D_ID = 8 THEN 33
        #                WHEN V_GS_DB_LIST_D_ID = 9 THEN 7
        #                WHEN V_GS_DB_LIST_D_ID = 10 THEN 9
        #                WHEN V_GS_DB_LIST_D_ID = 11 THEN 31
        #                WHEN V_GS_DB_LIST_D_ID = 12 THEN 36
        #                WHEN V_GS_DB_LIST_D_ID = 13 THEN 27
        #                WHEN V_GS_DB_LIST_D_ID = 14 THEN 4
        #                WHEN V_GS_DB_LIST_D_ID = 15 THEN 14
        #                WHEN V_GS_DB_LIST_D_ID = 16 THEN 23
        #                WHEN V_GS_DB_LIST_D_ID = 17 THEN 21
        #                WHEN V_GS_DB_LIST_D_ID = 19 THEN 29
        #                WHEN V_GS_DB_LIST_D_ID = 21 THEN 26
        #                WHEN V_GS_DB_LIST_D_ID = 22 THEN 35
        #                WHEN V_GS_DB_LIST_D_ID = 23 THEN 22
        #                WHEN V_GS_DB_LIST_D_ID = 24 THEN 16
        #                WHEN V_GS_DB_LIST_D_ID = 25 THEN 17
        #                WHEN V_GS_DB_LIST_D_ID = 26 THEN 2
        #                WHEN V_GS_DB_LIST_D_ID = 27 THEN 1
        #                WHEN V_GS_DB_LIST_D_ID = 28 THEN 5
        #                WHEN V_GS_DB_LIST_D_ID = 29 THEN 8
        #                WHEN V_GS_DB_LIST_D_ID = 30 THEN 20
        #                WHEN V_GS_DB_LIST_D_ID = 31 THEN 34
        #                WHEN V_GS_DB_LIST_D_ID = 32 THEN 30
        #                WHEN V_GS_DB_LIST_D_ID = 34 THEN 32
        #                WHEN V_GS_DB_LIST_D_ID = 35 THEN 12
        #                WHEN V_GS_DB_LIST_D_ID = 36 THEN 45
        #          END id_dblist,
        #          '' id_datalist,
        #          CASE WHEN STATUS_ID = 7 THEN 5
        #               WHEN STATUS_ID = 1 THEN 1
        #               WHEN STATUS_ID = 2 THEN 1
        #               WHEN STATUS_ID = 3 THEN 2
        #               WHEN STATUS_ID = 4 THEN 3
        #               WHEN STATUS_ID = 5 THEN 4
        #               WHEN STATUS_ID = 6 THEN 4
        #               WHEN STATUS_ID = 27 THEN 5
        #           END id_tuningstatus,
        #           REQUEST_COMMENTS sql_info,
        #           TUNNER_COMMENTS tuning_info,
        #           '' asis_plan,
        #           '' tobe_plan,
        #           REQUEST_SQL asis_sql_text,
        #           ASIS_ELAPSED asis_elapsed_time,
        #           '' expect_elapsed_time,
        #           '' avg_rows,
        #           '' daily_exec_cnt,
        #           '' bind_value,
        #           TUNNER_SQL tobe_sql_text,
        #           TOBE_ELAPSED tobe_elapsed_time,
        #           NVL(( SELECT B.ID FROM DBMON.MIG_SABUN_MAPPING B WHERE B.SABUN = A.EMP_NO ),18243) id_reg_user,
        #           NVL(case when TUNNER_USERNAME = '이치호' then 2
        #             when TUNNER_USERNAME = '구경서' then 4
        #             when TUNNER_USERNAME = '정관호' then 21
        #             when TUNNER_USERNAME = '이선경' then 7569
        #             when TUNNER_USERNAME = '이정아' then 12963
        #           end,18243) id_tuning_user,
        #           to_date(TUNNER_DT, 'YYYY-MM-DD HH24:MI:SS')  tuning_dtm,
        #           DISTRIBUTE_DT dist_dtm,
        #           to_date(CREATED, 'YYYY-MM-DD HH24:MI:SS') reg_dtm
        #   FROM DBMON.DJANGO_TUNNE_BBS A
        #   -- where id = 2480
        #     """)

        #   columnNames = [d[0] for d in cur.description]
        #   datasets = [dict(zip(columnNames, row)) for row in cur]

        #   for dataset in datasets :

        #     tuninglist = TuningList (
        #                   title = dataset['TITLE'],
        #                   sql_id = dataset['SQL_ID'],
        #                   sql_type = dataset['SQL_TYPE'],
        #                   id_domain = Domain.objects.get(id=dataset['ID_DOMAIN']),
        #                   id_dblist = DbList.objects.get(id=dataset['ID_DBLIST']),
        #                   id_datalist = None,
        #                   id_tuningstatus = TuningStatus.objects.get(id=dataset['ID_TUNINGSTATUS']),
        #                   sql_info = dataset['SQL_INFO'],
        #                   tuning_info = dataset['TUNING_INFO'],
        #                   asis_plan = dataset['ASIS_PLAN'],
        #                   tobe_plan = dataset['TOBE_PLAN'],
        #                   asis_sql_text = dataset['ASIS_SQL_TEXT'],
        #                   asis_elapsed_time = dataset['ASIS_ELAPSED_TIME'],
        #                   expect_elapsed_time = dataset['EXPECT_ELAPSED_TIME'],
        #                   avg_rows = dataset['AVG_ROWS'],
        #                   daily_exec_cnt = dataset['DAILY_EXEC_CNT'],
        #                   bind_value = dataset['BIND_VALUE'],
        #                   tobe_sql_text = dataset['TOBE_SQL_TEXT'],
        #                   tobe_elapsed_time = dataset['TOBE_ELAPSED_TIME'],
        #                   id_reg_user = User.objects.get(id=dataset['ID_REG_USER']),
        #                   id_tuning_user = User.objects.get(id=dataset['ID_TUNING_USER']),
        #                   tuning_dtm = dataset['TUNING_DTM'],
        #                   dist_dtm = dataset['DIST_DTM'],
        #                   reg_dtm = dataset['REG_DTM'],
        #                 )
        #     tuninglist.save()

        if self.form.is_valid():

            table_name = self.form.cleaned_data['table_name']
            grantee = self.form.cleaned_data['grantee']
            id_dblist = self.form.cleaned_data['id_dblist']
            oper_cd = self.form.cleaned_data['oper_cd']

            ###################################################
            ## 화면에 바인딩 된 값을 유지하기 위한 처리
            ###################################################
            self.request.session['table_name'] = table_name if table_name is not None else ""
            self.request.session['grantee'] = grantee if grantee is not None else ""
            self.request.session['id_dblist'] = id_dblist.id if id_dblist is not None else ""
            self.request.session['oper_cd'] = oper_cd.id if oper_cd is not None else ""

            ###################################################
            ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            self.request.session['click_search'] = 'Y'
            ###################################################

        else:
            ###################################################
            ## 1. 처음 진입
            ## 2. 페이징 버튼 클릭 시
            ###################################################
            page_kwarg = self.page_kwarg
            page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
            if page == 1:
                self.request.session['table_name'] = "ORD_ORD_M"
                self.request.session['grantee'] = ""
                self.request.session['id_dblist'] = "10"
                self.request.session['oper_cd'] = ""

            table_name = self.request.session['table_name'] if 'table_name' in self.request.session else "ORD_ORD_M"
            grantee = self.request.session['grantee'] if 'grantee' in self.request.session else ""
            id_dblist = self.request.session['id_dblist'] if 'id_dblist' in self.request.session else "10"
            oper_cd = self.request.session['oper_cd'] if 'oper_cd' in self.request.session else "3"

        obj = GrantList.objects.filter(
            *((Q(table_name__istartswith=table_name.upper()),) if table_name else ()),
            *((Q(grantee__istartswith=grantee.upper()),) if grantee else ()),
            *((Q(id_dblist=id_dblist),) if id_dblist else ()),
            *((Q(oper_cd=oper_cd),) if oper_cd else ()),
        ).order_by('id_dblist', 'oper_cd', 'grantee', 'owner', 'table_name')

        return obj

    def paginate_queryset(self, queryset, page_size):

        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty())
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            if page == 'last':
                page_number = paginator.num_pages
            else:
                raise Http404(_("Page is not 'last', nor can it be converted to an int."))
        try:

            ###################################################
            ## SEARCH 버튼 클릭 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            if 'click_search' in self.request.session:
                if self.request.session['click_search'] == "N":
                    page = paginator.page(page_number)
                else:
                    page = paginator.page(1)
            else:
                page = paginator.page(page_number)
            ###################################################

            ###################################################
            ## COLUMNLIST 에서 목록으로 클릭 시 원래 페이지 번호를 찾아가기 위한 코드
            ###################################################
            self.request.session['page_number'] = page_number

            return (paginator, page, page.object_list, page.has_other_pages())



        except:
            page = paginator.page(1)
            return (paginator, page, page.object_list, page.has_other_pages())

    def post(self, request, *args, **kwargs):

        # self.object = None
        # keyword = None
        # self.form = self.get_form(self.form_class)
        # if self.form.is_valid():
        #   # self.object = self.form.save()
        #   keyword = self.form.cleaned_data['keyword']

        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        header = []

        context = super().get_context_data(**kwargs)

        page_range = ""
        paginator = context['paginator']
        page_numbers_range = 5  # Display only 5 page numbers
        max_index = len(paginator.page_range)

        ###################################################
        ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
        ###################################################
        if 'click_search' in self.request.session:
            if self.request.session['click_search'] == "N":
                page = self.request.GET.get('page')
            else:
                page = 1
                self.request.session['click_search'] = "N"
        else:
            page = self.request.GET.get('page')
        ###################################################

        current_page = int(page) if page else 1

        start_index = int((current_page - 1) / page_numbers_range) * page_numbers_range
        end_index = start_index + page_numbers_range
        if end_index >= max_index:
            end_index = max_index

        page_range = paginator.page_range[start_index:end_index]
        context['page_range'] = page_range

        header = [
            'DB용도',
            '개발/운영',
            'USERNAME',
            'OWNER',
            '테이블명',
            '권한',
        ]

        context['template_parent'] = "dev"
        context['template_child'] = "GrantListLV"
        context['header'] = header

        return context


class DBABatchLV(LoginRequiredMixin, ListView):
    model = GrantList
    template_name = 'DBABatchLV.html'
    context_object_name = 'objects'

    def get_queryset(self):

        bb = self.request.POST.get('btn_batch', '')

        if bb == "click":
            user = "gsdba"
            pw = "inomuski"
            ip = "  165.243.205.158"
            port = "1521"
            tns = "METADB"

            conn = cx_Oracle.connect(user, pw, ip + ":" + port + "/" + tns)
            cur = conn.cursor()
            cur.execute("""
      SELECT
             ID,
             TITLE,
             SQL_ID,
             CASE WHEN NEW_SQL IS NULL THEN 1
                  WHEN NEW_SQL = 1 THEN 1
                  WHEN NEW_SQL = 2 THEN 2
                  WHEN NEW_SQL = 0 THEN 1
             END SQL_TYPE,
             CASE WHEN CATEGORY_DETAIL = 'MCPC_JBP' THEN 18
                  WHEN CATEGORY_DETAIL = 'MCPC_고객/제휴' THEN 1
                  WHEN CATEGORY_DETAIL = 'MCPC_채널주문' THEN 3
                  WHEN CATEGORY_DETAIL = 'MCPC_제휴' THEN 7
                  WHEN CATEGORY_DETAIL = 'MCPC_여행몰' THEN 21
                  WHEN CATEGORY_DETAIL = 'MCPC_캠페인' THEN 29
                  WHEN CATEGORY_DETAIL = '기간계_업무지원' THEN 20
                  WHEN CATEGORY_DETAIL = '기간계_상품' THEN 2
                  WHEN CATEGORY_DETAIL = '기간계_방송' THEN 15
                  WHEN CATEGORY_DETAIL = '기간계_주문/결제' THEN 3
                  WHEN CATEGORY_DETAIL = '기간계_물류/배송' THEN 14
                  WHEN CATEGORY_DETAIL = '기간계_회계' THEN 31
                  WHEN CATEGORY_DETAIL = '기간계_분석' THEN 17
                  WHEN CATEGORY_DETAIL = '기간계_그룹웨어' THEN 20
                  WHEN CATEGORY_DETAIL = 'MCPC_보험몰' THEN 16
                  WHEN CATEGORY_DETAIL = 'MCPC_단품' THEN 2
                  WHEN CATEGORY_DETAIL = 'MCPC_매장' THEN 13
                  WHEN CATEGORY_DETAIL = '기간계_고객/SR' THEN 1
                  WHEN CATEGORY_DETAIL = 'MCPC_공통' THEN 10
                  WHEN CATEGORY_DETAIL = 'MCPC_이벤트' THEN 25
                  WHEN CATEGORY_DETAIL = '통합보험' THEN 16
                  WHEN CATEGORY_DETAIL = '기간계_공통' THEN 10
                  WHEN CATEGORY_DETAIL = '연동형TC' THEN 15
                  WHEN CATEGORY_DETAIL = '기간계_데이터홈쇼핑' THEN 15
                  WHEN CATEGORY_DETAIL = '종료된서비스' THEN 38
                  WHEN CATEGORY_DETAIL = 'MCPC_업무지원' THEN 20
                  WHEN CATEGORY_DETAIL = '기간계_위드넷/상품' THEN 47
                  WHEN CATEGORY_DETAIL = 'MCPC_위드넷' THEN 47
                  WHEN CATEGORY_DETAIL = '기간계_위드넷/물류' THEN 47
             ELSE null END ID_DOMAIN,
             case when category_detail = '인터파크&쿠팡 입점' then 13
             when category_detail = '구찌코리아 연동'  then 4
             when category_detail = '방송심의업무개선'  then 6
             when category_detail = '신물류센터(군포) 오픈 TF' then 9
             when category_detail = '유통법개정2차' then 12
             when category_detail = 'ebay제휴입점'  then 3
             when category_detail = '신물류시스템 마이크로서비스 전환' then 10
             when category_detail = '신물류 TF 기간계 I/F API 전환' then 8
             when category_detail = 'GS슈퍼'  then 1
             when category_detail = '여행딜' then 11
             when category_detail = '카카오페이_토스 결제수단 추가'  then 16
             when category_detail = 'PG사 이중화' then 2
             when category_detail = '기간계구조개선' then 5
             when category_detail = '51L' then 18
             when category_detail = '51T' then 18
             when category_detail = '51D' then 18
             else null end ID_PROJECTLIST,
             CASE  WHEN V_GS_DB_LIST_D_ID = 1 THEN 10
                   WHEN V_GS_DB_LIST_D_ID = 2 THEN 3
                   WHEN V_GS_DB_LIST_D_ID = 20 THEN 40
                   WHEN V_GS_DB_LIST_D_ID = 18 THEN 39
                   WHEN V_GS_DB_LIST_D_ID = 3 THEN 19
                   WHEN V_GS_DB_LIST_D_ID = 4 THEN 25
                   WHEN V_GS_DB_LIST_D_ID = 5 THEN 24
                   WHEN V_GS_DB_LIST_D_ID = 6 THEN 13
                   WHEN V_GS_DB_LIST_D_ID = 8 THEN 33
                   WHEN V_GS_DB_LIST_D_ID = 9 THEN 7
                   WHEN V_GS_DB_LIST_D_ID = 10 THEN 9
                   WHEN V_GS_DB_LIST_D_ID = 11 THEN 31
                   WHEN V_GS_DB_LIST_D_ID = 12 THEN 36
                   WHEN V_GS_DB_LIST_D_ID = 13 THEN 27
                   WHEN V_GS_DB_LIST_D_ID = 14 THEN 4
                   WHEN V_GS_DB_LIST_D_ID = 15 THEN 14
                   WHEN V_GS_DB_LIST_D_ID = 16 THEN 23
                   WHEN V_GS_DB_LIST_D_ID = 17 THEN 21
                   WHEN V_GS_DB_LIST_D_ID = 19 THEN 29
                   WHEN V_GS_DB_LIST_D_ID = 21 THEN 26
                   WHEN V_GS_DB_LIST_D_ID = 22 THEN 35
                   WHEN V_GS_DB_LIST_D_ID = 23 THEN 22
                   WHEN V_GS_DB_LIST_D_ID = 24 THEN 16
                   WHEN V_GS_DB_LIST_D_ID = 25 THEN 17
                   WHEN V_GS_DB_LIST_D_ID = 26 THEN 2
                   WHEN V_GS_DB_LIST_D_ID = 27 THEN 1
                   WHEN V_GS_DB_LIST_D_ID = 28 THEN 5
                   WHEN V_GS_DB_LIST_D_ID = 29 THEN 8
                   WHEN V_GS_DB_LIST_D_ID = 30 THEN 20
                   WHEN V_GS_DB_LIST_D_ID = 31 THEN 34
                   WHEN V_GS_DB_LIST_D_ID = 32 THEN 30
                   WHEN V_GS_DB_LIST_D_ID = 34 THEN 32
                   WHEN V_GS_DB_LIST_D_ID = 35 THEN 12
                   WHEN V_GS_DB_LIST_D_ID = 36 THEN 45
             END id_dblist,
             '' id_datalist,
             CASE WHEN STATUS_ID = 7 THEN 5
                  WHEN STATUS_ID = 1 THEN 1
                  WHEN STATUS_ID = 2 THEN 1
                  WHEN STATUS_ID = 3 THEN 2
                  WHEN STATUS_ID = 4 THEN 3
                  WHEN STATUS_ID = 5 THEN 4
                  WHEN STATUS_ID = 6 THEN 4
                  WHEN STATUS_ID = 27 THEN 6
              END id_tuningstatus,
              REQUEST_COMMENTS sql_info,
              TUNNER_COMMENTS tuning_info,
              '' asis_plan,
              '' tobe_plan,
              REQUEST_SQL asis_sql_text,
              ASIS_ELAPSED asis_elapsed_time,
              '' expect_elapsed_time,
              '' avg_rows,
              '' daily_exec_cnt,
              '' bind_value,
              TUNNER_SQL tobe_sql_text,
              TOBE_ELAPSED tobe_elapsed_time,
              NVL(( SELECT BB.ID FROM DBMON.MIG_SABUN_MAPPING BB,
                                      DBMON.AUTH_USER CC
                                WHERE BB.SABUN = A.EMP_NO
                                  AND BB.SABUN = CC.USERNAME  ),18243) id_reg_user,
              case when TUNNER_USERNAME = '이치호' then 2
                when TUNNER_USERNAME = '구경서' then 4
                when TUNNER_USERNAME = '정관호' then 21
                when TUNNER_USERNAME = '이선경' then 7569
                when TUNNER_USERNAME = '이정아' then 12963
                else 4
              end id_tuning_user,
              to_date(decode(TUNNER_DT,'N/A',null), 'YYYY-MM-DD HH24:MI:SS')  tuning_dtm,
              DISTRIBUTE_DT dist_dtm,
              to_date(CREATED, 'YYYY-MM-DD HH24:MI:SS') reg_dtm
      FROM DBMON.DJANGO_TUNNE_BBS A
      WHERE ID <> '999999999'
      -- where id = 2480
        """)

            columnNames = [d[0] for d in cur.description]
            datasets = [dict(zip(columnNames, row)) for row in cur]

            TuningList.objects.all().delete()

            for dataset in datasets:
                asis_sql_text = str(dataset['ASIS_SQL_TEXT']).replace(
                    """<p></p><h4><b><font color="red">★변경된 부분은 <span style="background-color:yellow;">노란색 음영</span>으로 반드시 표시해주세요★</font></b></h4><p></p>""",
                    "")
                tuninglist = TuningList(
                    id=dataset['ID'],
                    title=dataset['TITLE'],
                    sql_id=dataset['SQL_ID'],
                    sql_type=dataset['SQL_TYPE'],
                    id_domain=Domain.objects.get(id=dataset['ID_DOMAIN']) if dataset['ID_DOMAIN'] is not None else None,
                    id_dblist=DbList.objects.get(id=dataset['ID_DBLIST']),
                    id_projectlist=ProjectList.objects.get(id=dataset['ID_PROJECTLIST']) if dataset[
                                                                                                'ID_PROJECTLIST'] is not None else None,
                    id_datalist=None,
                    id_tuningstatus=TuningStatus.objects.get(id=dataset['ID_TUNINGSTATUS']),
                    sql_info=dataset['SQL_INFO'],
                    tuning_info=dataset['TUNING_INFO'],
                    asis_plan=dataset['ASIS_PLAN'],
                    tobe_plan=dataset['TOBE_PLAN'],
                    asis_sql_text=asis_sql_text,
                    asis_elapsed_time=dataset['ASIS_ELAPSED_TIME'],
                    expect_elapsed_time=dataset['EXPECT_ELAPSED_TIME'],
                    avg_rows=dataset['AVG_ROWS'],
                    daily_exec_cnt=dataset['DAILY_EXEC_CNT'],
                    bind_value=dataset['BIND_VALUE'],
                    tobe_sql_text=dataset['TOBE_SQL_TEXT'],
                    tobe_elapsed_time=dataset['TOBE_ELAPSED_TIME'],
                    id_reg_user=User.objects.get(id=dataset['ID_REG_USER']) if dataset[
                                                                                   'ID_REG_USER'] != 18243 else None,
                    id_tuning_user=User.objects.get(id=dataset['ID_TUNING_USER']) if User.objects.get(
                        id=dataset['ID_TUNING_USER']) else None,
                    tuning_dtm=dataset['TUNING_DTM'],
                    dist_dtm=dataset['DIST_DTM'],
                    reg_dtm=dataset['REG_DTM'],
                )
                tuninglist.save()

        obj = []

        return obj

    def post(self, request, *args, **kwargs):

        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        context['template_parent'] = "dba"
        context['template_child'] = "DBABatchLV"

        return context


class ModelSMTCPrd(LoginRequiredMixin, FormMixin, ListView):
    model = StdWord
    template_name = 'ModelSMTCPrd.html'
    context_object_name = 'objects'
    form_class = StdWordForm

    def get_queryset(self):
        self.form = self.get_form(self.form_class)
        obj = []
        return obj

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        req_attr = StdAttr.objects.filter(accept_yn__gt=0, use_yn=1).order_by('std_attr_kor')

        context['template_parent'] = "stdmodel"
        context['template_child'] = "ModelSMTCPrd"

        return context


class ModelLV(LoginRequiredMixin, ListView):
    model = StdWord
    template_name = 'ModelLV.html'
    context_object_name = 'objects'

    def post(self, request, *args, **kwargs):

        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        _loggingVisit(self.request, 1)

        context = super().get_context_data(**kwargs)

        data_model = DataModel.objects.all().order_by('id_domainanddblist__id_dblist__exp_order','exp_order')

        print(data_model.query)


        model = {}

        for data in data_model:

            db_use = data.id_domainanddblist.id_dblist.db_use

            if db_use in model.keys():
                model[db_use][1].append(data)

            else:
                model[db_use] = [db_use, []]
                model[db_use][1].append(data)

        context['template_parent'] = "ModelLV"
        context['template_child'] = "ModelLV"        
        context['model'] = model
        

        return context


class TableListLV(LoginRequiredMixin, FormMixin, ListView):
    model = TableList
    template_name = 'TableListLV.html'
    context_object_name = 'objects'
    paginate_by = 10
    form_class = TableListForm
    keyword = ""

    def get_initial(self):

        ###################################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################
        table_name = ""
        owner = ""
        id_dblist = ""
        oper_cd = ""
        choice_tab_or_col = ""
        choice_option = ""

        table_name = self.request.session['table_name'] if 'table_name' in self.request.session else ""
        owner = self.request.session['owner'] if 'owner' in self.request.session else ""
        id_dblist = self.request.session['id_dblist'] if 'id_dblist' in self.request.session else ""
        oper_cd = self.request.session['oper_cd'] if 'oper_cd' in self.request.session else ""
        choice_tab_or_col = self.request.session[
            'choice_tab_or_col'] if 'choice_tab_or_col' in self.request.session else "1"
        choice_option = self.request.session['choice_option'] if 'choice_option' in self.request.session else ""

        return {'table_name': table_name,
                'owner': owner,
                'id_dblist': id_dblist,
                'oper_cd': oper_cd,
                'choice_tab_or_col': choice_tab_or_col,
                'choice_option': choice_option,
                }
        ###################################################

    def get_queryset(self):

        _loggingVisit(self.request, 22)

        self.form = self.get_form(self.form_class)

        if self.form.is_valid():

            table_name = self.form.cleaned_data['table_name']
            owner = self.form.cleaned_data['owner']
            if owner:
                owner = owner.id
            id_dblist = self.form.cleaned_data['id_dblist']
            oper_cd = self.form.cleaned_data['oper_cd']
            choice_tab_or_col = self.form.cleaned_data['choice_tab_or_col']
            choice_option = self.form.cleaned_data['choice_option']

            ###################################################
            ## 화면에 바인딩 된 값을 유지하기 위한 처리
            ###################################################
            self.request.session['table_name'] = table_name if table_name is not None else ""
            self.request.session['owner'] = owner if owner is not None else ""
            self.request.session['id_dblist'] = id_dblist.id if id_dblist is not None else ""
            self.request.session['oper_cd'] = oper_cd.id if oper_cd is not None else ""
            self.request.session['choice_tab_or_col'] = choice_tab_or_col if choice_tab_or_col is not None else "1"
            self.request.session['choice_option'] = choice_option if choice_option is not None else ""

            ###################################################
            ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            self.request.session['click_search'] = 'Y'
            ###################################################

        else:
            ###################################################
            ## 1. 처음 진입
            ## 2. 페이징 버튼 클릭 시
            ###################################################
            page_kwarg = self.page_kwarg
            page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
            if page == 1:
                self.request.session['table_name'] = ""
                self.request.session['owner'] = ""
                self.request.session['id_dblist'] = ""
                self.request.session['oper_cd'] = ""
                self.request.session['choice_tab_or_col'] = "1"
                self.request.session['choice_option'] = ""

            table_name = self.request.session['table_name'] if 'table_name' in self.request.session else ""
            owner = self.request.session['owner'] if 'owner' in self.request.session else ""
            id_dblist = self.request.session['id_dblist'] if 'id_dblist' in self.request.session else ""
            oper_cd = self.request.session['oper_cd'] if 'oper_cd' in self.request.session else ""
            choice_tab_or_col = self.request.session[
                'choice_tab_or_col'] if 'choice_tab_or_col' in self.request.session else "1"
            choice_option = self.request.session['choice_option'] if 'choice_option' in self.request.session else ""

            # if self.request.method == "POST" :

        dbowner = DBOwner.objects.all()
        list_owner = []

        for o in dbowner:
            list_owner.append(o.owner)

        cdc_yn = ""
        secu_yn = ""

        if choice_option == "CDC":
            cdc_yn = 1
        elif choice_option == "개인정보":
            secu_yn = 1

        if choice_tab_or_col == "1":
            obj = TableList.objects.filter(
                *((Q(table_name__istartswith=table_name.upper()),) if table_name else ()),
                *((Q(owner=DBOwner.objects.filter(id=owner).values('owner').first()['owner']),) if owner else ()),
                *((Q(id_dblist=id_dblist),) if id_dblist else ()),
                *((Q(oper_cd=oper_cd),) if oper_cd else ()),
                *((Q(cdc_yn=cdc_yn),) if cdc_yn else ()),
                *((Q(secu_yn=secu_yn),) if secu_yn else ()),
                Q(owner__in=list_owner),
                drop_yn=0,
            ).order_by('db_use', 'oper_cd', 'owner', 'table_name')

        else:
            obj = ColumnList.objects.filter(
                *((Q(column_name__istartswith=table_name.upper()),) if table_name else ()),
                *((Q(owner=DBOwner.objects.filter(id=owner).values('owner').first()['owner']),) if owner else ()),
                *((Q(id_dblist=id_dblist),) if id_dblist else ()),
                *((Q(oper_cd=oper_cd),) if oper_cd else ()),
                *((Q(secu_yn=secu_yn),) if secu_yn else ()),
                Q(owner__in=list_owner),
                drop_yn=0,
            ).order_by('db_use', 'oper_cd', 'owner', 'table_name', 'column_id')

        return obj

    def paginate_queryset(self, queryset, page_size):

        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty())
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            if page == 'last':
                page_number = paginator.num_pages
            else:
                raise Http404(_("Page is not 'last', nor can it be converted to an int."))
        try:

            ###################################################
            ## SEARCH 버튼 클릭 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            if 'click_search' in self.request.session:
                if self.request.session['click_search'] == "N":
                    page = paginator.page(page_number)
                else:
                    page = paginator.page(1)
            else:
                page = paginator.page(page_number)
            ###################################################

            ###################################################
            ## COLUMNLIST 에서 목록으로 클릭 시 원래 페이지 번호를 찾아가기 위한 코드
            ###################################################
            self.request.session['page_number'] = page_number

            return (paginator, page, page.object_list, page.has_other_pages())



        except:
            page = paginator.page(1)
            return (paginator, page, page.object_list, page.has_other_pages())

    def post(self, request, *args, **kwargs):

        # self.object = None
        # keyword = None
        # self.form = self.get_form(self.form_class)
        # if self.form.is_valid():
        #   # self.object = self.form.save()
        #   keyword = self.form.cleaned_data['keyword']

        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        header = []

        context = super().get_context_data(**kwargs)

        page_range = ""
        paginator = context['paginator']
        page_numbers_range = 5  # Display only 5 page numbers
        max_index = len(paginator.page_range)

        ###################################################
        ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
        ###################################################
        if 'click_search' in self.request.session:
            if self.request.session['click_search'] == "N":
                page = self.request.GET.get('page')
            else:
                page = 1
                self.request.session['click_search'] = "N"
        else:
            page = self.request.GET.get('page')
        ###################################################

        current_page = int(page) if page else 1

        start_index = int((current_page - 1) / page_numbers_range) * page_numbers_range
        end_index = start_index + page_numbers_range
        if end_index >= max_index:
            end_index = max_index

        page_range = paginator.page_range[start_index:end_index]
        context['page_range'] = page_range

        choice_tab_or_col = self.request.session[
            'choice_tab_or_col'] if 'choice_tab_or_col' in self.request.session else "1"

        if choice_tab_or_col == "1":
            header = [
                'DB용도',
                '개발/<br>운영',
                '도메인',
                'OWNER',
                '테이블<br>영문명',
                '테이블<br>한글명',
                'CDC<br>여부',
                '개인정보<br>여부',
                '스페이스명',
                '사이즈<br>(MB)',
                '담당자<br>(Ops)',
                '담당자<br>(Dev)',
                '권한',
                '권한<br>신청',
            ]
        else:
            header = [
                'DB용도',
                '개발/<br>운영',
                'OWNER',
                '테이블<br>한글명',
                '테이블<br>영문명',
                '컬럼<br>영문명',
                '컬럼<br>한글명',
                '데이터<br>타입',
                'PK',
                '개인정보<br>여부',
            ]

        my_privs = UserRequestTabPrivHist.objects.filter(id_reg_user_id=self.request.user.id,
                                                         approv_yn="0",
                                                         object_type="TABLE",
                                                         use_yn='1')

        my_privs_list = []
        for priv in my_privs:
            obj = ObjectList.objects.get(id=priv.id_objectlist.id)
            tab = TableList.objects.filter(id_dblist=obj.id_dblist.id,
                                           oper_cd=obj.oper_cd.id,
                                           owner=obj.owner,
                                           table_name=obj.object_name,
                                           )

            my_privs_list.append(tab[0].id)

        context['template_parent'] = "dev"
        context['template_child'] = "TableListLV"
        context['header'] = header
        context['my_privs_list'] = my_privs_list

        return context


class DataRequestLV(LoginRequiredMixin, ListView):
    model = DataRequest
    template_name = 'DataRequestLV.html'
    context_object_name = 'objects'
    form_class = DataRequestForm

    def get_queryset(self):
        # return DataRequest.objects.filter( Q(id_reg_user=self.request.user.id) | Q(id_prov_user=self.request.user.id)).order_by('-mod_dtm')
        return DataRequest.objects.all().order_by('-reg_dtm')

    def get_context_data(self, **kwargs):
        context = super(ListView, self).get_context_data(**kwargs)

        context['template_parent'] = "data"
        context['template_child'] = "DataRequestLV"

        _loggingVisit(self.request, 14)

        return context


class DataRequestCV(LoginRequiredMixin, CreateView):
    model = DataRequest
    form_class = DataRequestForm
    template_name = 'DataRequestCV.html'

    # def save(self, commit=True):
    #   if self.instance.id is not None :
    #       self.instance.id_mod_user = self.request.user

    #   else :
    #       self.instance.id_reg_user = self.request.user
    #       self.instance.id_mod_user = self.request.user

    #   if self.instance.prov_yn == '1' :
    #     self.instance.prov_yn_dtm = datetime.datetime.now()

    #     dt = datetime.datetime.now() + datetime.timedelta(days=1)
    #     dt = dt.replace(hour=23, minute=59, second=59)
    #     self.instance.poss_view_dtm = dt

    #   if self.instance.id_datalist is not None :
    #     self.instance.id_prov_user = User.objects.get(pk= DomainAndDbList.objects.filter(Q(id_domain=self.instance.id_datalist.id_domain),
    #                                   Q(id_dblist=self.instance.id_datalist.id_dblist)).values_list('it_manager', flat=True)[0])
    #     ds_emp = _query_dict('iam',iamSql.format(emp_no=self.instance.id_prov_user))[0]

    #   ds_emp = _query_dict('iam',iamSql.format(emp_no=self.request.user))[0]

    #   return super(DataRequestForm, self).save(commit=True)

    def get_success_url(self):
        return reverse_lazy('DataRequestLV')

    def get_form_kwargs(self):
        # add request for form to validate
        kwargs = super(DataRequestCV, self).get_form_kwargs()
        kwargs.update({"request": self.request})
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(CreateView, self).get_context_data(**kwargs)

        context['template_parent'] = "data"
        context['template_child'] = "DataRequestLV"
        _loggingVisit(self.request, 14)

        return context


class DataRequestUV(LoginRequiredMixin, UpdateView):
    model = DataRequest
    form_class = DataRequestForm
    template_name = 'DataRequestUV.html'

    context_object_name = 'objects'

    def get_success_url(self):
        return reverse_lazy('DataRequestLV')

    def get_object(self):
        return get_object_or_404(DataRequest, id=self.kwargs['id'])

    def get_form_kwargs(self):
        # add request for form to validate
        kwargs = super(DataRequestUV, self).get_form_kwargs()
        kwargs.update({"request": self.request})
        return kwargs

    def get_context_data(self, **kwargs):
        # IT개발팀 인원만 가능하도록. 하드코딩 ^^;
        prov_poss_member = User.objects.filter(last_name='IT개발팀')
        prov_poss_id = []
        for m in prov_poss_member:
            prov_poss_id.append(m.id)

        context = super(UpdateView, self).get_context_data(**kwargs)

        context['template_parent'] = "data"
        context['template_child'] = "DataRequestLV"

        context['prov_poss_id'] = prov_poss_id

        _loggingVisit(self.request, 14)

        return context


class DataListCreateView(LoginRequiredMixin, CreateView):
    model = DataList
    form_class = DataListForm
    template_name = 'DataListCreateView.html'

    # success_url = reverse_lazy('sqlTuningMain')

    def get_success_url(self):
        return reverse_lazy('dataView')

    def get_form_kwargs(self):
        # add request for form to validate
        kwargs = super(DataListCreateView, self).get_form_kwargs()
        kwargs.update({"request": self.request})
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(CreateView, self).get_context_data(**kwargs)

        context['template_parent'] = "data"
        context['template_child'] = "DataListCreateView"

        _loggingVisit(self.request, 4)

        return context


class DataListUpdateView(LoginRequiredMixin, UpdateView):
    model = DataList
    form_class = DataListForm
    template_name = 'DataListUpdateView.html'

    context_object_name = 'datalist'

    def get_success_url(self):
        return reverse_lazy('dataView')

    def get_object(self):
        datalist = get_object_or_404(DataList, id=self.kwargs['id'])

        return datalist

    def get_form_kwargs(self):
        # add request for form to validate
        kwargs = super(DataListUpdateView, self).get_form_kwargs()
        kwargs.update({"request": self.request})
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(UpdateView, self).get_context_data(**kwargs)

        context['template_parent'] = "dataView"
        context['template_child'] = "dataView"

        _loggingVisit(self.request, 4)

        return context


def DataListDelView(request):
    # Login 한 user 만 접속 허용
    # Login 페이지를 통하지 않았을 경우 Login 페이지로 redirect
    if not request.user.is_authenticated:
        return redirect('%s?next=%s' % (settings.LOGIN_URL, request.path))

    _loggingVisit(request, 4)

    sql_id = request.POST.get('div_sql_id', '')

    DataList.objects.filter(id=sql_id).delete()

    return redirect('dataView')


def DataListSelectView_Approve(request):
    template_parent = "DataListSelectView_Approve"
    template_child = "DataListSelectView_Approve"

    # id_dblist=39 : 데이터허브
    ds_data_list = DataList.objects.filter(~Q(id_dblist=39), ~Q(id_req_users=None), Q(exp_yn=1),
                                           (Q(prov_yn1=0) | Q(prov_yn2=0))).order_by("-reg_dtm")

    return render(request, 'DataListSelectView_Approve.html', {
        'template_parent': template_parent,
        'template_child': template_child,
        'ds_data_list': ds_data_list,
    })


class DataListUpdateView_Approve(LoginRequiredMixin, UpdateView):
    model = DataList
    form_class = DataListUpdateView_Approve_Form
    template_name = 'DataListUpdateView_Approve.html'

    context_object_name = 'datalist_approve'

    def get_success_url(self):
        return reverse_lazy('DataListSelectView_Approve')

    def get_object(self):
        datalist_approve = get_object_or_404(DataList, id=self.kwargs['id'])

        return datalist_approve

    def get_form_kwargs(self):
        # add request for form to validate
        kwargs = super(DataListUpdateView_Approve, self).get_form_kwargs()
        kwargs.update({"request": self.request})
        return kwargs

    def get_context_data(self, **kwargs):
        context = super(UpdateView, self).get_context_data(**kwargs)

        context['template_parent'] = "DataListSelectView_Approve"
        context['template_child'] = "DataListSelectView_Approve"

        return context


class Autocomplete_User_Search(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return User.objects.none()
        qs = User.objects.all().filter(Q(is_active=1),
                                       ~Q(last_name='인사팀휴직'),
                                       ~Q(last_name='GS SHOP'),
                                       ~Q(last_name='공용계정그룹')).order_by("-username")
        if self.q:
            qs = qs.filter(first_name__istartswith=self.q)
        return qs


# 팀변경
# 변경관리 시 사용되는 팀 필터
class Autocomplete_MetaReq_User_Search(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return User.objects.none()
        qs = User.objects.all().filter(Q(is_active=1), ~Q(username__istartswith='D'),
                                       (Q(last_name='상품정산IT팀') |
                                        Q(last_name='주문방송IT팀') |
                                        Q(last_name='물류기획팀') |
                                        Q(last_name='물류IT팀') |
                                        Q(last_name='빅데이터팀') |
                                        Q(last_name='회원Product팀') |
                                        Q(last_name='전시Product팀') |
                                        Q(last_name='Mobile Native-TFT') |
                                        Q(last_name='MSA-TFT') |
                                        Q(last_name='클라우드팀'))).order_by("first_name")
        if self.q:
            qs = qs.filter(first_name__istartswith=self.q)
        return qs


class Autocomplete_Domain(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return Domain.objects.none()
        qs = Domain.objects.all().order_by('domain_name')
        if self.q:
            qs = qs.filter(domain_name__istartswith=self.q)
        return qs


class Autocomplete_DbList(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return DbList.objects.none()
        qs = DbList.objects.filter(use_yn="1", gather_meta_yn="1").order_by('exp_order')

        oper_cd = self.forwarded.get('oper_cd', None)

        if oper_cd:
            qs = qs.extra(tables=['cust_db_detail'],
                          where=[
                              'cust_db_detail.id_dblist = cust_db_list.id and cust_db_detail.oper_cd = ' + oper_cd]).distinct()

        if self.q:
            qs = qs.filter(db_use__icontains=self.q.upper())

        return qs


class Autocomplete_MetaReq_DbList(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return DbList.objects.none()
        qs = DbList.objects.filter(use_yn="1", metareq_yn="1").order_by('exp_order', 'db_use')

        oper_cd = self.forwarded.get('oper_cd', None)

        if oper_cd:
            qs = qs.extra(tables=['cust_db_detail'],
                          where=[
                              'cust_db_detail.id_dblist = cust_db_list.id and cust_db_detail.oper_cd = ' + oper_cd]).distinct()

        if self.q:
            qs = qs.filter(db_use__icontains=self.q.upper())

        return qs


class Autocomplete_ExecuteSQL_DbList(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return DbList.objects.none()
        qs = DbDetail.objects.filter(
            Q(id=344) | Q(id=343) | Q(id=318)
        ).order_by('id_dblist__db_use')

        if self.q:
            qs = qs.filter(id_dblist__db_use__icontains=self.q.upper())

        return qs


class Autocomplete_DbList_Part(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return DbList.objects.none()
        qs = DbList.objects.all().filter(Q(id=10) |
                                         Q(id=3) |
                                         Q(id=39))

        if self.q:
            qs = qs.filter(db_use__istartswith=self.q)

        return qs


class Autocomplete_OperCd(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return OperCd.objects.none()
        qs = OperCd.objects.filter(id__in=[3, 4])
        if self.q:
            qs = qs.filter(oper_cd__istartswith=self.q)
        return qs


class Autocomplete_DataList(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return DataList.objects.none()
        qs = DataList.objects.filter(privacy_yn="1").order_by("-reg_dtm")
        if self.q:
            qs = qs.filter(data_title__istartswith=self.q)
        return qs


class Autocomplete_DBOwner(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !

        if not self.request.user.is_authenticated:
            return DBOwner.objects.none()
        qs = DBOwner.objects.all()

        id_dblist = self.forwarded.get('id_dblist', None)

        if id_dblist:
            qs = qs.filter(id_dblist=id_dblist).order_by('-table_count')
            # Transaction.objects.all().values('actor').annotate(total=Count('actor')).order_by('total')

            if self.q:
                qs = qs.filter(owner__contains=self.q.upper())
            return qs
        else:
            return None


class Autocomplete_TableList(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return TableList.objects.none()

        # qs = TableList.objects.all()
        qs = TableList.objects.filter(oper_cd=3, drop_yn="0").extra(tables=['cust_owner'], where=[
            'cust_owner.owner = cust_table_list.owner and cust_owner.id_dblist = cust_table_list.id_dblist']).order_by(
            "owner", "table_name")

        id_dblist = self.forwarded.get('id_dblist', None)

        if id_dblist:
            qs = qs.filter(id_dblist=id_dblist)

            if self.q:
                qs = qs.filter(table_name__istartswith=self.q)

            return qs
        else:
            return None


class Autocomplete_TableList_Meta(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return TableList.objects.none()

        # 운영만 존재하는 DB
        # 39 : 데이터허브
        qs = TableList.objects.filter(drop_yn="0").extra(tables=['cust_owner'],
                                                         where=['cust_owner.owner = cust_table_list.owner \
                                                              and cust_owner.id_dblist = cust_table_list.id_dblist \
                                                              and ( cust_table_list.oper_cd = 4 \
                                                                 or cust_table_list.id_dblist = 39 )']).order_by(
            "owner", "table_name")

        id_dblist = self.forwarded.get('id_dblist', None)
        owner = self.forwarded.get('owner', None)

        if id_dblist:
            qs = qs.filter(id_dblist=id_dblist)

            if owner:
                owner = DBOwner.objects.get(id=owner)

                qs = qs.filter(owner=owner.owner)

                if self.q:
                    qs = qs.filter(table_name__istartswith=self.q)

                return qs
        else:
            return None


class Autocomplete_DomainAndDbList(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return DomainAndDbList.objects.none()
        qs = DomainAndDbList.objects.all().order_by('id_dblist__db_use', 'id_domain__domain_name')

        id_dblist = self.forwarded.get('id_dblist', None)

        if id_dblist:
            qs = qs.filter(id_dblist=id_dblist)

            if self.q:
                qs = qs.filter(id_domain__domain_name__icontains=self.q.upper())

            return qs
        else:
            return None


class Autocomplete_DBTablespace(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return DBTablespace.objects.none()

        qs = DBTablespace.objects.all().order_by('tablespace_name')

        owner = self.forwarded.get('owner', None)

        if owner:
            qs = qs.filter(id_owner=owner)

            if self.q:
                qs = qs.filter(tablespace_name__icontains=self.q.upper())

            return qs
        else:
            return None


class Autocomplete_MetaGrantList(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return MetaGrantList.objects.none()

        qs = MetaGrantList.objects.all()
        # qs = MetaGrantList.objects.values('app_service').annotate(id=Max('id'),id_dblist=Max('id_dblist')).values('id','app_service','id_dblist')

        id_dblist = self.forwarded.get('id_dblist', None)

        if id_dblist:
            qs = qs.filter(id_dblist=id_dblist)

            if self.q:
                qs = qs.filter(app_service__icontains=self.q.upper())

            return qs
        else:
            return None


class Autocomplete_ProjectList(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return ProjectList.objects.none()
        qs = ProjectList.objects.all().order_by("-start_dtm")
        if self.q:
            qs = qs.filter(title__icontains=self.q)
        return qs


totalColor = [
    'rgba(255, 99, 132, 0.6)',
    'rgba(54, 162, 235, 0.6)',
    'rgba(255, 206, 86, 0.6)',
    'rgba(75, 192, 192, 0.6)',
    'rgba(153, 102, 255, 0.6)',
    'rgba(255, 159, 64, 0.6)',
]

totalBorderColor = [
    'rgba(255, 99, 132, 0.6)',
    'rgba(54, 162, 235, 0.6)',
    'rgba(255, 206, 86, 0.6)',
    'rgba(75, 192, 192, 0.6)',
    'rgba(153, 102, 255, 0.6)',
    'rgba(255, 159, 64, 0.6)',

]


def chart_team_visit_cnt(request):
    labels = []
    data = []
    backgroundColor = []

    sql = """
      select team_name, count(distinct name) cnt
      from cust_user_visit
      where reg_dtm >  date_add(now(),interval -7 day)
      group by team_name
      order by cnt desc
    """

    queryset = _query_dict("default", sql)

    ind = 0
    for team in queryset:
        labels.append(team['team_name'])
        data.append(team['cnt'])
        backgroundColor.append(totalColor[ind])

        if len(totalColor) == ind + 1:
            ind = 0
        else:
            ind += 1

    return JsonResponse(data={
        'labels': labels,
        'data': data,
        'backgroundColor': backgroundColor,
    })


def chart_db_size(request):
    label = []
    labels = []
    data = defaultdict(list)
    backgroundColor = []

    sql = """
      select reg_dtm, db_use, size
      from cust_db_size_aggr
      order by reg_dtm;
    """

    queryset = _query_dict("default", sql)

    ind = 0

    for team in queryset:
        label.append(team['db_use'])
        labels.append(team['reg_dtm'])
        data[team['db_use']].append(team['size'])
        backgroundColor.append(totalBorderColor[ind])

        if len(totalColor) == ind + 1:
            ind = 0
        else:
            ind += 1

    labels = list(set(labels))
    labels.sort()
    label = list(set(label))
    backgroundColor = list(set(backgroundColor))

    item = []

    ind = 0
    for t in label:
        inner_dict = {}
        inner_dict["data"] = data[t]
        inner_dict["label"] = t
        inner_dict["borderColor"] = backgroundColor[ind]
        inner_dict["fill"] = False
        item.append(inner_dict)
        ind += 1

    return JsonResponse(data={
        'item': item,
        'labels': labels,

    })


# def chart_monitor(request):
#     id_dbdetail = request.POST.get('id_dbdetail', '')
#     item_nm = request.POST.get('item', '')

#     label = []
#     labels = []
#     data = defaultdict(list)
#     backgroundColor = []

#     sql = """
#       SELECT  A.MONITOR_VALUE,
#               DATE_FORMAT(A.REG_DTM,'%H:%i') REG_DTM,
#               E.ITEM_NM,
#               D.OPER_CD,
#               B.DB_USE,
#               D.DB_ORDER,
#               D.HOST_NM,
#               D.SVC_IP
#       FROM CUST_MONITOR_ITEM_LOG A USE INDEX (ix_cust_monitor_item_log_01),
#          CUST_DB_LIST B,
#          CUST_MONITOR_ITEM_LIST C,
#          CUST_DB_DETAIL D,
#          CUST_MONITOR_ITEM E
#       WHERE A.REG_DTM >= DATE_ADD(STR_TO_DATE(DATE_FORMAT(NOW(), '%Y%m%d %H%i'),'%Y%m%d %H%i'), INTERVAL -1 hour)
#       AND D.ID = {id_dbdetail}
#       AND A.id_monitoritemlist = C.ID
#       AND A.ID_DBDETAIL = D.ID
#       AND A.ID_DBLIST = B.ID
#       AND C.ITEM_NM = '{item}'
#       AND E.ID = C.ID_MONITORITEM      
#       ORDER BY A.REG_DTM
#     """.format(id_dbdetail=id_dbdetail,
#                item=item_nm)
#     queryset = _query_dict("default", sql)

#     ind = 0

#     for team in queryset:
#         label.append(team['ITEM_NM'])
#         labels.append(team['REG_DTM'])
#         data[team['ITEM_NM']].append(team['MONITOR_VALUE'])
#         backgroundColor.append(totalBorderColor[ind])

#         if len(totalColor) == ind + 1:
#             ind = 0
#         else:
#             ind += 1

#     labels = list(set(labels))
#     labels.sort()
#     label = list(set(label))
#     backgroundColor = list(set(backgroundColor))

#     item = []

#     ind = 0

#     for t in label:
#         inner_dict = {}
#         inner_dict["data"] = data[t]
#         inner_dict["label"] = t
#         inner_dict["borderColor"] = backgroundColor[ind]
#         inner_dict["fill"] = False
#         item.append(inner_dict)
#         ind += 1

#     return JsonResponse(data={
#         'item': item,
#         'labels': labels,

#     })


##############################################################################
## 용어 표준
##############################################################################
class StdDivAttrLV(LoginRequiredMixin, FormMixin, ListView):
    model = StdWord
    template_name = 'StdDivAttrLV.html'
    context_object_name = 'objects'
    form_class = StdWordForm

    def get_queryset(self):
        self.form = self.get_form(self.form_class)
        obj = []
        return obj

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        _loggingVisit(self.request, 19)

        context = super().get_context_data(**kwargs)

        is_superuser = self.request.user.is_superuser

        if is_superuser :
            req_attr = StdAttr.objects.filter(accept_yn__gt=0,
                                              reg_dtm__gt=datetime.datetime.now() - datetime.timedelta(days=14)).order_by(
                '-reg_dtm')
        else :
            req_attr = StdAttr.objects.filter(accept_yn__gt=0,
                                              id_reg_user=self.request.user.id,
                                              reg_dtm__gt=datetime.datetime.now() - datetime.timedelta(days=14)).order_by(
                '-reg_dtm')


        context['template_parent'] = "stddata"
        context['template_child'] = "StdDivAttrLV"
        context['req_attr'] = req_attr

        return context


# 용어 나누기 클릭
def ajaxDivAttr(request):

    attr = request.POST.get('attr', '')
    attr = attr.upper()

    clean_attr = (attr.replace('[', '')).replace(']', '')

    qs1 = StdAttr.objects.filter(std_attr_kor__iexact=clean_attr, accept_yn=0, use_yn=1)
    qs2 = StdAttr.objects.filter(std_attr_kor__icontains=clean_attr, accept_yn=0, use_yn=1).order_by('std_attr_kor')

    exact_datatype = ""

    if len(qs1) > 0:
        exact_datatype = qs1[0].id_stddomain.info_type

    dataset = qs1.union(qs2)

    attr_data = []

    final_data_type = []

    for data in dataset:
        if data.std_attr_kor == clean_attr:
            sql = """SELECT {id} ID, data_type DATA_TYPE, count(*) CNT
                  FROM CUST_COLUMN_LIST
                  WHERE COLUMN_NAME = '{column_name}'
                  AND DATA_TYPE = '{data_type}'
                  AND DB_USE = 'SMTC'
                  AND DROP_YN = 0
                  group by data_type
                  order by 2 desc
              """.format(column_name=data.std_attr_eng,
                         data_type=data.id_stddomain.oracle_datatype,
                         id=data.id)

            
            cnt_ds = _query_dict("default", sql)

            if len(cnt_ds) > 0:
                final_data_type.append([cnt_ds[0]['ID'], cnt_ds[0]['CNT'], len(cnt_ds[0]['DATA_TYPE'])])


    if len(final_data_type) > 0:
        final_data_type = sorted(final_data_type, key=itemgetter(1, 2), reverse=True)

        final_data_type_id = final_data_type[0][0]

        in_dataset = StdAttr.objects.get(id=final_data_type_id)



        attr_data.append(
            [
                in_dataset.std_attr_kor.replace(clean_attr, '<b><font color=red>' + clean_attr + '</font></b>'),
                in_dataset.std_attr_eng,
                f_get_datatype(in_dataset.id, 'ORACLE'),
                f_get_datatype(in_dataset.id, 'MYSQL'),
                in_dataset.id_stddomain.info_type,
                in_dataset.expl,
                "DBA" if in_dataset.id_reg_user.is_superuser else in_dataset.id_reg_user.first_name
            ]
        )

    for data in dataset:

        if data.std_attr_kor != clean_attr:


            reg_user = ""
            if data.id_reg_user is None :
                reg_user = "DBA"
            elif data.id_reg_user.is_superuser :
                reg_user = "DBA"
            else :
                reg_user = data.id_reg_user.first_name



            attr_data.append(
                [
                    data.std_attr_kor.replace(clean_attr, '<b><font color=red>' + clean_attr + '</font></b>'),
                    data.std_attr_eng,
                    f_get_datatype(data.id, 'ORACLE'),
                    f_get_datatype(data.id, 'MYSQL'),
                    data.id_stddomain.info_type,
                    data.expl,
                    reg_user
                ]
            )

    l_lkor, \
    l_leng, \
    l_domain, \
    l_eng_attr_name, \
    l_domain_flag, \
    l_word_find_flag, \
    l_last_num_flag, \
    lcnt, \
    total_attr, \
    max_length, \
    l_attr_find, \
    l_aleary_req, \
    list_attr = f_div_attr(attr)

    context = {
        'lkor': l_lkor,
        'leng': l_leng,
        'domain': l_domain,
        'eng_attr_name': l_eng_attr_name,
        'domain_flag': l_domain_flag,
        'word_find_flag': l_word_find_flag,
        'last_num_flag': l_last_num_flag,
        'lcnt': lcnt,
        'total_attr': total_attr,
        'max_length': max_length,
        'list_attr': list_attr,
        'l_attr_find': l_attr_find,
        'l_aleary_req': l_aleary_req,
        'attr_data': attr_data,
        'exact_datatype': exact_datatype
    }



    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxRequestNewAttr(request):
    id_stddomain = request.POST.get('id_stddomain', '')
    kor = request.POST.get('kor', '')
    eng = request.POST.get('eng', '')
    csr_no = request.POST.get('csr_no', '')

    error_msg = ""

    obj, flag = StdAttr.objects.get_or_create(
        std_attr_kor=kor,
        std_attr_eng=eng,
        id_stddomain=StdDomain.objects.get(id=id_stddomain),
        use_yn=1,
    )

    if flag:  ## 생성
        obj.id_stddomain = StdDomain.objects.get(id=id_stddomain)
        obj.reg_dtm = datetime.datetime.now()
        obj.csr_no = csr_no
        obj.id_reg_user = request.user
        obj.accept_yn = 1

        obj.save()
    else:  # 존재
        # 반려 중일 경우
        if obj.accept_yn == 2:
            obj.id_stddomain = StdDomain.objects.get(id=id_stddomain)
            obj.reg_dtm = datetime.datetime.now()
            obj.csr_no = csr_no
            obj.id_reg_user = request.user
            obj.accept_yn = 1
            obj.reject_exp = ""

            obj.save()
        else:
            error_msg = "* 오류 : 이미 [{username}]에 의해 신청되었으며, 승인 대기 중입니다.".format(username=obj.id_reg_user.first_name)

    context = {
        'result': 'OK',
        'error_msg': error_msg
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


#################################################################################
## 도메인 표준
#################################################################################

class StdDomainLV(LoginRequiredMixin, FormMixin, ListView):
    model = StdDomain
    template_name = 'StdDomainLV.html'
    context_object_name = 'objects'
    paginate_by = 10
    form_class = StdDomainForm
    keyword = ""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # 반려건 조회
        req_domain = StdDomain.objects.filter(accept_yn__gt=0,
                                              reg_dtm__gt=datetime.datetime.now() - datetime.timedelta(days=14)).order_by('domain_name')

        context['template_parent'] = "stddata"
        context['template_child'] = "StdDomainLV"
        context['req_domain'] = req_domain

        return context


# 키워드 조회
def ajaxSelectDomain(request):
    domain_name = request.POST.get('word', '')
    id_std_domaintype = request.POST.get('id_id_std_domaintype', '')

    obj = []
    word_data = []

    if len(domain_name) > 0 or len(id_std_domaintype):

        qs1 = StdDomain.objects.filter(Q(accept_yn=0),
                                       Q(use_yn=1),
                                       *((Q(domain_name__iexact=domain_name.upper()),) if domain_name else ()),
                                       *((Q(id_std_domaintype=id_std_domaintype),) if id_std_domaintype else ()),
                                       ).order_by('domain_name', )
        qs2 = StdDomain.objects.filter(Q(accept_yn=0),
                                       Q(use_yn=1),
                                       *((Q(domain_name__icontains=domain_name.upper()),) if domain_name else ()),
                                       *((~Q(domain_name__iexact=domain_name.upper()),) if domain_name else ()),
                                       *((Q(id_std_domaintype=id_std_domaintype),) if id_std_domaintype else ()),
                                       ).order_by('domain_name', )

        obj = list(chain(qs1, qs2))




        for data in obj:
            word_data.append([
                data.domain_name.replace(domain_name.upper(),
                                         '<b><font color=red>' + domain_name.upper() + '</font></b>'),
                data.info_type,
                data.oracle_datatype,
                data.mysql_datatype,
                data.expl,
            ]
            )

    context = {
        'word_data': word_data,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


def ajaxRequestNewDomain(request):
    id_domain_name = request.POST.get('id_domain_name', '')
    id_data_type_text = request.POST.get('id_data_type_text', '')
    id_data_type_value = request.POST.get('id_data_type_value', '')
    id_leng = request.POST.get('id_leng', '')
    id_decimal_leng = request.POST.get('id_decimal_leng', '')
    id_group_code = request.POST.get('id_group_code', '')
    id_group_code_yn_value = request.POST.get('id_group_code_yn_value', '')
    expl = request.POST.get('id_expl', '')

    error_msg = ""
    ok_div = True
    attr_div = ""

    last_number = ""

    try :
        last_number = re.search("([a-zA-Z_가-힣])(\d+$)", id_domain_name).group(2)
    except :
        pass

    
    if id_domain_name.replace("]","").replace("[","")[-2:] in ('코드', '번호'):
        l_lkor, \
        l_leng, \
        l_domain, \
        l_eng_attr_name, \
        l_domain_flag, \
        l_word_find_flag, \
        l_last_num_flag, \
        lcnt, \
        total_attr, \
        max_length, \
        l_attr_find, \
        l_aleary_req, \
        list_attr = f_div_attr(id_domain_name)

        ok_div = l_word_find_flag[0]
        attr_div = ' '.join(l_lkor[0])

        if ok_div == False:
            error_msg = "* 오류 : 구성 단어가 없습니다. [ " + attr_div + " ]"

        if StdAttr.objects.filter(std_attr_kor=id_domain_name, use_yn=1).exists():
            ok_div = False
            error_msg = "* 오류 : 동일한 이름의 용어가 존재합니다. [ " + attr_div + " ]"


    elif len(last_number) > 0 :
        ok_div = False
        error_msg = "* 오류 : 도메인에는 숫자가 포함될 수 없습니다."

    else:
        if StdWord.objects.filter(std_wd_kor=id_domain_name).exists() == False:
            ok_div = False
            error_msg = "* 오류 : 미등록 단어 이거나 OO코드 / OO번호 도메인이 아닙니다. [ " + id_domain_name + " ]"

    if ok_div:

        # id_domain_name = id_domain_name.replace("]","").replace("[","")

        oracle_datatype = ""

        # 1 : 문자열
        if id_data_type_value == "1" and int(id_leng) <= 4000:
            oracle_datatype = "VARCHAR2(" + id_leng + ")"
        # 2 : 숫자
        elif id_data_type_value == "2":
            oracle_datatype = "NUMBER" + \
                              ("(" + str(id_leng) if id_leng != "" else "") + \
                              ("," + str(id_decimal_leng) if id_decimal_leng != "" and id_decimal_leng != "0" else "") + \
                              (")" if id_leng != "" else "")
        # 3 : 날짜
        elif id_data_type_value == "3":
            oracle_datatype = "DATE"

        elif id_data_type_value == "1" and int(id_leng) > 4000:
            oracle_datatype = 'CLOB'
        else:
            oracle_datatype = id_data_type_text

        mysql_datatype = ""

        # 1 : 문자열
        if id_data_type_value == "1" and int(id_leng) <= 16000:
            mysql_datatype = "VARCHAR(" + id_leng + ")"
        # 2 : 숫자
        elif id_data_type_value == "2":
            if id_decimal_leng != "" and id_decimal_leng != "0":
                mysql_datatype = "DECIMAL(" + str(id_leng) + ","+str(id_decimal_leng)+")"
            elif id_leng != "" and int(id_leng) >= 1 and int(id_leng) < 3:
                mysql_datatype = "TINYINT"
            elif id_leng != "" and int(id_leng) >= 3 and int(id_leng) < 5:
                mysql_datatype = "SMALLINT"
            elif id_leng != "" and int(id_leng) >= 5 and int(id_leng) < 9:
                mysql_datatype = "INT"
            elif id_leng != "" and int(id_leng) >= 9 and int(id_leng) < 19:
                mysql_datatype = "BIGINT"
            elif id_leng != "" and int(id_leng) >= 19:
                mysql_datatype = "DECIMAL"
        # 3 : 날짜
        elif id_data_type_value == "3":
            mysql_datatype = "DATETIME"
        elif id_data_type_value == "1" and int(id_leng) > 16000:
            mysql_datatype = 'TEXT'
        else:
            mysql_datatype = id_data_type_text

        # [, ] 제거
        data_type = id_domain_name.replace("]","").replace("[","") + "_" + id_data_type_text + str(id_leng) + (
            "," + str(id_decimal_leng) if id_decimal_leng != "" else "")

        obj, flag = StdDomain.objects.get_or_create(
            info_type=data_type,
        )

        if flag:  ## 생성
            obj.domain_name = id_domain_name
            obj.id_stddatatype = StdDataType.objects.get(id=id_data_type_value)
            obj.oracle_datatype = oracle_datatype
            obj.mysql_datatype = mysql_datatype
            obj.group_code = id_group_code
            obj.group_code_yn = id_group_code_yn_value
            obj.expl = expl
            obj.reg_dtm = datetime.datetime.now()
            obj.id_reg_user = request.user
            obj.accept_yn = 1

            obj.save()
        else:  # 존재
            # 반려상태에서 다시 신청할 경우
            if obj.accept_yn == "2":
                obj.domain_name = id_domain_name
                obj.id_stddatatype = StdDataType.objects.get(id=id_data_type_value)
                obj.oracle_datatype = oracle_datatype
                obj.mysql_datatype = mysql_datatype
                obj.group_code = id_group_code
                obj.group_code_yn = id_group_code_yn_value
                obj.expl = expl
                obj.reg_dtm = datetime.datetime.now()
                obj.id_reg_user = request.user
                obj.accept_yn = 1
                obj.reject_exp = ""

                obj.save()
            else:
                error_msg = "* 오류 : 이미 존재하는 도메인입니다."

    context = {
        'result': 'OK',
        'error_msg': error_msg
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


class Autocomplete_StdDomainType(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return StdDomainType.objects.none()
        qs = StdDomainType.objects.all().order_by('domain_type')
        if self.q:
            qs = qs.filter(domain_type__contains=self.q.upper())
        return qs


#########################################################################
## 단어 표준
#########################################################################

# 단어 첫 화면 진입
class StdWordLV(LoginRequiredMixin, FormMixin, ListView):
    model = StdWord
    template_name = 'StdWordLV.html'
    context_object_name = 'objects'
    paginate_by = 10
    form_class = StdWordForm
    keyword = ""

    # 신청된 단어 리스트
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        req_word = StdWord.objects.filter(accept_yn__gt=0,
                                          reg_dtm__gt=datetime.datetime.now() - datetime.timedelta(days=14)).order_by(
            '-reg_dtm', 'std_wd_kor')

        context['template_parent'] = "stddata"
        context['template_child'] = "StdWordLV"
        context['req_word'] = req_word

        return context


# 키워드 입력, 조회 시
def ajaxSelectWord(request):
    std_wd_kor = request.POST.get('word', '')
    choice_tab_or_col = request.POST.get('choice_tab_or_col', '')

    obj = []

    word_data = []

    if choice_tab_or_col == "1":
        qs1 = StdWord.objects.filter(
            *((Q(std_wd_kor__iexact=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
        ).order_by('std_wd_kor', )
        qs2 = StdWord.objects.filter(
            *((Q(std_wd_kor__icontains=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
            *((~Q(std_wd_kor__iexact=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
        ).order_by('std_wd_kor', )

        obj = list(chain(qs1, qs2))

        for data in obj:
            word_data.append([
                data.std_wd_kor.replace(std_wd_kor, '<b><font color=red>' + std_wd_kor + '</font></b>'),
                data.std_wd_eng,
                data.std_wd_eng_ful,
                data.expl,
            ]
            )


    elif choice_tab_or_col == "2":

        qs1 = StdWord.objects.filter(
            *((Q(std_wd_eng__iexact=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
        ).order_by('std_wd_eng', )
        qs2 = StdWord.objects.filter(
            *((Q(std_wd_eng__icontains=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
            *((~Q(std_wd_eng__iexact=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
        ).order_by('std_wd_eng', )

        obj = list(chain(qs1, qs2))

        for data in obj:
            word_data.append([
                data.std_wd_kor,
                data.std_wd_eng.replace(std_wd_kor, '<b><font color=red>' + std_wd_kor + '</font></b>'),
                data.std_wd_eng_ful,
                data.expl,
            ]
            )

    elif choice_tab_or_col == "3":

        qs1 = StdWord.objects.filter(
            *((Q(std_wd_eng_ful__iexact=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
        ).order_by('std_wd_eng_ful', )
        qs2 = StdWord.objects.filter(
            *((Q(std_wd_eng_ful__icontains=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
            *((~Q(std_wd_eng_ful__iexact=std_wd_kor.upper(), accept_yn=0),) if std_wd_kor else ()),
        ).order_by('std_wd_eng_ful', )

        obj = list(chain(qs1, qs2))

        for data in obj:
            word_data.append([
                data.std_wd_kor,
                data.std_wd_eng,
                data.std_wd_eng_ful.replace(std_wd_kor, '<b><font color=red>' + std_wd_kor + '</font></b>'),
                data.expl,
            ]
            )

    context = {
        'word_data': word_data,
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


# 새로운 단어 요청
def ajaxRequestNewWord(request):
    std_wd_kor = request.POST.get('std_wd_kor', '')
    std_wd_eng = request.POST.get('std_wd_eng', '')
    std_wd_eng_ful = request.POST.get('std_wd_eng_ful', '')
    expl = request.POST.get('expl', '')

    std_wd_kor = std_wd_kor.upper()
    std_wd_eng = std_wd_eng.upper()
    std_wd_eng_ful = std_wd_eng_ful.upper()

    error_msg = ""

    chk_dup_eng = StdWord.objects.filter(~Q(std_wd_kor=std_wd_kor),
                                         Q(std_wd_eng=std_wd_eng),
                                         accept_yn="0"
                                         )

    if chk_dup_eng:
        chk_dup_eng = chk_dup_eng[0]
        error_msg = "오류 : 영문약어명이 중복됩니다. 다른 중복 약어를 입력해주세요<br>* 중복단어 : {std_wd_kor} / {std_wd_eng}".format(
            std_wd_kor=chk_dup_eng.std_wd_kor,
            std_wd_eng=chk_dup_eng.std_wd_eng)

    else:

        obj, flag = StdWord.objects.get_or_create(
            std_wd_kor=std_wd_kor,
        )

        if flag:  ## 생성
            obj.std_wd_eng = std_wd_eng
            obj.std_wd_eng_ful = std_wd_eng_ful
            obj.expl = expl
            obj.reg_dtm = datetime.datetime.now()
            obj.id_reg_user = request.user
            obj.accept_yn = 1

            obj.save()
        else:  # 존재
            # 반려
            if obj.accept_yn == "2":
                obj.std_wd_eng = std_wd_eng
                obj.std_wd_eng_ful = std_wd_eng_ful
                obj.expl = expl
                obj.reg_dtm = datetime.datetime.now()
                obj.id_reg_user = request.user
                obj.accept_yn = 1
                obj.reject_exp = ""

                obj.save()
            else:
                error_msg = "오류 : [{std_wd_kor}]는 이미 존재하는 단어입니다.".format(std_wd_kor=std_wd_kor)

    context = {
        'result': 'OK',
        'error_msg': error_msg
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


# 단어/도메인/용어 승인/반려/취소
def ajaxRequestStdModify(request):
    req_id = request.POST.get('req_id', '')
    # 0 : 단어
    # 1 : 도메인
    # 2 : 용어
    std_div = request.POST.get('std_div', '')
    req_div = request.POST.get('req_div', '')
    comment = request.POST.get('comment', '')

    error_yn = "N"
    error_msg = ""

    if std_div == "0":  # 단어
        if req_div == "1":  # 취소
            data = StdWord.objects.get(id=req_id).delete()
            error_msg = "정상적으로 취소되었습니다."

        elif req_div == "0":  # 승인
            data = StdWord.objects.get(id=req_id)
            data.accept_yn = "0"
            data.reject_exp = "승인되었습니다" if comment == "" else comment

            data.save()

            ds_emp = _query_dict('iam', iamSql.format(emp_no=data.id_reg_user.username))[0]
            msg = """[SmartDBA] 단어 {nm} 가 승인 되었습니다""".format(nm=data.std_wd_kor)
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

            error_msg = "정상적으로 승인되었습니다."


        elif req_div == "2":  # 반려
            data = StdWord.objects.get(id=req_id)
            data.accept_yn = "2"
            data.reject_exp = comment
            data.save()

            ds_emp = _query_dict('iam', iamSql.format(emp_no=data.id_reg_user.username))[0]
            msg = """[SmartDBA] 단어 {nm} 가 반려 되었습니다""".format(nm=data.std_wd_kor)
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])


            error_msg = "정상적으로 반려되었습니다."

    elif std_div == "1":  # 도메인
        if req_div == "1":  # 취소
            data = StdDomain.objects.get(id=req_id).delete()
            error_msg = "정상적으로 취소되었습니다."

        elif req_div == "0":  # 승인
            data = StdDomain.objects.get(id=req_id)

            id_reg_user = data.id_reg_user
            id_stddomain = data
            domain_name = data.domain_name

            data.accept_yn = "0"
            data.reject_exp = "승인되었습니다" if comment == "" else comment
            data.save()

            try :
                last_number = re.search("([a-zA-Z_가-힣])(\d+$)", domain_name).group(2)
                clear_number_of_domain_name = domain_name.replace(last_number,'')
            except :
                clear_number_of_domain_name = domain_name

            if domain_name.replace("]","").replace("[","")[-2:] in ('코드', '번호'):

                print(domain_name)
                l_lkor, \
                l_leng, \
                l_domain, \
                l_eng_attr_name, \
                l_domain_flag, \
                l_word_find_flag, \
                l_last_num_flag, \
                lcnt, \
                total_attr, \
                max_length, \
                l_attr_find, \
                l_aleary_req, \
                list_attr = f_div_attr(domain_name)

                ok_div = l_word_find_flag[0]

                attr_div_eng = '_'.join(l_leng[0])

                domain_name = domain_name.replace("]","").replace("[","")

                if ok_div:
                    attr_data = StdAttr(
                        id_stddomain=id_stddomain,
                        std_attr_kor=domain_name,
                        std_attr_eng=attr_div_eng,
                        reg_dtm=datetime.datetime.now(),
                        id_reg_user=id_reg_user,
                        accept_yn=0,
                        use_yn=1
                    )
                    attr_data.save()

                    stddomin = StdDomain.objects.get(id=id_stddomain.id)

                    stddomin.domain_name = domain_name

                    stddomin.save()

            ds_emp = _query_dict('iam', iamSql.format(emp_no=data.id_reg_user.username))[0]
            msg = """[SmartDBA] 도메인 {nm} 가 승인 되었습니다""".format(nm=data.domain_name)
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

            error_msg = "정상적으로 승인되었습니다."

        elif req_div == "2":  # 반려
            data = StdDomain.objects.get(id=req_id)
            data.accept_yn = "2"
            data.reject_exp = comment
            data.save()

            ds_emp = _query_dict('iam', iamSql.format(emp_no=data.id_reg_user.username))[0]
            msg = """[SmartDBA] 도메인 {nm} 가 반려 되었습니다""".format(nm=data.domain_name)
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

            error_msg = "정상적으로 반려되었습니다."

    elif std_div == "2":  # 용어
        if req_div == "1":  # 취소
            data = StdAttr.objects.get(id=req_id).delete()
            error_msg = "정상적으로 취소되었습니다."

        elif req_div == "0":  # 승인
            data = StdAttr.objects.get(id=req_id)
            data.accept_yn = "0"
            data.reject_exp = "승인되었습니다" if comment == "" else comment
            data.save()

            ds_emp = _query_dict('iam', iamSql.format(emp_no=data.id_reg_user.username))[0]
            msg = """[SmartDBA] 용어 {nm} 가 승인 되었습니다""".format(nm=data.std_attr_kor)
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

            error_msg = "정상적으로 승인되었습니다."

        elif req_div == "2":  # 반려
            data = StdAttr.objects.get(id=req_id)
            data.accept_yn = "2"
            data.reject_exp = comment
            data.save()

            error_msg = "정상적으로 반려되었습니다."

            ds_emp = _query_dict('iam', iamSql.format(emp_no=data.id_reg_user.username))[0]
            msg = """[SmartDBA] 용어 {nm} 가 반려 되었습니다""".format(nm=data.std_attr_kor)
            _query_sms(ds_emp['MOBILE'], msg, team_name=ds_emp['NAME'], email=ds_emp['SUB_MAIL'])

    context = {
        'result': 'OK',
        'error_yn': error_yn,
        'error_msg': error_msg
    }

    param = json.dumps(context, default=json_default)

    return HttpResponse(param, content_type="application/json")


#######################################################################################
## 튜닝
#######################################################################################


class TuningListLV(LoginRequiredMixin, FormMixin, ListView):
    model = TuningList
    template_name = 'TuningListLV.html'
    context_object_name = 'objects'
    form_class = TuningListLVForm
    paginate_by = 10
   

    def get_initial(self):

        ###############################################
        ## 화면에 바인딩 된 값을 유지하기 위한 처리
        ###################################################
        title = ""

        keyword = self.request.session['keyword'] if 'keyword' in self.request.session else ""
        id_dblist = self.request.session['id_dblist'] if 'id_dblist' in self.request.session else ""
        id_domain = self.request.session['id_domain'] if 'id_domain' in self.request.session else ""
        id_projectlist = self.request.session['id_projectlist'] if 'id_projectlist' in self.request.session else ""
        id_tuningstatus = self.request.session['id_tuningstatus'] if 'id_tuningstatus' in self.request.session else ""
        choice_status = self.request.session['choice_status'] if 'choice_status' in self.request.session else ""

        return {
            'keyword': keyword,
            'id_dblist': id_dblist,
            'id_domain': id_domain,
            'id_projectlist': id_projectlist,
            'id_tuningstatus': id_tuningstatus,
            'choice_status': choice_status,
        }

    def get_queryset(self):

        self.form = self.get_form(self.form_class)

        keyword = self.request.POST.get('keyword', '')

        ###################################################
        ## SEARCH 버튼 클리 시
        ###################################################
        if self.form.is_valid():

            id_dblist = self.form.cleaned_data['id_dblist']
            id_domain = self.form.cleaned_data['id_domain']
            id_projectlist = self.form.cleaned_data['id_projectlist']
            id_tuningstatus = self.form.cleaned_data['id_tuningstatus']
            choice_status = self.form.cleaned_data['choice_status']

            ###################################################
            ## 화면에 바인딩 된 값을 유지하기 위한 처리
            ###################################################
            self.request.session['keyword'] = keyword if keyword is not None else ""
            self.request.session['id_dblist'] = id_dblist.id if id_dblist is not None else ""
            self.request.session['id_domain'] = id_domain.id if id_domain is not None else ""
            self.request.session['id_projectlist'] = id_projectlist.id if id_projectlist is not None else ""
            self.request.session['id_tuningstatus'] = id_tuningstatus.id if id_tuningstatus is not None else ""
            self.request.session['choice_status'] = choice_status if choice_status is not None else ""

            ###################################################
            ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            self.request.session['click_search'] = 'Y'
            ###################################################

        ###################################################
        ## 1. 처음 진입
        ## 2. 페이징 버튼 클릭 시
        ###################################################
        else:

            page_kwarg = self.page_kwarg
            page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
            if page == 1:
                self.request.session['keyword'] = ""
                self.request.session['id_dblist'] = ""
                self.request.session['id_domain'] = ""
                self.request.session['id_projectlist'] = ""
                self.request.session['id_tuningstatus'] = ""
                self.request.session['choice_status'] = ""

            keyword = self.request.session['keyword'] if 'keyword' in self.request.session else ""
            id_dblist = self.request.session['id_dblist'] if 'id_dblist' in self.request.session else ""
            id_domain = self.request.session['id_domain'] if 'id_domain' in self.request.session else ""
            id_projectlist = self.request.session['id_projectlist'] if 'id_projectlist' in self.request.session else ""
            id_tuningstatus = self.request.session[
                'id_tuningstatus'] if 'id_tuningstatus' in self.request.session else ""
            choice_status = self.request.session['choice_status'] if 'choice_status' in self.request.session else ""

        _loggingVisit(self.request, 13)




        if choice_status == '': # 그 외 검색
            if keyword.isdigit():
                obj = TuningList.objects.exclude(id=999999999).filter(
                    (
                        Q(id=int(keyword))
                    ),
                    *((Q(id_dblist=id_dblist),) if id_dblist else ()),
                    *((Q(id_domain=id_domain),) if id_domain else ()),
                    *((Q(id_projectlist=id_projectlist),) if id_projectlist else ()),
                    *((Q(id_tuningstatus=id_tuningstatus),) if id_tuningstatus else ()),
                ).order_by("-id")
            else:
                obj = TuningList.objects.exclude(id=999999999).filter(
                    (Q(title__icontains=keyword) |
                     Q(id_reg_user__first_name=keyword)
                     ),
                    *((Q(id_dblist=id_dblist),) if id_dblist else ()),
                    *((Q(id_domain=id_domain),) if id_domain else ()),
                    *((Q(id_projectlist=id_projectlist),) if id_projectlist else ()),
                    *((Q(id_tuningstatus=id_tuningstatus),) if id_tuningstatus else ()),
                ).order_by("-id")
        elif choice_status == '1':  # SQL ID 검색
            obj = TuningList.objects.exclude(id=999999999).filter(
                (Q(sql_id__icontains=keyword)),
                *((Q(id_dblist=id_dblist),) if id_dblist else ()),
                *((Q(id_domain=id_domain),) if id_domain else ()),
                *((Q(id_projectlist=id_projectlist),) if id_projectlist else ()),
                *((Q(id_tuningstatus=id_tuningstatus),) if id_tuningstatus else ()),
            ).order_by("-id")
        elif choice_status == '2':  # SQL TEXT 검색
            obj = TuningList.objects.exclude(id=999999999).filter(
                (Q(asis_sql_text__icontains=keyword) |
                 Q(tobe_sql_text__icontains=keyword)
                 ),
                *((Q(id_dblist=id_dblist),) if id_dblist else ()),
                *((Q(id_domain=id_domain),) if id_domain else ()),
                *((Q(id_projectlist=id_projectlist),) if id_projectlist else ()),
                *((Q(id_tuningstatus=id_tuningstatus),) if id_tuningstatus else ()),
            ).order_by("-id")

        return obj

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    ###################################################
    ## PAGING 처리
    ###################################################
    def paginate_queryset(self, queryset, page_size):

        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty())
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = int(page)
        except ValueError:
            if page == 'last':
                page_number = paginator.num_pages
            else:
                raise Http404(_("Page is not 'last', nor can it be converted to an int."))
        try:

            ###################################################
            ## SEARCH 버튼 클릭 시 페이징이 1번으로 가기 위한 코드
            ###################################################
            if 'click_search' in self.request.session:
                if self.request.session['click_search'] == "N":
                    page = paginator.page(page_number)
                else:
                    page = paginator.page(1)
            else:
                page = paginator.page(page_number)
            ###################################################
            
            self.request.session['page_number'] = page_number

            return (paginator, page, page.object_list, page.has_other_pages())

        except:
            page = paginator.page(1)
            return (paginator, page, page.object_list, page.has_other_pages())

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)

        page_range = ""
        paginator = context['paginator']
        page_numbers_range = 5  # Display only 5 page numbers
        max_index = len(paginator.page_range)

        ###################################################
        ## SEARCH 버튼 클리 시 페이징이 1번으로 가기 위한 코드
        ###################################################
        if 'click_search' in self.request.session:
            if self.request.session['click_search'] == "N":
                page = self.request.GET.get('page')
            else:
                page = 1
                self.request.session['click_search'] = "N"
        else:
            page = self.request.GET.get('page')
        ###################################################

        current_page = int(page) if page else 1

        start_index = int((current_page - 1) / page_numbers_range) * page_numbers_range
        end_index = start_index + page_numbers_range
        if end_index >= max_index:
            end_index = max_index

        page_range = paginator.page_range[start_index:end_index]
        context['page_range'] = page_range

        context['template_parent'] = "TuningListLV"
        context['template_child'] = "TuningListLV"

        return context



class TuningListCV(LoginRequiredMixin, CreateView):
    model = TuningList
    form_class = TuningListForm
    template_name = 'TuningListCV.html'



    def get_success_url(self):
        return reverse_lazy('TuningListDV', kwargs={'id': self.object.id})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({"request": self.request})
        return kwargs

    def get_context_data(self, **kwargs):


        context = super(CreateView, self).get_context_data(**kwargs)

        context['template_parent'] = "TuningListLV"
        context['template_child'] = "TuningListLV"
        context['dba_users'] = "dba_users"

        return context


class TuningListUV(LoginRequiredMixin, UpdateView):
    model = TuningList
    form_class = TuningListForm
    template_name = 'TuningListUV.html'

    context_object_name = 'objects'

    def get_success_url(self):
        return reverse_lazy('TuningListDV', kwargs={'id': self.object.id})

    def get_object(self):
        obj = get_object_or_404(TuningList, id=self.kwargs['id'])

        return obj

    def get_form_kwargs(self):
        # add request for form to validate
        kwargs = super().get_form_kwargs()
        kwargs.update({"request": self.request})
        return kwargs

    def get_context_data(self, **kwargs):


        context = super().get_context_data(**kwargs)

        context['template_parent'] = "TuningListLV"
        context['template_child'] = "TuningListLV"

        return context


class TuningListDeleteView(LoginRequiredMixin, DeleteView):
    model = TuningList
    success_message = 'Your Photo has been deleted successfully.'

    def get_success_url(self):
        return reverse_lazy('TuningListLV')


class TuningListDV(LoginRequiredMixin, FormMixin, DetailView):
    model = TuningList
    template_name = 'TuningListDV.html'
    context_object_name = 'objects'
    form_class = TuningListDVForm

    def get_object(self):
        tuninglist = get_object_or_404(TuningList, id=self.kwargs['id'])

        return tuninglist

    # form 에 request 를 넘겨준다
    def get_form_kwargs(self):
        # add request for form to validate
        kwargs = super().get_form_kwargs()
        kwargs.update({"request": self.request})
        kwargs.update({"obj_id": self.kwargs['id']})
        return kwargs

    def post(self, request, *args, **kwargs):
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):


        dba_member = User.objects.filter(is_superuser="1")
        dba_users = []

        for m in dba_member:
            dba_users.append(m.id)

        context = super().get_context_data(**kwargs)

        context['template_parent'] = "TuningListLV"
        context['template_child'] = "TuningListLV"
        context['dba_users'] = dba_users

        return context



